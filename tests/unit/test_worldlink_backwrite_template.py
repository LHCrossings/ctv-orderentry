"""
Guard the WorldLink backwrite template against re-bloat.

The template once shipped with a 16 MB data sheet + 8 MB pivot cache for two
sheets the transformer deletes at runtime ("Run Sheet", "Monthly Totals").
openpyxl re-parsed all of it on every generate → ~14s per export. The template
was slimmed to only the two sheets the output keeps. These checks fail if a
future Excel re-export reintroduces pivot caches or the unused heavy sheets.
"""

import zipfile
from pathlib import Path

_TEMPLATE = (
    Path(__file__).parent.parent.parent
    / "src" / "backwrite" / "templates" / "worldlink_template.xlsx"
)


def test_template_has_no_pivot_caches():
    """Pivot caches were the 14s load cost — there must be none."""
    names = zipfile.ZipFile(_TEMPLATE).namelist()
    pivots = [n for n in names if "pivot" in n.lower()]
    assert pivots == [], f"template re-bloated with pivot parts: {pivots}"


def test_template_is_slim():
    """Bloated template was 2.4 MB; slim is ~50 KB. Keep it well under 500 KB."""
    size = _TEMPLATE.stat().st_size
    assert size < 500_000, f"template is {size:,} bytes — likely re-bloated"


def test_template_keeps_only_the_filled_sheets():
    """Output keeps exactly the two sheets the transformer fills."""
    from openpyxl import load_workbook
    wb = load_workbook(_TEMPLATE, read_only=True)
    assert set(wb.sheetnames) == {"Sales Confirmation", "MONTHLY LINES AND BROKER FEES"}, \
        f"unexpected sheets: {wb.sheetnames}"
