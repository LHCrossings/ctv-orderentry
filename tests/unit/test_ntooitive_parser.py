"""
Ntooitive parser tests — driven by the real LA Care CRC 2026 REV 2 proposal
(committed as a fixture; it carries its own totals, so it is its own oracle).

The negative tests TAMPER with the document (a mutated workbook copy, or a
mutated pdfplumber word stream) rather than asserting guards exist — a dropped
cell, a blanked rate, a renamed column and a rate on a bonus row must each
refuse, and a no-op mutation must still parse (or the negative tests pass for
the wrong reason).
"""

import sys
from datetime import date
from pathlib import Path

import pytest

_root = Path(__file__).parent.parent.parent
for p in [str(_root), str(_root / "src")]:
    if p not in sys.path:
        sys.path.insert(0, p)

from browser_automation.parsers.ntooitive_parser import (  # noqa: E402
    list_ntooitive_options,
    parse_ntooitive,
    parse_ntooitive_pdf,
    parse_ntooitive_xlsx,
)

FIXTURE_DIR = _root / "tests" / "fixtures" / "ntooitive"
XLSX = str(FIXTURE_DIR / "ntooitive_lacare_crc2026_rev2.xlsx")
PDF = str(FIXTURE_DIR / "ntooitive_lacare_crc2026_rev2.pdf")


@pytest.fixture(scope="module")
def order():
    return parse_ntooitive(XLSX)


# ── Header / option selection ────────────────────────────────────────────────


def test_header_fields(order):
    assert order.source_format == "xlsx"
    assert order.agency == "Ntooitive"
    assert "L.A. Care" in order.advertiser
    assert order.market_code == "LAX"
    assert order.billing_cycle == "Broadcast"
    assert order.rates_are_net is False


def test_default_sheet_is_the_newest_option(order):
    # Option 1 is REVISED 8/14/2026; Option 2 is the stale 6/17 proposal.
    assert order.option_label == "Option 1"
    assert order.order_date == date(2026, 8, 14)


def test_option_listing():
    opts = list_ntooitive_options(XLSX)
    assert [o["sheet"] for o in opts] == ["Option 1", "Option 2"]
    assert opts[0]["date"] == date(2026, 8, 14)
    assert opts[1]["date"] == date(2026, 6, 17)
    assert opts[0]["gross"] == pytest.approx(12800.0)


def test_explicit_sheet_selection():
    o2 = parse_ntooitive_xlsx(XLSX, sheet_name="Option 2")
    assert o2.option_label == "Option 2"
    assert o2.flight_start == "06/29/2026"
    assert len(o2.week_dates) == 23
    assert o2.gross_airtime == pytest.approx(12820.0)
    assert o2.charges == []


def test_unknown_sheet_refuses():
    with pytest.raises(ValueError, match="sheet 'Option 9' not found"):
        parse_ntooitive_xlsx(XLSX, sheet_name="Option 9")


# ── Flight dates ─────────────────────────────────────────────────────────────


def test_flight_dates_resolve_the_yearless_header(order):
    # Header says "8/18 -11/30" with no year; the week grid opens Mon 8/17/26.
    assert order.flight_start == "08/18/2026"
    assert order.flight_end == "11/30/2026"
    assert order.week_dates[0] == date(2026, 8, 17)
    assert order.week_dates[-1] == date(2026, 11, 23)
    assert len(order.week_dates) == 15


# ── Lines ────────────────────────────────────────────────────────────────────


def test_line_shapes(order):
    assert len(order.lines) == 4
    paid, bonus = order.paid_lines, order.bonus_lines
    assert [ln.total_spots for ln in paid] == [60, 48]
    assert [ln.total_spots for ln in bonus] == [48, 48]
    assert [ln.rate for ln in paid] == [120.0, 100.0]
    assert all(ln.length_sec == 30 for ln in order.lines)


def test_dual_window_daypart_survives_verbatim(order):
    mand = order.paid_lines[0]
    assert mand.language_block == "Mandarin News"
    assert "6a-7a" in mand.daypart and "8p-9p" in mand.daypart


def test_bonus_language_comes_from_the_ros_daypart(order):
    # The Language Block cell says 'Mandarin' but the ROS runs in the whole
    # Chinese block ('Chinese ROS') — base_language must say Chinese.
    blocks = {ln.base_language for ln in order.bonus_lines}
    assert blocks == {"Chinese", "Korean"}


def test_money_reconciles(order):
    assert order.gross_airtime == pytest.approx(12000.0)
    assert order.gross_contract == pytest.approx(12800.0)
    assert order.implied_commission == pytest.approx(0.15)
    assert len(order.charges) == 1
    assert order.charges[0].amount == pytest.approx(800.0)


# ── PDF print reader — must reproduce the workbook exactly ───────────────────


@pytest.fixture(scope="module")
def pdf_order(real_pdfplumber):
    return parse_ntooitive_pdf(PDF)


def test_pdf_matches_xlsx(order, pdf_order):
    assert pdf_order.source_format == "pdf"
    assert pdf_order.market_code == order.market_code
    assert pdf_order.flight_start == order.flight_start
    assert pdf_order.flight_end == order.flight_end
    assert pdf_order.gross_airtime == order.gross_airtime
    assert pdf_order.gross_contract == order.gross_contract
    assert pdf_order.implied_commission == pytest.approx(order.implied_commission)

    def key(ln):
        return (
            ln.language_block,
            ln.spot_type,
            ln.rate,
            ln.length_sec,
            ln.total_spots,
            tuple(ln.week_spots),
        )

    assert [key(ln) for ln in pdf_order.lines] == [key(ln) for ln in order.lines]
    assert [(c.description, c.amount) for c in pdf_order.charges] == [
        (c.description, c.amount) for c in order.charges
    ]


def test_dispatcher_routes_both_formats(real_pdfplumber):
    assert parse_ntooitive(XLSX).source_format == "xlsx"
    assert parse_ntooitive(PDF).source_format == "pdf"
    with pytest.raises(ValueError, match="unsupported file type"):
        parse_ntooitive("order.docx")


# ── Tampered-workbook guards ─────────────────────────────────────────────────
#
# Fixture coordinates on 'Option 1' (frozen with the committed file):
#   row 18 = column headers;  V18 = 'Promo Unit Cost (Gross)'
#   row 19 = Mandarin News:   F19 = first week cell (5), V19 = rate (120)
#   row 21 = BONUS Mandarin:  V21 = rate ('-')
#   E13 = Gross (Airtime), E14 = Gross Translation Fees, E15 = Gross Amount


def _tampered(tmp_path, mutate) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    mutate(wb)
    out = str(tmp_path / "tampered.xlsx")
    wb.save(out)
    return out


def test_noop_mutation_still_parses(tmp_path):
    def mutate(wb):
        wb["Option 1"]["K6"] = "Somebody Else"

    o = parse_ntooitive_xlsx(_tampered(tmp_path, mutate))
    assert o.gross_airtime == pytest.approx(12000.0)
    assert len(o.lines) == 4


def test_blanked_paid_rate_refuses(tmp_path):
    def mutate(wb):
        ws = wb["Option 1"]
        ws["V19"] = None
        ws["W19"] = None  # keep rate×units from tripping first

    with pytest.raises(ValueError, match="no readable rate"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_dropped_week_cell_refuses(tmp_path):
    # A vanished cell shrinks the line's spot sum; the rate × units == line
    # total reconciliation is the first guard to see it.
    def mutate(wb):
        wb["Option 1"]["F19"] = None

    with pytest.raises(ValueError, match="sheet says"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_wrong_total_unit_cell_refuses(tmp_path):
    # Total Unit # disagreeing with the week cells while all money still foots.
    def mutate(wb):
        wb["Option 1"]["U19"] = 61

    with pytest.raises(ValueError, match="Total Unit #"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_rate_column_flipping_to_net_refuses(tmp_path):
    def mutate(wb):
        wb["Option 1"]["V18"] = "Promo Unit Cost (NET)"

    with pytest.raises(ValueError, match="not gross"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_renamed_rate_column_refuses(tmp_path):
    def mutate(wb):
        wb["Option 1"]["V18"] = "Unit Price"

    with pytest.raises(ValueError, match="missing 'rate' column"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_rate_on_a_bonus_row_refuses(tmp_path):
    def mutate(wb):
        wb["Option 1"]["V21"] = 50

    with pytest.raises(ValueError, match="BONUS line"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_wrong_line_total_refuses(tmp_path):
    def mutate(wb):
        wb["Option 1"]["W19"] = 7100  # rate 120 × 60 = 7200

    with pytest.raises(ValueError, match="sheet says"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_wrong_header_airtime_refuses(tmp_path):
    def mutate(wb):
        ws = wb["Option 1"]
        ws["E13"] = 11000
        ws["E15"] = 11800  # keep airtime+translation==contract from masking it

    with pytest.raises(ValueError, match="Gross \\(Airtime\\)"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_unclassified_charge_refuses(tmp_path):
    # Remove the translation-fee row but leave the contract total carrying the
    # extra $800 — money that is neither airtime nor a recognised charge.
    def mutate(wb):
        ws = wb["Option 1"]
        ws["C14"] = None
        ws["E14"] = None

    with pytest.raises(ValueError, match="non-airtime money"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


def test_dropped_line_trips_the_footer(tmp_path):
    # Blank the Korean News row entirely: its spots vanish but the footer
    # 'Total Paid' still says 108 — the parse must refuse.
    def mutate(wb):
        ws = wb["Option 1"]
        for col in "ABCDEFGHIJKLMNOPQRSTUVWX":
            ws[f"{col}20"] = None

    with pytest.raises(ValueError, match="Total Paid"):
        parse_ntooitive_xlsx(_tampered(tmp_path, mutate))


# ── Tampered PDF word stream (the layer that fails in the wild) ─────────────


def _mutated_pdfplumber(real, mutate):
    class _Page:
        def __init__(self, p):
            self._p = p

        def extract_words(self, *a, **k):
            return mutate([dict(w) for w in self._p.extract_words(*a, **k)])

        def extract_text(self, *a, **k):
            return self._p.extract_text(*a, **k)

    class _Pdf:
        def __init__(self, p):
            self._p = p
            self.pages = [_Page(x) for x in p.pages]

        def __enter__(self):
            return self

        def __exit__(self, *a):
            self._p.__exit__(*a)

    class _Mod:
        @staticmethod
        def open(path):
            return _Pdf(real.open(path))

    return _Mod


def test_pdf_dropped_cell_refuses(real_pdfplumber, monkeypatch):
    dropped = {"n": 0}

    def mutate(words):
        out = []
        for w in words:
            # Drop the first Mandarin-row week cell ('5' at ~x226, top ~207.7)
            if (
                not dropped["n"]
                and w["text"] == "5"
                and 200 < w["top"] < 215
                and 215 < w["x0"] < 240
            ):
                dropped["n"] = 1
                continue
            out.append(w)
        return out

    monkeypatch.setitem(sys.modules, "pdfplumber", _mutated_pdfplumber(real_pdfplumber, mutate))
    with pytest.raises(ValueError, match="Total Unit #"):
        parse_ntooitive_pdf(PDF)
    assert dropped["n"] == 1


def test_pdf_noop_mutation_still_parses(real_pdfplumber, monkeypatch):
    monkeypatch.setitem(
        sys.modules, "pdfplumber", _mutated_pdfplumber(real_pdfplumber, lambda w: w)
    )
    o = parse_ntooitive_pdf(PDF)
    assert o.gross_airtime == pytest.approx(12000.0)
    assert len(o.lines) == 4


# ── Automation helpers ───────────────────────────────────────────────────────


def test_split_daypart_union():
    from browser_automation.ntooitive_automation import split_daypart_union

    assert split_daypart_union("M-F 6a-7a & 8p-9p") == ("M-F", "6a-7a; 8p-9p")
    assert split_daypart_union("M- F 8a-9a") == ("M-F", "8a-9a")
    assert split_daypart_union("M-Su 11a-1p") == ("M-Su", "11a-1p")
    assert split_daypart_union("Chinese ROS") == ("Chinese ROS", "")


def test_line_plan_delivers_every_spot_on_the_io_start(order):
    # The proposal's own start (Tue 8/18) already truncates the first week:
    # Mandarin's 5 spots over 4 remaining weekdays need 2/day — its own line.
    from browser_automation.ntooitive_automation import _line_plan

    plan = _line_plan(order, date(2026, 8, 18), order.flight_end)
    entered = sum(
        r["spots_per_week"] * r["weeks"] for _ln, _d, _t, _desc, ranges, _n in plan for r in ranges
    )
    assert entered == sum(ln.total_spots for ln in order.lines) == 204

    mand_ranges = plan[0][4]
    assert mand_ranges[0]["date_from"] == date(2026, 8, 18)
    assert mand_ranges[0]["weeks"] == 1
    assert mand_ranges[0]["max_daily"] == 2  # 5 spots / 4 days left
    assert mand_ranges[1]["max_daily"] == 1  # full weeks


def test_line_plan_late_start_drops_and_notes(order):
    from browser_automation.ntooitive_automation import _line_plan

    plan = _line_plan(order, date(2026, 8, 26), order.flight_end)
    entered = sum(
        r["spots_per_week"] * r["weeks"] for _ln, _d, _t, _desc, ranges, _n in plan for r in ranges
    )
    notes = [n for _ln, _d, _t, _desc, _r, ns in plan for n in ns]
    assert entered == 187  # 204 − 17 undeliverable
    assert len(notes) == 4  # one dropped week per line
    assert all("precede the start date" in n for n in notes)
