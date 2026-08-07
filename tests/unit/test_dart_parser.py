"""DART xlsx parsing — column mapping, week dates, and the reconciliation guard.

The Aug/Sept 2026 order entered as a $3,000 buy with **$0 on every line** and a
flight of 8/17-9/6 instead of 8/17-9/27. Two positional assumptions:

* the rate was read from a hardcoded col D, but DART had inserted a "Length"
  column, so D held ":15s"; `Decimal(":15s")` raised and a bare `except` turned
  it into `Decimal("0")`;
* only the FIRST week date was read and the rest synthesized as `first + 7i`,
  silently rewriting the sheet's real 8/17, 9/14, 9/21 columns to 8/17, 8/24, 8/31.

Both were silent. Every test here is built from the real sheet's layout.
"""

import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

_root = Path(__file__).parent.parent.parent
for _p in (_root, _root / "browser_automation"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from browser_automation.parsers.dart_parser import parse_dart_xlsx  # noqa: E402

# Real header layout: Length sits between Schedule and Rate, and the week columns
# are NOT consecutive (8/17 then a jump to 9/14).
_HEADER = ["Programming", "Schedule", "Length", "Rate  ",
           date(2026, 8, 17), date(2026, 9, 14), date(2026, 9, 21),
           "Total Units", "Value", "Total Cost"]
_ROWS = [
    ["Cantonese News/ Talk", "M-F 5p-6p",      ":15s", 34, 8, 8, 7, 23, 782, 782],
    ["Mandarin News",        "M-Sun 6p-7p",    ":15s", 34, 10, 10, 9, 29, 986, 986],
    ["Mandarin Drama",       "M-F 7p-8p",      ":15s", 34, 5, 5, 4, 14, 476, 476],
    ["Vietnamese Drama",     "M-Sun  10a-11a", ":15s", 21, 12, 12, 12, 36, 756, 756],
    ["Chinese",              "ROS  Bonus schedule", ":30s", 68, 6, 6, 6, 18, 1224, 0],
    ["Vietnamese",           "ROS  Bonus schedule", ":30s", 42, 5, 5, 5, 15, 630, 0],
]
_PAID_SUMMARY = ["PAID ", None, None, None, 35, 35, 32, 135, 4854, 3000]


def _write_sheet(tmp_path, header=None, rows=None, paid_summary=None,
                 header_at_row=14) -> str:
    """Build a DART-shaped workbook. Grid starts in col B, as the real sheet does."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DART"
    ws["D2"] = "Dallas Area Rapid Transit "
    ws["D4"] = "The Asian Channel  KLEG44.3 & KFWD 52.4"
    ws["D5"] = "Carmen D. Hillebrand"
    ws["B13"] = "FLIGHT SCHEDULE (:15 seconds/:30s) -  3 weeks flight "

    def put(row_idx, values):
        for offset, value in enumerate(values):
            if value is not None:
                ws.cell(row=row_idx, column=2 + offset, value=value)

    put(header_at_row, header if header is not None else _HEADER)
    for i, row in enumerate(rows if rows is not None else _ROWS):
        put(header_at_row + 1 + i, row)
    summary = _PAID_SUMMARY if paid_summary is None else paid_summary
    if summary:
        put(header_at_row + 1 + len(rows if rows is not None else _ROWS), summary)

    path = tmp_path / "dart.xlsx"
    wb.save(path)
    return str(path)


def test_rate_comes_from_the_labeled_column_not_a_fixed_index(tmp_path):
    """The bug: col D is Length (':15s'), the rate lives in col E."""
    order = parse_dart_xlsx(_write_sheet(tmp_path))
    assert [ln.rate for ln in order.paid_lines] == [
        Decimal("34"), Decimal("34"), Decimal("34"), Decimal("21")
    ]
    assert order.total_cost == Decimal("3000")


def test_every_week_date_is_read_including_a_gap(tmp_path):
    """8/17 -> 9/14 is a 4-week jump; synthesizing first+7i gave 8/24 and 8/31."""
    order = parse_dart_xlsx(_write_sheet(tmp_path))
    assert order.week_start_dates == [
        date(2026, 8, 17), date(2026, 9, 14), date(2026, 9, 21)
    ]
    assert order.flight_start == date(2026, 8, 17)
    assert order.flight_end == date(2026, 9, 27)


def test_spot_length_is_per_line(tmp_path):
    """DART mixes lengths on one order: paid :15s, bonus :30s."""
    order = parse_dart_xlsx(_write_sheet(tmp_path))
    assert [ln.spot_length for ln in order.paid_lines] == [15, 15, 15, 15]
    assert [ln.spot_length for ln in order.bonus_lines] == [30, 30]


def test_bonus_rate_is_zeroed_regardless_of_the_sheet(tmp_path):
    order = parse_dart_xlsx(_write_sheet(tmp_path))
    assert all(ln.rate == 0 for ln in order.bonus_lines)


def test_older_layout_without_a_length_column_still_parses(tmp_path):
    """DART 2604 (April 2026) had Rate in col D and no Length column, and entered
    correctly at the same $34/$21 rates. Label-based mapping must keep that working,
    with spot_length falling back to the order-level duration."""
    header = ["Programming", "Schedule", "Rate",
              date(2026, 8, 17), date(2026, 8, 24), date(2026, 8, 31),
              "Total Units", "Value", "Total Cost"]
    rows = [
        ["Cantonese News/ Talk", "M-F 5p-6p",           34, 8, 8, 7, 23, 782, 782],
        ["Chinese", "ROS  Bonus schedule",              68, 6, 6, 6, 18, 1224, 0],
    ]
    summary = ["PAID ", None, None, 14, 14, 13, 41, 2006, 782]

    order = parse_dart_xlsx(
        _write_sheet(tmp_path, header=header, rows=rows, paid_summary=summary))

    assert order.paid_lines[0].rate == Decimal("34")
    assert order.total_cost == Decimal("782")
    assert order.week_start_dates == [
        date(2026, 8, 17), date(2026, 8, 24), date(2026, 8, 31)
    ]
    # No Length column -> None, so the automation uses order.duration_seconds.
    assert all(ln.spot_length is None for ln in order.lines)
    assert order.duration_seconds == 15


def test_header_is_found_even_when_it_moves(tmp_path):
    """A banner or extra intro row shifts the grid — the header is located, not assumed."""
    order = parse_dart_xlsx(_write_sheet(tmp_path, header_at_row=17))
    assert order.total_cost == Decimal("3000")
    assert len(order.lines) == 6


def test_shifted_rate_column_now_raises_instead_of_entering_zero(tmp_path):
    """Reproduce the original failure mode: rate read as ':15s' -> 0.

    Simulated by relabeling Rate so it can't be found — the parse must refuse
    rather than fall back to a positional guess.
    """
    header = list(_HEADER)
    header[3] = "Unit Cost"          # a rename the map doesn't know
    with pytest.raises(ValueError, match="missing required column"):
        parse_dart_xlsx(_write_sheet(tmp_path, header=header))


def test_a_non_numeric_paid_rate_raises(tmp_path):
    rows = [list(r) for r in _ROWS]
    rows[0][3] = ":15s"              # rate cell holding a length
    with pytest.raises(ValueError, match="not a number"):
        parse_dart_xlsx(_write_sheet(tmp_path, rows=rows))


def test_line_total_cost_mismatch_raises(tmp_path):
    rows = [list(r) for r in _ROWS]
    rows[0][9] = 999                 # Total Cost no longer equals rate x units
    with pytest.raises(ValueError, match="Total Cost"):
        parse_dart_xlsx(_write_sheet(tmp_path, rows=rows))


def test_total_units_mismatch_raises(tmp_path):
    rows = [list(r) for r in _ROWS]
    rows[1][7] = 99                  # Total Units no longer equals the week columns
    with pytest.raises(ValueError, match="Total Units"):
        parse_dart_xlsx(_write_sheet(tmp_path, rows=rows))


def test_paid_summary_mismatch_raises(tmp_path):
    """The whole-order guard: a dropped or misread line must refuse to enter."""
    summary = list(_PAID_SUMMARY)
    summary[9] = 4000
    with pytest.raises(ValueError, match="PAID Total Cost"):
        parse_dart_xlsx(_write_sheet(tmp_path, paid_summary=summary))


def test_missing_paid_summary_still_parses(tmp_path):
    """Reconciliation is best-effort per target: no PAID row, no whole-order check."""
    order = parse_dart_xlsx(_write_sheet(tmp_path, paid_summary=[]))
    assert order.total_cost == Decimal("3000")


def test_spot_counts_and_totals_match_the_sheet(tmp_path):
    order = parse_dart_xlsx(_write_sheet(tmp_path))
    assert [ln.spot_counts for ln in order.paid_lines] == [
        [8, 8, 7], [10, 10, 9], [5, 5, 4], [12, 12, 12]
    ]
    assert sum(ln.total_spots for ln in order.paid_lines) == 102
    assert sum(ln.total_spots for ln in order.bonus_lines) == 33
