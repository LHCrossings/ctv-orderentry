"""
Wallrich xlsx parser tests — driven by the real SMUD 3Q26 Est 769 Strata
Spot Schedule export (committed as a fixture; it carries its own totals, so
it is its own oracle).

The negative tests TAMPER with a mutated workbook copy rather than asserting
guards exist — a blanked rate, a slid week cell, a renamed column, a wrong
header total and a trade line must each refuse, and a no-op mutation must
still parse (or the negative tests pass for the wrong reason).
"""

import sys
from pathlib import Path

import pytest

_root = Path(__file__).parent.parent.parent
for p in [str(_root), str(_root / "src")]:
    if p not in sys.path:
        sys.path.insert(0, p)

from browser_automation.parsers import wallrich_parser  # noqa: E402
from browser_automation.parsers.wallrich_parser import (  # noqa: E402
    consolidate_wallrich_weeks,
    parse_wallrich,
    parse_wallrich_xlsx,
)

FIXTURE = str(_root / "tests" / "fixtures" / "wallrich" / "smud_3q26_est769.xlsx")
REAL_FILENAME = "SMUD 3Q26 POM Community Impact Spot Schedule (Crossings TV) Est 769.xlsx"


@pytest.fixture(scope="module")
def est():
    estimates = parse_wallrich_xlsx(FIXTURE)
    assert len(estimates) == 1
    return estimates[0]


# ── Header ───────────────────────────────────────────────────────────────────


def test_header_fields(est):
    assert est.estimate_number == "769"
    assert est.client == "SD15"
    assert est.market == "Sacramento"
    assert est.flight_start == "8/24/2026"
    assert est.flight_end == "11/22/2026"
    assert est.separation == 30
    assert est.description == "Clean Energy Vision 2030"
    assert est.rates_are_net is False
    assert est.pdf_path == FIXTURE


def test_week_starts(est):
    assert len(est.week_starts) == 13
    assert est.week_starts[0] == "8/24"
    assert est.week_starts[-1] == "11/16"


# ── Lines ────────────────────────────────────────────────────────────────────


def test_lines(est):
    assert len(est.lines) == 6
    ln1 = est.lines[0]
    assert ln1.days == "MTuWThF"
    assert ln1.time == "7:00p-8:00p"  # leading zeros stripped from hours
    assert ln1.program == "Cantonese"
    assert ln1.duration == 30
    assert ln1.rate == 50.0
    assert ln1.total_spots == 27
    assert sum(ln1.weekly_spots) == 27
    assert ln1.is_bonus is False
    hmong = est.lines[4]
    assert hmong.days == "SaSu"
    assert hmong.program == "Hmong"


def test_order_totals_reconcile(est):
    assert sum(ln.total_spots for ln in est.lines) == 270
    assert sum(ln.rate * ln.total_spots for ln in est.lines) == 13500.0


def test_consolidation_splits_on_dark_weeks(est):
    # Line 1 weekly pattern: 0,3,3,3,0,3,3,3,3,0,0,3,3 → three ranges
    ranges = consolidate_wallrich_weeks(
        est.lines[0].weekly_spots, est.week_starts, est.flight_end, 2026
    )
    assert [
        (r["start_date"], r["end_date"], r["spots_per_week"], r["total_spots"]) for r in ranges
    ] == [
        ("08/31/2026", "09/20/2026", 3, 9),
        ("09/28/2026", "10/25/2026", 3, 12),
        ("11/09/2026", "11/22/2026", 3, 6),
    ]
    assert sum(r["total_spots"] for r in ranges) == est.lines[0].total_spots


# ── Dispatcher ───────────────────────────────────────────────────────────────


def test_dispatcher_routes_xlsx_to_xlsx_reader():
    estimates = parse_wallrich(FIXTURE)
    assert estimates and estimates[0].estimate_number == "769"


def test_dispatcher_routes_pdf_to_pdf_reader(monkeypatch):
    seen = {}
    monkeypatch.setattr(
        wallrich_parser,
        "parse_wallrich_pdf",
        lambda p: seen.setdefault("path", p) and [] or [],
    )
    parse_wallrich("/some/order.pdf")
    assert seen["path"] == "/some/order.pdf"


# ── Detection ────────────────────────────────────────────────────────────────


def test_scanner_content_detection_routes_to_wallrich():
    from domain.enums import OrderType
    from orchestration.order_scanner import _detect_xlsx_content

    assert _detect_xlsx_content(Path(FIXTURE)) == OrderType.WALLRICH


def test_station_and_layout_are_not_the_definer(tmp_path):
    # Lee (2026-08-21): "anyone can use Strata layouts" — a Strata-looking
    # KBTV workbook that is not for SMUD/SD15 must NOT misroute to Wallrich.
    # The client is the definer.
    import openpyxl

    from domain.enums import OrderType
    from orchestration.order_scanner import _detect_xlsx_content

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Client:"
    ws["B1"] = "SomeOtherClient"
    ws["A2"] = "Estimate:"
    ws["B2"] = "123-T26_Other Campaign"
    ws["A3"] = "Station:"
    ws["B3"] = "KBTV (CROSSINGS)-TV-Cash"
    out = tmp_path / "kbtv_other_client.xlsx"
    wb.save(str(out))
    assert _detect_xlsx_content(out) == OrderType.UNKNOWN


def test_smud_filename_detects_wallrich():
    # SMUD in the filename is enough — the workbook never needs opening.
    from business_logic.services.order_detection_service import detect_from_filename
    from domain.enums import OrderType

    assert detect_from_filename(REAL_FILENAME) == OrderType.WALLRICH


def test_sd15_client_cell_detects_wallrich(tmp_path):
    # A future export whose filename/estimate drops "SMUD" still detects via
    # the exact client code SD15 (exact cell match — "SD15" inside a longer
    # string does not count, SMUD covers those).
    import openpyxl

    from domain.enums import OrderType
    from orchestration.order_scanner import _detect_xlsx_content

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Client:"
    ws["B1"] = "SD15"
    out = tmp_path / "sd15_client.xlsx"
    wb.save(str(out))
    assert _detect_xlsx_content(out) == OrderType.WALLRICH


# ── Tamper tests ─────────────────────────────────────────────────────────────


def _tampered(tmp_path, mutate):
    """Copy the fixture, apply mutate(ws), return the mutated file's path.

    Loads data_only so formula cells (the Ln column is `=FIXED(n,0,TRUE)`)
    are materialized as values — a plain load-and-save drops the cached
    values and the copy parses as an empty grid.
    """
    import openpyxl

    wb = openpyxl.load_workbook(FIXTURE, data_only=True)
    mutate(wb.active)
    out = tmp_path / "tampered.xlsx"
    wb.save(str(out))
    wb.close()
    return str(out)


def _find(ws, text):
    """(row, col) of the first cell whose value equals text."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell.row, cell.column
    raise AssertionError(f"cell {text!r} not found in fixture")


def test_noop_mutation_still_parses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "Ln")
        ws.cell(row=r, column=c, value="Ln")  # rewrite same value

    estimates = parse_wallrich_xlsx(_tampered(tmp_path, mutate))
    assert estimates[0].estimate_number == "769"


def test_blanked_rate_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "Rate")
        # ws.cell(..., value=None) is a NO-OP in openpyxl — assign directly
        ws.cell(row=r + 1, column=c).value = None

    with pytest.raises(ValueError, match="no rate"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))


def test_changed_week_cell_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "[8/31]")
        ws.cell(row=r + 1, column=c, value=4)  # was 3

    with pytest.raises(ValueError, match="week cells sum"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))


def test_renamed_rate_column_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "Rate")
        ws.cell(row=r, column=c, value="Cost")

    with pytest.raises(ValueError, match="missing grid column"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))


def test_wrong_header_total_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "Total Spots:")
        ws.cell(row=r, column=c + 1, value=271)

    with pytest.raises(ValueError, match="header says 271"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))


def test_trade_line_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "C/T")
        ws.cell(row=r + 1, column=c, value="t")

    with pytest.raises(ValueError, match="only cash supported"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))


def test_trade_spots_header_refuses(tmp_path):
    def mutate(ws):
        r, c = _find(ws, "Trade Spots:")
        ws.cell(row=r, column=c + 1, value=5)

    with pytest.raises(ValueError, match="Trade Spots"):
        parse_wallrich_xlsx(_tampered(tmp_path, mutate))
