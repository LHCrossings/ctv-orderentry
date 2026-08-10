"""
Crispin LLC order parser (advertiser: Bay Area AQMD).

Crispin is an AGENCY parser: the agency is fixed (Crispin LLC → ANAGRAF agency
446) and the advertiser is looked up in ANAGRAF. Bay Area AQMD exists TWICE in
Etere (183 = Allison & Partners / AGENZIA 187, 448 = Crispin / AGENZIA 446); the
correct customer is disambiguated by the agency link (see crispin_automation).

TWO source formats, one `CrispinOrder` shape — use `parse_crispin(path)`:

1. **Proposal workbook** (.xlsx/.xlsm) — `parse_crispin_xlsx`. What we quote:
   - header block (Agency / Advertiser / Contact / …) as label→value cell pairs
   - a market banner row ("San Francisco Bay Area - Xfinity Channel 3131 …")
   - a column-header row: Language | Daypart | Unit Value | Discounted Rate |
     Length | <week-date columns…> | Total Spots | Total Value | Proposed Amount
   - airtime rows, then Total Paid / Total Bonuses / Total footer rows
   Rate rule (Lee): use the **Discounted Rate** column, never Unit Value. A
   discounted rate of 0 marks a bonus line (the :15s ROS added-value spots).
   Proposal rates are NET (no commission was attached when we quoted).

2. **Official IO** (.pdf) — `parse_crispin_pdf`. A "Brand Time Schedule", the
   same agency-system layout Daviselen and Intertrend send. Rates here are
   **GROSS** (net ÷ 0.85) and the ANAGRAF agency commission nets them back down,
   so they enter verbatim — see `crispin_automation` for the money discipline.

   Structural traps this reader exists to survive (all live in the 8/10/26 IO):
   - A long flight is split into **column regimes** of ~13 weeks each. Every
     line repeats once per regime and must be merged by line number.
   - A single page can carry **two regimes** (a regime's summary block, a
     `-----` divider, then the next regime's header mid-page), so column
     mapping is per-REGION, never per-page.
   - **Zero cells are not printed.** Spot counts must be read by x-coordinate
     against the day-number header, not by splitting text.
   - Flight dates are **per line** (the DATES column), not per order.
   - Non-airtime rows ride in the grid as ordinary lines (`TRANSLATION COST`)
     and must become CONTRATTISPESE charges, never contract lines.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

import openpyxl

# ─── Market detection ────────────────────────────────────────────────────────
_MARKET_MAP = [
    (("san francisco", "bay area", "sfo"), "SFO"),
    (("sacramento", "central valley", "cvc"), "CVC"),
    (("seattle", "sea"), "SEA"),
    (("los angeles", "lax"), "LAX"),
    (("houston", "hou"), "HOU"),
    (("chicago", "minneapolis", "cmp"), "CMP"),
    (("washington", "wdc"), "WDC"),
    (("new york", "nyc"), "NYC"),
]

# Base languages used to normalise a line's language block for ROS mapping.
_BASE_LANGUAGES = ["Cantonese", "Mandarin", "Filipino", "Vietnamese",
                   "Korean", "Hmong", "Punjabi", "Japanese", "Hindi",
                   "South Asian", "Chinese"]

# Time range at the END of a daypart string, e.g. "M-F 7p-8p" → "7p-8p".
_TIME_RE = re.compile(
    r"(\d{1,2}(?::\d{2})?\s*[apAP]?\s*[-–]\s*\d{1,2}(?::\d{2})?\s*[apAP])"
)


def split_daypart(daypart: str) -> tuple[str, str]:
    """'M-F 7p-8p' → ('M-F', '7p-8p'). Returns (days, '') if no time found."""
    dp = (daypart or "").strip()
    m = _TIME_RE.search(dp)
    if not m:
        return dp, ""
    return dp[: m.start()].strip(), m.group(1).strip()


def _num(v) -> float:
    """Coerce a money/count cell to float. '$120.00', '42.5', '-', None, '' → float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if s in ("", "-", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ─── Data model ──────────────────────────────────────────────────────────────

@dataclass
class CrispinLine:
    language_block: str          # raw, e.g. "Cantonese News" or "Cantonese"
    daypart: str                 # raw, e.g. "M-F 7p-8p" or "ROS"
    unit_value: float            # standard rate card (informational)
    rate: float                  # the billed rate (0 ⇒ bonus)
    length_sec: int              # 30 / 15 from the Length column
    week_dates: List[date]       # Monday of each flight week
    week_spots: List[int]        # spots per week (parallel to week_dates)
    # ── PDF-IO only (the proposal has no per-line flight) ──
    line_number: str = ""        # the IO's own LINE# ("006")
    date_from: Optional[date] = None   # this line's flight start (DATES column)
    date_to: Optional[date] = None     # this line's flight end
    total_spots_stated: Optional[int] = None  # the IO's own TOT, for reconciling

    @property
    def is_bonus(self) -> bool:
        return round(self.rate, 4) == 0.0

    @property
    def total_spots(self) -> int:
        return sum(self.week_spots)

    # Aliases so the generic parser_bridge normalizer (web preview) picks these up.
    @property
    def weekly_spots(self) -> List[int]:
        return self.week_spots

    @property
    def length(self) -> int:
        return self.length_sec

    @property
    def base_language(self) -> str:
        """Normalised language for ROS mapping ('Cantonese News' → 'Cantonese')."""
        low = self.language_block.strip().lower()
        for lang in _BASE_LANGUAGES:
            if low.startswith(lang.lower()):
                return lang
        return self.language_block.strip()


@dataclass
class CrispinCharge:
    """A non-airtime row riding in the IO grid (translation / production cost).

    Enters Etere as a CONTRATTISPESE row attached to an airtime line, never as a
    contract line of its own — see `crispin_automation._write_production_charges`.
    """
    description: str             # as printed, e.g. "TRANSLATION COST"
    amount: float                # IMPORTO (same gross/net basis as the rates)
    line_number: str = ""        # the IO's own LINE#
    date_from: Optional[date] = None
    date_to: Optional[date] = None


@dataclass
class CrispinOrder:
    agency: str
    advertiser: str
    contact: str
    email: str
    market_code: str
    market_label: str
    order_date: Optional[date]
    lines: List[CrispinLine] = field(default_factory=list)
    rates_are_net: bool = False   # PDF IO quotes GROSS; proposal quotes net
    source_path: str = ""
    charges: List[CrispinCharge] = field(default_factory=list)
    # ── PDF-IO header fields (blank for the proposal) ──
    source_format: str = "xlsx"   # 'xlsx' | 'pdf'
    order_number: str = ""        # "212735" (leading zeros stripped)
    estimate: str = ""            # "0001"
    estimate_detail: str = ""     # "BAAQ 2026 TV SUMMERCAMPAIGN"
    revision: str = ""            # "2"
    client_code: str = ""         # "BAAQ"
    station_code: str = ""        # "3131CA"

    @property
    def paid_lines(self) -> List[CrispinLine]:
        return [ln for ln in self.lines if not ln.is_bonus]

    @property
    def bonus_lines(self) -> List[CrispinLine]:
        return [ln for ln in self.lines if ln.is_bonus]

    @property
    def week_dates(self) -> List[date]:
        """Union of every line's week columns, ascending.

        The proposal has one shared week grid, so this is `lines[0]`'s columns.
        The PDF IO splits a long flight across column regimes and each line
        carries only the weeks it appears under, so the union is the real grid.
        """
        seen = {d for ln in self.lines for d in ln.week_dates}
        return sorted(seen)

    @property
    def flight_start(self) -> Optional[str]:
        """Earliest line flight start — the DATES column when the IO gives one,
        else the first week column."""
        starts = [ln.date_from for ln in self.lines if ln.date_from]
        if starts:
            return min(starts).strftime("%m/%d/%Y")
        wd = self.week_dates
        return wd[0].strftime("%m/%d/%Y") if wd else None

    @property
    def flight_end(self) -> Optional[str]:
        """Latest line flight end, else the Sunday of the last week column."""
        ends = [ln.date_to for ln in self.lines if ln.date_to]
        if ends:
            return max(ends).strftime("%m/%d/%Y")
        wd = self.week_dates
        return (wd[-1] + timedelta(days=6)).strftime("%m/%d/%Y") if wd else None


# ─── Parser ──────────────────────────────────────────────────────────────────

_HEADER_LABELS = {"agency", "advertiser", "contact", "email", "station",
                  "languages", "payment terms", "revision"}


def _pick_sheet(wb):
    """Return the worksheet holding the airtime grid (Language + Daypart header)."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip().lower() for c in row if c is not None]
            if "language" in cells and "daypart" in cells:
                return ws
    return wb.active


def parse_crispin_xlsx(path: str) -> CrispinOrder:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))

    # ── Header label→value pairs (label cell, value = next non-empty cell) ──
    hdr: dict[str, str] = {}
    for r in rows:
        for i, c in enumerate(r):
            if c is None:
                continue
            label = str(c).strip().lower()
            # 'Revision ( revised the start date)' → key on the leading word
            key = next((k for k in _HEADER_LABELS if label.startswith(k)), None)
            if key and key not in hdr:
                for c2 in r[i + 1:]:
                    if c2 is not None and str(c2).strip():
                        hdr[key] = str(c2).strip() if not isinstance(c2, (datetime, date)) else c2
                        break

    agency = str(hdr.get("agency", "")).strip()
    advertiser = str(hdr.get("advertiser", "")).strip()
    contact = str(hdr.get("contact", "")).strip()
    email = str(hdr.get("email", "")).strip()
    order_date = _to_date(hdr.get("revision"))

    # ── Market banner + column-header row ──
    market_label = ""
    market_locked = False
    header_ri = None
    col: dict[str, int] = {}
    week_cols: List[int] = []
    week_dates: List[date] = []

    for ri, r in enumerate(rows):
        cells_l = [str(c).strip().lower() if c is not None else "" for c in r]
        if header_ri is None and "language" in cells_l and "daypart" in cells_l:
            header_ri = ri
            for ci, cl in enumerate(cells_l):
                if cl == "language":
                    col["lang"] = ci
                elif cl == "daypart":
                    col["daypart"] = ci
                elif cl.startswith("unit value"):
                    col["unit"] = ci
                elif cl.startswith("discounted"):
                    col["disc"] = ci
                elif cl == "length":
                    col["length"] = ci
                elif cl.startswith("total spots"):
                    col["total_spots"] = ci
            # week columns = header cells that are real dates
            for ci, cval in enumerate(r):
                d = _to_date(cval)
                if d is not None:
                    week_cols.append(ci)
                    week_dates.append(d)
            break
        # Market banner (before the grid): a row naming a market. The real
        # banner ("San Francisco Bay Area - Xfinity Channel 3131 / KQTA 15.3")
        # carries a channel marker — prefer and lock onto it, since the
        # Advertiser header row can also contain a city name.
        if not market_locked:
            joined = " ".join(cells_l)
            is_banner = any(k in joined for k in ("xfinity", "kqta", "channel"))
            for keys, code in _MARKET_MAP:
                if any(k in joined for k in keys):
                    market_label = " ".join(str(c).strip() for c in r if c is not None)
                    market_locked = is_banner
                    break

    if header_ri is None:
        raise ValueError("Crispin parser: could not find the 'Language/Daypart' column header row")
    if not week_cols:
        raise ValueError("Crispin parser: no weekly date columns found in the header row")
    for req in ("lang", "daypart", "disc", "length"):
        if req not in col:
            raise ValueError(f"Crispin parser: missing '{req}' column in header row")

    market_code = "SFO"
    for keys, code in _MARKET_MAP:
        if any(k in market_label.lower() for k in keys):
            market_code = code
            break

    # ── Airtime rows ──
    # The airtime block is followed by a Total Paid / Total Bonuses / Total
    # footer and then unrelated sections (translation costs, impressions table,
    # T&Cs). Capture the two footer totals for reconciliation, then stop; and
    # only accept a row as a line when its daypart is real (a time range or ROS)
    # so stray text/impression rows can never masquerade as airtime.
    lines: List[CrispinLine] = []
    footer_paid = footer_bonus = None
    in_totals = False
    for r in rows[header_ri + 1:]:
        joined_l = " ".join(str(c).strip().lower() for c in r if c is not None)
        if "total paid" in joined_l:
            if "total_spots" in col and col["total_spots"] < len(r):
                footer_paid = int(_num(r[col["total_spots"]]))
            in_totals = True
            continue
        if "total bonus" in joined_l:
            if "total_spots" in col and col["total_spots"] < len(r):
                footer_bonus = int(_num(r[col["total_spots"]]))
            in_totals = True
            continue
        if in_totals:
            break  # past the airtime block — footers captured, stop

        lang_raw = r[col["lang"]] if col["lang"] < len(r) else None
        if lang_raw is None or not str(lang_raw).strip():
            continue

        daypart = str(r[col["daypart"]]).strip() if col["daypart"] < len(r) and r[col["daypart"]] else ""
        # Airtime guard: a real line has a time range (paid) or "ROS" (bonus).
        if not (daypart.upper() == "ROS" or _TIME_RE.search(daypart)):
            continue
        unit_value = _num(r[col["unit"]]) if "unit" in col and col["unit"] < len(r) else 0.0
        rate = _num(r[col["disc"]]) if col["disc"] < len(r) else 0.0
        length_cell = str(r[col["length"]]) if col["length"] < len(r) else ""
        m = re.search(r"(\d+)", length_cell)
        length_sec = int(m.group(1)) if m else 30

        spots = [int(_num(r[ci])) if ci < len(r) else 0 for ci in week_cols]

        lines.append(CrispinLine(
            language_block=str(lang_raw).strip(),
            daypart=daypart,
            unit_value=unit_value,
            rate=rate,
            length_sec=length_sec,
            week_dates=list(week_dates),
            week_spots=spots,
        ))

    if not lines:
        raise ValueError("Crispin parser: no airtime lines found")

    order = CrispinOrder(
        agency=agency,
        advertiser=advertiser,
        contact=contact,
        email=email,
        market_code=market_code,
        market_label=market_label,
        order_date=order_date,
        lines=lines,
        source_path=path,
    )

    # ── Reconcile against the footer (Brentan/SCWA totals lesson) ──
    paid_sum = sum(ln.total_spots for ln in order.paid_lines)
    bonus_sum = sum(ln.total_spots for ln in order.bonus_lines)
    if footer_paid is not None and paid_sum != footer_paid:
        raise ValueError(
            f"Crispin parser: paid spot total {paid_sum} != footer 'Total Paid' {footer_paid} "
            f"— a line was likely dropped; refusing to enter."
        )
    if footer_bonus is not None and bonus_sum != footer_bonus:
        raise ValueError(
            f"Crispin parser: bonus spot total {bonus_sum} != footer 'Total Bonuses' {footer_bonus}."
        )

    return order


# ─── Brand Time Schedule PDF (the official IO) ───────────────────────────────
#
# Read by coordinate, mapped by HEADER LABEL. Zero cells are not printed, so the
# only way to know which week a "3" belongs to is its x-position against the
# day-number header — and the day-number header changes partway down page 3.

_ROW_TOL = 2.0      # cluster words into rows on RAW top floats. Intra-row jitter
                    # is ~0.5pt (28.6/28.8 twin title rows); row pitch is ~8.9pt.
                    # Never round() first — that manufactures phantom rows.
_COL_TOL = 8.0      # week-cell centre vs column centre. Column pitch is ~19.3pt,
                    # so this sits well inside half-pitch and cannot cross columns.
_BAND_PAD = 3.0     # a value sits at, or a hair right of, its header label's x0.
_GRID_X_MIN = 290.0  # left edge of the week grid — text left of this is never a cell

_MONTHS = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
           'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}

# Column labels of the schedule grid, in print order. Mapping by label (never by
# index) is the DART rule: the agency adding a column must be a no-op, and a
# RENAMED column must fail loudly instead of falling back to a positional guess.
_PDF_LABELS = ['LINE#', 'DAY(S)', 'TIME', 'DATES', 'PROGRAM', 'LEN', 'DP']

_DAYS_NORMAL = {'M-SU': 'M-Su', 'M-SA': 'M-Sa', 'M-F': 'M-F', 'SA-SU': 'Sa-Su',
                'SA': 'Sa', 'SU': 'Su', 'M-TH': 'M-Th', 'SU-SA': 'Su-Sa'}

# A grid row naming no language is NOT airtime. It is only accepted as a charge
# when it also names a recognisable non-airtime cost; anything else raises,
# because silently entering it either way would be wrong.
_CHARGE_RE = re.compile(
    r'\b(TRANSLATION|PRODUCTION|DUBBING|DUB|VOICE\s*-?\s*OVER|VOICEOVER|TALENT'
    r'|POST|EDIT|CAPTION|SUBTITL\w*|FEE|COST|CHARGE)\b', re.IGNORECASE)


def _pdf_num(text: str) -> Optional[Decimal]:
    """'2447.06' / '12' / '23,529.94' → Decimal. Non-numeric → None."""
    s = (text or '').replace(',', '').replace('$', '').strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _cluster_rows(words: List[dict]) -> List[Tuple[float, List[dict]]]:
    """Group words into visual rows on RAW `top`, ascending. See _ROW_TOL."""
    rows: List[Tuple[float, List[dict]]] = []
    for w in sorted(words, key=lambda w: w['top']):
        if rows and w['top'] - rows[-1][0] <= _ROW_TOL:
            rows[-1][1].append(w)
        else:
            rows.append((w['top'], [w]))
    return [(top, sorted(ws, key=lambda w: w['x0'])) for top, ws in rows]


def _band_text(words: List[dict], lo: float, hi: float) -> str:
    """Join the words whose x0 falls in [lo, hi)."""
    return ' '.join(w['text'] for w in words if lo <= w['x0'] < hi).strip()


def _centre(w: dict) -> float:
    return (w['x0'] + w['x1']) / 2.0


def _parse_pdf_time(raw: str) -> Optional[str]:
    """'0600A' → '6:00a', '0100P' → '1:00p', '1200A' → '12:00a'."""
    m = re.fullmatch(r'(\d{1,2})(\d{2})\s*([AP])', (raw or '').strip(), re.IGNORECASE)
    if not m:
        return None
    return f"{int(m.group(1))}:{m.group(2)}{m.group(3).lower()}"


def _compact_range(t_from: str, t_to: str) -> str:
    """'6:00a'+'12:00a' → '6a-12a'; '11:00a'+'1:00p' → '11a-1p'; '7:00p'+'8:00p' → '7-8p'.

    Matches how Lee writes line descriptions in Etere ("Cantonese 7-8p"), and
    still parses back through EtereClient.parse_time_range.
    """
    def part(t: str) -> Tuple[str, str]:
        m = re.fullmatch(r'(\d{1,2}):(\d{2})([ap])', t)
        if not m:
            return t, ''
        hh, mm, mer = m.group(1), m.group(2), m.group(3)
        return (hh if mm == '00' else f"{hh}:{mm}"), mer

    a, mer_a = part(t_from)
    b, mer_b = part(t_to)
    if not mer_a or not mer_b:
        return f"{t_from}-{t_to}"
    return f"{a}-{b}{mer_b}" if mer_a == mer_b else f"{a}{mer_a}-{b}{mer_b}"


def _parse_pdf_date(raw: str, anchor_year: int, anchor_month: int) -> Optional[date]:
    """'AUG10' → date(2026, 8, 10), rolling the year when the month wraps back
    past the estimate period's start month."""
    m = re.fullmatch(r'([A-Z]{3})(\d{1,2})', (raw or '').strip().upper())
    if not m or m.group(1) not in _MONTHS:
        return None
    month = _MONTHS[m.group(1)]
    year = anchor_year if month >= anchor_month else anchor_year + 1
    try:
        return date(year, month, int(m.group(2)))
    except ValueError:
        return None


def _pdf_days(raw: str) -> str:
    return _DAYS_NORMAL.get((raw or '').strip().upper(), (raw or '').strip())


def parse_crispin_pdf(path: str) -> CrispinOrder:
    """Parse a Crispin 'Brand Time Schedule' insertion order PDF."""
    import pdfplumber

    from browser_automation.line_language import guess_language

    with pdfplumber.open(path) as pdf:
        pages = [(_cluster_rows(pg.extract_words()), pg.extract_text() or '')
                 for pg in pdf.pages]

    all_text = '\n'.join(t for _, t in pages)

    # ── Header fields ────────────────────────────────────────────────────────
    station_code = ''
    m = re.search(r'Brand Time Schedule\s*-\s*(\S+)', all_text)
    if m:
        station_code = m.group(1).strip()

    m = re.search(r'PERIOD FROM\s+([A-Z]{3})(\d{1,2})/(\d{2})', all_text)
    if not m:
        raise ValueError("Crispin PDF: no 'PERIOD FROM' line — not a Brand Time Schedule?")
    anchor_month = _MONTHS.get(m.group(1).upper(), 1)
    anchor_year = 2000 + int(m.group(3))

    advertiser, client_code = '', ''
    m = re.search(r'CLIENT\s+([A-Z0-9]+)\s+(.+?)\s*(?:Market\b|$)', all_text, re.MULTILINE)
    if m:
        client_code = m.group(1).strip()
        advertiser = m.group(2).strip()

    market_label = ''
    m = re.search(r'\bMarket\s+([A-Z]{2,4})\s+([A-Z]{2})\s+(.+?)(?:\s+RTG\b|$)',
                  all_text, re.MULTILINE)
    if m:
        market_label = m.group(3).strip()
    market_code = ''
    for keys, code in _MARKET_MAP:
        if any(k in market_label.lower() for k in keys):
            market_code = code
            break
    if not market_code:
        raise ValueError(
            f"Crispin PDF: unrecognised market {market_label!r} — add it to _MARKET_MAP "
            f"rather than defaulting, so the spots cannot land on the wrong station."
        )

    order_number = ''
    m = re.search(r'Order#?\s*(\d+)', all_text)
    if m:
        order_number = m.group(1).lstrip('0') or '0'

    estimate, estimate_detail = '', ''
    m = re.search(r'ESTIMATE\s+(\d+)\s*(.*)', all_text)
    if m:
        estimate = m.group(1).strip()
        estimate_detail = m.group(2).strip()

    revision = ''
    m = re.search(r'REVISION:\s*(\S+)', all_text)
    if m:
        revision = m.group(1).strip()

    # "CRISPIN AGENCY" — a single all-caps token, never a greedy run: the page
    # title ("Brand Time Schedule - 3131CA") shares this text line in extraction.
    agency = 'Crispin LLC'
    m = re.search(r'\b([A-Z][A-Z&.\-]{2,})\s+AGENCY\b', all_text)
    if m:
        agency = f"{m.group(1).title()} LLC"

    # ── Walk every page: column regimes, grid rows, summary rows ─────────────
    bands: Dict[str, Tuple[float, float]] = {}
    tot_x0: Optional[float] = None
    regimes: List[dict] = []      # {'page','top','weeks': [(centre, date)]}
    grid: Dict[str, dict] = {}    # LINE# → row record
    grid_order: List[str] = []
    summaries: List[dict] = []    # per-regime PTS/WEEK reconciliation rows
    grand: Optional[dict] = None

    for pi, (rows, _) in enumerate(pages):
        pending_months: Optional[List[dict]] = None

        for ri, (top, words) in enumerate(rows):
            texts = [w['text'] for w in words]

            # -- the grid's column-header row: learn the field bands once --
            if 'LINE#' in texts and 'DAY(S)' in texts:
                if not bands:
                    xs = {}
                    for w in words:
                        if w['text'] in _PDF_LABELS or w['text'] == 'TOT':
                            xs.setdefault(w['text'], w['x0'])
                    missing = [lb for lb in _PDF_LABELS + ['TOT'] if lb not in xs]
                    if missing:
                        raise ValueError(
                            f"Crispin PDF: schedule header is missing column(s) {missing} — "
                            f"the layout changed; refusing to guess by position."
                        )
                    tot_x0 = xs['TOT']
                    edges = [(lb, xs[lb]) for lb in _PDF_LABELS]
                    for i, (lb, x) in enumerate(edges):
                        nxt = edges[i + 1][1] if i + 1 < len(edges) else _GRID_X_MIN
                        bands[lb] = (x - _BAND_PAD, nxt - _BAND_PAD)

            # -- month row of a column regime --
            months = [w for w in words
                      if w['x0'] >= _GRID_X_MIN and w['text'].upper() in _MONTHS]
            if len(months) >= 5:
                pending_months = months
                # The month row may also BE the row above the main header; the day
                # numbers arrive on the next row either way.
                continue

            # -- day-number row: pairs with the pending month row --
            days_hdr = [w for w in words
                        if w['x0'] >= _GRID_X_MIN
                        and re.fullmatch(r'\d{1,2}', w['text'])
                        and (tot_x0 is None or w['x0'] < tot_x0 - _COL_TOL)]
            if pending_months and len(days_hdr) >= 5:
                if len(days_hdr) != len(pending_months):
                    raise ValueError(
                        f"Crispin PDF p{pi + 1}: {len(pending_months)} month headers but "
                        f"{len(days_hdr)} day numbers — cannot align the week grid."
                    )
                weeks: List[Tuple[float, date]] = []
                prev_month = None
                year = anchor_year
                for mo_w, dy_w in zip(sorted(pending_months, key=lambda w: w['x0']),
                                      sorted(days_hdr, key=lambda w: w['x0'])):
                    month = _MONTHS[mo_w['text'].upper()]
                    if prev_month is None:
                        year = anchor_year if month >= anchor_month else anchor_year + 1
                    elif month < prev_month:
                        year += 1
                    prev_month = month
                    weeks.append((_centre(dy_w), date(year, month, int(dy_w['text']))))
                for a, b in zip(weeks, weeks[1:]):
                    if (b[1] - a[1]).days != 7:
                        raise ValueError(
                            f"Crispin PDF p{pi + 1}: week columns {a[1]} → {b[1]} are "
                            f"{(b[1] - a[1]).days} days apart, expected 7 — the month/day "
                            f"header did not align."
                        )
                regimes.append({'page': pi, 'top': top, 'weeks': weeks})
                pending_months = None
                continue

            if not bands:
                continue

            def cell_map(ws: List[dict], regime: dict) -> Dict[date, int]:
                """Week cells of one row, matched to columns by centre distance."""
                out: Dict[date, int] = {d: 0 for _, d in regime['weeks']}
                for w in ws:
                    if w['x0'] < _GRID_X_MIN:
                        continue
                    if tot_x0 is not None and w['x0'] >= tot_x0 - _COL_TOL:
                        continue
                    val = _pdf_num(w['text'])
                    if val is None or val != val.to_integral_value():
                        continue
                    cx, best = _centre(w), None
                    for col_x, d in regime['weeks']:
                        dist = abs(cx - col_x)
                        if dist <= _COL_TOL and (best is None or dist < best[0]):
                            best = (dist, d)
                    if best:
                        out[best[1]] = int(val)
                return out

            def tail(ws: List[dict]) -> Tuple[Optional[int], Optional[Decimal]]:
                """(TOT, money) — the two right-hand columns. Money is the value
                carrying a decimal point, so a shifted right margin can't swap them."""
                spots = money = None
                for w in ws:
                    if tot_x0 is None or w['x0'] < tot_x0 - _COL_TOL:
                        continue
                    val = _pdf_num(w['text'])
                    if val is None:
                        continue
                    if '.' in w['text']:
                        if money is None:
                            money = val
                    elif spots is None:
                        spots = int(val)
                return spots, money

            def regime_for(page: int, y: float) -> dict:
                for reg in reversed(regimes):
                    if reg['page'] < page or (reg['page'] == page and reg['top'] < y):
                        return reg
                raise ValueError(
                    f"Crispin PDF p{page + 1}: a grid row at y={y:.0f} has no week-column "
                    f"header above it."
                )

            # -- station summary rows (reconciliation targets) --
            prog = _band_text(words, *bands['PROGRAM'])
            if station_code and prog.split()[:1] == [station_code]:
                rest = prog.split()[1:]
                spots, money = tail(words)
                if rest[:1] == ['TOT']:
                    grand = {'spots': spots, 'money': money}
                else:
                    reg = regime_for(pi, top)
                    summaries.append({
                        'label': rest[0] if rest else '',
                        'weeks': [d for _, d in reg['weeks']],
                        'cells': cell_map(words, reg),
                        'spots': spots, 'money': money,
                    })
                continue

            # -- a schedule line --
            line_no = _band_text(words, *bands['LINE#'])
            if not re.fullmatch(r'\d{3}', line_no):
                continue

            reg = regime_for(pi, top)
            spots, money = tail(words)
            if spots is None or money is None:
                raise ValueError(
                    f"Crispin PDF p{pi + 1} line {line_no}: could not read the TOT/COST "
                    f"columns (got spots={spots}, cost={money})."
                )

            # The continuation row underneath carries the end time and end date.
            end_time = end_date = ''
            if ri + 1 < len(rows):
                nxt = rows[ri + 1][1]
                if not _band_text(nxt, *bands['LINE#']):
                    end_time = _band_text(nxt, *bands['TIME'])
                    end_date = _band_text(nxt, *bands['DATES'])

            t_from = _parse_pdf_time(_band_text(words, *bands['TIME']))
            t_to = _parse_pdf_time(end_time)
            d_from = _parse_pdf_date(_band_text(words, *bands['DATES']),
                                     anchor_year, anchor_month)
            d_to = _parse_pdf_date(end_date, anchor_year, anchor_month)
            if d_from and d_to and d_to < d_from:      # flight crosses New Year
                d_to = date(d_to.year + 1, d_to.month, d_to.day)

            len_m = re.search(r'(\d+)', _band_text(words, *bands['LEN']))
            rec = {
                'line_number': line_no,
                'days': _pdf_days(_band_text(words, *bands['DAY(S)'])),
                'time_from': t_from, 'time_to': t_to,
                'date_from': d_from, 'date_to': d_to,
                'program': prog,
                'length_sec': int(len_m.group(1)) if len_m else 30,
                'dp': _band_text(words, *bands['DP']),
                'rate': money,
                'stated_total': spots,
                'cells': cell_map(words, reg),
            }

            if line_no in grid:                        # merge the other regime
                prev = grid[line_no]
                for key in ('days', 'time_from', 'time_to', 'date_from', 'date_to',
                            'program', 'length_sec', 'rate'):
                    if prev[key] != rec[key]:
                        raise ValueError(
                            f"Crispin PDF line {line_no}: {key} differs between column "
                            f"regimes ({prev[key]!r} vs {rec[key]!r}) — these are not the "
                            f"same line."
                        )
                for d, v in rec['cells'].items():
                    if d in prev['cells'] and prev['cells'][d] and prev['cells'][d] != v:
                        raise ValueError(
                            f"Crispin PDF line {line_no}: week {d} read twice with "
                            f"different values ({prev['cells'][d]} vs {v})."
                        )
                    prev['cells'][d] = max(prev['cells'].get(d, 0), v)
                prev['stated_total'] += spots
            else:
                grid[line_no] = rec
                grid_order.append(line_no)

    if not bands:
        raise ValueError("Crispin PDF: no 'LINE# DAY(S) …' schedule header found")
    if not grid:
        raise ValueError("Crispin PDF: no schedule lines found")

    # ── Reconcile against the IO's own arithmetic BEFORE trusting anything ───
    for no in grid_order:
        rec = grid[no]
        got = sum(rec['cells'].values())
        if got != rec['stated_total']:
            raise ValueError(
                f"Crispin PDF line {no} ({rec['program']}): week columns sum to {got} "
                f"but the line's own TOT says {rec['stated_total']} — a cell was misread; "
                f"refusing to enter."
            )

    for s in summaries:
        for d in s['weeks']:
            want = s['cells'].get(d, 0)
            got = sum(rec['cells'].get(d, 0) for rec in grid.values())
            if got != want:
                raise ValueError(
                    f"Crispin PDF summary {s['label']}: week {d} sums to {got} across the "
                    f"lines but the PTS/WEEK row says {want}."
                )
        got_spots = sum(sum(rec['cells'].get(d, 0) for d in s['weeks'])
                        for rec in grid.values())
        if s['spots'] is not None and got_spots != s['spots']:
            raise ValueError(
                f"Crispin PDF summary {s['label']}: {got_spots} spots across the lines "
                f"but the summary says {s['spots']}."
            )
        got_money = sum(rec['rate'] * sum(rec['cells'].get(d, 0) for d in s['weeks'])
                        for rec in grid.values())
        if s['money'] is not None and abs(got_money - s['money']) > Decimal('0.01'):
            raise ValueError(
                f"Crispin PDF summary {s['label']}: lines total ${got_money} but the "
                f"summary says ${s['money']}."
            )

    if grand:
        got_spots = sum(rec['stated_total'] for rec in grid.values())
        if grand['spots'] is not None and got_spots != grand['spots']:
            raise ValueError(
                f"Crispin PDF: {got_spots} spots parsed but '{station_code} TOT' says "
                f"{grand['spots']} — a line was dropped; refusing to enter."
            )
        got_money = sum(rec['rate'] * rec['stated_total'] for rec in grid.values())
        if grand['money'] is not None and abs(got_money - grand['money']) > Decimal('0.01'):
            raise ValueError(
                f"Crispin PDF: parsed total ${got_money} but '{station_code} TOT' says "
                f"${grand['money']} — refusing to enter."
            )

    # ── Split airtime from non-airtime charges ──────────────────────────────
    lines: List[CrispinLine] = []
    charges: List[CrispinCharge] = []
    for no in grid_order:
        rec = grid[no]
        if guess_language(rec['program']) is None:
            if not _CHARGE_RE.search(rec['program']):
                raise ValueError(
                    f"Crispin PDF line {no}: program {rec['program']!r} names no language "
                    f"and no recognisable production cost — cannot tell whether it is "
                    f"airtime or a charge. Add it to _CHARGE_RE or fix the IO."
                )
            charges.append(CrispinCharge(
                description=rec['program'],
                amount=float(rec['rate'] * rec['stated_total']),
                line_number=no,
                date_from=rec['date_from'],
                date_to=rec['date_to'],
            ))
            continue

        # Keep only the week columns this line's own flight actually covers. A
        # line appears under one regime but is zero-padded across the other's 13
        # columns too, so without this every line reports the full 26-week grid.
        # A dropped column holding spots is a contradiction, not padding.
        weeks = sorted(rec['cells'])
        if rec['date_from'] and rec['date_to']:
            keep = [d for d in weeks
                    if d <= rec['date_to'] and d + timedelta(days=6) >= rec['date_from']]
            stray = {d: rec['cells'][d] for d in weeks
                     if d not in keep and rec['cells'][d]}
            if stray:
                raise ValueError(
                    f"Crispin PDF line {no}: {sum(stray.values())} spot(s) in week(s) "
                    f"{sorted(stray)} fall outside the line's stated flight "
                    f"{rec['date_from']}–{rec['date_to']}."
                )
            weeks = keep

        is_bonus = rec['rate'] == 0
        if is_bonus:
            daypart = 'ROS'
        elif rec['time_from'] and rec['time_to']:
            daypart = f"{rec['days']} {_compact_range(rec['time_from'], rec['time_to'])}"
        else:
            raise ValueError(
                f"Crispin PDF line {no}: paid line has no readable time range "
                f"({rec['time_from']!r}–{rec['time_to']!r})."
            )

        lines.append(CrispinLine(
            language_block=rec['program'],
            daypart=daypart,
            unit_value=0.0,
            rate=float(rec['rate']),
            length_sec=rec['length_sec'],
            week_dates=weeks,
            week_spots=[rec['cells'][d] for d in weeks],
            line_number=no,
            date_from=rec['date_from'],
            date_to=rec['date_to'],
            total_spots_stated=rec['stated_total'],
        ))

    if not lines:
        # A production-ONLY order takes the zero-spot carrier-line pattern
        # (N_PASSAGGI=1, IMPORTO=0, HIDEFROMSCHEDULER=1, description
        # 'PRODUCTION ONLY - no airtime' + the charge) rather than the Production
        # box, which needs a paid airtime line to hang off. That path is not
        # built — say so instead of half-entering the order.
        total = sum(c.amount for c in charges)
        raise ValueError(
            f"Crispin PDF: no airtime lines — this is a production-only order "
            f"(${total:,.2f} in {len(charges)} charge(s)). Those enter as a zero-spot "
            f"carrier line, which this parser does not build yet; enter it by hand."
        )

    return CrispinOrder(
        agency=agency,
        advertiser=advertiser,
        contact='',
        email='',
        market_code=market_code,
        market_label=market_label,
        order_date=None,
        lines=lines,
        # The IO quotes GROSS rates (net ÷ 0.85); the ANAGRAF agency commission
        # nets them back down, so no gross-up is owed downstream.
        rates_are_net=False,
        source_path=path,
        charges=charges,
        source_format='pdf',
        order_number=order_number,
        estimate=estimate,
        estimate_detail=estimate_detail,
        revision=revision,
        client_code=client_code,
        station_code=station_code,
    )


def parse_crispin(path: str) -> CrispinOrder:
    """Parse a Crispin order — the official Brand Time Schedule IO (.pdf) or the
    proposal workbook (.xlsx/.xlsm). One order shape either way."""
    return parse_crispin_pdf(path) if str(path).lower().endswith('.pdf') \
        else parse_crispin_xlsx(path)
