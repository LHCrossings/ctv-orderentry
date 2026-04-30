"""
3 Olives Media Order Parser

Parses both PDF and Excel (.xlsm) variants of the 3 Olives Media insertion order
for Riverside County Voter Registration and Elections (and similar clients).

Document Layout:
  Header block (text above table):
    Riverside County Voter Registration & Elections
    Client: <Client Name>
    Contact: <Name>
    Email: <name>@3olivesmedia.com
    Channel: Spectrum 1519
    Date: M/D/YYYY

  Data table columns:
    [BONUS] | Insertion | Time | Value | Discounted | [week dates...] | Units | Value | GROSS
    - col 0 (PDF) / col 1 (Excel): "BONUS" flag or None
    - col 1 (PDF) / col 2 (Excel): Program/Insertion name
    - col 2 (PDF) / col 3 (Excel): Time/daypart string
    - col 3 (PDF) / col 4 (Excel): Value/card rate (not used for Etere)
    - col 4 (PDF) / col 5 (Excel): Discounted/billing rate  ← use this
    - next N cols: per-week spot counts (N = number of weeks)
    - then: Units (total spots), total Value, total GROSS

Key Business Rules:
  - Rate: use "Discounted" column — the actual billing rate we enter in Etere
  - Bonus lines: Discounted = 0; col 0/1 = "BONUS"
  - Time: "/" separates multiple time ranges (treated same as ";")
  - "M-Sun" / "M-Sunday" normalised to "M-Su"; "Sat-Sun" → "Sa-Su"
  - "ROS Bonus" → days M-Su, time 6a-11:59p
  - Multiline PDF cells (newlines) are joined with a space before parsing
  - Week dates: datetime objects (Excel) or "D-Mon" strings (PDF) → "Mon D" format
  - Flight start = first week date; flight end = last week date + 6 days
  - Market: not in document — operator provides at input-gathering time (default LAX)
  - Excel has "Option 1" and "Option 2" sheets; Option 1 is the primary
"""

import re
import sys
import datetime
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Tuple

_project_root = Path(__file__).parent.parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))


# ─────────────────────────────────────────────────────────────────────────────
# DAYPART PARSING
# ─────────────────────────────────────────────────────────────────────────────

_DAY_RE = re.compile(r'\b(M-Su|M-F|Sa-Su|M-Sa|Sa|Su)\b', re.IGNORECASE)


def _normalise_day_tokens(s: str) -> str:
    s = re.sub(r'\bM-Sun\b|\bM-Sunday\b|\bMon-Sun\b', 'M-Su', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSat-Sun\b|\bSat-Sunday\b', 'Sa-Su', s, flags=re.IGNORECASE)
    return s


def _broadest_days(patterns: List[str]) -> str:
    lowers = [p.lower() for p in patterns]
    if any('m-su' in d for d in lowers):
        return 'M-Su'
    has_weekday = any(d in ('m-f', 'm-sa') for d in lowers)
    has_weekend = any('sa-su' in d for d in lowers)
    if has_weekday and has_weekend:
        return 'M-Su'
    if has_weekday:
        return 'M-F'
    if has_weekend:
        return 'Sa-Su'
    return patterns[0] if patterns else 'M-Su'


def parse_daypart(daypart: str) -> Tuple[str, str]:
    """
    Parse a 3Olives daypart string into (etere_days, etere_time_str).

    Examples:
        "M-F 7p-8p/ 11:30p-12a"    → ("M-F",  "7p-8p; 11:30p-12a")
        "M-Sun 8p-9p/ M-F 9p-10p"  → ("M-Su", "8p-9p; 9p-10p")
        "M-Sun 11a-12p"             → ("M-Su", "11a-12p")
        "Sat-Sun 4p-6p"             → ("Sa-Su", "4p-6p")
        "ROS Bonus"                 → ("M-Su", "6a-11:59p")

    Returns semicolons in time_str so EtereClient.parse_time_range() can
    apply its earliest-start + latest-end consolidation rule.
    """
    dp = _normalise_day_tokens(daypart.strip())

    if re.search(r'\bROS\b', dp, re.IGNORECASE):
        return ('M-Su', '6a-11:59p')

    day_matches = _DAY_RE.findall(dp)

    # Replace "/" with ";" to form multi-range time string
    time_str = re.sub(r'\s*/\s*', '; ', dp)
    # Strip out all day tokens so only times remain
    time_str = _DAY_RE.sub('', time_str).strip()
    time_str = re.sub(r'^[\s;]+|[\s;]+$', '', time_str)
    time_str = re.sub(r'\s*;\s*', '; ', time_str)
    time_str = re.sub(r'\s+', ' ', time_str).strip()

    etere_days = _broadest_days(day_matches) if day_matches else 'M-Su'
    return (etere_days, time_str)


# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE / BLOCK HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _block_prefixes(program: str) -> List[str]:
    p = program.lower()
    if 'cantonese' in p:
        return ['C']
    if 'mandarin' in p:
        return ['M']
    if 'chinese' in p:
        return ['C', 'M']
    if 'vietnamese' in p:
        return ['V']
    if 'korean' in p:
        return ['K']
    if 'filipino' in p or 'tagalog' in p:
        return ['T']
    return []


def _ros_key(program: str) -> Optional[str]:
    p = program.lower()
    if 'chinese' in p or 'cantonese' in p or 'mandarin' in p:
        return 'Chinese'
    if 'vietnamese' in p:
        return 'Vietnamese'
    if 'korean' in p:
        return 'Korean'
    if 'filipino' in p or 'tagalog' in p:
        return 'Filipino'
    return None


# ─────────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class ThreeOlivesLine:
    """A single program line from a 3 Olives Media insertion order."""
    program: str            # e.g. "Chinese(Cantonese) News"
    time_str: str           # raw daypart, e.g. "M-F 7p-8p/ 11:30p-12a"
    gross_rate: Decimal     # card rate (Value column) — for reference only
    rate: Decimal           # billing rate (Discounted column) — use for Etere
    weekly_spots: List[int]
    total_spots: int
    week_start_dates: List[str]   # "May 4" format, one per week
    is_bonus: bool

    def get_etere_days(self) -> str:
        days, _ = parse_daypart(self.time_str)
        return days

    def get_etere_time(self) -> str:
        _, time = parse_daypart(self.time_str)
        return time

    @property
    def days(self) -> str:
        """Parsed day pattern — exposed for the parser bridge normalizer."""
        return self.get_etere_days()

    @property
    def time(self) -> str:
        """Parsed time range — exposed for the parser bridge normalizer."""
        return self.get_etere_time()

    def get_block_prefixes(self) -> List[str]:
        return _block_prefixes(self.program)

    def get_ros_key(self) -> Optional[str]:
        return _ros_key(self.program)

    def get_description(self, etere_days: str, etere_time: str) -> str:
        prog = self.program.strip()
        if self.is_bonus:
            return f"{etere_days} {etere_time} BNS {prog}"
        return f"{etere_days} {etere_time} {prog}"


@dataclass
class ThreeOlivesOrder:
    """Complete parsed 3 Olives Media insertion order."""
    client: str
    contact: str
    email: str
    channel: str
    order_date: str       # as printed (e.g. "4/13/2026")
    flight_start: str     # MM/DD/YYYY
    flight_end: str       # MM/DD/YYYY
    lines: List[ThreeOlivesLine]
    source_sheet: str = 'Option 1'   # Excel sheet name or 'PDF'
    rates_are_net: bool = False      # Discounted rate is already gross from our perspective


# ─────────────────────────────────────────────────────────────────────────────
# DATE / RATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

_MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _dt_to_mon_d(dt: datetime.datetime) -> str:
    """datetime(2026,5,4) → "May 4" """
    return f"{_MONTH_ABBR[dt.month - 1]} {dt.day}"


def _pdf_col_to_mon_d(s: str) -> str:
    """PDF column header "4-May" → "May 4" """
    m = re.match(r'^(\d+)-([A-Za-z]+)$', s.strip())
    if m:
        return f"{m.group(2)} {m.group(1)}"
    return s


def _week_dates_to_flight(week_dates: List[str]) -> Tuple[str, str]:
    """
    Derive MM/DD/YYYY flight start and end from "Mon D" week start dates.
    Flight end = last week start + 6 days.
    """
    parsed = []
    for wd in week_dates:
        m = re.match(r'^([A-Za-z]+)\s+(\d+)$', wd.strip())
        if m:
            try:
                month = datetime.datetime.strptime(m.group(1), '%b').month
                parsed.append(datetime.datetime(2026, month, int(m.group(2))))
            except ValueError:
                pass

    if not parsed:
        return ('', '')

    starts = parsed
    flight_start = starts[0]
    flight_end = starts[-1] + datetime.timedelta(days=6)
    return (flight_start.strftime('%-m/%-d/%Y'), flight_end.strftime('%-m/%-d/%Y'))


def _parse_dollar(val) -> Decimal:
    """Parse a rate from Excel (int/float) or PDF (string with $ and space artifacts)."""
    if val is None:
        return Decimal('0')
    if isinstance(val, (int, float)):
        v = float(val)
        return Decimal(str(int(v))) if v == int(v) else Decimal(str(v))
    cleaned = re.sub(r'[$,\s]', '', str(val)).strip()
    if not cleaned or cleaned == '-':
        return Decimal('0')
    try:
        return Decimal(cleaned)
    except Exception:
        return Decimal('0')


def _clean_cell(val) -> str:
    """Normalise a table cell: join embedded newlines, strip whitespace."""
    if val is None:
        return ''
    s = str(val)
    # Newline mid-time-range: "9p-\n10p" → "9p-10p" (no space)
    s = re.sub(r'([ap])-\s*\n\s*(\d)', r'\1-\2', s, flags=re.IGNORECASE)
    # Newline mid-program-name: join with space
    s = re.sub(r'\s*\n\s*', ' ', s)
    return s.strip()


# ─────────────────────────────────────────────────────────────────────────────
# SKIP LOGIC
# ─────────────────────────────────────────────────────────────────────────────

_SKIP_SUBSTRINGS = {
    'total paid', 'total bonuses', 'total airtime', 'grand total',
    'asian groups', 'estimated impressions', 'approved by', 'discount rate',
    'translations', 'thank you',
}


def _should_skip(program: str, time_str: str) -> bool:
    pl = program.lower()
    tl = time_str.lower()
    return (
        any(k in pl for k in _SKIP_SUBSTRINGS)
        or any(k in tl for k in _SKIP_SUBSTRINGS)
        or tl in {'dma', 'riverside', ''}
    )


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────────────────────────────────────

def _parse_excel_header(rows: list) -> dict:
    result = {'client': '', 'contact': '', 'email': '', 'channel': '', 'order_date': ''}
    for row in rows:
        for cell in row:
            s = _clean_cell(cell)
            for key, pat in (
                ('client',     r'Client:\s*(.+)'),
                ('contact',    r'Contact:\s*(.+)'),
                ('email',      r'Email:\s*(.+)'),
                ('channel',    r'Channel:\s*(.+)'),
                ('order_date', r'Date:\s*(.+)'),
            ):
                m = re.match(pat, s, re.IGNORECASE)
                if m:
                    result[key] = m.group(1).strip()
    return result


def parse_threeolives_excel(path: str, sheet_name: str = 'Option 1') -> ThreeOlivesOrder:
    """
    Parse a 3 Olives Media .xlsm/.xlsx file.

    Args:
        path: Path to .xlsm or .xlsx file
        sheet_name: Sheet to parse (default "Option 1")
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    used_sheet = ws.title

    all_rows = list(ws.iter_rows(values_only=True))

    header = _parse_excel_header(all_rows[:12])

    # Find the column header row and extract week date datetimes
    header_row_idx = None
    week_dates: List[str] = []

    for i, row in enumerate(all_rows):
        if any(_clean_cell(c) == 'Insertion' for c in row):
            header_row_idx = i
            found_disc = False
            for cell in row:
                s = _clean_cell(cell)
                if s == 'Discounted':
                    found_disc = True
                    continue
                if s == 'Units':
                    break
                if found_disc and isinstance(cell, datetime.datetime):
                    week_dates.append(_dt_to_mon_d(cell))
            break

    if header_row_idx is None:
        raise ValueError(f"Could not find 'Insertion' header row in {used_sheet}")

    n_weeks = len(week_dates)
    lines: List[ThreeOlivesLine] = []

    for row in all_rows[header_row_idx + 1:]:
        if not any(c is not None for c in row):
            continue

        # Excel layout: [None, BONUS_flag, program, time, gross, disc, w1..wN, units, val, gross_total]
        is_bonus = _clean_cell(row[1]).upper() == 'BONUS'
        program = _clean_cell(row[2])
        time_str = _clean_cell(row[3])

        if not program or not time_str:
            continue
        if _should_skip(program, time_str):
            continue

        gross_rate = _parse_dollar(row[4])
        disc_rate = _parse_dollar(row[5])

        weekly = []
        for j in range(6, 6 + n_weeks):
            v = row[j] if j < len(row) else None
            weekly.append(int(v) if isinstance(v, (int, float)) and v > 0 else 0)

        units_col = 6 + n_weeks
        total_spots = int(row[units_col]) if units_col < len(row) and isinstance(row[units_col], (int, float)) else sum(weekly)

        if total_spots == 0 and sum(weekly) == 0:
            continue

        lines.append(ThreeOlivesLine(
            program=program,
            time_str=time_str,
            gross_rate=gross_rate,
            rate=disc_rate,
            weekly_spots=weekly,
            total_spots=total_spots,
            week_start_dates=week_dates[:],
            is_bonus=is_bonus,
        ))

    flight_start, flight_end = _week_dates_to_flight(week_dates)

    return ThreeOlivesOrder(
        client=header['client'],
        contact=header['contact'],
        email=header['email'],
        channel=header['channel'],
        order_date=header['order_date'],
        flight_start=flight_start,
        flight_end=flight_end,
        lines=lines,
        source_sheet=used_sheet,
    )


# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_threeolives_pdf(path: str) -> ThreeOlivesOrder:
    """Parse a 3 Olives Media PDF insertion order."""
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ''
        table = page.extract_table()

    if not table:
        raise ValueError("pdfplumber could not extract a table from the PDF")

    def _field(pat: str) -> str:
        m = re.search(pat, text, re.IGNORECASE)
        return m.group(1).strip() if m else ''

    client = _field(r'Client:\s*(.+)')
    contact = _field(r'Contact:\s*(.+)')
    email = _field(r'Email:\s*(.+)')
    channel = _field(r'Channel:\s*(.+)')
    order_date = _field(r'Date:\s*(\S+)')

    # Locate column header row and collect week date strings
    header_row_idx = None
    week_dates: List[str] = []

    for i, row in enumerate(table):
        if any(_clean_cell(c) == 'Insertion' for c in row):
            header_row_idx = i
            found_disc = False
            for cell in row:
                s = _clean_cell(cell)
                if s == 'Discounted':
                    found_disc = True
                    continue
                if s == 'Units':
                    break
                if found_disc and s and s not in ('Value', 'GROSS'):
                    week_dates.append(_pdf_col_to_mon_d(s))
            break

    if header_row_idx is None:
        raise ValueError("Could not find 'Insertion' header row in PDF table")

    n_weeks = len(week_dates)
    lines: List[ThreeOlivesLine] = []

    for row in table[header_row_idx + 1:]:
        if not any(c for c in row):
            continue

        # PDF layout: [BONUS_flag_or_None, program, time, gross, disc, w1..wN, units, val, gross_total]
        is_bonus = _clean_cell(row[0]).upper() == 'BONUS'
        program = _clean_cell(row[1])
        time_str = _clean_cell(row[2])

        if not program or not time_str:
            continue
        if _should_skip(program, time_str):
            continue

        gross_rate = _parse_dollar(_clean_cell(row[3]))
        disc_raw = _clean_cell(row[4]) if len(row) > 4 else ''
        disc_rate = _parse_dollar(disc_raw)

        weekly = []
        for j in range(5, 5 + n_weeks):
            v = _clean_cell(row[j]) if j < len(row) else '0'
            try:
                weekly.append(int(v))
            except (ValueError, TypeError):
                weekly.append(0)

        units_col = 5 + n_weeks
        try:
            total_spots = int(_clean_cell(row[units_col])) if units_col < len(row) else sum(weekly)
        except (ValueError, TypeError):
            total_spots = sum(weekly)

        if total_spots == 0 and sum(weekly) == 0:
            continue

        lines.append(ThreeOlivesLine(
            program=program,
            time_str=time_str,
            gross_rate=gross_rate,
            rate=disc_rate,
            weekly_spots=weekly,
            total_spots=total_spots,
            week_start_dates=week_dates[:],
            is_bonus=is_bonus,
        ))

    flight_start, flight_end = _week_dates_to_flight(week_dates)

    return ThreeOlivesOrder(
        client=client,
        contact=contact,
        email=email,
        channel=channel,
        order_date=order_date,
        flight_start=flight_start,
        flight_end=flight_end,
        lines=lines,
        source_sheet='PDF',
    )


# ─────────────────────────────────────────────────────────────────────────────
# DISPATCH
# ─────────────────────────────────────────────────────────────────────────────

def parse_threeolives(path: str, sheet_name: str = 'Option 1') -> ThreeOlivesOrder:
    """Parse a 3 Olives Media order from either a PDF or Excel file."""
    ext = Path(path).suffix.lower()
    if ext in {'.xlsx', '.xlsm'}:
        return parse_threeolives_excel(path, sheet_name=sheet_name)
    elif ext == '.pdf':
        return parse_threeolives_pdf(path)
    else:
        raise ValueError(f"Unsupported format: {ext} — expected .pdf, .xlsx, or .xlsm")


# ─────────────────────────────────────────────────────────────────────────────
# STANDALONE TEST
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys as _sys

    path = _sys.argv[1] if len(_sys.argv) > 1 else \
        'incoming/3Olives Media_2026 Primary_Riverside CountyVoters.pdf'
    sheet = _sys.argv[2] if len(_sys.argv) > 2 else 'Option 1'

    order = parse_threeolives(path, sheet_name=sheet)

    print(f"\n3 OLIVES MEDIA ORDER  [{order.source_sheet}]")
    print('=' * 70)
    print(f"Client:  {order.client}")
    print(f"Contact: {order.contact}")
    print(f"Email:   {order.email}")
    print(f"Channel: {order.channel}")
    print(f"Date:    {order.order_date}")
    print(f"Flight:  {order.flight_start} – {order.flight_end}")
    print(f"Lines:   {len(order.lines)}")
    print()

    paid_total = sum(l.total_spots for l in order.lines if not l.is_bonus)
    bonus_total = sum(l.total_spots for l in order.lines if l.is_bonus)
    cost_total = sum(l.rate * l.total_spots for l in order.lines)

    print(f"Paid spots:  {paid_total}")
    print(f"Bonus spots: {bonus_total}")
    print(f"Total cost:  ${cost_total:,.2f}")
    print()

    for line in order.lines:
        days, time = parse_daypart(line.time_str)
        tag = '[BONUS] ' if line.is_bonus else '        '
        print(f"{tag}{line.program!r}")
        print(f"        Days={days}  Time={time!r}")
        print(f"        Rate=${line.rate}  Spots={line.total_spots}  Weekly={line.weekly_spots}")
        print(f"        Blocks={line.get_block_prefixes()}")
