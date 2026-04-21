"""
Polaris Media Group Order Parser

Parses Excel (.xlsx) insertion orders from Polaris Media Group for Crossings TV.

Expected xlsx structure (Sheet: "Crossings TV"):
  Header block (anywhere near the top):
    Label rows: col C = label, col D = value
      AGENCY              → "Polaris"
      Advertiser          → advertiser name
      PREPARED BY:        → preparer name
      Flight Date:        → e.g. "4/16 THROUGH 4/20"
      TOTAL GROSS BUDGET: → numeric budget

  Column-header row: col C = "Media /MARKET", col D = "DAYS", col E = "Time", etc.
  Data rows begin immediately after the column-header row.

Data row columns (0-indexed):
  2: Market header cell — only present on first row of each market section,
     e.g. "CROSSINGS TV                          SAN FRANCISCO"
  3: Days  (e.g. "M-F", "Sat", "Sat- Sun ", "Sat& Sun")
  4: Time range  (e.g. "6a-7a", "7p-7:30p", "11:30p-12a")
  5: Programming / program name  (e.g. "Mandarin News", "Cantonese Talk")
  6: Gross Rate per :30s  (numeric) — "TOTAL " string signals end of data
  7: Units / total spot count  (integer)
  8: Gross cost  (for verification only)
"""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import List
import re
import sys
from pathlib import Path

_project_root = Path(__file__).parent.parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from browser_automation.day_utils import to_etere


# ─────────────────────────────────────────────────────────────────────────────
# MARKET DETECTION
# ─────────────────────────────────────────────────────────────────────────────

_MARKET_KEYWORDS = [
    ("SAN FRANCISCO", "SFO"),
    ("SACRAMENTO",    "CVC"),
    ("SEATTLE",       "SEA"),
    ("LOS ANGELES",   "LAX"),
    ("CHICAGO",       "CMP"),
    ("HOUSTON",       "HOU"),
    ("WASHINGTON",    "WDC"),
    ("NEW YORK",      "NYC"),
    ("DALLAS",        "DAL"),
]


def _detect_market(cell_text: str) -> str:
    """Extract market code from a market header cell like 'CROSSINGS TV   SAN FRANCISCO'."""
    upper = cell_text.upper()
    for keyword, code in _MARKET_KEYWORDS:
        if keyword in upper:
            return code
    print(f"[POLARIS PARSER] ⚠ Unrecognised market cell: {cell_text!r} — defaulting to SFO")
    return "SFO"


# ─────────────────────────────────────────────────────────────────────────────
# FLIGHT DATE PARSING
# ─────────────────────────────────────────────────────────────────────────────

def _parse_flight_dates(raw: str) -> tuple[str, str]:
    """
    Parse "4/16 THROUGH 4/20" → ("4/16/<year>", "4/20/<year>").

    Year is inferred as the current year.  If the computed end date has
    already passed, advances to the following year.
    """
    parts = re.findall(r'\d{1,2}/\d{1,2}', raw)
    today = date.today()
    year = today.year

    if len(parts) < 2:
        return (f"{today.month}/{today.day}/{year}", f"{today.month}/{today.day}/{year}")

    def _to_date(md: str, yr: int) -> date:
        m, d = md.split('/')
        return date(yr, int(m), int(d))

    start = _to_date(parts[0], year)
    end   = _to_date(parts[1], year)

    if end < today:
        year += 1
        start = _to_date(parts[0], year)
        end   = _to_date(parts[1], year)

    return (f"{start.month}/{start.day}/{year}", f"{end.month}/{end.day}/{year}")



# ─────────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class PolarisLine:
    """Single line item from a Polaris insertion order."""
    days: str           # Etere-format day pattern, e.g. "M-F", "Sa", "Sa-Su"
    time_str: str       # Raw time string as printed, e.g. "7p-7:30p"
    program: str        # Program name, e.g. "Mandarin News"
    rate: Decimal       # Gross rate per :30s spot
    total_spots: int    # Total spots for the flight (the "Unit" column)
    market: str         # Market code, e.g. "SFO"

    @property
    def is_bonus(self) -> bool:
        return self.rate == Decimal("0") and self.total_spots > 0

    def get_time_from_to(self) -> tuple[str, str]:
        """Return (time_from, time_to) in HH:MM 24-hour format."""
        from browser_automation.etere_client import EtereClient
        return EtereClient.parse_time_range(self.time_str)

    def get_description(self) -> str:
        """Build Etere line description: '[BNS ]Days Program'."""
        label = "BNS " if self.is_bonus else ""
        return f"{label}{self.days} {self.program}"


@dataclass
class PolarisOrder:
    """Complete Polaris insertion order parsed from xlsx."""
    advertiser: str
    prepared_by: str
    flight_start: str   # M/D/YYYY
    flight_end: str     # M/D/YYYY
    gross_budget: Decimal
    lines: List[PolarisLine]

    @property
    def markets(self) -> List[str]:
        """Unique markets in order-of-appearance."""
        seen: dict[str, None] = {}
        for ln in self.lines:
            seen[ln.market] = None
        return list(seen)

    @property
    def total_spots(self) -> int:
        return sum(ln.total_spots for ln in self.lines)

    def lines_for_market(self, market: str) -> List[PolarisLine]:
        return [ln for ln in self.lines if ln.market == market]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_polaris_pdf(path: str) -> PolarisOrder:
    """
    Parse a Polaris insertion order PDF file.

    The PDF layout mirrors the Excel format: a header block followed by a
    table with columns Media/MARKET | DAYS | Time | Programming |
    Gross Rate per :30s | Unit | GROSS.
    """
    import pdfplumber

    print(f"\n[POLARIS PARSER] Reading PDF: {path}")

    with pdfplumber.open(str(path)) as pdf:
        page = pdf.pages[0]
        raw_text = page.extract_text() or ""
        tables = page.extract_tables()

    # ── Header fields from raw text ────────────────────────────────────────
    advertiser   = ""
    prepared_by  = ""
    flight_raw   = ""
    gross_budget = Decimal("0")

    for line in raw_text.splitlines():
        stripped = line.strip()
        upper = stripped.upper()
        if upper.startswith("ADVERTISER"):
            advertiser = re.sub(r"(?i)^ADVERTISER\s*", "", stripped).strip()
        elif upper.startswith("PREPARED BY"):
            prepared_by = re.sub(r"(?i)^PREPARED BY:?\s*", "", stripped).strip()
        elif upper.startswith("FLIGHT DATE"):
            flight_raw = re.sub(r"(?i)^FLIGHT DATE:?\s*", "", stripped).strip()
        elif "TOTAL GROSS BUDGET" in upper:
            m = re.search(r"[\$]?([\d,]+\.?\d*)", stripped)
            if m:
                try:
                    gross_budget = Decimal(m.group(1).replace(",", "")).quantize(
                        Decimal("0.01"), ROUND_HALF_UP
                    )
                except Exception:
                    pass

    flight_start, flight_end = _parse_flight_dates(flight_raw)

    print(f"[POLARIS PARSER] Advertiser:  {advertiser}")
    print(f"[POLARIS PARSER] Prepared by: {prepared_by}")
    print(f"[POLARIS PARSER] Flight:      {flight_start} – {flight_end}")
    print(f"[POLARIS PARSER] Budget:      ${gross_budget:,}")

    # ── Table rows ─────────────────────────────────────────────────────────
    lines: List[PolarisLine] = []
    current_market = "SFO"

    if not tables:
        raise ValueError(f"No tables found in Polaris PDF: {path}")

    table = tables[0]
    for row in table:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        cells = [str(c or "").strip() for c in row]

        # Market header cell (col 0) — update market, but still process row data
        if cells[0] and "CROSSINGS TV" in cells[0].upper():
            current_market = _detect_market(cells[0])

        # Column-header row
        if cells[1].upper() in ("DAYS", "DAY"):
            continue

        # TOTAL row signals end
        if "TOTAL" in cells[4].upper() or "TOTAL" in cells[1].upper():
            break

        days_str  = cells[1]
        time_str  = cells[2]
        units_str = cells[5]

        if not days_str or not time_str:
            continue

        # Program name may overflow into the rate cell — recombine and extract
        program = cells[3]
        rate_cell = cells[4]
        m_rate = re.search(r'\$\s*([\d,]+(?:\.\d+)?)', rate_cell)
        if not m_rate:
            continue
        if not program:
            continue
        overflow = rate_cell[:m_rate.start()].strip()
        if overflow:
            program = program + overflow

        try:
            rate = Decimal(m_rate.group(1).replace(",", "")).quantize(
                Decimal("0.01"), ROUND_HALF_UP
            )
        except Exception:
            continue

        try:
            total_spots = int(units_str.replace(",", "").strip() or "0")
        except (TypeError, ValueError):
            continue

        if total_spots <= 0:
            continue

        days = to_etere(days_str)
        line = PolarisLine(
            days=days,
            time_str=time_str,
            program=program,
            rate=rate,
            total_spots=total_spots,
            market=current_market,
        )
        lines.append(line)

        rate_label = "BONUS" if line.is_bonus else f"${rate}"
        print(f"[POLARIS PARSER]   {days:<8s}  {time_str:<15s}  "
              f"{program:<30s}  {rate_label}/spot  {total_spots} spots  [{current_market}]")

    if not lines:
        raise ValueError(f"No lines parsed from Polaris PDF: {path}")

    print(f"[POLARIS PARSER] Total: {len(lines)} lines, "
          f"{sum(l.total_spots for l in lines)} spots")

    return PolarisOrder(
        advertiser=advertiser,
        prepared_by=prepared_by,
        flight_start=flight_start,
        flight_end=flight_end,
        gross_budget=gross_budget,
        lines=lines,
    )


def parse_polaris_file(path: str) -> PolarisOrder:
    """Dispatch to PDF or xlsx parser based on file extension."""
    if str(path).lower().endswith(".pdf"):
        return parse_polaris_pdf(path)
    return parse_polaris_xlsx(path)


def parse_polaris_xlsx(path: str) -> PolarisOrder:
    """
    Parse a Polaris insertion order Excel file.

    Header rows are located by scanning for their label text (col C) rather
    than by hardcoded row indices, so the parser is robust to leading blank
    rows or minor layout shifts.

    Args:
        path: Path to the .xlsx file.

    Returns:
        PolarisOrder with all line items populated.

    Raises:
        ValueError: If no lines can be parsed or required fields are missing.
    """
    import openpyxl

    print(f"\n[POLARIS PARSER] Reading: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["Crossings TV"] if "Crossings TV" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # ── Scan header rows by label ──────────────────────────────────────────
    advertiser   = ""
    prepared_by  = ""
    flight_raw   = ""
    gross_budget = Decimal("0")
    data_start   = len(rows)  # will be updated when column-header row is found

    for i, row in enumerate(rows):
        label = str(row[2] or "").strip().upper()
        value = row[3]

        if label == "ADVERTISER":
            advertiser = str(value or "").strip()
        elif label in ("PREPARED BY:", "PREPARED BY"):
            prepared_by = str(value or "").strip()
        elif label in ("FLIGHT DATE:", "FLIGHT DATE"):
            flight_raw = str(value or "").strip()
        elif "TOTAL GROSS BUDGET" in label:
            try:
                gross_budget = Decimal(str(value or 0)).quantize(
                    Decimal("0.01"), ROUND_HALF_UP
                )
            except Exception:
                pass
        elif label in ("MEDIA /MARKET", "MEDIA/MARKET") and str(value or "").strip().upper() == "DAYS":
            # This is the column-header row; data begins on the next row
            data_start = i + 1
            break

    flight_start, flight_end = _parse_flight_dates(flight_raw)

    print(f"[POLARIS PARSER] Advertiser:  {advertiser}")
    print(f"[POLARIS PARSER] Prepared by: {prepared_by}")
    print(f"[POLARIS PARSER] Flight:      {flight_start} – {flight_end}")
    print(f"[POLARIS PARSER] Budget:      ${gross_budget:,}")

    # ── Data rows ─────────────────────────────────────────────────────────
    lines: List[PolarisLine] = []
    current_market = "SFO"  # fallback if no market header is seen

    for row in rows[data_start:]:
        # Col 2: optional market header
        market_cell = str(row[2] or "").strip()
        if market_cell and "CROSSINGS TV" in market_cell.upper():
            current_market = _detect_market(market_cell)

        # Col 3: days — skip blank or label rows
        days_raw = row[3]
        if days_raw is None:
            continue
        days_str = str(days_raw).strip()
        if not days_str or days_str.upper() in ("DAYS",):
            continue

        # Col 6: rate — "TOTAL" string signals end of data
        rate_raw = row[6]
        if isinstance(rate_raw, str) and "TOTAL" in rate_raw.upper():
            break

        # Col 4: time
        time_raw = row[4]
        if time_raw is None:
            continue
        time_str = str(time_raw).strip()
        if not time_str:
            continue

        # Col 5: program name
        program = str(row[5] or "").strip()
        if not program:
            continue

        # Col 6: rate (numeric)
        try:
            rate = Decimal(str(rate_raw or 0)).quantize(Decimal("0.01"), ROUND_HALF_UP)
        except Exception:
            continue

        # Col 7: units / total spots
        try:
            total_spots = int(row[7] or 0)
        except (TypeError, ValueError):
            continue

        if total_spots <= 0:
            continue

        days = to_etere(days_str)
        line = PolarisLine(
            days=days,
            time_str=time_str,
            program=program,
            rate=rate,
            total_spots=total_spots,
            market=current_market,
        )
        lines.append(line)

        rate_label = "BONUS" if line.is_bonus else f"${rate}"
        print(f"[POLARIS PARSER]   {days:<8s}  {time_str:<15s}  "
              f"{program:<30s}  {rate_label}/spot  {total_spots} spots  [{current_market}]")

    if not lines:
        raise ValueError(f"No lines parsed from Polaris xlsx: {path}")

    print(f"[POLARIS PARSER] Total: {len(lines)} lines, "
          f"{sum(l.total_spots for l in lines)} spots")

    return PolarisOrder(
        advertiser=advertiser,
        prepared_by=prepared_by,
        flight_start=flight_start,
        flight_end=flight_end,
        gross_budget=gross_budget,
        lines=lines,
    )


# ─────────────────────────────────────────────────────────────────────────────
# STANDALONE TEST
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys as _sys

    if len(_sys.argv) < 2:
        print("Usage: python polaris_parser.py <xlsx_path>")
        _sys.exit(1)

    try:
        order = parse_polaris_file(_sys.argv[1])

        print("\n" + "=" * 70)
        print("POLARIS ORDER SUMMARY")
        print("=" * 70)
        print(f"Advertiser:  {order.advertiser}")
        print(f"Prepared by: {order.prepared_by}")
        print(f"Flight:      {order.flight_start} – {order.flight_end}")
        print(f"Budget:      ${order.gross_budget:,}")
        print(f"Markets:     {order.markets}")
        print(f"Total Lines: {len(order.lines)}")
        print(f"Total Spots: {order.total_spots}")

        print("\n" + "=" * 70)
        print("LINES")
        print("=" * 70)
        for ln in order.lines:
            tf, tt = ln.get_time_from_to()
            print(f"\n  [{ln.market}] {ln.days}  {ln.time_str}  →  {tf}–{tt}")
            print(f"  Program:  {ln.program}")
            print(f"  Rate:     {'BONUS' if ln.is_bonus else f'${ln.rate}'}")
            print(f"  Spots:    {ln.total_spots}")
            print(f"  Desc:     {ln.get_description()}")

    except Exception as exc:
        print(f"\n✗ Error: {exc}")
        import traceback
        traceback.print_exc()
        _sys.exit(1)
