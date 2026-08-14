"""
Ntooitive order parser (advertiser: L.A. Care).

Ntooitive is an AGENCY parser: the agency is fixed (Ntooitive → ANAGRAF agency
299, Commissione 15%) and the advertiser is looked up in ANAGRAF via the agency
link (LA Care Health Plan = 300 / AGENZIA 299).

The source document is Charmaine's house "Crossings TV Media Proposal" template
(she is the AE), sent back by the agency as the order. TWO source formats, one
`NtooitiveOrder` shape — use `parse_ntooitive(path)`:

1. **Proposal workbook** (.xlsx/.xlsm) — `parse_ntooitive_xlsx`. The workbook
   can carry SEVERAL "Option" sheets (competing proposals); only one is the
   order. `list_ntooitive_options()` describes each grid-bearing sheet so the
   gather can prompt; the parser defaults to the sheet with the latest
   revised/proposed date. Grid columns (mapped by HEADER LABEL, never index):
   Language Block | Day Part/Program | Spot Type | Length | <week dates as real
   datetimes> | Total Unit # | Promo Unit Cost (Gross) | Line Total Cost
   (Gross) | Line Total Cost (NET)

2. **PDF print of the workbook** (.pdf) — `parse_ntooitive_pdf`. Read by word
   coordinates (SAGENT lesson: never text-flow a columnar grid); week cells are
   matched to the week-label header by x-centre distance.

**Rates are GROSS.** The NET column is gross × (1 − commission); the ANAGRAF
agency commission nets the entered gross back down at billing (Crispin lesson:
never multiply rates in a parser — if the money basis looks wrong, the fix is
ANAGRAF). `rates_are_net` is always False here.

**Translation fees never become a contract line.** The header block's
"Gross Translation Fees" money becomes an `NtooitiveCharge` that the automation
puts in the first paid line's Production box (→ CONTRATTISPESE 'Production').

Reconciliation (the document is its own oracle — raise, never enter short):
  - per line: sum(week cells) == Total Unit #
  - per paid line: rate × units == Line Total Cost (Gross)
  - footer Total Paid / Total Bonuses unit counts
  - sum(paid line gross) == footer gross == header "Gross (Airtime)"
  - header "Gross Amount of Contract" == airtime + charges
  - the implied commission (1 − net/gross) must be identical on every paid line
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional

import openpyxl

# ─── Market detection (label from the "Market:" header cell) ─────────────────
# Beware: the live template misspells Spectrum ("Sepctrum") — match city names.
_MARKET_MAP = [
    (("los angeles", "spectrum 1519", "lax"), "LAX"),
    (("san francisco", "bay area", "sfo"), "SFO"),
    (("sacramento", "central valley", "cvc"), "CVC"),
    (("seattle", "sea"), "SEA"),
    (("houston", "hou"), "HOU"),
    (("chicago", "minneapolis", "cmp"), "CMP"),
    (("washington", "wdc"), "WDC"),
    (("new york", "nyc"), "NYC"),
]

_BASE_LANGUAGES = ["Cantonese", "Mandarin", "Filipino", "Vietnamese",
                   "Korean", "Hmong", "Punjabi", "Japanese", "Hindi",
                   "South Asian", "Chinese"]

# One time range, e.g. "6a-7a", "8:30p-9p". A daypart may contain several
# (dual-window news buys: "M-F 6a-7a & 8p-9p").
_TIME_RE = re.compile(
    r"(\d{1,2}(?::\d{2})?\s*[apAP]?\s*[-–]\s*\d{1,2}(?::\d{2})?\s*[apAP])"
)


def find_time_ranges(daypart: str) -> list[str]:
    """All time ranges in a daypart string, in print order."""
    return [m.group(1).strip() for m in _TIME_RE.finditer(daypart or "")]


def _money(v) -> Optional[float]:
    """Money cell → float, or None when unreadable. Never silently 0 (DART
    lesson): a zero is a legitimate value (bonus), so it cannot double as
    'could not read it'. '-' is Excel's accounting-format zero."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if s in ("-", "–"):
        return 0.0
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _norm_label(v) -> str:
    """Normalise a header/label cell for matching: lowercase, collapse
    whitespace (labels wrap with newlines inside cells)."""
    return " ".join(str(v or "").split()).lower()


def _market_code(label: str) -> str:
    low = (label or "").lower()
    for keys, code in _MARKET_MAP:
        if any(k in low for k in keys):
            return code
    return ""


# ─── Data model ──────────────────────────────────────────────────────────────

@dataclass
class NtooitiveLine:
    language_block: str          # "Mandarin News", "Korean" (bonus)
    daypart: str                 # "M-F 6a-7a & 8p-9p", "Chinese ROS"
    spot_type: str               # "COM" | "BONUS"
    rate: float                  # GROSS per-spot rate (0 ⇒ bonus)
    length_sec: int
    week_dates: List[date]       # Monday of each flight week
    week_spots: List[int]        # spots per week (parallel to week_dates)
    gross_total: Optional[float] = None   # the sheet's own Line Total (Gross)
    net_total: Optional[float] = None     # the sheet's own Line Total (NET)

    @property
    def is_bonus(self) -> bool:
        return round(self.rate, 4) == 0.0

    @property
    def total_spots(self) -> int:
        return sum(self.week_spots)

    # Aliases so the generic parser_bridge normalizer (web preview) picks these up.
    @property
    def weekly_spots(self) -> List[int]:
        return self.week_spots

    @property
    def length(self) -> int:
        return self.length_sec

    @property
    def base_language(self) -> str:
        """Normalised language for ROS/desc mapping. A bonus row's daypart
        ('Chinese ROS') names the block to use; the Language Block cell may
        say 'Mandarin' while the ROS runs in the whole Chinese block."""
        for src in ((self.daypart if self.is_bonus else ""), self.language_block):
            low = (src or "").strip().lower()
            for lang in _BASE_LANGUAGES:
                if low.startswith(lang.lower()):
                    return lang
        return self.language_block.strip()


@dataclass
class NtooitiveCharge:
    """Non-airtime money from the header block (translation fees). Enters as a
    CONTRATTISPESE 'Production' charge on the first paid line, never a line."""
    description: str
    amount: float


@dataclass
class NtooitiveOrder:
    agency: str
    advertiser: str
    contact: str
    email: str
    market_code: str
    market_label: str
    order_date: Optional[date]           # REVISED / Date Proposed
    option_label: str                    # sheet title ("Option 1") or "PDF"
    billing_cycle: str = ""
    flight_start_text: str = ""          # header "Flight schedule" start, e.g. "8/18"
    lines: List[NtooitiveLine] = field(default_factory=list)
    charges: List[NtooitiveCharge] = field(default_factory=list)
    rates_are_net: bool = False          # GROSS quoted — always False
    implied_commission: float = 0.0      # 1 − net/gross from the paid lines
    gross_airtime: Optional[float] = None
    gross_contract: Optional[float] = None
    source_format: str = "xlsx"          # 'xlsx' | 'pdf'
    source_path: str = ""
    _flight_start_date: Optional[date] = None
    _flight_end_date: Optional[date] = None

    @property
    def paid_lines(self) -> List[NtooitiveLine]:
        return [ln for ln in self.lines if not ln.is_bonus]

    @property
    def bonus_lines(self) -> List[NtooitiveLine]:
        return [ln for ln in self.lines if ln.is_bonus]

    @property
    def week_dates(self) -> List[date]:
        seen = {d for ln in self.lines for d in ln.week_dates}
        return sorted(seen)

    @property
    def flight_start(self) -> Optional[str]:
        """The header's flight start (e.g. 8/18 on a week grid that opens
        Monday 8/17) — this is what feeds the start-date prompt and the
        max-per-day planner. Falls back to the first week column."""
        if self._flight_start_date:
            return self._flight_start_date.strftime("%m/%d/%Y")
        wd = self.week_dates
        return wd[0].strftime("%m/%d/%Y") if wd else None

    @property
    def flight_end(self) -> Optional[str]:
        if self._flight_end_date:
            return self._flight_end_date.strftime("%m/%d/%Y")
        wd = self.week_dates
        return (wd[-1] + timedelta(days=6)).strftime("%m/%d/%Y") if wd else None


# ─── Header key/value labels (both formats) ──────────────────────────────────

_HEADER_LABELS = {
    "media buying agency": "agency",
    "advertiser": "advertiser",
    "contact": "contact",
    "email": "email",
    "billing cycle": "billing_cycle",
    "market": "market",
    "flight schedule": "flight",
    "revised": "revised",
    "date proposed": "proposed",
    "gross (airtime)": "gross_airtime",
    "gross translation fees": "translation",
    "gross amount of contract": "gross_contract",
}


def _match_header_label(label: str) -> Optional[str]:
    for key, name in _HEADER_LABELS.items():
        if label.startswith(key):
            return name
    return None


def _resolve_flight_dates(flight_text: str,
                          week_dates: List[date]) -> tuple[Optional[date], Optional[date]]:
    """Parse the header's year-less '8/18 -11/30' against the week grid.

    The year comes from the week columns (real dates), trying the grid's own
    year and its neighbours so a flight crossing New Year still resolves.
    """
    if not flight_text or not week_dates:
        return None, None
    pairs = re.findall(r"(\d{1,2})\s*/\s*(\d{1,2})", flight_text)
    if len(pairs) < 2:
        return None, None
    grid_start, grid_end = week_dates[0], week_dates[-1] + timedelta(days=6)

    def _fit(m: int, d: int, lo: date, hi: date) -> Optional[date]:
        for y in (lo.year - 1, lo.year, lo.year + 1):
            try:
                cand = date(y, m, d)
            except ValueError:
                continue
            if lo - timedelta(days=7) <= cand <= hi + timedelta(days=7):
                return cand
        return None

    start = _fit(int(pairs[0][0]), int(pairs[0][1]), grid_start, grid_end)
    end = _fit(int(pairs[-1][0]), int(pairs[-1][1]), grid_start, grid_end)
    return start, end


# ─── Workbook option discovery ───────────────────────────────────────────────

def _find_grid_header(rows: list[tuple]) -> Optional[int]:
    """Row index of the column-header row ('Language Block' + 'Total Unit')."""
    for ri, r in enumerate(rows):
        labels = [_norm_label(c) for c in r if c is not None]
        if any(lbl.startswith("language block") for lbl in labels) and \
           any(lbl.startswith("total unit") for lbl in labels):
            return ri
    return None


def list_ntooitive_options(path: str) -> list[dict]:
    """Describe every grid-bearing sheet: {'sheet', 'date', 'flight', 'gross'}.

    Lets the gather prompt which Option to enter; `parse_ntooitive_xlsx`
    defaults to the sheet with the latest revised/proposed date."""
    wb = openpyxl.load_workbook(path, data_only=True)
    options: list[dict] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if _find_grid_header(rows) is None:
            continue
        hdr = _header_pairs(rows)
        options.append({
            "sheet": ws.title,
            "date": _to_date(hdr.get("revised")) or _to_date(hdr.get("proposed")),
            "flight": str(hdr.get("flight", "") or "").strip(),
            "gross": _money(hdr.get("gross_contract")) or _money(hdr.get("gross_airtime")),
        })
    wb.close()
    return options


def _header_pairs(rows: list[tuple]) -> dict:
    """Label→value pairs from the header block (value = next non-empty cell).
    Only the FIRST occurrence of a label counts — the airtime summary further
    down reuses words like 'Gross Investment'."""
    hdr: dict = {}
    for r in rows:
        for i, c in enumerate(r):
            if c is None:
                continue
            name = _match_header_label(_norm_label(c))
            if name and name not in hdr:
                for c2 in r[i + 1:]:
                    if c2 is not None and str(c2).strip():
                        hdr[name] = c2
                        break
    return hdr


# ─── Workbook parser ─────────────────────────────────────────────────────────

def parse_ntooitive_xlsx(path: str, sheet_name: Optional[str] = None) -> NtooitiveOrder:
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Pick the sheet: explicit > latest revised/proposed date > first grid ──
    grid_sheets = [ws for ws in wb.worksheets
                   if _find_grid_header(list(ws.iter_rows(values_only=True))) is not None]
    if not grid_sheets:
        raise ValueError("Ntooitive parser: no sheet with a 'Language Block' grid found")
    if sheet_name:
        matches = [ws for ws in grid_sheets if ws.title == sheet_name]
        if not matches:
            raise ValueError(f"Ntooitive parser: sheet '{sheet_name}' not found "
                             f"(grid sheets: {[w.title for w in grid_sheets]})")
        ws = matches[0]
    elif len(grid_sheets) == 1:
        ws = grid_sheets[0]
    else:
        def _sheet_date(w) -> date:
            hdr = _header_pairs(list(w.iter_rows(values_only=True)))
            return (_to_date(hdr.get("revised")) or _to_date(hdr.get("proposed"))
                    or date.min)
        ws = max(grid_sheets, key=_sheet_date)

    rows = list(ws.iter_rows(values_only=True))
    hdr = _header_pairs(rows)
    header_ri = _find_grid_header(rows)

    # ── Column map by HEADER LABEL (DART rule: added column = no-op, renamed
    # column = loud failure) ──
    col: dict[str, int] = {}
    week_cols: List[int] = []
    week_dates: List[date] = []
    for ci, c in enumerate(rows[header_ri]):
        d = _to_date(c)
        if d is not None:
            week_cols.append(ci)
            week_dates.append(d)
            continue
        lbl = _norm_label(c)
        if not lbl:
            continue
        if lbl.startswith("language block"):
            col["lang"] = ci
        elif lbl.startswith("day part"):
            col["daypart"] = ci
        elif lbl.startswith("spot type"):
            col["spot_type"] = ci
        elif lbl == "length":
            col["length"] = ci
        elif lbl.startswith("total unit"):
            col["total_units"] = ci
        elif lbl.startswith("promo unit cost"):
            # The money basis is part of the label — a template that flips to
            # NET must refuse here, not enter 17.6% light (Crispin lesson).
            if "(gross)" not in lbl:
                raise ValueError(f"Ntooitive parser: rate column is not gross: {lbl!r}")
            col["rate"] = ci
        elif lbl.startswith("line total cost (gross)"):
            col["gross"] = ci
        elif lbl.startswith("line total cost (net)"):
            col["net"] = ci

    for req in ("lang", "daypart", "spot_type", "length", "total_units", "rate", "gross"):
        if req not in col:
            raise ValueError(f"Ntooitive parser: missing '{req}' column in header row "
                             f"— the template changed; refusing to guess positions")
    if not week_cols:
        raise ValueError("Ntooitive parser: no weekly date columns in the header row")

    # ── Airtime rows (until the Total Paid / Total Bonuses footers) ──
    lines: List[NtooitiveLine] = []
    footer: dict[str, Optional[float]] = {}
    for r in rows[header_ri + 1:]:
        joined = _norm_label(" ".join(str(c) for c in r if c is not None))
        if joined.startswith("total paid"):
            footer["paid_units"] = _money(r[col["total_units"]]) if col["total_units"] < len(r) else None
            footer["paid_gross"] = _money(r[col["gross"]]) if col["gross"] < len(r) else None
            continue
        if joined.startswith("total bonus"):
            footer["bonus_units"] = _money(r[col["total_units"]]) if col["total_units"] < len(r) else None
            break

        spot_type = str(r[col["spot_type"]] or "").strip().upper() if col["spot_type"] < len(r) else ""
        if spot_type not in ("COM", "BONUS"):
            continue
        lang = str(r[col["lang"]] or "").strip() if col["lang"] < len(r) else ""
        daypart = " ".join(str(r[col["daypart"]] or "").split()) if col["daypart"] < len(r) else ""
        if not lang:
            continue

        rate = _money(r[col["rate"]]) if col["rate"] < len(r) else None
        if spot_type == "COM":
            if not rate:
                raise ValueError(f"Ntooitive parser: paid line '{lang} {daypart}' has "
                                 f"no readable rate — refusing to enter $0 airtime")
        else:
            rate = rate or 0.0
            if rate:
                raise ValueError(f"Ntooitive parser: BONUS line '{lang}' carries a "
                                 f"rate of ${rate} — contradictory row; refusing")

        m = re.search(r"(\d+)", str(r[col["length"]] or ""))
        length_sec = int(m.group(1)) if m else 30
        spots = [int(_money(r[ci]) or 0) if ci < len(r) else 0 for ci in week_cols]

        lines.append(NtooitiveLine(
            language_block=lang,
            daypart=daypart,
            spot_type=spot_type,
            rate=float(rate),
            length_sec=length_sec,
            week_dates=list(week_dates),
            week_spots=spots,
            gross_total=_money(r[col["gross"]]) if col["gross"] < len(r) else None,
            net_total=_money(r[col["net"]]) if "net" in col and col["net"] < len(r) else None,
        ))

    if not lines:
        raise ValueError("Ntooitive parser: no airtime lines found")

    order = _assemble_order(
        hdr=hdr, lines=lines, week_dates=week_dates, footer=footer,
        option_label=ws.title, source_format="xlsx", source_path=path,
    )
    wb.close()

    # Per-line Total Unit # check needs the raw cells — do it here where we
    # still know the column.
    for ln, r in zip(lines, _data_rows(rows, header_ri, col)):
        stated_units = _money(r[col["total_units"]]) if col["total_units"] < len(r) else None
        if stated_units is not None and int(stated_units) != ln.total_spots:
            raise ValueError(
                f"Ntooitive parser: line '{ln.language_block} {ln.daypart}' week "
                f"cells sum to {ln.total_spots} but Total Unit # says "
                f"{int(stated_units)} — a cell was dropped; refusing to enter")
    return order


def _data_rows(rows: list[tuple], header_ri: int, col: dict) -> list[tuple]:
    """The airtime rows, in the same order `parse_ntooitive_xlsx` read them."""
    out = []
    for r in rows[header_ri + 1:]:
        joined = _norm_label(" ".join(str(c) for c in r if c is not None))
        if joined.startswith("total paid") or joined.startswith("total bonus"):
            break
        spot_type = str(r[col["spot_type"]] or "").strip().upper() if col["spot_type"] < len(r) else ""
        lang = str(r[col["lang"]] or "").strip() if col["lang"] < len(r) else ""
        if spot_type in ("COM", "BONUS") and lang:
            out.append(r)
    return out


# ─── Shared assembly + reconciliation ────────────────────────────────────────

def _assemble_order(hdr: dict, lines: List[NtooitiveLine], week_dates: List[date],
                    footer: dict, option_label: str, source_format: str,
                    source_path: str) -> NtooitiveOrder:
    gross_airtime = _money(hdr.get("gross_airtime"))
    gross_contract = _money(hdr.get("gross_contract"))
    translation = _money(hdr.get("translation"))

    market_label = str(hdr.get("market", "") or "").strip()
    flight_text = str(hdr.get("flight", "") or "").strip()
    f_start, f_end = _resolve_flight_dates(flight_text, week_dates)

    # ── Charges: translation fees → Production box, never a line ──
    charges: List[NtooitiveCharge] = []
    if translation:
        charges.append(NtooitiveCharge("Translation fees", float(translation)))

    # ── Reconcile (raise — a parse that cannot foot must refuse to enter) ──
    paid = [ln for ln in lines if not ln.is_bonus]
    bonus = [ln for ln in lines if ln.is_bonus]

    for ln in paid:
        if ln.gross_total is not None:
            expect = round(ln.rate * ln.total_spots, 2)
            if abs(expect - ln.gross_total) > 0.01:
                raise ValueError(
                    f"Ntooitive parser: line '{ln.language_block} {ln.daypart}' "
                    f"rate ${ln.rate} × {ln.total_spots} = ${expect:,.2f} but the "
                    f"sheet says ${ln.gross_total:,.2f}")

    if footer.get("paid_units") is not None:
        got = sum(ln.total_spots for ln in paid)
        if got != int(footer["paid_units"]):
            raise ValueError(f"Ntooitive parser: paid spots {got} != footer "
                             f"'Total Paid' {int(footer['paid_units'])} — a line "
                             f"was likely dropped; refusing to enter")
    if footer.get("bonus_units") is not None:
        got = sum(ln.total_spots for ln in bonus)
        if got != int(footer["bonus_units"]):
            raise ValueError(f"Ntooitive parser: bonus spots {got} != footer "
                             f"'Total Bonuses' {int(footer['bonus_units'])}")

    line_gross = round(sum(ln.rate * ln.total_spots for ln in paid), 2)
    for label, stated in (("footer 'Total Paid'", footer.get("paid_gross")),
                          ("header 'Gross (Airtime)'", gross_airtime)):
        if stated is not None and abs(line_gross - stated) > 0.01:
            raise ValueError(f"Ntooitive parser: line gross ${line_gross:,.2f} != "
                             f"{label} ${stated:,.2f}")

    if gross_contract is not None and gross_airtime is not None:
        other = round(gross_contract - gross_airtime, 2)
        have = round(sum(c.amount for c in charges), 2)
        if abs(other - have) > 0.01:
            raise ValueError(
                f"Ntooitive parser: header says ${other:,.2f} of non-airtime money "
                f"but only ${have:,.2f} was classified (translation fees) — "
                f"an unrecognised charge; refusing to guess")

    # ── Implied commission: identical on every paid line, or the doc is mixed ──
    implied = 0.0
    ratios = [1.0 - (ln.net_total / ln.gross_total)
              for ln in paid
              if ln.net_total is not None and ln.gross_total]
    if ratios:
        implied = round(ratios[0], 4)
        if any(abs(r - implied) > 0.001 for r in ratios):
            raise ValueError("Ntooitive parser: the NET/gross ratio differs "
                             f"between paid lines ({sorted(set(round(r, 4) for r in ratios))}) "
                             "— mixed money basis; refusing")

    return NtooitiveOrder(
        agency=str(hdr.get("agency", "") or "").strip(),
        advertiser=str(hdr.get("advertiser", "") or "").strip(),
        contact=str(hdr.get("contact", "") or "").strip(),
        email=str(hdr.get("email", "") or "").strip(),
        market_code=_market_code(market_label),
        market_label=market_label,
        order_date=_to_date(hdr.get("revised")) or _to_date(hdr.get("proposed")),
        option_label=option_label,
        billing_cycle=str(hdr.get("billing_cycle", "") or "").strip(),
        flight_start_text=flight_text,
        lines=lines,
        charges=charges,
        implied_commission=implied,
        gross_airtime=gross_airtime,
        gross_contract=gross_contract,
        source_format=source_format,
        source_path=source_path,
        _flight_start_date=f_start,
        _flight_end_date=f_end,
    )


# ─── PDF parser (a print of one Option sheet) ────────────────────────────────

_ROW_TOL = 2.0      # cluster words into rows on RAW top floats (never round —
                    # round() manufactures phantom rows at .5 boundaries)
_COL_TOL = 10.0     # week-cell centre vs week-label centre; column pitch ~23pt


def _cluster_rows(words: list[dict]) -> list[list[dict]]:
    ws = sorted(words, key=lambda w: w["top"])
    rows: list[list[dict]] = []
    cur: list[dict] = []
    last = None
    for w in ws:
        if last is not None and w["top"] - last > _ROW_TOL:
            rows.append(cur)
            cur = []
        cur.append(w)
        last = w["top"]
    if cur:
        rows.append(cur)
    return [sorted(r, key=lambda w: w["x0"]) for r in rows]


def _centre(w: dict) -> float:
    return (w["x0"] + w["x1"]) / 2.0


_WEEK_LABEL_RE = re.compile(r"^(\d{1,2})-([A-Za-z]{3})$")   # "17-Aug"

# A token that begins the daypart half of the combined Language/Daypart region:
# a day pattern ("M-F", "M-", "Sa-Su", "Sat", "M-Su") or a time ("6a-7a", "8p").
_DAYPART_START_RE = re.compile(
    r"^(?:\d|(?:M|Tu|W|Th|F|Sa|Su|Mon|Tue|Wed|Thu|Fri|Sat|Sun)(?:[-–]|$))",
    re.IGNORECASE,
)


def _split_lang_daypart(region: str) -> tuple[str, str]:
    """Split 'Mandarin News M-F 6a-7a & 8p-9p' / 'Mandarin Chinese ROS' into
    (language block, daypart) by content."""
    tokens = region.split()
    for i, t in enumerate(tokens):
        if i and _DAYPART_START_RE.match(t):
            return " ".join(tokens[:i]), " ".join(tokens[i:])
    # Bonus rows: '<block> <ROS-language> ROS' — the daypart is the block name
    # immediately before 'ROS' when that token is a language word.
    upper = [t.upper() for t in tokens]
    if "ROS" in upper:
        ri = upper.index("ROS")
        start = ri
        if ri and tokens[ri - 1].capitalize() in _BASE_LANGUAGES:
            start = ri - 1
        return " ".join(tokens[:start]), " ".join(tokens[start:])
    return region, ""
_MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
           'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def parse_ntooitive_pdf(path: str) -> NtooitiveOrder:
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        words: list[dict] = []
        text = ""
        for page in pdf.pages:
            words.extend(page.extract_words())
            text += (page.extract_text() or "") + "\n"

    # ── Header fields from the page text ──
    hdr: dict = {}

    def _rx(pattern: str) -> str:
        m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        return m.group(1).strip() if m else ""

    hdr["agency"] = _rx(r"Media Buying Agency\s*\n?\s*(?:\(if any\):\s*)?([A-Za-z][\w .&-]*?)(?:\s+Address|\s*$)")
    hdr["advertiser"] = _rx(r"Advertiser:\s*(.+?)\s*(?:Billing Cycle|$)")
    hdr["contact"] = _rx(r"Contact:\s*(.+?)\s*$")
    hdr["email"] = _rx(r"Email:\s*(\S+@\S+)")
    hdr["billing_cycle"] = _rx(r"Billing Cycle\s+(\w+)")
    hdr["market"] = _rx(r"Market:\s*(.+?)\s*$")
    hdr["flight"] = _rx(r"Flight schedule\s*([\d/]+\s*-\s*[\d/]+)")
    revised = _rx(r"REVISED\s+(\d{1,2}/\d{1,2}/\d{4})")
    proposed = _rx(r"Date Proposed:\s*(\d{1,2}/\d{1,2}/\d{4})")
    for key, raw in (("revised", revised), ("proposed", proposed)):
        if raw:
            hdr[key] = datetime.strptime(raw, "%m/%d/%Y").date()
    # Money amounts print with stray spaces when the label wraps ("$ 8 00.00"),
    # so strip ALL whitespace inside the number.
    for key, pat in (("gross_airtime", r"Gross \(Airtime\)\s*\$\s*([\d ,.]+)"),
                     ("gross_contract", r"Gross Amount of Contract\s*\$\s*([\d ,.]+)")):
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            hdr[key] = float(re.sub(r"[ ,]", "", m.group(1)))
    if re.search(r"Gross Translation Fees", text, re.IGNORECASE) and \
            hdr.get("gross_contract") is not None and hdr.get("gross_airtime") is not None:
        # The fee amount itself prints garbled across the wrapped label; the
        # header's own arithmetic (contract − airtime) is the reliable copy,
        # and _assemble_order re-checks it.
        fee = round(hdr["gross_contract"] - hdr["gross_airtime"], 2)
        if fee > 0:
            hdr["translation"] = fee

    # ── Grid: header row = the row carrying week labels + 'Language' ──
    rows = _cluster_rows(words)
    year_anchor = (hdr.get("revised") or hdr.get("proposed") or date.today()).year

    header_row = None
    for r in rows:
        texts = [w["text"] for w in r]
        if any(_WEEK_LABEL_RE.match(t) for t in texts) and "Language" in texts:
            header_row = r
            break
    if header_row is None:
        raise ValueError("Ntooitive parser (pdf): could not find the week-label header row")

    # Week columns: label → date (year from the revised date, +1y on wrap)
    week_cells: list[tuple[float, date]] = []
    prev: Optional[date] = None
    y = year_anchor
    for w in header_row:
        m = _WEEK_LABEL_RE.match(w["text"])
        if not m:
            continue
        d = date(y, _MONTHS[m.group(2).lower()], int(m.group(1)))
        if prev and d < prev:
            y += 1
            d = date(y, d.month, d.day)
        week_cells.append((_centre(w), d))
        prev = d
    week_dates = [d for _, d in week_cells]

    # Field bands from the header labels' x-positions (label-mapped, not fixed)
    def _label_x(txt: str) -> float:
        for w in header_row:
            if w["text"] == txt:
                return w["x0"]
        raise ValueError(f"Ntooitive parser (pdf): header label {txt!r} missing "
                         "— the template changed; refusing to guess positions")

    x_lang = _label_x("Language")
    _label_x("Program")   # validate the daypart column still exists
    x_st = _label_x("Spot")
    x_len = _label_x("Length")
    x_units = _label_x("#")            # 'Total Unit #'
    first_week_x = min(c for c, _ in week_cells)
    # Money bands from the (Gross)/(Gross)/(NET) labels on the header row.
    money_marks = sorted(w["x0"] for w in header_row if w["text"] in ("(Gross)", "(NET)"))
    if len(money_marks) < 3:
        raise ValueError("Ntooitive parser (pdf): money column labels missing")
    x_rate, x_gross, x_net = money_marks[:3]

    grid_top = header_row[0]["top"]

    def _band(row: list[dict], lo: float, hi: float) -> str:
        return " ".join(w["text"] for w in row if lo <= w["x0"] < hi).strip()

    lines: List[NtooitiveLine] = []
    footer: dict[str, Optional[float]] = {}
    pending_money: Optional[str] = None   # 'paid' footer money can print one row lower

    for r in rows:
        if r[0]["top"] <= grid_top:
            continue
        rowtext = " ".join(w["text"] for w in r)
        low = rowtext.lower()

        if pending_money is not None:
            m = re.search(r"\$\s*([\d,]+\.\d{2})", rowtext)
            if m and footer.get("paid_gross") is None:
                footer["paid_gross"] = float(m.group(1).replace(",", ""))
            pending_money = None

        if low.startswith("total paid"):
            units = _band(r, x_units - 12, x_rate - 4)
            footer["paid_units"] = _money(re.sub(r"[^\d]", "", units) or None)
            m = re.search(r"\$\s*([\d,]+\.\d{2})", rowtext)
            if m:
                footer["paid_gross"] = float(m.group(1).replace(",", ""))
            else:
                pending_money = "paid"
            continue
        if low.startswith("total bonuses"):
            units = _band(r, x_units - 12, x_rate - 4)
            footer["bonus_units"] = _money(re.sub(r"[^\d]", "", units) or None)
            break

        spot_type = _band(r, x_st - 6, x_len - 4).upper()
        if spot_type not in ("COM", "BONUS"):
            continue
        # The Language Block and Day Part columns physically overlap by ~1pt in
        # this print (title text runs under the daypart), so a geometric split
        # is unreliable — read the combined region and split by CONTENT: the
        # daypart starts at the first day-pattern/time token (or at the block
        # name preceding 'ROS' on a bonus row). The bonus marker column sits
        # LEFT of the Language band and is excluded by the x_lang bound.
        region = _band(r, x_lang - 4, x_st - 6)
        lang, daypart = _split_lang_daypart(region)
        if not lang:
            continue

        length_txt = _band(r, x_len - 4, first_week_x - 12)
        m = re.search(r"(\d+)", length_txt)
        length_sec = int(m.group(1)) if m else 30

        # Week cells by x-centre distance against the week-label centres.
        # Zero cells may not print — the Total Unit # reconciliation below is
        # what catches a genuinely dropped cell.
        spots = [0] * len(week_cells)
        for w in r:
            if not w["text"].isdigit():
                continue
            c = _centre(w)
            if c < first_week_x - _COL_TOL or c >= x_units - 4:
                continue
            best_i, best_d = None, None
            for i, (cc, _) in enumerate(week_cells):
                dist = abs(c - cc)
                if best_d is None or dist < best_d:
                    best_i, best_d = i, dist
            if best_d is not None and best_d <= _COL_TOL:
                spots[best_i] += int(w["text"])

        units_txt = _band(r, x_units - 12, x_rate - 4)
        stated_units = _money(re.sub(r"[^\d]", "", units_txt) or None)

        def _money_band(lo: float, hi: float) -> Optional[float]:
            txt = _band(r, lo, hi).replace("$", "").strip()
            return _money(txt or None)

        rate = _money_band(x_rate - 4, x_gross - 4)
        gross_total = _money_band(x_gross - 4, x_net - 4)
        net_total = _money_band(x_net - 4, 10_000.0)

        if spot_type == "COM" and not rate:
            raise ValueError(f"Ntooitive parser (pdf): paid line '{lang} {daypart}' "
                             f"has no readable rate — refusing to enter $0 airtime")
        if spot_type == "BONUS" and rate:
            raise ValueError(f"Ntooitive parser (pdf): BONUS line '{lang}' carries "
                             f"a rate of ${rate} — contradictory row; refusing")

        ln = NtooitiveLine(
            language_block=lang,
            daypart=daypart,
            spot_type=spot_type,
            rate=float(rate or 0.0),
            length_sec=length_sec,
            week_dates=list(week_dates),
            week_spots=spots,
            gross_total=gross_total,
            net_total=net_total,
        )
        if stated_units is not None and int(stated_units) != ln.total_spots:
            raise ValueError(
                f"Ntooitive parser (pdf): line '{lang} {daypart}' week cells sum "
                f"to {ln.total_spots} but Total Unit # says {int(stated_units)} — "
                f"a cell was dropped or mis-columned; refusing to enter")
        lines.append(ln)

    if not lines:
        raise ValueError("Ntooitive parser (pdf): no airtime lines found")

    return _assemble_order(
        hdr=hdr, lines=lines, week_dates=week_dates, footer=footer,
        option_label="PDF", source_format="pdf", source_path=path,
    )


# ─── Dispatcher ──────────────────────────────────────────────────────────────

def parse_ntooitive(path: str, sheet_name: Optional[str] = None) -> NtooitiveOrder:
    """Route on extension: workbook (.xlsx/.xlsm) or PDF print. One order shape."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_ntooitive_xlsx(path, sheet_name=sheet_name)
    if ext == ".pdf":
        return parse_ntooitive_pdf(path)
    raise ValueError(f"Ntooitive parser: unsupported file type {ext!r}")


if __name__ == "__main__":
    import sys
    o = parse_ntooitive(sys.argv[1], sheet_name=sys.argv[2] if len(sys.argv) > 2 else None)
    print(f"{o.agency} / {o.advertiser}  [{o.option_label}]  {o.market_code} "
          f"({o.market_label})")
    _fmt = lambda v: f"${v:,.2f}" if v is not None else "—"  # noqa: E731
    print(f"Flight {o.flight_start} → {o.flight_end}   "
          f"gross airtime {_fmt(o.gross_airtime)}  contract {_fmt(o.gross_contract)}  "
          f"implied commission {o.implied_commission:.1%}")
    for ln in o.lines:
        tag = "BNS " if ln.is_bonus else "    "
        print(f"  {tag}:{ln.length_sec}s {ln.language_block:<16} {ln.daypart:<24} "
              f"${ln.rate:>7.2f}  {ln.total_spots:>3} spots  {ln.week_spots}")
    for ch in o.charges:
        print(f"  CHARGE {ch.description}: ${ch.amount:,.2f}")
