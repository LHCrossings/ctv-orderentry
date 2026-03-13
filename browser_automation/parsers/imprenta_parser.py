"""
imprenta_parser.py — Imprenta / PG&E XLSX Order Parser

File format (single-sheet XLSX):
    Row 2:  Campaign title
    Row 3:  "IMPRENTA: <client>"
    Row 4:  Flight date range
    Row 5:  Column headers — B=Net Amount, C=Length, D–G=week dates (datetime), H=Total Spots
    Row 6+: One market-label row, then program data rows
            Col A: program name with embedded days + time
            Col B: net rate
            Col C: duration (':15', ':30')
            Cols D–G: spots per week (aligns with dates in row 5)
    Notes section (no spot data) follows the summary rows.
    A18 area: "Media type: 15-second bookends" if order is bookend

When multiple time ranges appear in col A (e.g. "7p-8p, 11:30p-12a"),
they are joined with "; " so EtereClient merges them to earliest-start/
latest-end via its semicolon rule.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

# ── Market name → Etere code ─────────────────────────────────────────────────
_MARKET_MAP: dict[str, str] = {
    "sacramento": "CVC",
    "san francisco": "SFO",
    "sf": "SFO",
    "new york": "NYC",
    "nyc": "NYC",
    "los angeles": "LAX",
    "seattle": "SEA",
    "houston": "HOU",
    "chicago": "CMP",
    "washington": "WDC",
}


def _market_code(raw: str) -> str:
    key = raw.strip().lower()
    for k, v in _MARKET_MAP.items():
        if k in key:
            return v
    return raw.strip().upper()


# ── Duration ──────────────────────────────────────────────────────────────────
def _parse_duration(s: str) -> int:
    """':15' → 15, ':30' → 30."""
    m = re.search(r':(\d+)', str(s))
    return int(m.group(1)) if m else 30


# ── Days extraction ───────────────────────────────────────────────────────────
_DAY_PATTERNS: list[tuple[str, str]] = [
    (r'\bM-?Su(?:n(?:day)?)?\b',   'M-Su'),
    (r'\bM-?Sat(?:urday)?\b',       'M-Sa'),
    (r'\bM-?F(?:ri(?:day)?)?\b',    'M-F'),
    (r'\bSat(?:urday)?-?Su(?:n)?\b','Sa-Su'),
    (r'\bSa-?Su\b',                 'Sa-Su'),
    (r'\bM(?:on(?:day)?)?\b',       'M'),
    (r'\bF(?:ri(?:day)?)?\b',       'F'),
    (r'\bSat(?:urday)?\b',          'Sa'),
    (r'\bSu(?:n(?:day)?)?\b',       'Su'),
]


def _extract_days(program: str) -> str:
    for pattern, result in _DAY_PATTERNS:
        if re.search(pattern, program, re.IGNORECASE):
            return result
    return "M-Su"


# ── Time extraction ───────────────────────────────────────────────────────────
# Matches "7p-8p", "11:30p-12a", "8:00AM-9:00PM", etc.
_TIME_RE = re.compile(
    r'\d{1,2}(?::\d{2})?\s*[aApP]\.?[mM]?\.?\s*[-–]\s*\d{1,2}(?::\d{2})?\s*[aApP]\.?[mM]?\.?'
)


def _extract_time(program: str) -> str:
    """Return all time ranges found, joined with '; ' for EtereClient."""
    matches = _TIME_RE.findall(program)
    return '; '.join(m.strip() for m in matches)


# ── Program name cleanup ──────────────────────────────────────────────────────
def _clean_program(raw: str) -> str:
    """Strip embedded days and times from the program name cell."""
    s = raw
    s = _TIME_RE.sub('', s)
    for pattern, _ in _DAY_PATTERNS:
        s = re.sub(pattern, '', s, flags=re.IGNORECASE)
    # Remove leading/trailing punctuation and whitespace (preserve balanced parens)
    s = re.sub(r'[\s\-–,;/]+$', '', s.strip())
    s = re.sub(r'^[\s\-–,;/]+', '', s.strip())
    s = re.sub(r'\s{2,}', ' ', s)
    return s.strip()


# ── Data classes ──────────────────────────────────────────────────────────────
@dataclass
class ImprentaLine:
    program: str
    days: str
    time: str                        # semicolon-joined if multiple ranges
    duration: int                    # seconds
    rate_net: float
    rate_gross: float
    spots_by_week: list[int]
    week_date_ranges: list[tuple[date, date]]
    market: str                      # Etere market code
    is_bonus: bool
    is_bookend: bool                 # True for all paid lines if order is bookend


@dataclass
class ImprentaParseResult:
    lines: list[ImprentaLine]
    flight_start: date
    flight_end: date
    campaign: str
    client: str
    market: str                      # primary Etere market code
    week_start_dates: list[date]
    is_bookend: bool
    gross_up_factor: float


# ── Main parser ───────────────────────────────────────────────────────────────
def parse_imprenta_file(
    file_path: Path,
    gross_up_factor: float = 1 / 0.85,
) -> ImprentaParseResult:
    """Parse an Imprenta XLSX broadcast order."""
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active

    # ── Scan header area for metadata ────────────────────────────────────────
    is_bookend = False
    campaign = ""
    client = ""
    flight_start: Optional[date] = None
    flight_end: Optional[date] = None

    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "").strip()
            if not v:
                continue
            vl = v.lower()
            if "bookend" in vl:
                is_bookend = True
            if "campaign" in vl and not campaign:
                campaign = re.sub(r'(?i)^campaign\s*:', '', v).strip()
            if "client" in vl and not client:
                client = re.sub(r'(?i)^client\s*:', '', v).strip()
            if "flight" in vl and not flight_start:
                dates = re.findall(r'\d{1,2}/\d{1,2}/\d{4}', v)
                if len(dates) >= 2:
                    try:
                        flight_start = datetime.strptime(dates[0], "%m/%d/%Y").date()
                        flight_end   = datetime.strptime(dates[1], "%m/%d/%Y").date()
                    except ValueError:
                        pass
                elif len(dates) == 1:
                    try:
                        flight_start = datetime.strptime(dates[0], "%m/%d/%Y").date()
                    except ValueError:
                        pass

    # ── Find week-date header row ─────────────────────────────────────────────
    # Locate the row containing sequentially weekly datetime values.
    header_row_idx: Optional[int] = None
    week_start_dates: list[date] = []
    week_cols: list[int] = []  # 1-based column indices

    for row in ws.iter_rows():
        date_cells = [
            (cell.column, cell.value.date())
            for cell in row
            if isinstance(cell.value, datetime)
        ]
        if len(date_cells) >= 2:
            date_cells.sort(key=lambda x: x[0])
            dates_only = [d for _, d in date_cells]
            if all((dates_only[i+1] - dates_only[i]).days == 7 for i in range(len(dates_only)-1)):
                header_row_idx = row[0].row
                week_start_dates = dates_only
                week_cols = [c for c, _ in date_cells]
                break

    if not week_start_dates:
        raise ValueError("Could not find weekly date header row in XLSX")

    # Week end dates snap to broadcast-week Sunday
    week_date_ranges: list[tuple[date, date]] = [
        (s, s + timedelta(days=(6 - s.weekday()) % 7))
        for s in week_start_dates
    ]

    if flight_start is None:
        flight_start = week_date_ranges[0][0]
    if flight_end is None:
        flight_end = week_date_ranges[-1][1]

    # ── Parse data rows ───────────────────────────────────────────────────────
    lines: list[ImprentaLine] = []
    current_market = ""

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        col_a = str(row[0].value or "").strip()
        if not col_a or col_a == "``":
            continue

        # Stop at standalone metadata labels that follow the data section
        if col_a.upper() in ("TV",):
            break

        # Skip summary rows
        if col_a.lower().startswith("total"):
            continue

        # Read spot counts for this row
        spot_vals: list[int] = []
        for wc in week_cols:
            try:
                spot_vals.append(int(row[wc - 1].value or 0))
            except (TypeError, ValueError):
                spot_vals.append(0)

        if not any(s > 0 for s in spot_vals):
            # No spots → market label row or notes
            col_b = row[1].value
            if col_b is None:
                current_market = _market_code(col_a)
            continue

        # Net rate
        try:
            rate_net = float(row[1].value or 0)
        except (TypeError, ValueError):
            continue

        duration = _parse_duration(str(row[2].value or ""))
        is_bonus = rate_net == 0 or "bonus" in col_a.lower()
        days = _extract_days(col_a)
        time_str = _extract_time(col_a)
        program = _clean_program(col_a)

        rate_gross = 0.0 if is_bonus else round(rate_net * gross_up_factor, 2)

        lines.append(ImprentaLine(
            program=program,
            days=days,
            time=time_str,
            duration=duration,
            rate_net=rate_net,
            rate_gross=rate_gross,
            spots_by_week=list(spot_vals),
            week_date_ranges=list(week_date_ranges),
            market=current_market,
            is_bonus=is_bonus,
            is_bookend=is_bookend and not is_bonus,
        ))

    primary_market = current_market or (lines[0].market if lines else "")

    return ImprentaParseResult(
        lines=lines,
        flight_start=flight_start,
        flight_end=flight_end,
        campaign=campaign,
        client=client,
        market=primary_market,
        week_start_dates=week_start_dates,
        is_bookend=is_bookend,
        gross_up_factor=gross_up_factor,
    )
