"""Marathon Ventures (MVMS) post-log export — one workbook per calendar quarter.

Spot-level record of every AIRED WorldLink spot in the requested range, in Marathon's
17-column template (their sample: "MVMS Data Request - SAMPLE Crossings TV.xlsx"), plus a
few extra columns on the right that Marathon did not ask for but that make the file
self-explaining (Rep, Spot Type, Market, Line Descr).

Decisions (Lee, 2026-09-02):
  * WorldLink agency only (ANAGRAF 133). "Agency" in Marathon's sense = the client agency
    behind the order (Tatari, Marketing Architects, Direct Donor ...), which we parse from
    the parenthetical in the ANAGRAF client name; a Rep column carries "Worldlink".
  * COM + BNS only — no per-inquiry (PER) rows.
  * 800 Number stays blank: Etere holds no phone numbers.
  * Aired = TPALINSE.STATUS 'Q' (settled). 'A' is ABORTED, not aired.

Usage:
    uv run python3 scripts/mvms_post_log_export.py --from 2025-01-01 --to 2026-09-30 \\
        --out-dir "/mnt/k/!Archives/MVMS Data Request"
    uv run python3 scripts/mvms_post_log_export.py --from 2026-06-01 --to 2026-06-30 --dry-run
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from browser_automation.etere_direct_client import connect  # noqa: E402

FPS = 29.97
DAY = 24 * 3600
WORLDLINK_ID = 133

MARKETS = {
    1: ("NYC", "Crossings TV - New York"),
    2: ("CMP", "Crossings TV - Chicago/Minneapolis"),
    3: ("HOU", "Crossings TV - Houston"),
    4: ("SFO", "Crossings TV - San Francisco"),
    5: ("SEA", "Crossings TV - Seattle"),
    6: ("LAX", "Crossings TV - Los Angeles"),
    7: ("CVC", "Crossings TV - Sacramento/Central Valley"),
    8: ("WDC", "Crossings TV - Washington DC"),
    9: ("MMT", "Crossings TV - National Multimarket"),
    10: ("DAL", "The Asian Channel - Dallas"),
}

# Client-agency aliases as they appear in the ANAGRAF client-name parenthetical.
AGENCY_ALIASES = {
    "ma": "Marketing Architects",
    "marketing architects": "Marketing Architects",
    "tatari": "Tatari Inc",
    "tatari inc": "Tatari Inc",
    "dd": "Direct Donor",
    "direct": "Direct Donor",
    "direct donor": "Direct Donor",
    "icon": "Icon Media Direct, Inc.",
    "imd": "Icon Media Direct, Inc.",
    "icon media direct, inc": "Icon Media Direct, Inc.",
    "icon media direct, inc.": "Icon Media Direct, Inc.",
    "kcll": "Key Contacts - Legal Leads",
}

# DR daypart codes on broadcast-day seconds (06:00 = 21600 ... 30:00). Boundaries follow
# Marathon's own sample, where 9:00 AM-6:00 PM is "DA".
DAYPARTS = [
    ("EM", 6 * 3600, 9 * 3600),  # early morning
    ("DA", 9 * 3600, 18 * 3600),  # daytime
    ("PR", 18 * 3600, 23 * 3600),  # prime (incl. early fringe)
    ("LF", 23 * 3600, 26 * 3600),  # late fringe (23:00-02:00)
    ("ON", 26 * 3600, 30 * 3600),  # overnight (02:00-06:00)
]

MARATHON_COLUMNS = [
    "Network",
    "Agency",
    "Advertiser",
    "Brand",
    "Product",
    "Date",
    "Air Time",
    "Length",
    "Rate",
    "Order",
    "Estimate #",
    "Program (Aired)",
    "Ad-ID",
    "800 Number",
    "Inv. Descr.",
    "Daypart (Placed*)",
    "Time Period",
]
EXTRA_COLUMNS = ["Rep", "Spot Type", "Market", "Line Descr"]

SPOT_SQL = """
SELECT t.ID_TPALINSE, t.COD_USER, t.DATA, t.ORA, t.DURATION, t.NEWTYPE,
       RTRIM(t.COD_PROGRA), RTRIM(f.DESCRIZIO),
       RTRIM(c.COD_CONTRATTO), RTRIM(c.CUSTOMERREF),
       RTRIM(cu.RAG_SOCIAL), RTRIM(ag.RAG_SOCIAL),
       RTRIM(r.DESCRIZIONE), r.IMPORTO, r.ORA_INIZIO, r.ORA_FINE, r.ID_BOOKINGCODE
FROM TPALINSE t
JOIN trafficPalinse tp ON tp.id_tpalinse = t.ID_TPALINSE
JOIN CONTRATTIRIGHE r ON r.ID_CONTRATTIRIGHE = tp.ID_ContrattiRighe
JOIN CONTRATTITESTATA c ON c.ID_CONTRATTITESTATA = r.ID_CONTRATTITESTATA
LEFT JOIN ANAGRAF cu ON cu.ID_ANAGRAF = c.COMMITTENTE
LEFT JOIN ANAGRAF ag ON ag.ID_ANAGRAF = c.AGENZIA
LEFT JOIN FILMATI f ON f.ID_FILMATI = t.ID_FILMATI
WHERE t.DATA BETWEEN %s AND %s AND t.LIVELLO = 0 AND t.STATUS = 'Q'
  AND t.COD_USER BETWEEN 1 AND 10 AND t.NEWTYPE IN ({types}) AND c.AGENZIA = %s
ORDER BY t.COD_USER, t.DATA, t.ORA, t.XORDER
"""

COUNT_SQL = """
SELECT COUNT(*)
FROM TPALINSE t
JOIN trafficPalinse tp ON tp.id_tpalinse = t.ID_TPALINSE
JOIN CONTRATTIRIGHE r ON r.ID_CONTRATTIRIGHE = tp.ID_ContrattiRighe
JOIN CONTRATTITESTATA c ON c.ID_CONTRATTITESTATA = r.ID_CONTRATTITESTATA
WHERE t.DATA BETWEEN %s AND %s AND t.LIVELLO = 0 AND t.STATUS = 'Q'
  AND t.COD_USER BETWEEN 1 AND 10 AND t.NEWTYPE IN ({types}) AND c.AGENZIA = %s
"""

PGM_SQL = """
SELECT COD_USER, DATA, ORA, RTRIM(TITLE)
FROM TPALINSE
WHERE DATA BETWEEN %s AND %s AND LIVELLO = 0 AND NEWTYPE = 'PGM' AND COD_USER BETWEEN 1 AND 10
ORDER BY COD_USER, DATA, ORA
"""


def frames_to_seconds(frames: int) -> float:
    return frames / FPS


def air_datetime(day: date, ora: int) -> tuple[date, time]:
    """Broadcast day runs 06:00 → 30:00 on one DATA; 24:00+ is the next calendar date."""
    secs = int(round(frames_to_seconds(ora)))
    if secs >= DAY:
        return day + timedelta(days=1), time(*_hms(secs - DAY))
    return day, time(*_hms(secs))


def _hms(secs: int) -> tuple[int, int, int]:
    return secs // 3600, (secs % 3600) // 60, secs % 60


def length_text(frames: int) -> str:
    secs = int(round(frames_to_seconds(frames)))
    return f"{secs // 60}:{secs % 60:02d}"


def ampm(secs: int, with_seconds: bool) -> str:
    secs %= DAY
    h, m, s = _hms(secs)
    suffix = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12}:{m:02d}:{s:02d} {suffix}" if with_seconds else f"{h12}:{m:02d} {suffix}"


def window_secs(ora_ini: int | None, ora_fin: int | None) -> tuple[int, int] | None:
    """Line window in broadcast-day seconds; an end at or before the start wraps past midnight."""
    if ora_ini is None or ora_fin is None:
        return None
    lo = int(round(frames_to_seconds(ora_ini)))
    hi = int(round(frames_to_seconds(ora_fin)))
    if lo < 6 * 3600:
        lo += DAY
    if hi < 6 * 3600 or hi <= lo:
        hi += DAY
    return lo, hi


def daypart_code(win: tuple[int, int] | None) -> str:
    if win is None:
        return ""
    lo, hi = win
    hit = [code for code, a, b in DAYPARTS if lo < b and hi > a]
    if len(hit) >= 3:
        return "ROS"
    return "/".join(hit)


def time_period(win: tuple[int, int] | None) -> tuple[str, str]:
    """('9:00 AM-6:00 PM', '(9:00:00 AM-6:00:00 PM)') like Marathon's sample."""
    if win is None:
        return "", ""
    lo, hi = win
    return (
        f"{ampm(lo, False)}-{ampm(hi, False)}",
        f"({ampm(lo, True)}-{ampm(hi, True)})",
    )


_PAREN = re.compile(r"^(?P<adv>.*?)\s*\((?P<ag>[^)]*)\)\s*$")


def split_client(name: str | None, contract_code: str | None) -> tuple[str, str]:
    """ANAGRAF client 'Woof (Tatari Inc)' → ('Woof', 'Tatari Inc'). Fallback: the code token
    after 'WL' ('WL MA 209183' → Marketing Architects); last resort 'Worldlink'."""
    name = (name or "").strip()
    m = _PAREN.match(name)
    if m and m.group("ag").strip():
        raw = m.group("ag").strip()
        return m.group("adv").strip(), AGENCY_ALIASES.get(raw.lower(), raw)
    toks = (contract_code or "").split()
    if len(toks) >= 3 and toks[0].upper() == "WL":
        raw = " ".join(toks[1:-1])
        return name, AGENCY_ALIASES.get(raw.lower(), raw)
    return name, "Worldlink"


def quarters(start: date, end: date):
    q_start = date(start.year, 3 * ((start.month - 1) // 3) + 1, 1)
    while q_start <= end:
        m = q_start.month + 3
        nxt = date(q_start.year + (m > 12), m if m <= 12 else m - 12, 1)
        yield (
            max(q_start, start),
            min(nxt - timedelta(days=1), end),
            f"{q_start.year}Q{(q_start.month - 1) // 3 + 1}",
        )
        q_start = nxt


def load_programs(cur, lo: date, hi: date) -> dict[tuple[int, date], list[tuple[int, str]]]:
    cur.execute(PGM_SQL, (lo, hi))
    out: dict[tuple[int, date], list[tuple[int, str]]] = defaultdict(list)
    for cod_user, d, ora, title in cur.fetchall():
        out[(cod_user, d.date() if isinstance(d, datetime) else d)].append((ora, title or ""))
    return out


def program_at(programs, cod_user: int, d: date, ora: int) -> str:
    rows = programs.get((cod_user, d))
    if not rows:
        return ""
    title = ""
    for p_ora, p_title in rows:  # sorted by ORA
        if p_ora <= ora:
            title = p_title
        else:
            break
    return title


def build_rows(cur, lo: date, hi: date, agency_id: int, types: list[str]):
    programs = load_programs(cur, lo, hi)
    placeholders = ",".join("%s" for _ in types)
    cur.execute(SPOT_SQL.format(types=placeholders), (lo, hi, *types, agency_id))
    rows = []
    for (
        _id,
        cod_user,
        d,
        ora,
        dur,
        newtype,
        cod_progra,
        creative,
        code,
        custref,
        client,
        agency_name,
        line_desc,
        importo,
        ora_ini,
        ora_fin,
        _booking,
    ) in cur.fetchall():
        d = d.date() if isinstance(d, datetime) else d
        air_date, air_time = air_datetime(d, ora)
        advertiser, client_agency = split_client(client, code)
        is_bonus = newtype == "BNS"
        rate = 0.0 if is_bonus else float(importo or 0)
        win = window_secs(ora_ini, ora_fin)
        inv_descr, period = time_period(win)
        brand = advertiser
        product = (
            creative.split(":", 1)[1].strip() if creative and ":" in creative else (creative or "")
        )
        rows.append(
            [
                MARKETS[cod_user][1],
                client_agency,
                advertiser,
                brand,
                product,
                air_date,
                air_time,
                length_text(dur),
                round(rate, 2),
                code,
                custref or "",
                program_at(programs, cod_user, d, ora),
                cod_progra,
                "",  # 800 Number — not held in Etere
                inv_descr,
                daypart_code(win),
                period,
                agency_name or "Worldlink",
                "Bonus" if is_bonus else "Paid",
                MARKETS[cod_user][0],
                line_desc or "",
            ]
        )
    return rows


def write_workbook(path: Path, rows: list[list], label: str) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Post Log")
    ws.append(MARATHON_COLUMNS + EXTRA_COLUMNS)
    for r in rows:
        ws.append(r)
    summary = wb.create_sheet("Summary")
    summary.append([f"Crossings TV / The Asian Channel — WorldLink aired spots, {label}"])
    summary.append(["Rows", len(rows)])
    summary.append([])
    summary.append(["Market", "Paid", "Bonus"])
    by_market: dict[str, Counter] = defaultdict(Counter)
    by_agency: Counter = Counter()
    by_month: Counter = Counter()
    for r in rows:
        by_market[r[19]][r[18]] += 1
        by_agency[r[1]] += 1
        by_month[r[5].strftime("%Y-%m")] += 1
    for mkt in sorted(by_market):
        summary.append([mkt, by_market[mkt]["Paid"], by_market[mkt]["Bonus"]])
    summary.append([])
    summary.append(["Agency", "Rows"])
    for ag, n in by_agency.most_common():
        summary.append([ag, n])
    summary.append([])
    summary.append(["Month", "Rows"])
    for mo in sorted(by_month):
        summary.append([mo, by_month[mo]])
    summary.append([])
    summary.append(["Notes"])
    summary.append(["Aired = Etere status Q (settled). Aborted/never-run spots excluded."])
    summary.append(["800 Number is blank: not held in the traffic system."])
    summary.append(
        [
            "Rate = rate as entered in Etere. WorldLink prices a Crossings TV unit once, on the New York "
            "line; the other market rows of the same order line are entered at 0.00. Bonus rows carry 0.00."
        ]
    )
    summary.append(
        ["Air Time is station local time; spots after midnight are dated the next calendar day."]
    )
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--from", dest="date_from", required=True, help="YYYY-MM-DD")
    ap.add_argument("--to", dest="date_to", required=True, help="YYYY-MM-DD")
    ap.add_argument("--out-dir", default=".", help="folder for the per-quarter workbooks")
    ap.add_argument("--agency-id", type=int, default=WORLDLINK_ID)
    ap.add_argument("--types", default="COM,BNS", help="TPALINSE NEWTYPE list (default COM,BNS)")
    ap.add_argument("--dry-run", action="store_true", help="query + reconcile, write nothing")
    a = ap.parse_args()
    start = date.fromisoformat(a.date_from)
    end = date.fromisoformat(a.date_to)
    types = [t.strip().upper() for t in a.types.split(",") if t.strip()]
    out_dir = Path(a.out_dir)
    if not a.dry_run:
        out_dir.mkdir(parents=True, exist_ok=True)

    conn = connect()
    cur = conn.cursor()
    total = 0
    agencies: Counter = Counter()
    ok = True
    for lo, hi, label in quarters(start, end):
        rows = build_rows(cur, lo, hi, a.agency_id, types)
        placeholders = ",".join("%s" for _ in types)
        cur.execute(COUNT_SQL.format(types=placeholders), (lo, hi, *types, a.agency_id))
        expected = cur.fetchone()[0]
        match = "OK" if expected == len(rows) else "MISMATCH"
        if match != "OK":
            ok = False
        agencies.update(r[1] for r in rows)
        total += len(rows)
        print(f"[{label}] {lo} → {hi}: {len(rows):,} rows (db count {expected:,}) {match}")
        if rows and not a.dry_run:
            path = out_dir / f"CrossingsTV_Worldlink_PostLog_{label}.xlsx"
            write_workbook(path, rows, f"{lo:%b %d, %Y} – {hi:%b %d, %Y}")
            print(f"        wrote {path}")
    conn.close()
    print(f"\nTotal rows: {total:,}")
    print("Client agencies found:")
    for ag, n in agencies.most_common():
        print(f"   {ag:<40} {n:>9,}")
    if not ok:
        print("\nROW COUNT MISMATCH — do not ship these files.")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
