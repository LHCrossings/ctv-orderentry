"""Backfill CTV_LineLanguage from historical billing books.

Reads the CLEANED tab of every Master Billing Sheet workbook (col J = Language,
col M = Line = ID_CONTRATTIRIGHE) and upserts source='billing-book' rows.
A billing-book write NEVER overwrites an 'entry'/'user' row (enforced by
upsert_line_languages); later books win over earlier ones, so a line billed
across months keeps its most recent language.

Usage:
    uv run python3 scripts/backfill_line_languages.py [--since YYMM] [--dry-run]

Books are discovered under BILLING_DIR (live book + Miscellany archives).
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

_project_root = Path(__file__).parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from browser_automation.etere_direct_client import connect, upsert_line_languages  # noqa: E402
from browser_automation.line_language import LANGUAGE_CODES  # noqa: E402

BILLING_DIR = Path("/mnt/c/Work Temp/Billing")
_BOOK_RE = re.compile(r"Master Billing Sheet (\d{4})\.xlsm$", re.IGNORECASE)
_VALID = {c.upper(): c for c in LANGUAGE_CODES}


def find_books(since: str | None) -> list[tuple[str, Path]]:
    """All (yymm, path) billing books, sorted oldest→newest so later books win."""
    books: dict[str, Path] = {}
    for p in BILLING_DIR.rglob("Master Billing Sheet *.xlsm"):
        if p.name.startswith("~$"):
            continue
        m = _BOOK_RE.search(p.name)
        if not m:
            continue
        yymm = m.group(1)
        if since and yymm < since:
            continue
        # Duplicate months (live + archive copy): keep the live-dir one
        if yymm not in books or p.parent == BILLING_DIR:
            books[yymm] = p
    return sorted(books.items())


def read_cleaned(path: Path) -> dict[int, str]:
    """{line_id: lang} from the CLEANED tab. Skips blank/manual/invalid rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        name = next((s for s in wb.sheetnames if s.strip().upper() == "CLEANED"), None)
        if not name:
            return {}
        ws = wb[name]
        out: dict[int, str] = {}
        # Header row 1: J (10) = Language, M (13) = Line
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=13, values_only=True):
            lang_raw, line_raw = row[0], row[3]
            if lang_raw is None or line_raw in (None, ""):
                continue
            lang = _VALID.get(str(lang_raw).strip().upper())
            if not lang:
                continue
            try:
                line_id = int(float(line_raw))
            except (ValueError, TypeError):
                continue
            if line_id > 0:
                out[line_id] = lang
        return out
    finally:
        wb.close()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", default="2201", help="earliest book YYMM (default 2201)")
    ap.add_argument("--dry-run", action="store_true", help="parse books, write nothing")
    args = ap.parse_args()

    books = find_books(args.since)
    print(f"Found {len(books)} billing book(s) since {args.since}", flush=True)

    # Parse from a local (ext4) copy — openpyxl over the /mnt/c 9P mount is
    # several times slower than one sequential file copy + local parse.
    import shutil
    import tempfile

    catalog: dict[int, str] = {}
    with tempfile.TemporaryDirectory(prefix="billing_books_") as tmpdir:
        for yymm, path in books:
            try:
                local = Path(tmpdir) / path.name
                shutil.copyfile(path, local)
                rows = read_cleaned(local)
                local.unlink()
            except Exception as exc:  # noqa: BLE001 - one bad book must not kill the run
                print(f"  {yymm}  ✗ {path.name}: {exc}", flush=True)
                continue
            catalog.update(rows)  # later book wins
            print(f"  {yymm}  {len(rows):>6} lines  (catalog: {len(catalog)})", flush=True)

    print(f"\nTotal unique lines: {len(catalog)}")
    print("Language distribution:", dict(Counter(catalog.values()).most_common()))

    if args.dry_run:
        print("\n--dry-run: nothing written")
        return

    # Set-based bulk write: stage into a temp table with multi-row VALUES
    # inserts, then one UPDATE + one INSERT per batch. The row-by-row
    # upsert_line_languages helper is right for entry-time (1-10 lines) but
    # takes hours at 47k rows; this takes ~a minute. Same precedence rule:
    # never touch rows whose SOURCE isn't 'billing-book'.
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("CREATE TABLE #langstage (ID_CONTRATTIRIGHE int PRIMARY KEY, LANG nvarchar(8))")
        items = list(catalog.items())
        for i in range(0, len(items), 500):
            chunk = items[i:i + 500]
            values = ",".join(f"({int(lid)}, N'{lang}')" for lid, lang in chunk)  # lang is validated against LANGUAGE_CODES
            cur.execute(f"INSERT INTO #langstage (ID_CONTRATTIRIGHE, LANG) VALUES {values}")
            if (i // 500) % 20 == 19:
                print(f"  staged {i + len(chunk)}/{len(items)}…", flush=True)
        cur.execute("""
            UPDATE t SET t.LANG = s.LANG, t.UPDATED_AT = GETDATE()
            FROM CTV_LineLanguage t JOIN #langstage s
              ON t.ID_CONTRATTIRIGHE = s.ID_CONTRATTIRIGHE
            WHERE t.SOURCE = 'billing-book' AND t.LANG <> s.LANG
        """)
        updated = cur.rowcount
        cur.execute("""
            INSERT INTO CTV_LineLanguage (ID_CONTRATTIRIGHE, LANG, SOURCE)
            SELECT s.ID_CONTRATTIRIGHE, s.LANG, 'billing-book'
            FROM #langstage s
            WHERE NOT EXISTS (SELECT 1 FROM CTV_LineLanguage t
                              WHERE t.ID_CONTRATTIRIGHE = s.ID_CONTRATTIRIGHE)
        """)
        inserted = cur.rowcount
        conn.commit()
        cur.execute("SELECT SOURCE, COUNT(*) FROM CTV_LineLanguage GROUP BY SOURCE")
        print(f"\n✓ {inserted} inserted, {updated} updated. "
              f"Catalog by source: {dict(cur.fetchall())}")


if __name__ == "__main__":
    main()
