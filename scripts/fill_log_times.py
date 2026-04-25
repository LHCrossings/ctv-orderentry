"""
Fill missing spot air times in a weekly traffic log (.xlsm) from TPALINSE.

For each day sheet, finds COM/BNS/AV rows (any non-PRG spot type) where the
Comments column (I) is blank, then looks up ORA from TPALINSE using the
asset code (parsed from Show Name) + date + market station ID.

TPALINSE.ORA is stored in frames at 29.97fps; converted to HH:MM:SS on write.

Usage:
    uv run python scripts/fill_log_times.py "path/to/LAX Log - 260420.xlsm"

The file is updated in-place.  Market is inferred from the filename
(e.g. "LAX" -> station 6).
"""
import sys
import re
import datetime
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from browser_automation.etere_direct_client import connect

# Market code → TPALINSE.COD_USER
MARKET_IDS = {
    "NYC": 1, "CMP": 2, "HOU": 3, "SFO": 4,
    "SEA": 5, "LAX": 6, "CVC": 7, "WDC": 8, "DAL": 10,
}

# Column indices (1-based for openpyxl)
COL_DATE     = 2   # B: Start Date
COL_SHOW     = 8   # H: Show Name  ("ASSETCODE: Description")
COL_COMMENTS = 9   # I: Comments / spot time
COL_TYPE     = 14  # N: Type (PRG, COM, BNS, AV, ...)

FPS = 29.97


def ora_to_timedelta(ora: int) -> datetime.timedelta:
    total_seconds = round(ora / FPS)
    return datetime.timedelta(seconds=total_seconds)


def asset_code_from_show_name(show_name: str) -> str | None:
    """Extract 'ASSETCODE' from 'ASSETCODE: Description (EtereCode)'."""
    if not show_name:
        return None
    m = re.match(r"^([^:]+):", show_name.strip())
    return m.group(1).strip() if m else None


def detect_market(filepath: Path) -> int | None:
    name = filepath.name.upper()
    for code, mid in MARKET_IDS.items():
        if code in name:
            return mid
    return None


def fetch_ora_map(conn, date: datetime.date, asset_code: str, market_id: int) -> list[int]:
    """Return sorted list of ORA values for matching spots on this date/market."""
    cur = conn.cursor(as_dict=True)
    cur.execute(
        "SELECT ORA FROM TPALINSE"
        " WHERE DATA = %s AND TITLE LIKE %s AND COD_USER = %d"
        " ORDER BY ORA",
        (date, f"%{asset_code}%", market_id),
    )
    return [r["ORA"] for r in cur.fetchall()]


def main():
    if len(sys.argv) < 2:
        print("Usage: fill_log_times.py <path-to-log.xlsm>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    market_id = detect_market(path)
    if market_id is None:
        print(f"Could not detect market from filename '{path.name}'.")
        print(f"Expected one of: {', '.join(MARKET_IDS)}")
        sys.exit(1)

    market_code = next(k for k, v in MARKET_IDS.items() if v == market_id)
    print(f"Market: {market_code} (COD_USER={market_id})")

    wb = openpyxl.load_workbook(path, keep_vba=True)

    with connect() as conn:
        total_filled = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Collect rows needing a time fill, grouped by (date, asset_code)
            # so we can batch-fetch ORA values and match in order.
            # Structure: {(date, asset_code): [row_number, ...]}
            pending: dict[tuple, list[int]] = defaultdict(list)

            for row in ws.iter_rows(min_row=2):
                spot_type = row[COL_TYPE - 1].value
                if not spot_type or spot_type.upper() == "PRG":
                    continue
                comment = row[COL_COMMENTS - 1].value
                if comment:  # already filled
                    continue
                date_val = row[COL_DATE - 1].value
                if not isinstance(date_val, datetime.datetime):
                    continue
                show_name = row[COL_SHOW - 1].value or ""
                asset_code = asset_code_from_show_name(show_name)
                if not asset_code:
                    continue
                key = (date_val.date(), asset_code)
                pending[key].append(row[0].row)

            if not pending:
                continue

            # Fetch ORA lists once per (date, asset_code) pair
            ora_lists: dict[tuple, list[int]] = {}
            for (date, asset_code) in pending:
                ora_lists[(date, asset_code)] = fetch_ora_map(conn, date, asset_code, market_id)

            # Write times back row by row in the order they appear in the sheet
            sheet_filled = 0
            for key, row_nums in pending.items():
                oras = ora_lists.get(key, [])
                date, asset_code = key
                if not oras:
                    print(f"  [{sheet_name}] No TPALINSE match: {asset_code} on {date}")
                    continue
                if len(oras) < len(row_nums):
                    print(f"  [{sheet_name}] WARNING: {asset_code} on {date} — "
                          f"{len(row_nums)} log rows but only {len(oras)} TPALINSE entries")

                for i, row_num in enumerate(row_nums):
                    if i >= len(oras):
                        break
                    td = ora_to_timedelta(oras[i])
                    ws.cell(row_num, COL_COMMENTS).value = td
                    sheet_filled += 1

            if sheet_filled:
                print(f"  {sheet_name}: filled {sheet_filled} spot time(s)")
            total_filled += sheet_filled

    if total_filled == 0:
        print("Nothing to fill — all spot times already present.")
    else:
        wb.save(path)
        print(f"\nSaved. Total filled: {total_filled}")


if __name__ == "__main__":
    main()
