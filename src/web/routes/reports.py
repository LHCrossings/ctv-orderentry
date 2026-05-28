"""
Reports routes: placement-by-week, as-run, and future report pages.
"""

import asyncio
import io
from collections import defaultdict
from datetime import date, timedelta

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from src.web.auth import require_export_token

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _week_start(d) -> date:
    if isinstance(d, str):
        d = date.fromisoformat(d)
    elif hasattr(d, "date"):
        d = d.date()
    return d - timedelta(days=d.weekday())


def _fmt_week(ws: date) -> str:
    we = ws + timedelta(days=6)
    return f"{ws.strftime('%b')} {ws.day}–{we.strftime('%b')} {we.day}"


def _fetch_placements_sync(contract_id: int):
    from browser_automation.etere_direct_client import connect as _db_connect

    with _db_connect() as conn:
        cur = conn.cursor(as_dict=True)
        cur.execute("""
            SELECT ID_CONTRATTITESTATA AS id,
                   COD_CONTRATTO       AS code,
                   DESCRIZIONE         AS description,
                   CONVERT(VARCHAR(10), DATA_INIZIO,  101) AS date_start,
                   CONVERT(VARCHAR(10), DATA_TERMINE, 101) AS date_end
            FROM CONTRATTITESTATA
            WHERE ID_CONTRATTITESTATA = %d
        """ % contract_id)
        hdr = cur.fetchone()
        if not hdr:
            return None, None
        cur.execute("""
            SELECT cr.DESCRIZIONE        AS description,
                   cr.OMAGGIO            AS is_bonus,
                   CAST(tp.DATA AS DATE) AS air_date,
                   COUNT(*)              AS spots
            FROM trafficPalinse tpp
            JOIN CONTRATTIRIGHE cr ON tpp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
            JOIN TPALINSE tp       ON tpp.id_tpalinse = tp.ID_TPALINSE
            WHERE cr.ID_CONTRATTITESTATA = %d
            GROUP BY cr.DESCRIZIONE, cr.OMAGGIO, CAST(tp.DATA AS DATE)
            ORDER BY CAST(tp.DATA AS DATE), cr.DESCRIZIONE
        """ % contract_id)
        return dict(hdr), cur.fetchall()


def _build_pivot(hdr: dict, daily_rows) -> dict:
    week_data: dict = defaultdict(lambda: defaultdict(int))
    bonus_set: set = set()
    desc_order: dict = {}

    for row in daily_rows:
        desc = row["description"]
        ws = _week_start(row["air_date"])
        week_data[ws][desc] += row["spots"]
        if row["is_bonus"]:
            bonus_set.add(desc)
        if desc not in desc_order:
            desc_order[desc] = len(desc_order)

    all_weeks = sorted(week_data.keys())
    desc_list = sorted(desc_order, key=lambda d: desc_order[d])
    week_labels = [_fmt_week(ws) for ws in all_weeks]

    rows_out = []
    week_totals = [0] * len(all_weeks)
    grand_total = 0

    for desc in desc_list:
        spots = [week_data[ws].get(desc, 0) for ws in all_weeks]
        total = sum(spots)
        grand_total += total
        for i, v in enumerate(spots):
            week_totals[i] += v
        rows_out.append({
            "description": desc,
            "spots": spots,
            "total": total,
            "is_bonus": desc in bonus_set,
        })

    return {
        "header": hdr,
        "weeks": week_labels,
        "rows": rows_out,
        "week_totals": week_totals,
        "grand_total": grand_total,
    }


def _build_excel(pivot: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hdr = pivot["header"]
    weeks = pivot["weeks"]
    rows = pivot["rows"]
    week_totals = pivot["week_totals"]
    grand_total = pivot["grand_total"]

    NAVY = "1F3864"
    BLUE = "2E75B6"
    LTBLUE = "D6E4F0"
    YELLOW = "FFF2CC"
    GREY = "F2F2F2"
    WHITE = "FFFFFF"

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def bdr():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Placement by Week"

    total_col = len(weeks) + 2
    last_col = get_column_letter(total_col)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = (
        f"{hdr.get('code', '')} — {hdr.get('description', '')} "
        f"· Order Placement by Week"
    )
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(NAVY)
    c.alignment = ctr
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6

    r = 3
    ws.row_dimensions[r].height = 18

    c = ws.cell(r, 1, "Description")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill(BLUE)
    c.alignment = lft

    for i, wl in enumerate(weeks):
        c = ws.cell(r, i + 2, wl)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill(BLUE)
        c.alignment = ctr

    c = ws.cell(r, total_col, "TOTAL")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill(NAVY)
    c.alignment = ctr

    for ri, row in enumerate(rows):
        dr = r + 1 + ri
        is_bns = row["is_bonus"]
        bg = YELLOW if is_bns else (GREY if ri % 2 == 0 else WHITE)

        c = ws.cell(dr, 1, row["description"])
        c.fill = fill(bg)
        c.alignment = lft
        c.border = bdr()
        if is_bns:
            c.font = Font(italic=True)

        for i, v in enumerate(row["spots"]):
            c = ws.cell(dr, i + 2, v if v else "—")
            c.fill = fill(bg)
            c.alignment = ctr
            c.border = bdr()
            if is_bns and v:
                c.font = Font(italic=True)

        c = ws.cell(dr, total_col, row["total"])
        c.font = Font(bold=True)
        c.fill = fill(LTBLUE)
        c.alignment = ctr
        c.border = bdr()

    tr = r + 1 + len(rows)

    c = ws.cell(tr, 1, "TOTAL")
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill(NAVY)
    c.alignment = lft
    c.border = bdr()

    for i, v in enumerate(week_totals):
        c = ws.cell(tr, i + 2, v)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill(NAVY)
        c.alignment = ctr
        c.border = bdr()

    c = ws.cell(tr, total_col, grand_total)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill(NAVY)
    c.alignment = ctr
    c.border = bdr()

    ws.column_dimensions["A"].width = 38
    for i in range(len(weeks)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14
    ws.column_dimensions[last_col].width = 8
    ws.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------

def build_reports_router(templates: Jinja2Templates) -> APIRouter:
    router = APIRouter()

    @router.get("/reports", response_class=HTMLResponse)
    async def reports_hub(request: Request):
        return templates.TemplateResponse(request, "reports.html")

    @router.get("/reports/placement-by-week", response_class=HTMLResponse)
    async def placement_by_week_page(request: Request):
        return templates.TemplateResponse(request, "reports/placement_by_week.html")

    @router.get("/api/reports/placement-by-week/{contract_id}")
    async def placement_by_week_data(contract_id: int):
        try:
            hdr, daily = await asyncio.get_running_loop().run_in_executor(
                None, lambda: _fetch_placements_sync(contract_id)
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))
        if hdr is None:
            raise HTTPException(status_code=404, detail="Contract not found")
        return JSONResponse(_build_pivot(hdr, daily))

    @router.get("/api/reports/placement-by-week/{contract_id}/excel")
    async def placement_by_week_excel(contract_id: int):
        try:
            hdr, daily = await asyncio.get_running_loop().run_in_executor(
                None, lambda: _fetch_placements_sync(contract_id)
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))
        if hdr is None:
            raise HTTPException(status_code=404, detail="Contract not found")
        pivot = _build_pivot(hdr, daily)
        excel_bytes = _build_excel(pivot)
        filename = f"{hdr['code'].replace(' ', '_')}_PlacementByWeek.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── As-Run Report ────────────────────────────────────────────────────────

    MARKET_MAP = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}
    FPS = 29.97

    def _frames_to_time(frames: int) -> str:
        total_sec = round(frames / FPS)
        h = total_sec // 3600
        m = (total_sec % 3600) // 60
        s = total_sec % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _fetch_asrun_sync(spot: str, date_from: date, date_to: date, market_ids: list[int]):
        from browser_automation.etere_direct_client import connect as _db_connect

        results = []
        like_pat = f"%{spot}%"
        with _db_connect() as conn:
            cur = conn.cursor(as_dict=True)
            for mkt_id in market_ids:
                cur.execute(
                    """
                    SELECT CAST(DATA AS DATE) AS air_date, ORA, TITLE
                    FROM TPALINSE
                    WHERE COD_USER = %s
                      AND DATA >= %s
                      AND DATA <= %s
                      AND TITLE LIKE %s
                    ORDER BY DATA, ORA
                    """,
                    (mkt_id, date_from.isoformat(), date_to.isoformat(), like_pat)
                )
                rows = cur.fetchall()
                airings = [
                    {
                        "date": str(r["air_date"]),
                        "time": _frames_to_time(r["ORA"]),
                        "title": r["TITLE"] or "",
                    }
                    for r in rows
                ]
                results.append({
                    "market": MARKET_MAP.get(mkt_id, str(mkt_id)),
                    "count": len(airings),
                    "airings": airings,
                })
        return results

    @router.get("/reports/as-run", response_class=HTMLResponse)
    async def as_run_page(request: Request):
        return templates.TemplateResponse(request, "reports/as_run.html")

    @router.get("/api/reports/as-run")
    async def as_run_data(
        spot: str = Query(...),
        date_from: str = Query(...),
        date_to: str = Query(...),
        markets: str = Query(...),
    ):
        try:
            d_from = date.fromisoformat(date_from)
            d_to   = date.fromisoformat(date_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        try:
            mkt_ids = [int(x) for x in markets.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid market IDs")

        try:
            results = await asyncio.get_running_loop().run_in_executor(
                None, lambda: _fetch_asrun_sync(spot, d_from, d_to, mkt_ids)
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

        return JSONResponse({
            "spot_query": spot,
            "date_from": date_from,
            "date_to": date_to,
            "results": results,
            "total": sum(m["count"] for m in results),
        })

    return router
