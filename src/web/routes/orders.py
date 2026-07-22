"""
Order queue routes: list, upload, move-to-used, history, restore, detail.
"""

import asyncio
import json
import os
import shutil
import subprocess
import sys
import threading as _threading
import time as _time
import uuid as _uuid
from datetime import date as _date_cls
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, List, Optional

from fastapi import (
    APIRouter,
    Body,
    File,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

_src_path = Path(__file__).parent.parent.parent
if str(_src_path) not in sys.path:
    sys.path.insert(0, str(_src_path))

# ── Etere web session cache ──────────────────────────────────────────────────
# One persistent login shared across all traffic-assign calls.
# Re-logs in only after 20 minutes of idle so we never burn multiple seats.
_etere_session_lock = _threading.Lock()
_etere_session_state: dict = {"session": None, "born_at": 0.0}
_ETERE_SESSION_TTL = 20 * 60  # seconds


def _get_etere_session():
    from browser_automation.etere_direct_client import etere_web_login
    now = _time.monotonic()
    with _etere_session_lock:
        s = _etere_session_state
        if s["session"] is not None and (now - s["born_at"]) < _ETERE_SESSION_TTL:
            return s["session"]
        session = etere_web_login()
        s["session"] = session
        s["born_at"] = now
        return session


def _invalidate_etere_session():
    with _etere_session_lock:
        _etere_session_state["session"] = None
        _etere_session_state["born_at"] = 0.0

from business_logic.services.pdf_order_detector import PDFOrderDetector
from orchestration.config import ApplicationConfig
from orchestration.order_scanner import OrderScanner
from web.parser_bridge import get_order_detail, list_parsers

_ALLOWED_EXTENSIONS = {".pdf", ".xml", ".xlsx", ".xlsm", ".jpg", ".jpeg", ".png"}

# Market code → COD_USER integer (must match EtereClient.MARKET_CODES)
_MARKET_CODES = {
    "NYC": 1, "CMP": 2, "HOU": 3, "SFO": 4, "SEA": 5,
    "LAX": 6, "CVC": 7, "WDC": 8, "MMT": 9, "DAL": 10,
}
_VALID_DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}
_FPS_GLOBAL = 29.97

# Crossings TV language → list of (days_set, time_from HH:MM, time_to HH:MM)
# DAL (The Asian Channel) has different mappings — add separately when needed.
_WD  = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}
_WE  = {"Saturday", "Sunday"}
_ALL = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}
_MSA = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"}

_CTV_LANG_WINDOWS: dict = {
    "Mandarin": [
        (_MSA, "06:00", "07:00"),
        (_ALL, "07:00", "08:00"),
        (_WD,  "20:00", "23:30"),
        (_WE,  "20:00", "23:59"),
    ],
    "Cantonese": [
        (_WD, "19:00", "20:00"),
        (_WD, "23:30", "23:59"),
    ],
    "Korean":     [(_ALL, "08:00", "10:00")],
    "Vietnamese": [(_ALL, "10:00", "13:00")],
    "Hindi": [
        (_WD, "13:00", "14:00"),
        (_WE, "13:00", "16:00"),
    ],
    "Punjabi":  [(_WD, "14:00", "16:00")],
    "Filipino": [
        (_WD, "16:00", "19:00"),
        (_WE, "16:00", "18:00"),
    ],
    "Hmong": [(_WE, "18:00", "20:00")],
}
_CTV_LANG_WINDOWS["Chinese"]    = _CTV_LANG_WINDOWS["Mandarin"] + _CTV_LANG_WINDOWS["Cantonese"]
_CTV_LANG_WINDOWS["SouthAsian"] = _CTV_LANG_WINDOWS["Hindi"]    + _CTV_LANG_WINDOWS["Punjabi"]

# Canonical list of supported traffic instruction formats — rendered as badges on the assign-assets page.
# Add one entry here whenever a new format is wired into either _detect_format() or a
# drop-first pre-check (e.g. Admerasia auto-color, TCAA auto-match).
_TRAFFIC_FORMAT_LABELS = [
    "Admerasia (McDonald's)",
    "TCAA (Toyota)",
    "Davis Elen",
    "IW Group (Lexus)",
    "Tatari (WorldLink)",
    "Direct Donor (WorldLink)",
    "Marketing Architects (WorldLink)",
    "Icon Media Direct (WorldLink)",
    "H&L Partners",
    "RPM",
]


# The Asian Channel (DAL) — broadcast day runs 0600–0559 (wraps past midnight)
_DAL_LANG_WINDOWS: dict = {
    "Mandarin": [
        (_WD,  "06:00", "09:30"),
        (_WE,  "06:00", "10:00"),
        (_ALL, "13:00", "17:00"),
        (_ALL, "18:00", "22:00"),
        (_ALL, "00:00", "01:00"),
        (_WD,  "02:00", "05:30"),
        (_WE,  "02:00", "05:59"),
    ],
    "Cantonese": [
        (_WD,  "09:30", "10:00"),
        (_ALL, "17:00", "18:00"),
        (_ALL, "01:00", "02:00"),
        (_WD,  "05:30", "05:59"),
    ],
    "Vietnamese": [(_ALL, "10:00", "11:00")],
    "Korean": [
        (_WD,  "11:00", "12:00"),
        (_WE,  "11:00", "13:00"),
        (_WD,  "22:00", "23:00"),
        (_WE,  "22:00", "23:59"),
    ],
}
_DAL_LANG_WINDOWS["Chinese"] = _DAL_LANG_WINDOWS["Mandarin"] + _DAL_LANG_WINDOWS["Cantonese"]


def _bcast_time_to_frames(t: str, fps: float) -> int:
    """Convert an 'HH:MM[:SS]' time-of-day to a broadcast-day frame count.

    The broadcast day runs 06:00 → 30:00, and Etere stores TPALINSE.ORA (and traffic
    block/segment offsets) on that scale — so the post-midnight tail 00:00–05:59
    lives at 24:00–29:59, on the same DATA. Any hour < 6 therefore belongs to that
    tail and must be shifted +24h; otherwise an ORA comparison lands at 0–6h where
    nothing exists (silent no-match). See tasks/lessons.md "Broadcast Day 06:00→30:00".

    This is the shared converter for every HH:MM→frame used against TPALINSE.ORA
    (traffic-assign filters/language windows, program-spot fill, break optimizer)."""
    parts = t.split(":")
    h = int(parts[0])
    mn = int(parts[1]) if len(parts) > 1 else 0
    s = int(parts[2]) if len(parts) > 2 else 0
    if h < 6:  # post-midnight tail of the broadcast day
        h += 24
    return round((h * 3600 + mn * 60 + s) * fps)


def _hhmm_to_frames(hhmm: str) -> int:
    return _bcast_time_to_frames(hhmm, _FPS_GLOBAL)


def _broadcast_month_folder(monday) -> tuple[int, str]:
    """Return (year, folder_name) for the broadcast month a week's Monday belongs to.

    Logs are filed by broadcast month, which starts on the Monday of the week
    containing the 1st of that calendar month — so the week of Dec 29 belongs
    to January if Jan 1 falls within it.
    """
    from datetime import timedelta
    for offset in range(7):
        day = monday + timedelta(days=offset)
        if day.day == 1:
            return day.year, day.strftime("%m %B %Y")
    first = monday.replace(day=1)
    bcast_start = first - timedelta(days=first.weekday())
    if bcast_start <= monday:
        return first.year, first.strftime("%m %B %Y")
    first_prev = (first - timedelta(days=1)).replace(day=1)
    return first_prev.year, first_prev.strftime("%m %B %Y")


# Weekly traffic-log tree; K: on Windows, the SMB mount elsewhere (override via env).
_TRAFFIC_LOG_ROOT = Path(
    os.environ.get(
        "TRAFFIC_LOG_ROOT",
        "K:/Traffic/logs" if sys.platform == "win32" else "/mnt/k/Traffic/logs",
    )
)


def _wb_load_fast(path: Path, **kw):
    """Load a workbook through one in-memory read. openpyxl's zipfile layer
    issues thousands of small reads, which is brutally slow on network
    shares (K:) — a single read_bytes() then BytesIO collapses that to one
    network transfer."""
    import io as _io

    import openpyxl
    return openpyxl.load_workbook(_io.BytesIO(path.read_bytes()), **kw)


def _wb_save_fast(wb, path: Path, transform=None) -> None:
    """Serialize the workbook in memory, then write it as ONE network
    transfer via a temp file + atomic replace (also avoids a torn file if
    the write is interrupted). Raises PermissionError if the target is
    locked open in Excel.

    `transform`, if given, is a `bytes -> bytes` hook applied to the serialized
    xlsx before writing — used to re-inject things openpyxl drops (e.g. the
    commercial log's custom-color picker swatches)."""
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if transform is not None:
        data = transform(data)
    tmp = path.with_name(path.name + ".tmp~")
    try:
        tmp.write_bytes(data)
        os.replace(tmp, path)
    except BaseException:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def _find_traffic_log(mkt: str, target: _date_cls) -> Optional[Path]:
    """Locate weekly traffic log .xlsm for a market and any date in that broadcast week."""
    monday = target - timedelta(days=target.weekday())
    year, month_folder = _broadcast_month_folder(monday)
    base = _TRAFFIC_LOG_ROOT / str(year) / month_folder
    for file_date in (monday.strftime("%y%m%d"), monday.strftime("%m%d%y")):
        for subfolder in ("", "done", "Done", "!Done"):
            folder = base / subfolder if subfolder else base
            p = folder / f"{mkt} Log - {file_date}.xlsm"
            if p.exists():
                return p
    return None


_MC_MARKET_IDS = {
    "NYC": 1, "CMP": 2, "HOU": 3, "SFO": 4,
    "SEA": 5, "LAX": 6, "CVC": 7, "WDC": 8, "DAL": 10,
}
_MC_FILL_FPS = 29.97


def _mc_fill_program_spots(
    ws,
    target: _date_cls,
    market_id: int,
    spots: list,
    time_in: str,
    time_out: str,
    cur,
    fps: float = _MC_FILL_FPS,
    program_language: str = "",
) -> list:
    """Fill column I for one program's spots from TPALINSE. Caller owns workbook save."""
    import datetime as _dt
    import re
    from collections import defaultdict

    # Language mismatch helpers
    _LANG_COMPAT = [frozenset({'M', 'C'}), frozenset({'SA', 'P'})]

    def _extract_asset_lang(code: str) -> str:
        """TVC30V12 → 'V', LEXUS15SA107 → 'SA', unrecognised → ''"""
        m = re.search(r'\d+([A-Za-z]+)\d+$', code)
        return m.group(1).upper() if m else ''

    def _lang_ok(asset_lang: str, prog_lang: str) -> bool:
        if not asset_lang or not prog_lang:
            return True
        if asset_lang == 'E':            # English airs anywhere
            return True
        if asset_lang == prog_lang.upper():
            return True
        pair = frozenset({asset_lang, prog_lang.upper()})
        return any(pair <= s for s in _LANG_COMPAT)

    def _time_to_frames(t: str) -> int:
        return _bcast_time_to_frames(t, fps)  # broadcast-day aware (post-midnight = 24:00–29:59)

    by_asset: dict = defaultdict(list)
    for spot in spots:
        show = spot.get("show_name", "")
        m = re.match(r"^([^:]+):", show.strip())
        asset_code = m.group(1).strip() if m else show.strip()
        by_asset[asset_code].append(spot)

    from_frames = _time_to_frames(time_in) if time_in else None
    # Allow up to 1 minute of show overrun — end time extended by 60 s worth of frames
    to_frames = (round(_time_to_frames(time_out) + 60 * fps)) if time_out else None

    results = []
    for asset_code, asset_spots in by_asset.items():
        if from_frames is not None and to_frames is not None:
            cur.execute(
                "SELECT ORA FROM TPALINSE"
                " WHERE DATA = %s AND TITLE LIKE %s AND COD_USER = %d"
                " AND ORA >= %d AND ORA < %d"
                " ORDER BY ORA",
                (target, f"%{asset_code}%", market_id, from_frames, to_frames),
            )
        else:
            cur.execute(
                "SELECT ORA FROM TPALINSE"
                " WHERE DATA = %s AND TITLE LIKE %s AND COD_USER = %d"
                " ORDER BY ORA",
                (target, f"%{asset_code}%", market_id),
            )
        oras = [r["ORA"] for r in cur.fetchall()]
        for i, spot in enumerate(asset_spots):
            if i >= len(oras):
                results.append({"excel_row": spot["excel_row"], "status": "no_match", "actual_time": ""})
                continue
            secs = int(oras[i] / fps)
            h, rem = divmod(secs, 3600)
            mn, s = divmod(rem, 60)
            time_str = f"{h}:{mn:02d}:{s:02d}"
            ws.cell(row=spot["excel_row"], column=9).value = _dt.timedelta(seconds=secs)
            asset_lang = _extract_asset_lang(asset_code)
            results.append({
                "excel_row": spot["excel_row"],
                "status": "filled",
                "actual_time": time_str,
                "lang_warning": not _lang_ok(asset_lang, program_language),
            })
    return results


def _build_spot_filter(filters: dict) -> str:
    """Return extra AND clauses for TPALINSE spot queries based on optional filter dict."""
    import re
    clauses = ["tp.DATA >= CAST(GETDATE() AS DATE)"]
    if filters.get("date_from") and re.match(r"^\d{4}-\d{2}-\d{2}$", filters["date_from"]):
        clauses.append(f"tp.DATA >= '{filters['date_from']}'")
    if filters.get("date_to") and re.match(r"^\d{4}-\d{2}-\d{2}$", filters["date_to"]):
        clauses.append(f"tp.DATA <= '{filters['date_to']}'")
    if filters.get("days"):
        safe = [d for d in filters["days"] if d in _VALID_DAYS]
        if safe:
            clauses.append(f"DATENAME(weekday, tp.DATA) IN ({','.join(repr(d) for d in safe)})")
    if filters.get("time_from"):
        try:
            clauses.append(f"tp.ORA >= {_hhmm_to_frames(filters['time_from'])}")
        except (ValueError, AttributeError):
            pass
    if filters.get("time_to"):
        try:
            clauses.append(f"tp.ORA <= {_hhmm_to_frames(filters['time_to'])}")
        except (ValueError, AttributeError):
            pass
    if filters.get("durations"):
        frames = [int(f) for f in filters["durations"] if str(f).lstrip("-").isdigit()]
        if frames:
            clauses.append(f"cr.DURATA IN ({','.join(str(f) for f in frames)})")
    if filters.get("markets"):
        ids = [_MARKET_CODES[k] for k in filters["markets"] if k in _MARKET_CODES]
        if ids:
            clauses.append(f"tp.COD_USER IN ({','.join(str(i) for i in ids)})")
    if filters.get("languages"):
        # Use DAL windows if DAL is the only market selected; otherwise CTV windows
        mkts = filters.get("markets", [])
        lang_table = _DAL_LANG_WINDOWS if mkts == ["DAL"] else _CTV_LANG_WINDOWS
        lang_subclauses = []
        for lang in filters["languages"]:
            for days_set, t_from, t_to in lang_table.get(lang, []):
                day_list = ",".join(f"'{d}'" for d in sorted(days_set))
                f_from = _hhmm_to_frames(t_from)
                f_to   = _hhmm_to_frames(t_to)
                lang_subclauses.append(
                    f"(DATENAME(weekday, tp.DATA) IN ({day_list})"
                    f" AND tp.ORA >= {f_from} AND tp.ORA <= {f_to})"
                )
        if lang_subclauses:
            clauses.append("(" + " OR ".join(lang_subclauses) + ")")
    if filters.get("spot_types"):
        safe = [t for t in filters["spot_types"] if re.match(r'^[A-Z]{1,6}$', t)]
        if safe:
            codes = ",".join(f"'{c}'" for c in safe)
            clauses.append(
                f"cr.ID_BOOKINGCODE IN (SELECT id_bookingcode FROM trf_bookingcode WHERE code IN ({codes}))"
            )
    return (" AND " + " AND ".join(clauses)) if clauses else ""


def _pi_filler_supporto(cur, filmati_id, fallback_desc: str = "") -> str:
    """Playout binding (TPALINSE.SUPPORTO) for a PI/PSA filler row.

    MUST be the channel prefix + the media FILE_ID (e.g. ``0ETX      PI-493-030``),
    never the human DESCRIZIO. Building it from the description (``0ETX      PI-493-030:
    Ship of ``) overruns the field, so the playout server can't resolve it to a file:
    the event errors (STATUS='E'), shows the red-X in EE, and never airs — while the
    same PI airs fine wherever its binding was built correctly. Mirrors the auto-assign
    convention (LEGACY_BASESUPP + FS_FILMATI.FILE_ID); FILE_ID for a PI equals the
    ``PI-nnn-nnn`` code, i.e. the DESCRIZIO up to the first colon.
    """
    supporto = ""
    if filmati_id:
        cur.execute(
            "SELECT TOP 1 ISNULL(d.LEGACY_BASESUPP,"
            " CAST(d.LEGACY_MEDIAID AS VARCHAR) + 'ETX      ') AS prefix, ff.FILE_ID"
            " FROM FS_FILMATI ff JOIN FS_METADEVICE d ON d.ID_METADEVICE = ff.ID_METADEVICE"
            " WHERE ff.ID_FILMATI = %d AND d.LEGACY_MEDIAID IS NOT NULL"
            " ORDER BY d.LEGACY_MEDIAID",
            (int(filmati_id),),
        )
        row = cur.fetchone()
        if row:
            # Callers pass either a dict cursor (filler insert) or a plain tuple
            # cursor (break-opt apply / bulk-apply) — handle both.
            if isinstance(row, dict):
                prefix, file_id = row.get("prefix"), row.get("FILE_ID")
            else:
                prefix, file_id = row[0], row[1]
            if file_id:
                supporto = (prefix or "") + file_id
    if not supporto:
        # Last resort only (FS_FILMATI row missing): use the FILE_ID-equivalent code
        # — the DESCRIZIO before the colon — never the full description.
        code = (fallback_desc or "").split(":", 1)[0].strip()
        supporto = "0ETX      " + code
    return supporto[:30]


def build_router(config: ApplicationConfig, templates: Jinja2Templates) -> APIRouter:
    router = APIRouter()

    used_dir = config.incoming_dir / "Used"
    # Entered-but-not-yet-backwritten IOs + their manifests live here
    # (tasks/backwrite-pipeline.md Phase 1). Manifests are written by
    # business_logic/services/backwrite_manifest.py at entry time.
    entered_dir = config.incoming_dir / "Entered"

    def _ensure_used_dir():
        used_dir.mkdir(parents=True, exist_ok=True)

    def _sweep_entered_strays() -> None:
        """Self-heal: an IO whose manifest is in Entered/ but whose file is
        still in the incoming root was locked (open in a viewer) when entry
        tried to move it — move it now. Best-effort; retried on every load."""
        try:
            if not entered_dir.exists():
                return
            for mf in entered_dir.glob("*.manifest.json"):
                stray = config.incoming_dir / mf.name[: -len(".manifest.json")]
                if not stray.is_file():
                    continue
                dest = entered_dir / stray.name
                try:
                    if dest.exists():
                        dest.unlink()
                    shutil.move(str(stray), str(dest))
                except OSError:
                    pass  # still locked — next load will retry
        except Exception:  # noqa: BLE001 - a sweep problem must never break the queue
            pass

    def _purge_used_folder(days: int = 30) -> int:
        """Delete files in Used/ older than `days` days. Returns count deleted."""
        if not used_dir.exists():
            return 0
        cutoff = datetime.now() - timedelta(days=days)
        deleted = 0
        for f in used_dir.iterdir():
            if f.is_file() and datetime.fromtimestamp(f.stat().st_mtime) < cutoff:
                f.unlink(missing_ok=True)
                deleted += 1
        return deleted

    def _make_scanner() -> OrderScanner:
        return OrderScanner(PDFOrderDetector(), config.incoming_dir)

    def _scan_dir(directory: Path) -> list[dict]:
        """Scan a directory for order files and return serializable list."""
        if not directory.exists():
            return []
        scanner = OrderScanner(PDFOrderDetector(), directory)
        orders = scanner.scan_for_orders()
        result = []
        listed = set()
        for order in orders:
            listed.add(order.pdf_path.name)
            stat = order.pdf_path.stat()
            ov = order.order_type.value if order.order_type else "Unknown"
            # Agency column label. For most parsers the order-type code IS the
            # agency; EQC is named after the client (Emerald Queen Casino), so
            # surface its real agency (TH Media) instead of the "eqc" code.
            agency_label = "TH Media" if ov == "eqc" else ov
            result.append({
                "filename": order.pdf_path.name,
                "order_type": ov,
                "agency_label": agency_label,
                "customer_name": order.customer_name or "Unknown",
                "estimate_number": order.estimate_number,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
        # Every file the tab badge counts must also get a row — detection
        # silently skips unrecognized xlsx/images, unparseable files, and
        # non-order extensions, which otherwise inflate the badge invisibly
        # and can only be cleaned up with shell access on the deploy host.
        # These rows render with an "Unknown" badge and the normal Mark
        # Done button, so strays can be cleared from the UI.
        for f in sorted(directory.iterdir()):
            if (not f.is_file() or f.name in listed
                    or f.name.startswith(('.', '~$'))
                    or f.name.endswith('.manifest.json')):
                continue
            stat = f.stat()
            result.append({
                "filename": f.name,
                "order_type": "Unknown",
                "agency_label": "Unknown",
                "customer_name": "Unrecognized file",
                "estimate_number": None,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
        return result

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    @router.get("/", response_class=HTMLResponse)
    async def portal(request: Request):
        return templates.TemplateResponse(request, "portal.html")

    @router.get("/orders", response_class=HTMLResponse)
    async def orders_hub(request: Request):
        return templates.TemplateResponse(request, "orders_hub.html")

    @router.get("/order-entry", response_class=HTMLResponse)
    async def index(request: Request):
        return templates.TemplateResponse(request, "index.html", {"parsers": list_parsers()})

    @router.get("/billing", response_class=HTMLResponse)
    async def billing(request: Request):
        return templates.TemplateResponse(request, "billing.html")

    # Master Control Bible — Maija's reference guides/flowcharts, served as
    # static pages from static/mc_bible/ (self-contained HTML files).
    @router.get("/master-control-bible", response_class=HTMLResponse)
    async def master_control_bible(request: Request):
        return templates.TemplateResponse(request, "mc_bible.html")

    @router.get("/billing/coop-invoicing", response_class=HTMLResponse)
    async def coop_invoicing(request: Request):
        return templates.TemplateResponse(request, "billing/coop_invoicing.html")

    @router.get("/billing/worldlink-placement", response_class=HTMLResponse)
    async def worldlink_placement(request: Request):
        return templates.TemplateResponse(request, "billing/worldlink_placement.html")

    @router.post("/billing/worldlink-placement/generate")
    async def worldlink_placement_generate(request: Request):
        import io as _io
        import re

        import pandas as pd
        from fastapi.responses import StreamingResponse

        from browser_automation.etere_direct_client import connect as _db_connect
        from src.backwrite.eterebridge_runner import (
            build_placement_csv_from_db,
            run_eterebridge_pipeline,
            save_to_excel_with_template,
        )

        body = await request.json()
        date_from    = body.get("date_from")
        date_to      = body.get("date_to")
        sales_person = body.get("sales_person", "House")
        agency_fee   = float(body.get("agency_fee", 0.15))

        if not date_from or not date_to:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "date_from and date_to are required"}, status_code=400)

        # Find all agency 133 contracts with active lines in the date range
        with _db_connect() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT DISTINCT ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                FROM CONTRATTITESTATA ct
                JOIN CONTRATTIRIGHE cr ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                WHERE ct.AGENZIA = 133
                  AND cr.DATA_INIZIO <= %s
                  AND cr.DATA_FINE   >= %s
                  AND ct.COD_CONTRATTO != 'TESTORDERFORLEE'
                ORDER BY ct.ID_CONTRATTITESTATA
            """, (date_to, date_from))
            contracts = cur.fetchall()

        if not contracts:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "No WorldLink contracts found for that date range"}, status_code=404)

        frames = []
        skipped = []
        for row in contracts:
            contract_id, cod_contratto = row[0], row[1]
            try:
                m = re.search(r'(\d+)\s*$', str(cod_contratto))
                wl_tracking = m.group(1) if m else ""

                csv_bytes = build_placement_csv_from_db(
                    contract_id,
                    date_from=date_from,
                    date_to=date_to,
                    isci_only=True,
                )
                user_inputs = {
                    "sales_person":  sales_person,
                    "billing_type":  "Broadcast",
                    "revenue_type":  "Direct Response Sales",
                    "agency_flag":   "Agency",
                    "agency_fee":    agency_fee,
                    "estimate":      wl_tracking,
                    "contract":      str(contract_id),
                    "affidavit":     "Y",
                    "is_worldlink":  True,
                }
                df = run_eterebridge_pipeline(csv_bytes, user_inputs)
                if df is not None and not df.empty:
                    frames.append(df)
            except Exception as exc:
                skipped.append(f"{cod_contratto}: {exc}")

        if not frames:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "No spot data found", "skipped": skipped}, status_code=404)

        combined   = pd.concat(frames, ignore_index=True)
        total_rows = len(combined)

        dal_mask      = combined["Market"] == "DAL"
        gross_dal     = combined.loc[dal_mask,  "Gross Rate"].sum()
        gross_nondal  = combined.loc[~dal_mask, "Gross Rate"].sum()
        dal_broker    = -(gross_dal    * (1 - agency_fee) * 0.10)
        nondal_broker = -(gross_nondal * (1 - agency_fee) * 0.10)

        xlsx_bytes = save_to_excel_with_template(combined, agency_fee=agency_fee)
        buf = _io.BytesIO(xlsx_bytes)

        filename = f"WL_Placement_{date_from}_{date_to}.xlsx"
        headers = {
            "Content-Disposition":  f"attachment; filename={filename}",
            "X-Contract-Count":     str(len(frames)),
            "X-Spot-Count":         str(total_rows),
            "X-DAL-Broker-Fees":    f"{dal_broker:.2f}",
            "X-NonDAL-Broker-Fees": f"{nondal_broker:.2f}",
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @router.get("/billing/compile-logs", response_class=HTMLResponse)
    async def billing_compile_logs_page(request: Request):
        return templates.TemplateResponse(request, "billing/compile_logs.html")

    @router.post("/api/billing/compile-logs/aggregate")
    async def billing_compile_logs_aggregate(
        billing_book: UploadFile = File(...),
        log_files: List[UploadFile] = File(...),
    ):
        import io as _io
        import re as _re

        import openpyxl

        MARKET_ORDER = ["NYC","CMP","HOU","SFO","SEA","LAX","CVC","WDC","MMT","DAL"]

        def _market_of(name: str) -> str:
            m = _re.match(r"^([A-Z]+)\s", name)
            return m.group(1) if m else name

        def _run(book_bytes: bytes, logs: list[tuple[str, bytes]]) -> bytes:
            # Sort logs by market order
            logs.sort(key=lambda x: (
                MARKET_ORDER.index(_market_of(x[0]))
                if _market_of(x[0]) in MARKET_ORDER else 99
            ))

            wb = openpyxl.load_workbook(_io.BytesIO(book_bytes), keep_vba=True)
            master_name = next((s for s in wb.sheetnames if s.upper() == "MASTER"), None)
            if master_name is None:
                raise ValueError("Billing book has no 'MASTER' tab")
            master = wb[master_name]

            for fname, log_bytes in logs:
                log_wb = openpyxl.load_workbook(_io.BytesIO(log_bytes), data_only=True, keep_vba=False)
                tab_name = next((s for s in log_wb.sheetnames if s.upper() == "MASTER FOR BILLING"), None)
                if tab_name is None:
                    continue
                log_ws = log_wb[tab_name]
                for row in log_ws.iter_rows(min_row=2, values_only=True):
                    if any(c is not None for c in row):
                        master.append(list(row))
                log_wb.close()

            out = _io.BytesIO()
            wb.save(out)
            return out.getvalue()

        book_bytes = await billing_book.read()
        log_data = [(f.filename or f"log_{i}.xlsm", await f.read()) for i, f in enumerate(log_files)]

        try:
            result = await asyncio.get_running_loop().run_in_executor(
                None, _run, book_bytes, log_data
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        fname = (billing_book.filename or "billing_book.xlsm").replace(
            ".xlsm", "_compiled.xlsm"
        ).replace(".xlsx", "_compiled.xlsx")
        return StreamingResponse(
            _io.BytesIO(result),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @router.get("/billing/monthly-logs", response_class=HTMLResponse)
    async def monthly_logs(request: Request):
        return templates.TemplateResponse(request, "billing/monthly_logs.html")

    @router.post("/api/billing/monthly-logs/check")
    async def monthly_logs_check(billing_book: UploadFile = File(...)):
        import io as _io
        from datetime import date as _date
        from datetime import timedelta

        import openpyxl

        data = await billing_book.read()
        try:
            wb = openpyxl.load_workbook(
                _io.BytesIO(data), read_only=True, keep_vba=False, data_only=True
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not open workbook: {exc}")

        if "MASTER" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="No MASTER tab found in workbook")

        ws = wb["MASTER"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="MASTER tab is empty")

        header = list(rows[0])
        try:
            date_col   = header.index("Start Date")
            market_col = header.index("Market")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Column not found in MASTER: {exc}")

        market_dates: dict[str, set] = {}
        for row in rows[1:]:
            if len(row) <= max(date_col, market_col):
                continue
            market  = row[market_col]
            raw_dt  = row[date_col]
            if not market or not raw_dt:
                continue
            if hasattr(raw_dt, "date"):
                d = raw_dt.date()
            elif isinstance(raw_dt, _date):
                d = raw_dt
            else:
                continue
            mkt = str(market).strip()
            if mkt not in market_dates:
                market_dates[mkt] = set()
            market_dates[mkt].add(d)

        # Infer billing month from filename YYMM (e.g. "2605" → May 2026)
        import calendar as _cal
        import re as _re
        _fm = _re.search(r'(\d{2})(0[1-9]|1[0-2])(?!\d)', billing_book.filename or "")
        month_info = None
        true_start = None
        true_end   = None
        if _fm:
            yy, mm = int(_fm.group(1)), int(_fm.group(2))
            year       = 2000 + yy
            cal_start  = _date(year, mm, 1)
            cal_end    = _date(year, mm, _cal.monthrange(year, mm)[1])
            bcast_start = cal_start - timedelta(days=cal_start.weekday())
            _next_mm, _next_yy = (mm + 1, year) if mm < 12 else (1, year + 1)
            _next1     = _date(_next_yy, _next_mm, 1)
            bcast_end  = _next1 - timedelta(days=_next1.weekday()) - timedelta(days=1)
            true_start, true_end = bcast_start, bcast_end
            _mnames = ['January','February','March','April','May','June',
                       'July','August','September','October','November','December']
            month_info = {
                "month_label":     f"{_mnames[mm - 1]} {year}",
                "broadcast_start": bcast_start.isoformat(),
                "broadcast_end":   bcast_end.isoformat(),
                "calendar_start":  cal_start.isoformat(),
                "calendar_end":    cal_end.isoformat(),
            }

        results = []
        for mkt in sorted(market_dates):
            dates    = market_dates[mkt]
            first    = true_start if true_start else min(dates)
            last     = true_end   if true_end   else max(dates)
            expected = set()
            cur = first
            while cur <= last:
                expected.add(cur)
                cur += timedelta(days=1)
            missing = sorted(expected - dates)
            results.append({
                "market":        mkt,
                "first_date":    first.isoformat(),
                "last_date":     last.isoformat(),
                "days_found":    len(dates),
                "days_expected": len(expected),
                "missing":       [d.isoformat() for d in missing],
            })

        from fastapi.responses import JSONResponse as _JSONResponse
        return _JSONResponse({"month_info": month_info, "results": results})

    # ------------------------------------------------------------------
    # Off-Air Gap Finder — spots stuck "never aired" after an AU crash +
    # EE "Reset transmitted events". Stuck signature: STATUS='I' AND
    # ASRUN_STATUS_M='I' (only end-of-day reconciliation writes
    # ASRUN_STATUS_M, so in-progress days never match). Marking verified
    # performs the exact write EE's "check manually the selected events"
    # does: STATUS='Q', ASRUN_STATUS_O='M'.
    # ------------------------------------------------------------------

    _OFFAIR_MARKETS = {1: "NYC", 2: "CMP", 3: "HOU", 4: "SFO", 5: "SEA",
                       6: "LAX", 7: "CVC", 8: "WDC", 9: "MMT", 10: "DAL"}
    _OFFAIR_FPS = 29.97

    def _offair_bcast_start(yr: int, mo: int) -> "_date_cls":
        first = _date_cls(yr, mo, 1)
        return first - timedelta(days=first.weekday())

    def _offair_ampm(frames: int) -> str:
        total_s = round(frames / _OFFAIR_FPS)
        h, rem = divmod(total_s, 3600)
        m = rem // 60
        nextday = "⁺¹" if h >= 24 else ""  # broadcast-day tail (24:00-29:59)
        h %= 24
        suffix = "a" if h < 12 else "p"
        return f"{h % 12 or 12}:{m:02d}{suffix}{nextday}"

    @router.get("/billing/offair-gaps", response_class=HTMLResponse)
    async def offair_gaps_page(request: Request):
        return templates.TemplateResponse(request, "billing/offair_gaps.html")

    @router.get("/api/billing/offair-gaps/scan")
    async def offair_gaps_scan(year: int, month: int):
        def _run():
            import calendar as _cal

            from browser_automation.etere_direct_client import connect as _connect

            cal_start = _date_cls(year, month, 1)
            cal_end = _date_cls(year, month, _cal.monthrange(year, month)[1])
            _ny, _nm = (year, month + 1) if month < 12 else (year + 1, 1)
            bcast_start = _offair_bcast_start(year, month)
            bcast_end = _offair_bcast_start(_ny, _nm) - timedelta(days=1)
            win_start = min(bcast_start, cal_start)
            win_end = max(bcast_end, cal_end)

            with _connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    """
                    SELECT t.ID_TPALINSE AS id, t.DATA, t.COD_USER, t.ORA,
                           t.NEWTYPE, t.TITLE,
                           ct.COD_CONTRATTO AS contract_code,
                           cl.RAG_SOCIAL    AS client_name
                    FROM TPALINSE t
                    LEFT JOIN trafficTPalinse tp ON tp.ID_TPalinse = t.ID_TPALINSE
                    LEFT JOIN CONTRATTITESTATA ct
                           ON ct.ID_CONTRATTITESTATA = tp.ID_CONTRATTITESTATA
                    LEFT JOIN ANAGRAF cl ON cl.ID_ANAGRAF = ct.COMMITTENTE
                    WHERE t.DATA BETWEEN %s AND %s
                      AND t.LIVELLO = 0
                      AND t.STATUS = 'I'
                      AND t.ASRUN_STATUS_M = 'I'
                      -- station IDs are 25s dynamic fillers placed last in a
                      -- program, INTENDED to be cut off — never billing-relevant
                      AND ISNULL(t.NEWTYPE, '') <> 'ID'
                    ORDER BY t.COD_USER, t.DATA, t.ORA
                    """,
                    (str(win_start), str(win_end)),
                )
                raw = cur.fetchall()

            # the traffic join can fan out on linked rows — keep first hit per spot
            rows, seen = [], set()
            for r in raw:
                if r["id"] in seen:
                    continue
                seen.add(r["id"])
                rows.append(r)

            split_frames = int(60 * 60 * _OFFAIR_FPS)  # >1h clean = separate gap
            gaps = []
            for r in rows:
                day = r["DATA"].date().isoformat()
                g = gaps[-1] if gaps else None
                if (g is None or g["cod_user"] != r["COD_USER"]
                        or g["date"] != day
                        or r["ORA"] - g["_last_ora"] > split_frames):
                    g = {
                        "cod_user": r["COD_USER"],
                        "market": _OFFAIR_MARKETS.get(
                            r["COD_USER"], f"st{r['COD_USER']}"),
                        "date": day,
                        "from": _offair_ampm(r["ORA"]),
                        "to": _offair_ampm(r["ORA"]),
                        "_last_ora": r["ORA"],
                        "ids": [],
                        "n_com": 0,
                        "clients": set(),
                        "spots": [],
                    }
                    gaps.append(g)
                g["_last_ora"] = r["ORA"]
                g["to"] = _offair_ampm(r["ORA"])
                g["ids"].append(r["id"])
                is_com = (r["NEWTYPE"] or "").strip().upper() in ("COM", "BNS", "TRD")
                if is_com:
                    g["n_com"] += 1
                    if r["client_name"]:
                        g["clients"].add(r["client_name"].strip())
                g["spots"].append({
                    "id": r["id"],
                    "time": _offair_ampm(r["ORA"]),
                    "type": (r["NEWTYPE"] or "").strip(),
                    "title": (r["TITLE"] or "").strip(),
                    "client": (r["client_name"] or "").strip(),
                    "contract": (r["contract_code"] or "").strip(),
                })
            for g in gaps:
                g["clients"] = sorted(g["clients"])
                del g["_last_ora"]

            return {
                "window": {
                    "start": win_start.isoformat(),
                    "end": win_end.isoformat(),
                    "bcast": [bcast_start.isoformat(), bcast_end.isoformat()],
                    "calendar": [cal_start.isoformat(), cal_end.isoformat()],
                },
                "gaps": gaps,
                "total_spots": sum(len(g["ids"]) for g in gaps),
                "total_com": sum(g["n_com"] for g in gaps),
            }

        result = await asyncio.get_running_loop().run_in_executor(None, _run)
        return JSONResponse(result)

    @router.post("/api/billing/offair-gaps/verify")
    async def offair_gaps_verify(payload: dict = Body(...)):
        try:
            ids = [int(i) for i in (payload.get("ids") or [])]
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="ids must be integers")
        if not ids:
            raise HTTPException(status_code=400, detail="No spot IDs supplied")

        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            updated = 0
            with _connect() as conn:
                cur = conn.cursor()
                for i in range(0, len(ids), 500):
                    chunk = ids[i:i + 500]
                    ph = ",".join(["%s"] * len(chunk))
                    # guarded: only rows still matching the stuck signature
                    cur.execute(
                        f"""
                        UPDATE TPALINSE
                        SET STATUS = 'Q', ASRUN_STATUS_O = 'M',
                            LASTUPDATE = GETDATE()
                        WHERE ID_TPALINSE IN ({ph})
                          AND LIVELLO = 0
                          AND STATUS = 'I'
                          AND ASRUN_STATUS_M = 'I'
                        """,
                        tuple(chunk),
                    )
                    updated += cur.rowcount
                conn.commit()
            return updated

        updated = await asyncio.get_running_loop().run_in_executor(None, _run)
        return JSONResponse({"updated": updated, "requested": len(ids)})

    @router.get("/scripts", response_class=HTMLResponse)
    async def scripts(request: Request):
        return templates.TemplateResponse(request, "scripts.html")

    @router.get("/scripts/block-refresh", response_class=HTMLResponse)
    async def block_refresh(request: Request):
        return templates.TemplateResponse(request, "scripts/block_refresh.html")

    @router.get("/scripts/import-edl", response_class=HTMLResponse)
    async def scripts_import_edl(request: Request):
        # Standalone EDL attach — reuses the daily-programming search-files and
        # edl-status APIs to find an asset and show its current EDL, then writes
        # the marks market-free via /api/scripts/import-edl below.
        return templates.TemplateResponse(request, "scripts/import_edl.html")

    @router.post("/api/scripts/import-edl")
    async def scripts_import_edl_apply(body: dict = Body(...)):
        """Write an EDIUS marker CSV to an asset's EDL, market-free. Unlike the
        Daily Programming importer, this does NOT explode against a channel — it
        just stores the marks on the asset (splits + EOM), committing on success.
        """
        from browser_automation.etere_direct_client import connect as _db_connect
        from src.business_logic.services.edl_import import apply_edl_from_csv, parse_edius_csv
        try:
            filmati = int(body.get("filmati"))
        except (TypeError, ValueError):
            return {"ok": False, "error": "Missing/invalid filmati"}
        try:
            splits, eom = parse_edius_csv(body.get("csv") or "")
        except ValueError as exc:
            return {"ok": False, "error": str(exc)}
        try:
            with _db_connect() as conn:
                res = apply_edl_from_csv(conn, filmati, splits, eom)  # cod_user=None → no explode
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"ok": False, "error": f"EDL import failed: {exc}"}
        res["splits"] = splits
        res["eom"] = eom
        res["count"] = len(splits) + 1
        return res

    @router.get("/scripts/separation", response_class=HTMLResponse)
    async def separation(request: Request):
        return templates.TemplateResponse(request, "scripts/separation.html")

    # ------------------------------------------------------------------
    # Scripts API — Separation
    # ------------------------------------------------------------------

    _FPS = 29.97

    def _frames_to_ampm(frames: int) -> str:
        if not frames:
            return "—"
        total_s = round(frames / _FPS)
        h, rem = divmod(total_s, 3600)
        m = rem // 60
        if h == 0:
            return f"12:{m:02d}a"
        if h < 12:
            return f"{h}:{m:02d}a"
        if h == 12:
            return f"12:{m:02d}p"
        return f"{h - 12}:{m:02d}p"

    def _frames_to_min(frames) -> int:
        return round(frames / _FPS / 60) if frames else 0

    def _frames_to_sec(frames) -> int:
        return round(frames / _FPS) if frames else 0

    def _days_str(lun, mar, mer, gio, ven, sab, dom) -> str:
        d = [bool(lun), bool(mar), bool(mer), bool(gio), bool(ven), bool(sab), bool(dom)]
        if all(d):
            return "M-Su"
        if d[:5] == [True] * 5 and not d[5] and not d[6]:
            return "M-F"
        if d[:6] == [True] * 6 and not d[6]:
            return "M-Sa"
        if not any(d[:5]) and d[5] and d[6]:
            return "Sa-Su"
        if d[5] and not any(d[:5]) and not d[6]:
            return "Sa"
        if d[6] and not any(d[:5]) and not d[5]:
            return "Su"
        abbr = ["M", "Tu", "W", "Th", "F", "Sa", "Su"]
        return "/".join(a for a, v in zip(abbr, d) if v) or "—"

    _MARKET_NAMES = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}

    @router.get("/api/scripts/separation/lines")
    async def get_separation_lines(contract_id: int = Query(..., gt=0)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT ID_CONTRATTIRIGHE, DESCRIZIONE,
                           COALESCE(DATESTART, DATA_INIZIO), COALESCE(DATEEND, DATA_FINE),
                           COALESCE(ORA_INIZIOF, ORA_INIZIO), COALESCE(ORA_FINEF, ORA_FINE),
                           LUNEDI, MARTEDI, MERCOLEDI, GIOVEDI, VENERDI, SABATO, DOMENICA,
                           DURATA, PASSAGGI_SETTIMANALI,
                           Interv_Committente, INTERVALLO, INTERV_CONTRATTO,
                           COD_USER
                    FROM   CONTRATTIRIGHE
                    WHERE  ID_CONTRATTITESTATA = %s
                    ORDER  BY ID_CONTRATTIRIGHE
                """, [contract_id])
                rows = cursor.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail=f"No lines found for contract {contract_id}.")

            lines = []
            for row in rows:
                (line_id, desc, date_from, date_to, ora_in, ora_out,
                 lun, mar, mer, gio, ven, sab, dom,
                 durata, spots_pw,
                 sep_cust, sep_ord, sep_evt,
                 cod_user) = row
                lines.append({
                    "line_id":      line_id,
                    "description":  desc or "",
                    "market":       _MARKET_NAMES.get(cod_user, str(cod_user) if cod_user else "—"),
                    "date_from":    f"{date_from.month}/{date_from.day}/{date_from.year}" if date_from else "",
                    "date_to":      f"{date_to.month}/{date_to.day}/{date_to.year}" if date_to else "",
                    "time_from":    _frames_to_ampm(ora_in),
                    "time_to":      _frames_to_ampm(ora_out),
                    "days":         _days_str(lun, mar, mer, gio, ven, sab, dom),
                    "duration_sec": _frames_to_sec(durata),
                    "spots_pw":     spots_pw or 0,
                    "sep_customer": _frames_to_min(sep_cust),
                    "sep_order":    _frames_to_min(sep_ord),
                    "sep_event":    _frames_to_min(sep_evt),
                })
            return JSONResponse({"contract_id": contract_id, "lines": lines})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/separation/apply")
    async def apply_separation_lines(payload: dict = Body(...)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            line_ids = [int(x) for x in payload.get("line_ids", [])]
            customer = int(payload.get("customer", 0))
            order    = int(payload.get("order", 0))
            event    = int(payload.get("event", 0))

            if not line_ids:
                raise HTTPException(status_code=400, detail="No lines selected.")

            def _to_frames(minutes: int) -> int:
                return round(minutes * 60 * _FPS)

            placeholders = ",".join(["%s"] * len(line_ids))
            with _db_connect() as conn:
                cursor = conn.cursor()
                # INTERVALLO = Order, INTERV_CONTRATTO = Event (old Etere web had these swapped)
                cursor.execute(f"""
                    UPDATE CONTRATTIRIGHE
                    SET    Interv_Committente = %s,
                           INTERVALLO         = %s,
                           INTERV_CONTRATTO   = %s
                    WHERE  ID_CONTRATTIRIGHE IN ({placeholders})
                """, [_to_frames(customer), _to_frames(order), _to_frames(event), *line_ids])
                conn.commit()
                updated = cursor.rowcount

            return JSONResponse({"updated": updated, "customer": customer, "order": order, "event": event})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/scripts/separation")
    async def run_separation(
        contract_id: int = Query(..., gt=0),
        customer: int = Query(0, ge=0),
        event: int = Query(0, ge=0),
        order: int = Query(0, ge=0),
    ):
        project_root = Path(__file__).parent.parent.parent.parent
        script_path = project_root / "scripts" / "update_separation_contract.py"

        python_exe = project_root / ".venv" / "Scripts" / "python.exe"
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                str(python_exe), "-u", str(script_path), str(contract_id),
                "--customer", str(customer),
                "--event", str(event),
                "--order", str(order),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(project_root),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.get("/scripts/reportsort", response_class=HTMLResponse)
    async def reportsort_page(request: Request):
        return templates.TemplateResponse(request, "scripts/reportsort.html")

    @router.get("/api/scripts/reportsort")
    async def run_reportsort(
        log_type: str = Query(...),
        date_from: str = Query(...),
        date_to: str = Query(...),
    ):
        project_root = Path(__file__).parent.parent.parent.parent
        script_path = project_root / "scripts" / "run_reportsort.py"

        python_exe = project_root / ".venv" / "Scripts" / "python.exe"
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                str(python_exe), "-u", str(script_path),
                log_type, date_from, date_to,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(project_root),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.get("/scripts/delete-spots", response_class=HTMLResponse)
    async def delete_spots_page(request: Request):
        return templates.TemplateResponse(request, "scripts/delete_spots.html")

    @router.get("/api/scripts/delete-spots")
    async def run_delete_spots(
        contract_id: int = Query(..., gt=0),
        date_from: str = Query(...),
        date_to: str = Query(...),
        confirm: str = Query("0"),
    ):
        project_root = Path(__file__).parent.parent.parent.parent
        script_path = project_root / "scripts" / "delete_scheduled_spots.py"

        python_exe = project_root / ".venv" / "Scripts" / "python.exe"
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        args = [str(python_exe), "-u", str(script_path), str(contract_id), date_from, date_to]
        if confirm == "1":
            args.append("--confirm")

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                *args,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(project_root),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.get("/scripts/unschedule", response_class=HTMLResponse)
    async def unschedule_page(request: Request):
        return templates.TemplateResponse(request, "scripts/unschedule.html")

    @router.get("/api/scripts/unschedule")
    async def run_unschedule(contract_id: int = Query(..., gt=0)):
        project_root = Path(__file__).parent.parent.parent.parent
        script_path = project_root / "scripts" / "unschedule_contract.py"

        python_exe = project_root / ".venv" / "Scripts" / "python.exe"
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                str(python_exe), "-u", str(script_path), str(contract_id),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(project_root),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    # ------------------------------------------------------------------
    # Fill Log Times
    # ------------------------------------------------------------------

    _fill_log_pending: dict = {}  # token -> (Path, filename)

    @router.get("/scripts/fill-log-times", response_class=HTMLResponse)
    async def fill_log_times_page(request: Request):
        return templates.TemplateResponse(request, "scripts/fill_log_times.html")

    @router.post("/api/scripts/fill-log-times")
    async def run_fill_log_times(file: UploadFile):
        import datetime as _dt
        import io
        import re
        import tempfile
        import uuid
        from collections import defaultdict

        MARKET_IDS = {
            "NYC": 1, "CMP": 2, "HOU": 3, "SFO": 4,
            "SEA": 5, "LAX": 6, "CVC": 7, "WDC": 8, "DAL": 10,
        }
        COL_DATE = 2
        COL_SHOW = 8
        COL_COMMENTS = 9
        COL_TYPE = 14
        FPS = 29.97

        filename = file.filename or "log.xlsm"
        name_upper = filename.upper()
        market_id = next((mid for code, mid in MARKET_IDS.items() if code in name_upper), None)
        if market_id is None:
            raise HTTPException(status_code=400,
                detail=f"Could not detect market from filename '{filename}'. "
                       f"Expected one of: {', '.join(MARKET_IDS)}.")

        raw = await file.read()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), keep_vba=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not open workbook: {exc}")

        market_code = next(k for k, v in MARKET_IDS.items() if v == market_id)
        messages = [f"Market: {market_code} (station {market_id})"]
        total_filled = 0

        project_root = Path(__file__).parent.parent.parent.parent
        sys.path.insert(0, str(project_root))
        from browser_automation.etere_direct_client import connect

        with connect() as conn:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                pending = defaultdict(list)

                for row in ws.iter_rows(min_row=2):
                    spot_type = row[COL_TYPE - 1].value
                    if not spot_type or str(spot_type).upper() == "PRG":
                        continue
                    comment = row[COL_COMMENTS - 1].value
                    if comment:
                        continue
                    date_val = row[COL_DATE - 1].value
                    if not isinstance(date_val, _dt.datetime):
                        continue
                    show_name = row[COL_SHOW - 1].value or ""
                    m = re.match(r"^([^:]+):", show_name.strip())
                    if not m:
                        continue
                    asset_code = m.group(1).strip()
                    pending[(date_val.date(), asset_code)].append(row[0].row)

                if not pending:
                    continue

                cur = conn.cursor(as_dict=True)
                sheet_filled = 0
                for (date, asset_code), row_nums in pending.items():
                    cur.execute(
                        "SELECT ORA FROM TPALINSE"
                        " WHERE DATA = %s AND TITLE LIKE %s AND COD_USER = %d"
                        " ORDER BY ORA",
                        (date, f"%{asset_code}%", market_id),
                    )
                    oras = [r["ORA"] for r in cur.fetchall()]
                    if not oras:
                        messages.append(f"  No TPALINSE match: {asset_code} on {date}")
                        continue
                    if len(oras) < len(row_nums):
                        messages.append(
                            f"  WARNING: {asset_code} on {date} — "
                            f"{len(row_nums)} log rows but only {len(oras)} Etere entries"
                        )
                    for i, row_num in enumerate(row_nums):
                        if i >= len(oras):
                            break
                        secs = round(oras[i] / FPS)
                        ws.cell(row_num, COL_COMMENTS).value = _dt.timedelta(seconds=secs)
                        sheet_filled += 1

                if sheet_filled:
                    messages.append(f"  {sheet_name}: filled {sheet_filled} spot time(s)")
                total_filled += sheet_filled

        if total_filled == 0:
            return JSONResponse({"filled": 0, "messages": messages})

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        token = str(uuid.uuid4())
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm")
        tmp.write(out.read())
        tmp.close()
        _fill_log_pending[token] = (Path(tmp.name), filename)

        return JSONResponse({"filled": total_filled, "messages": messages,
                             "token": token, "filename": filename})

    @router.get("/api/scripts/fill-log-times/download/{token}")
    async def download_filled_log(token: str):
        entry = _fill_log_pending.pop(token, None)
        if entry is None:
            raise HTTPException(status_code=404, detail="Download expired or not found.")
        tmp_path, filename = entry
        data = tmp_path.read_bytes()
        tmp_path.unlink(missing_ok=True)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── Worldlink Room Maker ─────────────────────────────────────────────────

    @router.get("/scripts/worldlink-room", response_class=HTMLResponse)
    async def worldlink_room_page(request: Request):
        return templates.TemplateResponse(request, "scripts/worldlink_room.html")

    @router.get("/api/scripts/worldlink-room/spots")
    async def worldlink_room_spots(
        market: str = Query(...),
        date_from: str = Query(...),
        date_to: str = Query(...),
        time_from: str = Query(...),
        time_to: str = Query(...),
    ):
        from browser_automation.etere_direct_client import connect as _edc
        from src.domain.enums import Market as _M
        FPS = 30
        def _to_frames(t):
            h, m = int(t[:2]), int(t[3:5])
            return (h * 3600 + m * 60) * FPS
        def _from_frames(f):
            s = f // FPS
            return f"{s // 3600:02d}:{(s % 3600) // 60:02d}"
        try:
            mkt_id = _M[market.upper()].etere_id
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Unknown market: {market}")
        conn = _edc()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT tp.ID_TPALINSE, tp.DATA, tp.ORA, tp.DURATION,
                       ttp.ID_ContrattiRighe,
                       cr.DESCRIZIONE, ct.COD_CONTRATTO, ct.DESCRIZIONE
                FROM TPALINSE tp
                JOIN trafficPalinse ttp ON ttp.id_tpalinse = tp.ID_TPALINSE
                JOIN CONTRATTIRIGHE cr ON ttp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                JOIN CONTRATTITESTATA ct ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                WHERE ct.AGENZIA = 133
                  AND tp.COD_USER = %s
                  AND tp.DATA >= %s AND tp.DATA <= %s
                  AND tp.ORA >= %s AND tp.ORA <= %s
                  AND tp.LIVELLO = 0
                ORDER BY tp.DATA, tp.ORA
            """, (mkt_id, date_from, date_to, _to_frames(time_from), _to_frames(time_to)))
            rows = cur.fetchall()
            def _dur(f):
                s = round(f / 29.97)
                return f":{s:02d}" if s < 60 else f"{s//60}:{s%60:02d}"
            return JSONResponse([{
                "id": r[0],
                "date": str(r[1])[:10],
                "time": _from_frames(r[2]),
                "duration": _dur(r[3]) if r[3] else "",
                "line_id": r[4],
                "line_desc": r[5] or "",
                "contract_code": r[6] or "",
                "contract_desc": r[7] or "",
            } for r in rows])
        finally:
            conn.close()

    @router.post("/api/scripts/worldlink-room/blacklist")
    async def worldlink_room_blacklist(body: dict = Body(...)):
        from browser_automation.etere_direct_client import connect as _edc
        spot_ids = body.get("spot_ids", [])
        if not spot_ids:
            return JSONResponse({"blacklisted": 0})
        conn = _edc()
        try:
            cur = conn.cursor()
            done = 0
            for tpa_id in spot_ids:
                cur.execute("""
                    SELECT ttp.ID_ContrattiRighe, cr.DATA_INIZIO, cr.DATA_FINE,
                           tp.ID_FILMATI
                    FROM trafficPalinse ttp
                    JOIN CONTRATTIRIGHE cr ON ttp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                    JOIN TPALINSE tp ON tp.ID_TPALINSE = ttp.id_tpalinse
                    WHERE ttp.id_tpalinse = %s
                """, (tpa_id,))
                row = cur.fetchone()
                if not row:
                    continue
                line_id, d_from, d_to, filmati_id = row
                # Preserve filmati ID so Etere (or our restore step) can re-apply it
                filmati_id = filmati_id if filmati_id and filmati_id > 0 else -1
                cur.execute("DELETE FROM trafficPalinse WHERE id_tpalinse = %s", (tpa_id,))
                cur.execute("DELETE FROM TPALINSE WHERE ID_TPALINSE = %s", (tpa_id,))
                cur.execute("""
                    SELECT ID_TrafficScheduleList FROM Traffic_ScheduleList
                    WHERE ID_ContrattiRighe = %s AND BlackList > 0
                """, (line_id,))
                if cur.fetchone() is None:
                    cur.execute("""
                        INSERT INTO Traffic_ScheduleList
                          (ID_ContrattiRighe, BlackList, PassageMiss,
                           ID_TRAFFICPALINSE, Date, ToDate,
                           Notes, Operator,
                           ID_FILMATI, ID_FILMATI_TAIL, ID_FILMATI_MIDDLE,
                           ID_FATTURAEMITTENTE, Split)
                        VALUES (%s,1,1,%s,%s,%s,%s,%s,%s,-1,-1,0,0)
                    """, (line_id, tpa_id, d_from, d_to, "WL room", "Portal", filmati_id))
                else:
                    cur.execute("""
                        UPDATE Traffic_ScheduleList
                        SET PassageMiss = PassageMiss + 1
                        WHERE ID_ContrattiRighe = %s AND BlackList > 0
                    """, (line_id,))
                conn.commit()
                done += 1
            return JSONResponse({"blacklisted": done})
        finally:
            conn.close()

    @router.get("/api/scripts/worldlink-room/blacklisted")
    async def worldlink_room_blacklisted(market: str = Query(...)):
        from browser_automation.etere_direct_client import connect as _edc
        from src.domain.enums import Market as _M
        try:
            mkt_id = _M[market.upper()].etere_id
        except KeyError:
            raise HTTPException(status_code=400, detail=f"Unknown market: {market}")
        conn = _edc()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT tsl.ID_TrafficScheduleList,
                       tsl.ID_ContrattiRighe,
                       tsl.PassageMiss,
                       tsl.Date, tsl.ToDate,
                       cr.DESCRIZIONE, ct.COD_CONTRATTO
                FROM Traffic_ScheduleList tsl
                JOIN CONTRATTIRIGHE cr ON tsl.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                JOIN CONTRATTITESTATA ct ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                WHERE ct.AGENZIA = 133
                  AND tsl.BlackList > 0
                  AND cr.COD_USER = %s
                ORDER BY tsl.Date, ct.COD_CONTRATTO
            """, (mkt_id,))
            rows = cur.fetchall()
            return JSONResponse([{
                "tsl_id": r[0],
                "line_id": r[1],
                "count": r[2],
                "date_from": str(r[3])[:10] if r[3] else "",
                "date_to": str(r[4])[:10] if r[4] else "",
                "line_desc": r[5] or "",
                "contract_code": r[6] or "",
            } for r in rows])
        finally:
            conn.close()

    @router.post("/api/scripts/worldlink-room/restore")
    async def worldlink_room_restore(body: dict = Body(...)):
        from browser_automation.etere_direct_client import connect as _edc
        line_ids = body.get("line_ids", [])
        if not line_ids:
            return JSONResponse({"restored": 0, "filmati": {}})
        conn = _edc()
        try:
            cur = conn.cursor()
            total = 0
            filmati = {}  # {line_id: {filmati_id, snapshot_max_tpa_id}}
            for line_id in line_ids:
                # Read filmati stored during blacklist
                cur.execute("""
                    SELECT ID_FILMATI FROM Traffic_ScheduleList
                    WHERE ID_ContrattiRighe = %s AND BlackList > 0
                """, (line_id,))
                tsl_row = cur.fetchone()
                stored_filmati = tsl_row[0] if tsl_row and tsl_row[0] and tsl_row[0] > 0 else None
                # Snapshot max TPALINSE ID — new spots will have higher IDs
                cur.execute("""
                    SELECT ISNULL(MAX(tp.ID_TPALINSE), 0)
                    FROM TPALINSE tp
                    JOIN trafficPalinse ttp ON ttp.id_tpalinse = tp.ID_TPALINSE
                    WHERE ttp.ID_ContrattiRighe = %s AND tp.LIVELLO = 0
                """, (line_id,))
                snapshot_max = cur.fetchone()[0]
                cur.execute("""
                    DELETE FROM Traffic_ScheduleList
                    WHERE ID_ContrattiRighe = %s AND BlackList > 0
                """, (line_id,))
                total += cur.rowcount
                cur.execute("""
                    UPDATE CONTRATTIRIGHE
                    SET ROWSTATUS = 0,
                        SCHEDULESTATUS = NULL,
                        SCHEDULELASTUPD = NULL,
                        SCHEDULEMSG = ''
                    WHERE ID_CONTRATTIRIGHE = %s
                """, (line_id,))
                conn.commit()
                if stored_filmati:
                    filmati[str(line_id)] = {"filmati_id": stored_filmati, "snapshot_max": snapshot_max}
            return JSONResponse({"restored": total, "filmati": filmati})
        finally:
            conn.close()

    @router.post("/api/scripts/worldlink-room/reapply-filmati")
    async def worldlink_room_reapply_filmati(body: dict = Body(...)):
        """Find new TPALINSE rows created after restore and apply stored filmati."""
        from browser_automation.etere_direct_client import connect as _edc
        filmati = body.get("filmati", {})  # {line_id: {filmati_id, snapshot_max}}
        if not filmati:
            return JSONResponse({"applied": 0, "pending": 0})
        conn = _edc()
        try:
            cur = conn.cursor()
            applied = 0
            pending = 0
            for line_id_str, info in filmati.items():
                line_id = int(line_id_str)
                filmati_id = info["filmati_id"]
                snapshot_max = info["snapshot_max"]
                # Find new TPALINSE rows for this line created after the snapshot
                cur.execute("""
                    SELECT tp.ID_TPALINSE
                    FROM TPALINSE tp
                    JOIN trafficPalinse ttp ON ttp.id_tpalinse = tp.ID_TPALINSE
                    WHERE ttp.ID_ContrattiRighe = %s
                      AND tp.ID_TPALINSE > %s
                      AND tp.LIVELLO = 0
                """, (line_id, snapshot_max))
                new_rows = [r[0] for r in cur.fetchall()]
                if new_rows:
                    for tpa_id in new_rows:
                        cur.execute(
                            "UPDATE TPALINSE SET ID_FILMATI = %s WHERE ID_TPALINSE = %s",
                            (filmati_id, tpa_id)
                        )
                        applied += 1
                    conn.commit()
                else:
                    pending += 1  # Etere hasn't scheduled this line yet
            return JSONResponse({"applied": applied, "pending": pending})
        finally:
            conn.close()

    @router.get("/api/scripts/block-refresh")
    async def run_block_refresh(contract_id: int = Query(..., gt=0)):
        project_root = Path(__file__).parent.parent.parent.parent
        script_path = project_root / "scripts" / "refresh_blocks_contract.py"

        python_exe = project_root / ".venv" / "Scripts" / "python.exe"
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                str(python_exe), "-u", str(script_path), str(contract_id),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(project_root),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    # ------------------------------------------------------------------
    # Order run logs
    # ------------------------------------------------------------------

    @router.get("/logs", response_class=HTMLResponse)
    async def logs_page(request: Request):
        return templates.TemplateResponse(request, "logs.html")

    @router.get("/api/logs")
    async def list_logs():
        logs_dir = Path(__file__).parent.parent.parent.parent / "logs"
        if not logs_dir.exists():
            return JSONResponse({"files": []})

        files = sorted(
            logs_dir.glob("*.log"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )

        result = []
        for f in files:
            stat = f.stat()
            # Parse date/time from filename: YYYY-MM-DD_HH-MM-SS.log
            try:
                dt = datetime.strptime(f.stem, "%Y-%m-%d_%H-%M-%S")
                display_date = dt.strftime("%b %d, %Y")
                display_time = dt.strftime("%I:%M %p").lstrip("0")
            except ValueError:
                display_date = f.stem
                display_time = ""
            size_bytes = stat.st_size
            if size_bytes >= 1024:
                size_label = f"{size_bytes / 1024:.1f} KB"
            else:
                size_label = f"{size_bytes} B"
            result.append({
                "filename":     f.name,
                "display_date": display_date,
                "display_time": display_time,
                "size_label":   size_label,
                "mtime":        stat.st_mtime,
            })
        return JSONResponse({"files": result})

    @router.delete("/api/logs/cleanup")
    async def cleanup_logs():
        import time
        logs_dir = Path(__file__).parent.parent.parent.parent / "logs"
        if not logs_dir.exists():
            return JSONResponse({"deleted": 0})
        cutoff = time.time() - 30 * 24 * 3600
        deleted = 0
        for f in logs_dir.glob("*.log"):
            if f.stat().st_mtime < cutoff:
                f.unlink()
                deleted += 1
        return JSONResponse({"deleted": deleted})

    @router.get("/api/logs/{filename}")
    async def get_log(filename: str):
        # Prevent path traversal — only .log files, no slashes or dots in name
        if "/" in filename or "\\" in filename or not filename.endswith(".log"):
            raise HTTPException(status_code=400, detail="Invalid filename.")
        logs_dir = Path(__file__).parent.parent.parent.parent / "logs"
        log_path = (logs_dir / filename).resolve()
        if not str(log_path).startswith(str(logs_dir.resolve())):
            raise HTTPException(status_code=403, detail="Forbidden.")
        if not log_path.exists():
            raise HTTPException(status_code=404, detail="Log not found.")
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(log_path.read_text(encoding="utf-8", errors="replace"))

    @router.get("/scripts/clean-asterisks", response_class=HTMLResponse)
    async def clean_asterisks_page(request: Request):
        return templates.TemplateResponse(request, "scripts/clean_asterisks.html")

    @router.get("/api/scripts/clean-asterisks/lines")
    async def get_asterisk_lines(contract_id: int = Query(..., gt=0)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                # pymssql uses %s paramstyle; pyodbc uses ?
                ph = '%s' if type(conn).__module__.startswith('pymssql') else '?'
                cursor = conn.cursor()
                cursor.execute(f"""
                    SELECT ID_CONTRATTIRIGHE, DESCRIZIONE
                    FROM   CONTRATTIRIGHE
                    WHERE  ID_CONTRATTITESTATA = {ph}
                      AND  DESCRIZIONE LIKE {ph}
                    ORDER  BY ID_CONTRATTIRIGHE
                """, [contract_id, '%*%'])
                rows = cursor.fetchall()

            lines = [
                {
                    "line_id":     row[0],
                    "description": row[1] or "",
                    "cleaned":     (row[1] or "").replace("*", "").strip(),
                }
                for row in rows
            ]
            return JSONResponse({"contract_id": contract_id, "lines": lines})

        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/clean-asterisks/apply")
    async def apply_clean_asterisks(payload: dict = Body(...)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            contract_id = int(payload.get("contract_id", 0))
            if not contract_id:
                raise HTTPException(status_code=400, detail="contract_id required.")

            with _db_connect() as conn:
                ph = '%s' if type(conn).__module__.startswith('pymssql') else '?'
                cursor = conn.cursor()
                cursor.execute(f"""
                    UPDATE CONTRATTIRIGHE
                    SET    DESCRIZIONE = RTRIM(LTRIM(REPLACE(DESCRIZIONE, '*', '')))
                    WHERE  ID_CONTRATTITESTATA = {ph}
                      AND  DESCRIZIONE LIKE {ph}
                """, [contract_id, '%*%'])
                conn.commit()
                updated = cursor.rowcount

            return JSONResponse({"updated": updated})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ------------------------------------------------------------------
    # Pending orders
    # ------------------------------------------------------------------

    @router.get("/api/orders")
    async def list_orders():
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, _sweep_entered_strays)
        result = await loop.run_in_executor(None, _scan_dir, config.incoming_dir)
        return JSONResponse(content=result)

    @router.get("/api/orders/counts")
    async def order_counts():
        """Cheap tab-badge counts for all three queue states — plain directory
        listings, no order detection, so all badges can populate on page load."""
        def _count_files(d: Path) -> int:
            if not d.exists():
                return 0
            return sum(
                1 for f in d.iterdir()
                if f.is_file() and not f.name.startswith(('.', '~$'))
                and not f.name.endswith('.manifest.json')
            )
        def _run():
            _sweep_entered_strays()
            return {
                "pending":  _count_files(config.incoming_dir),
                "awaiting": len(list(entered_dir.glob("*.manifest.json"))) if entered_dir.exists() else 0,
                "history":  _count_files(used_dir),
            }
        loop = asyncio.get_running_loop()
        return JSONResponse(content=await loop.run_in_executor(None, _run))

    @router.post("/api/orders/upload")
    async def upload_order(file: UploadFile):
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided.")

        suffix = Path(file.filename).suffix.lower()
        if suffix not in _ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"File type '{suffix}' not allowed. Accepted: {', '.join(_ALLOWED_EXTENSIONS)}"
            )

        dest = (config.incoming_dir / file.filename).resolve()
        if not str(dest).startswith(str(config.incoming_dir.resolve())):
            raise HTTPException(status_code=400, detail="Invalid filename.")

        contents = await file.read()
        dest.write_bytes(contents)
        return JSONResponse(content={"message": f"'{file.filename}' uploaded successfully.", "filename": file.filename})

    @router.delete("/api/orders/{filename:path}")
    async def move_to_used(filename: str):
        """Move a pending order to incoming/Used/ (soft delete)."""
        target = (config.incoming_dir / filename).resolve()
        if not str(target).startswith(str(config.incoming_dir.resolve())):
            raise HTTPException(status_code=400, detail="Invalid filename.")
        if not target.exists():
            raise HTTPException(status_code=404, detail="File not found.")

        _ensure_used_dir()
        dest = used_dir / target.name
        # If a file with the same name already exists in Used, append a timestamp
        if dest.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = used_dir / f"{target.stem}_{ts}{target.suffix}"

        try:
            shutil.move(str(target), str(dest))
        except OSError as exc:
            # Most common on Windows: the file is still held open by another
            # process (a PDF viewer, or a just-finished run console window that
            # still has a handle). Return JSON so the UI shows a real message
            # instead of an unparseable 500.
            return JSONResponse(
                status_code=409,
                content={"detail": (
                    f"Could not move '{filename}': {exc}. "
                    "If it's open in another program (a PDF viewer, or a run "
                    "window still showing 'Press Enter to close'), close it and try again."
                )},
            )

        # Best-effort cleanup — never let these fail a move that already succeeded.
        try:
            _purge_used_folder(days=30)
        except Exception as exc:
            print(f"[orders] Used/ purge skipped: {exc}")
        try:
            sidecar = Path(str(target) + ".ai.json")  # AI-fallback extraction cache, if any
            if sidecar.exists():
                sidecar.unlink()
        except Exception:
            pass

        return JSONResponse(content={"message": f"'{filename}' moved to Used."})

    # ------------------------------------------------------------------
    # Run queue
    # ------------------------------------------------------------------

    @router.post("/api/run")
    async def run_queue(body: Any = Body(default=[])):
        # Accept both legacy list format and new {files, overrides} format
        if isinstance(body, list):
            files: list[str] = body
            overrides: dict = {}
        else:
            files = body.get("files", [])
            overrides = body.get("overrides", {})

        _order_patterns = ("*.pdf", "*.xml", "*.xlsx", "*.xlsm", "*.jpg", "*.jpeg", "*.png")
        if not any(f for pat in _order_patterns for f in config.incoming_dir.glob(pat)):
            raise HTTPException(status_code=400, detail="No orders in queue.")

        # Write per-file override sidecars so the automation can consume them
        for filename, file_overrides in overrides.items():
            sidecar = config.incoming_dir / (filename + ".overrides.json")
            try:
                sidecar.write_text(json.dumps(file_overrides))
            except Exception:
                pass

        project_root = Path(__file__).parent.parent.parent.parent
        main_py = project_root / "main.py"

        # Prefer venv python, fall back to current interpreter
        python_exe = project_root / ".venv" / "Scripts" / "python.exe"  # Windows
        if not python_exe.exists():
            python_exe = project_root / ".venv" / "bin" / "python"       # Linux
        if not python_exe.exists():
            python_exe = Path(sys.executable)

        # Validate --files args if specific files selected
        safe_files = []
        if files:
            for f in files:
                p = (config.incoming_dir / f).resolve()
                if str(p).startswith(str(config.incoming_dir.resolve())) and p.exists():
                    safe_files.append(f)

        n = len(files) if files else "all"
        if sys.platform == "win32":
            # Use argument list + CREATE_NEW_CONSOLE — avoids shell=True which
            # causes cmd.exe to split filenames containing & at the shell level.
            args = [str(python_exe), str(main_py), "--pause"]
            if safe_files:
                args += ["--files"] + safe_files
            subprocess.Popen(args, cwd=str(project_root),
                             creationflags=subprocess.CREATE_NEW_CONSOLE)
            return JSONResponse({"message": f"Terminal opened — processing {n} order(s)."})
        else:
            return JSONResponse({"terminal": "sse", "files": safe_files,
                                 "message": f"Opening terminal — processing {n} order(s)."})

    # ------------------------------------------------------------------
    # Web terminal — SSE output + POST input (Linux/Mac)
    # ------------------------------------------------------------------

    _terminal_sessions: dict[str, asyncio.subprocess.Process] = {}

    @router.get("/api/terminal/stream")
    async def terminal_stream(files: str = Query("")):
        session_id = _uuid.uuid4().hex[:8]

        _project_root = Path(__file__).parent.parent.parent.parent
        _python_exe = _project_root / ".venv" / "bin" / "python"
        if not _python_exe.exists():
            _python_exe = Path(sys.executable)

        safe = [f for f in files.split(",") if f.strip()] if files else []
        validated: list[str] = []
        for f in safe:
            p = (config.incoming_dir / f).resolve()
            if str(p).startswith(str(config.incoming_dir.resolve())):
                validated.append(f)

        args = [str(_python_exe), "-u", str(_project_root / "main.py")]
        if validated:
            args += ["--files"] + validated

        proc = await asyncio.create_subprocess_exec(
            *args,
            stdin=asyncio.subprocess.PIPE,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            cwd=str(_project_root),
        )
        _terminal_sessions[session_id] = proc

        async def event_gen():
            yield f"data: {json.dumps({'type': 'session', 'id': session_id})}\n\n"
            try:
                while True:
                    chunk = await proc.stdout.read(4096)
                    if not chunk:
                        break
                    yield f"data: {json.dumps({'type': 'output', 'text': chunk.decode(errors='replace')})}\n\n"
            except Exception:
                pass
            finally:
                _terminal_sessions.pop(session_id, None)
            yield f"data: {json.dumps({'type': 'done'})}\n\n"

        return StreamingResponse(
            event_gen(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.post("/api/terminal/{session_id}/input")
    async def terminal_input(session_id: str, body: dict = Body(...)):
        proc = _terminal_sessions.get(session_id)
        if not proc or proc.stdin is None or proc.stdin.is_closing():
            raise HTTPException(404, "Session not found or closed")
        proc.stdin.write((body.get("text", "") + "\n").encode())
        await proc.stdin.drain()
        return JSONResponse({"ok": True})

    @router.post("/api/terminal/{session_id}/kill")
    async def terminal_kill(session_id: str):
        proc = _terminal_sessions.pop(session_id, None)
        if proc:
            try:
                proc.terminate()
            except Exception:
                pass
        return JSONResponse({"ok": True})

    # ------------------------------------------------------------------
    # History (Used folder)
    # ------------------------------------------------------------------

    @router.get("/api/history")
    async def list_history():
        return JSONResponse(content=_scan_dir(used_dir))

    @router.post("/api/history/{filename:path}/restore")
    async def restore_order(filename: str):
        """Move a file from Used back to incoming/."""
        target = (used_dir / filename).resolve()
        if not str(target).startswith(str(used_dir.resolve())):
            raise HTTPException(status_code=400, detail="Invalid filename.")
        if not target.exists():
            raise HTTPException(status_code=404, detail="File not found.")

        dest = config.incoming_dir / target.name
        if dest.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = config.incoming_dir / f"{target.stem}_{ts}{target.suffix}"

        shutil.move(str(target), str(dest))
        return JSONResponse(content={"message": f"'{filename}' restored to incoming."})

    # ------------------------------------------------------------------
    # Awaiting Backwrite (Entered folder — tasks/backwrite-pipeline.md Phase 1)
    # ------------------------------------------------------------------

    @router.get("/api/orders/awaiting-backwrite")
    async def list_awaiting_backwrite():
        """Entered orders awaiting backwrite — one row per manifest."""
        def _run():
            _sweep_entered_strays()
            rows = []
            if not entered_dir.exists():
                return rows
            for mf in entered_dir.glob("*.manifest.json"):
                try:
                    m = json.loads(mf.read_text(encoding="utf-8"))
                except Exception:
                    rows.append({
                        "filename": mf.name[: -len(".manifest.json")],
                        "order_type": "Unknown", "agency_label": "Unknown",
                        "customer_name": "(unreadable manifest)",
                        "entered_at": "", "contracts": [],
                        "io_parse_error": True, "io_present": False,
                    })
                    continue
                io_name = m.get("io_filename") or mf.name[: -len(".manifest.json")]
                ov = (m.get("order_type") or "Unknown")
                detail = m.get("io_detail") or {}
                rows.append({
                    "filename": io_name,
                    "order_type": ov,
                    "agency_label": "TH Media" if ov == "eqc" else ov,
                    "customer_name": m.get("customer_name") or detail.get("client") or "Unknown",
                    "entered_at": m.get("entered_at") or "",
                    "contracts": m.get("contracts") or [],
                    "rates_are_net": bool(m.get("rates_are_net")),
                    "io_parse_error": bool(m.get("io_parse_error")),
                    "io_present": (entered_dir / io_name).exists(),
                })
            rows.sort(key=lambda r: r["entered_at"], reverse=True)
            return rows
        loop = asyncio.get_running_loop()
        return JSONResponse(content=await loop.run_in_executor(None, _run))

    def _resolve_etere_id(c: dict) -> int | None:
        """The contract's Etere DB id. Prefer the manifest's stored etere_id;
        if absent — older manifests, or a multi-contract parser that recorded
        the id as the code and never set etere_id — resolve it live from the
        code via COD_CONTRATTO (exact then prefix, the same lookup the batch
        enricher uses), and finally treat an all-digit code as the id itself.
        Returns None only when nothing matches."""
        if c.get("etere_id"):
            return int(c["etere_id"])
        code = str(c.get("code") or "").strip()
        if not code:
            return None
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                ph = "%s" if type(conn).__module__.startswith("pymssql") else "?"
                cur = conn.cursor()
                cur.execute(
                    f"SELECT TOP 1 ID_CONTRATTITESTATA FROM CONTRATTITESTATA "
                    f"WHERE COD_CONTRATTO = {ph} ORDER BY ID_CONTRATTITESTATA DESC", (code,))
                row = cur.fetchone()
                if not row:
                    like = code.replace("[", "[[]").replace("%", "[%]").replace("_", "[_]") + "%"
                    cur.execute(
                        f"SELECT TOP 1 ID_CONTRATTITESTATA FROM CONTRATTITESTATA "
                        f"WHERE COD_CONTRATTO LIKE {ph} ORDER BY ID_CONTRATTITESTATA DESC", (like,))
                    row = cur.fetchone()
                if not row and code.isdigit():
                    cur.execute(
                        f"SELECT ID_CONTRATTITESTATA FROM CONTRATTITESTATA "
                        f"WHERE ID_CONTRATTITESTATA = {ph}", (int(code),))
                    row = cur.fetchone()
                return int(row[0]) if row else None
        except Exception:  # noqa: BLE001 - best-effort; caller reports "not found"
            return None

    def _contact_from_anagraf(cur, etere_id: int) -> dict:
        """Poll Etere ANAGRAF for the bill-to's contact block — the agency when
        the contract has one, else the client (committente), matching how the
        backwrite bill code is chosen. Every field falls back to '' when ANAGRAF
        has no value; the user fills gaps at review time (Phase 4). Nothing is
        persisted on our side — ANAGRAF is the single source of truth (customers.db
        was rejected: it drifts across machines)."""
        cur.execute(
            """SELECT LTRIM(RTRIM(ISNULL(bt.VIA, ''))),
                      LTRIM(RTRIM(ISNULL(bt.CITTA, ''))),
                      LTRIM(RTRIM(ISNULL(bt.PROVINCIA, ''))),
                      LTRIM(RTRIM(ISNULL(bt.CAP, ''))),
                      LTRIM(RTRIM(ISNULL(bt.TELEFONO, ''))),
                      LTRIM(RTRIM(ISNULL(bt.FAX, ''))),
                      LTRIM(RTRIM(ISNULL(bt.E_MAIL, ''))),
                      LTRIM(RTRIM(ISNULL(bt.Nome, '') + ' ' + ISNULL(bt.NomeDue, '')))
               FROM CONTRATTITESTATA ct
               LEFT JOIN ANAGRAF bt
                 ON bt.ID_ANAGRAF = COALESCE(NULLIF(ct.AGENZIA, 0), ct.COMMITTENTE)
               WHERE ct.ID_CONTRATTITESTATA = %s""",
            (int(etere_id),),
        )
        r = cur.fetchone() or []
        g = lambda i: (r[i] or "").strip() if i < len(r) else ""  # noqa: E731
        return {
            "address": g(0), "city": g(1), "state": g(2), "zip": g(3),
            "phone": g(4), "fax": g(5), "email_1": g(6), "contact_person": g(7),
        }

    _CONTACT_KEYS = ("contact_person", "address", "city", "state", "zip",
                     "phone", "fax", "email_1", "email_2", "email_3", "email_4")

    @router.get("/api/orders/awaiting-backwrite/{filename:path}/contact")
    async def awaiting_backwrite_contact(filename: str, contract_index: int = 0):
        """Poll ANAGRAF for the bill-to contact block so the Backwrite review
        modal can prefill it (Phase 4). Live Etere read — no persistence."""
        mf = (entered_dir / f"{filename}.manifest.json").resolve()
        if not str(mf).startswith(str(entered_dir.resolve())):
            return JSONResponse(status_code=400, content={"detail": "Invalid filename."})
        if not mf.exists():
            return JSONResponse(status_code=404, content={"detail": "Manifest not found."})
        m = json.loads(mf.read_text(encoding="utf-8"))
        contracts = m.get("contracts") or []
        if not contracts:
            return JSONResponse(status_code=409, content={"detail": "Manifest has no contracts."})
        c = contracts[min(contract_index, len(contracts) - 1)]

        def _run():
            etere_id = _resolve_etere_id(c)
            if not etere_id:
                return None, None, None, [], []
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                contact = _contact_from_anagraf(conn.cursor(), int(etere_id))
                cur = conn.cursor()
                cur.execute(
                    "SELECT COD_CONTRATTO, ISNULL(CUSTOMERREF, '') FROM CONTRATTITESTATA"
                    " WHERE ID_CONTRATTITESTATA = %s", (int(etere_id),))
                row = cur.fetchone() or ("", "")
                header_info = {"code": (row[0] or "").strip(),
                               "customer_ref": (row[1] or "").strip()}
            # Per-line detected languages for the review modal — same detection
            # the generate step runs, so what the user confirms is what applies.
            # Best-effort: no scheduled spots yet (or EtereBridge absent) just
            # hides the language section, it never blocks the contact prefill.
            languages, language_options = [], []
            try:
                from backwrite.eterebridge_runner import (
                    build_placement_csv_from_db,
                    get_language_details,
                    get_language_options,
                )
                csv_bytes = build_placement_csv_from_db(int(etere_id))
                languages = get_language_details(csv_bytes)
                language_options = list(dict.fromkeys(get_language_options()))
            except Exception:  # noqa: BLE001 - review modal degrades gracefully
                pass
            return etere_id, contact, header_info, languages, language_options

        loop = asyncio.get_running_loop()
        etere_id, contact, header_info, languages, language_options = await loop.run_in_executor(None, _run)
        if not etere_id:
            return JSONResponse(status_code=409, content={"detail": (
                f"Contract {c.get('code')} could not be matched in Etere "
                "(no stored ID and its code did not resolve). Use the legacy "
                "Backwrite page for this one.")})
        # Estimate defaults — same derivation the generate step uses, surfaced
        # so the operator can confirm/correct them in the modal (Maija 7/17):
        # 'estimate' fills the front/cover page, 'estimate_run' the Run Sheet
        # Estimate column.
        gi = (m.get("user_inputs") or [{}])[0] or {}
        gorder = gi.get("order") if isinstance(gi, dict) else {}
        gorder = gorder if isinstance(gorder, dict) else {}
        estimates = m.get("estimates") or []
        estimate = str(gorder.get("estimate_number") or (estimates[0] if estimates else "") or "")
        estimate_run = ""
        # Fall back to the Etere customer order ref (CUSTOMERREF) when the
        # manifest carries no estimate — iGraphix (and similar) store the
        # purchase/estimate number there at entry, not in the gathered inputs.
        # Authoritative: it's exactly what the parser pulled off the IO.
        if not estimate:
            estimate = (header_info or {}).get("customer_ref", "") or ""
        # Admerasia convention (Lee 7/17): front estimate = the Etere header's
        # customer order ref (e.g. '19-MD10-2607VT'); run-sheet estimate = the
        # short token from the contract code (e.g. 'Admerasia McD 19SE 2607'
        # → '19SE 2607').
        import re as _re
        hdr_code = (header_info or {}).get("code", "")
        if (m.get("order_type") or "").lower() == "admerasia" or hdr_code.lower().startswith("admerasia"):
            estimate = (header_info or {}).get("customer_ref", "") or estimate
            m_tok = _re.search(r"(\d+[A-Z]{2}\s+\d{4})\s*$", hdr_code)
            if m_tok:
                estimate_run = m_tok.group(1)

        return JSONResponse(content={
            "contract_code": str(c.get("code") or ""),
            "contact": contact,
            "estimate": estimate,
            "estimate_run": estimate_run,
            "languages": languages,
            "language_options": language_options,
        })

    def _archive_entered(filename: str) -> tuple[list, str | None]:
        """Move an entered IO + its manifest to Used/. Returns (moved, error)."""
        io_target = (entered_dir / filename).resolve()
        if not str(io_target).startswith(str(entered_dir.resolve())):
            return [], "Invalid filename."
        manifest = entered_dir / f"{filename}.manifest.json"
        if not io_target.exists() and not manifest.exists():
            return [], "File not found."
        _ensure_used_dir()
        moved = []
        for src in (io_target, manifest):
            if not src.exists():
                continue
            dest = used_dir / src.name
            if dest.exists():
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = used_dir / f"{src.stem}_{ts}{src.suffix}"
            try:
                shutil.move(str(src), str(dest))
            except OSError as exc:
                return moved, (
                    f"Could not move '{src.name}': {exc}. If it's open in "
                    "another program, close it and try again."
                )
            moved.append(src.name)
        return moved, None

    @router.post("/api/orders/awaiting-backwrite/{filename:path}/done")
    async def awaiting_backwrite_done(filename: str):
        """Archive an awaiting order (IO + manifest) to Used/.

        Manual escape hatch for the legacy-parallel period: the user
        backwrites through the legacy page, then archives here. The
        one-click backwrite endpoint archives automatically on success."""
        moved, err = _archive_entered(filename)
        if err:
            code = 400 if err == "Invalid filename." else 404 if err == "File not found." else 409
            return JSONResponse(status_code=code, content={"detail": err})
        return JSONResponse(content={"message": f"Archived {', '.join(moved)} to Used."})

    @router.post("/api/orders/awaiting-backwrite/{filename:path}/backwrite")
    async def awaiting_backwrite_generate(filename: str, body: dict = Body(default={})):
        """One-click backwrite (tasks/backwrite-pipeline.md Phase 2).

        Everything is derived — nothing re-keyed by a human:
          * manifest       → IO line structure, rates_are_net, gathered inputs
          * Etere contract → billing type (CENTROMEDIA — hard stop if unset),
                             agency % (P_AGENZIA), salesperson (AGENTE1)
          * Etere report   → the commercial-log CSV (same report the legacy
                             page fetches), which supplies the actual spots
        On success the IO + manifest are archived to Used/ automatically.
        """
        contract_idx = int(body.get("contract_index") or 0)
        # Per-line language overrides from the review modal ({description: code}).
        # Detected values echoed back unchanged are harmless no-op corrections.
        raw_langs = body.get("language_corrections") or {}
        lang_corrections = {
            str(k): str(v).strip()
            for k, v in raw_langs.items()
            if isinstance(raw_langs, dict) and str(v).strip()
        } if isinstance(raw_langs, dict) else {}

        def _run():
            mf = (entered_dir / f"{filename}.manifest.json").resolve()
            if not str(mf).startswith(str(entered_dir.resolve())):
                raise HTTPException(status_code=400, detail="Invalid filename.")
            if not mf.exists():
                raise HTTPException(status_code=404, detail="Manifest not found.")
            m = json.loads(mf.read_text(encoding="utf-8"))

            otype = (m.get("order_type") or "").lower()
            if otype == "worldlink":
                raise HTTPException(status_code=409, detail=(
                    "WorldLink orders use the dedicated WorldLink backwrite "
                    "(revision merge, MLBF tab) — open the Backwrite page."
                ))

            contracts = m.get("contracts") or []
            if not contracts:
                raise HTTPException(status_code=409, detail="Manifest has no contracts.")
            c = contracts[min(contract_idx, len(contracts) - 1)]
            etere_id = _resolve_etere_id(c)
            if not etere_id:
                raise HTTPException(status_code=409, detail=(
                    f"Contract {c.get('code')} could not be matched in Etere "
                    "(no stored ID and its code did not resolve) — cannot fetch "
                    "its commercial log. Use the legacy Backwrite page."
                ))

            # ── Live contract facts: billing window, agency %, AE ────────────
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    """SELECT ct.CENTROMEDIA, ct.P_AGENZIA,
                              LTRIM(RTRIM(CASE WHEN ae.Nome IS NOT NULL AND ae.Nome != ''
                                   THEN ae.Nome + ' ' + ae.RAG_SOCIAL ELSE ae.RAG_SOCIAL END)),
                              ct.AGENZIA
                       FROM CONTRATTITESTATA ct
                       LEFT JOIN ANAGRAF ae ON ae.ID_ANAGRAF = ct.AGENTE1
                       WHERE ct.ID_CONTRATTITESTATA = %s""",
                    (int(etere_id),),
                )
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail=f"Etere contract {etere_id} not found.")
                # Phase 4: prefill the contact block from ANAGRAF (live), then
                # apply any per-field overrides the user made in the review modal.
                contact = _contact_from_anagraf(cur, int(etere_id))
            body_contact = (body.get("contact") or {}) if isinstance(body, dict) else {}
            for k in _CONTACT_KEYS:
                if k in body_contact:
                    v = body_contact[k]
                    contact[k] = ("" if v is None else str(v)).strip()
            centromedia = int(row[0] or 0)
            p_agenzia = float(row[1] or 0)
            ae_name = (row[2] or "").strip()
            id_agenzia = int(row[3] or 0)
            if centromedia == 316:
                billing_type = "Broadcast"
            elif centromedia == 317:
                billing_type = "Calendar"
            else:
                # Never default silently — this is the template-drag error class.
                raise HTTPException(status_code=409, detail=(
                    f"Contract {c.get('code')} (ID {etere_id}) has NO billing type "
                    "set in Etere. Set Broadcast/Calendar on the contract "
                    "(Booked Business page has an editor), then retry."
                ))

            # ── Placement CSV straight from the Etere DB ─────────────────────
            # (Same data the legacy page's Etere-web report fetch returns, but
            # direct SQL: no web login, no license seat, no 3-minute report
            # queue. Same builder the WorldLink placement flow uses.)
            from backwrite.eterebridge_runner import build_placement_csv_from_db
            from backwrite.transformer import generate_excel, parse_csv
            try:
                csv_bytes = build_placement_csv_from_db(int(etere_id))
            except ValueError as exc:
                raise HTTPException(status_code=409, detail=str(exc))
            header, spots = parse_csv(csv_bytes)
            if not spots:
                raise HTTPException(status_code=409, detail=(
                    "No scheduled spots found for this contract — has it been "
                    "approved and scheduled yet?"
                ))

            # ── Inputs: manifest + live facts, zero re-keyed fields ──────────
            gi = (m.get("user_inputs") or [{}])[0] or {}
            gorder = gi.get("order") if isinstance(gi, dict) else {}
            gorder = gorder if isinstance(gorder, dict) else {}
            io_detail = m.get("io_detail") or None
            if io_detail and io_detail.get("error"):
                io_detail = None

            # Column Z ('Agency') is Agency vs Non-Agency = whether an AGENCY is
            # ATTACHED (ct.AGENZIA), NOT whether there's a commission. A 0%-
            # commission agency (e.g. Crispin / Bay Area AQMD) is still Agency.
            # Valid values are only "Agency" / "Non-Agency" — never "Direct".
            agency_flag = "Agency" if id_agenzia > 0 else "Non-Agency"
            agency_fee = p_agenzia / 100 if p_agenzia > 1 else p_agenzia
            gross_up = {}
            if m.get("rates_are_net") and agency_flag == "Agency" and io_detail:
                nets = {
                    round(float(ln.get("rate") or 0), 4)
                    for ln in io_detail.get("lines", []) if ln.get("rate")
                }
                # Map {Etere-rounded-gross: IO net} so BOTH the SC tab (reads IO
                # net) and the run sheet (reads Etere's rounded gross) gross up
                # to full precision → net lands penny-exact on the IO. {net: net}
                # only grossed the SC tab. Mirrors backwrite.py's fallback.
                if nets and (1 - agency_fee) > 0:
                    gross_up = {round(r / (1 - agency_fee), 2): r for r in nets}
                else:
                    gross_up = {r: r for r in nets}

            estimates = m.get("estimates") or []
            estimate = str(gorder.get("estimate_number") or (estimates[0] if estimates else "") or "")
            estimate_run = ""
            # The review modal shows both estimate fields (front page / run
            # sheet) — values confirmed there override the derived defaults.
            body_est = (body.get("estimates") or {}) if isinstance(body, dict) else {}
            if isinstance(body_est, dict):
                if body_est.get("estimate") is not None:
                    estimate = str(body_est["estimate"]).strip()
                if body_est.get("estimate_run") is not None:
                    estimate_run = str(body_est["estimate_run"]).strip()

            user_inputs = {
                "sales_person":   ae_name,
                "billing_type":   billing_type,
                "revenue_type":   "Internal Ad Sales",
                "agency_flag":    agency_flag,
                "agency_fee":     agency_fee,
                "estimate":       estimate,
                "estimate_run":   estimate_run,
                "contract":       str(etere_id),
                "affidavit":      "Y",
                "order_date":     "",
                "contact_person": contact.get("contact_person", ""),
                "phone":          contact.get("phone", ""),
                "fax":            contact.get("fax", ""),
                "email_1":        contact.get("email_1", ""),
                "email_2":        contact.get("email_2", ""),
                "email_3":        contact.get("email_3", ""),
                "email_4":        contact.get("email_4", ""),
                "address":        contact.get("address", ""),
                "city":           contact.get("city", ""),
                "state":          contact.get("state", ""),
                "zip":            contact.get("zip", ""),
                "notes":          str((gi.get("notes") if isinstance(gi, dict) else "") or ""),
                "gross_up_rates": gross_up,
                "language_corrections": lang_corrections,
                "revision":       "",
            }

            reconcile: dict = {}
            xlsx = generate_excel(header, spots, user_inputs, raw_csv=csv_bytes,
                                  io_detail=io_detail, validation_out=reconcile)

            # Phase 3 (tasks/backwrite-pipeline.md §2.4): compare what the IO
            # ordered (manifest) against what Etere actually scheduled. Catches
            # the July 2026 error class — a wrong INPUT (double gross-up, wrong
            # billing, revision gap) that yields a consistent-but-wrong Excel,
            # which the internal-totals check structurally cannot see.
            from backwrite.transformer import reconcile_io_vs_etere
            io_check = reconcile_io_vs_etere(
                io_detail, spots, agency_fee,
                bool(m.get("rates_are_net")), agency_flag == "Agency",
            )
            if io_check.get("messages"):
                reconcile["messages"] = list(reconcile.get("messages") or []) + io_check["messages"]
                reconcile["ok"] = bool(reconcile.get("ok", True)) and io_check["ok"]
                reconcile["io_check"] = io_check.get("detail", {})

            # The modal's language table was on screen when the user clicked
            # Generate — every value is user-verified. Persist to the
            # CTV_LineLanguage catalog so future backwrites never re-ask.
            if lang_corrections:
                from backwrite.eterebridge_runner import writeback_line_languages
                writeback_line_languages(csv_bytes, lang_corrections)

            import re as _re
            base = Path(m.get("io_filename") or filename).stem
            out_name = _re.sub(r'[\\/:*?"<>|]', "", base).strip() + ".xlsx"

            # Archive to Used/ only when everything reconciles. A flagged order
            # stays in the Awaiting queue (visible, logged) so the human can fix
            # Etere and retry — the same "never file something wrong silently"
            # philosophy as the CENTROMEDIA hard-stop above.
            archived = False
            archive_err = None
            if reconcile.get("ok", True):
                moved, archive_err = _archive_entered(filename)
                archived = bool(moved) and not archive_err
            else:
                print(f"[backwrite] {filename} NOT archived — reconciliation flagged: "
                      f"{'; '.join(reconcile.get('messages') or [])}")
            return xlsx, out_name, reconcile, archive_err, archived

        loop = asyncio.get_running_loop()
        xlsx, out_name, reconcile, archive_err, archived = await loop.run_in_executor(None, _run)

        import io as _io
        headers = {"Content-Disposition": f'attachment; filename="{out_name}"'}
        if reconcile:
            headers["X-Backwrite-Reconcile"] = json.dumps(reconcile)
        if archive_err:
            headers["X-Backwrite-Archive-Error"] = json.dumps(archive_err)
        headers["X-Backwrite-Archived"] = "1" if archived else "0"
        return StreamingResponse(
            _io.BytesIO(xlsx),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    # ------------------------------------------------------------------
    # Detail (works for both pending and history files)
    # ------------------------------------------------------------------

    def _resolve_file(filename: str) -> tuple[Path, str]:
        """Find the file in incoming/, Entered/ or Used/ and return (path, order_type)."""
        # Try incoming/ first
        for search_dir in [config.incoming_dir, entered_dir, used_dir]:
            candidate = (search_dir / filename).resolve()
            base = str(search_dir.resolve())
            if not str(candidate).startswith(base):
                continue
            if candidate.exists():
                scanner = OrderScanner(PDFOrderDetector(), search_dir)
                for o in scanner.scan_for_orders():
                    if o.pdf_path.resolve() == candidate:
                        return candidate, o.order_type.value if o.order_type else "UNKNOWN"
                return candidate, "UNKNOWN"
        raise HTTPException(status_code=404, detail="File not found.")

    @router.get("/api/orders/{filename:path}/detail")
    async def order_detail(filename: str):
        path, order_type = _resolve_file(filename)
        loop = asyncio.get_running_loop()
        detail = await loop.run_in_executor(None, get_order_detail, path, order_type)
        detail["filename"] = filename
        detail["order_type"] = order_type
        return JSONResponse(content=detail)

    @router.get("/api/history/{filename:path}/detail")
    async def history_detail(filename: str):
        target = (used_dir / filename).resolve()
        if not str(target).startswith(str(used_dir.resolve())):
            raise HTTPException(status_code=400, detail="Invalid filename.")
        if not target.exists():
            raise HTTPException(status_code=404, detail="File not found.")

        scanner = OrderScanner(PDFOrderDetector(), used_dir)
        order_type = "UNKNOWN"
        for o in scanner.scan_for_orders():
            if o.pdf_path.resolve() == target:
                order_type = o.order_type.value if o.order_type else "UNKNOWN"
                break

        detail = get_order_detail(target, order_type)
        detail["filename"] = filename
        detail["order_type"] = order_type
        return JSONResponse(content=detail)

    # ------------------------------------------------------------------
    # Traffic
    # ------------------------------------------------------------------

    _MARKET_NAMES = {
        1: "NYC", 2: "CMP", 3: "HOU", 4: "SFO", 5: "SEA",
        6: "LAX", 7: "CVC", 8: "WDC", 9: "MMT", 10: "DAL",
    }

    @router.get("/traffic", response_class=HTMLResponse)
    async def traffic_page(request: Request):
        return templates.TemplateResponse(request, "traffic.html")

    # ------------------------------------------------------------------
    # Traffic → Commercial Log sync (Lee 2026-07-17)
    #
    # After traffic is assigned or revised on a contract, the commercial
    # log's Spot Name column (H) goes stale — previously fixable only by
    # re-backwriting the whole order. Every log row carries the Etere line
    # number (column M) plus air date (B), so the sync is a deterministic
    # join: log rows grouped by (line, date) map 1:1 in time order onto
    # the line's scheduled TPALINSE spots that day; each row's name is
    # rewritten from the spot's current FILMATI creative. Rows are written
    # in _format_copy style — "TITLE (CODE)", but no suffix when the title
    # already starts with the code (redundant suffixes from older
    # EtereBridge output get cleaned). Count-mismatched groups are flagged
    # and skipped — that's a schedule revision, i.e. a true re-backwrite.
    # ------------------------------------------------------------------

    _LOG_SYNC_DEFAULT_PATH = r"K:\Traffic\Media library\Commercial Log.xlsx"
    _LOG_SYNC_SHEET = "Commercials"

    # Team's standard Custom Sort for the Commercials sheet (Lee 2026-07-17):
    # Start Date (B) → Market (AC) → Comments (I), all ascending. Excel's own
    # "last sort" memory kept drifting; log-sync now re-asserts this recipe on
    # every write so the Sort dialog always shows it. Columns are matched by
    # header text so a column move can't silently misalign the sort.
    # The team's full custom sort (Data ▸ Sort), matched by HEADER TEXT so a
    # column move can't misalign it. Recipe-only (writes sortState; never
    # physically reorders rows) — re-asserted on every write so the Sort dialog
    # always shows the real order. Captured from the reference log 2026-07-22.
    # kind: "value" = ascending cell values; ("list", "...") = custom-list order;
    # "fontColor" = sort by font color, Automatic on top.
    # NOTE: the team's real sort has a 7th-position "Time out → Font Color
    # (Automatic on top)" level. It is OMITTED here: a font-color sortCondition
    # needs a dxf with an auto/automatic color (`<dxf><font><color auto="1"/>`),
    # which Excel rejects as an invalid style and "repairs" — taking the whole
    # sortState (and mruColors) with it. Writing the other 6 levels is clean.
    # If that level is needed, it must be added manually in Excel (Data ▸ Sort);
    # a proper OOXML encoding is a TODO before re-adding it here.
    _LOG_SYNC_SORT_SPEC = (
        ("Start Date", "value"),
        ("Market",     "value"),
        ("Time In",    "value"),
        ("Type",       ("list", "PRG")),   # single-entry custom list: PRG on top, rest default
        ("Priority",   "value"),
        ("Comments",   "value"),
    )

    # Authoritative network row color per market (col AC), captured from the
    # team's reference log 2026-07-22. Used to restore a row's base network color
    # WITHOUT depending on col K having survived — makes color restore foolproof
    # even on a doc that arrived stripped. Aired rows (theme-5 orange), gray
    # non-airtime charges, bookend G/H pink, and NEED COPY red A-H are overrides
    # layered on top of this base color, not stored here.
    _LOG_SYNC_MARKET_COLORS = {
        "CVC": "FFFFFF00",  # yellow
        "SEA": "FFFFC000",  # gold/amber
        "LAX": "FFBDD7EE",  # light blue
        "SFO": "FFB2A1C7",  # lavender
        "WDC": "FFC4D331",  # lime
        "NYC": "FF00B0F0",  # sky blue
        "HOU": "FFA46A6A",  # dusty rose
        "DAL": "FF963634",  # maroon
        "MMT": "FFFFCCFF",  # light pink
        "CMP": "FF92D050",  # green
    }

    # Excel's EXACT dxf for the "Time out → Font Color (Automatic on top)" sort
    # level, captured from the team's Excel-authored log. Note it's a FULL font
    # spec — the bare "<font><color auto=1/></font>" that openpyxl produces is
    # rejected by Excel. openpyxl only models the <autoFilter> sortState, so the
    # font-color level (which Excel keeps in a worksheet-level <sortState>) is
    # dropped on every save and must be re-injected. See _log_sync_finalize_xlsx.
    _LOG_SYNC_FONTCOLOR_DXF = (
        '<dxf><font><b/><i val="0"/><strike val="0"/><condense val="0"/>'
        '<extend val="0"/><outline val="0"/><shadow val="0"/><u val="none"/>'
        '<vertAlign val="baseline"/><sz val="12"/><color auto="1"/>'
        '<name val="Arial"/><family val="2"/><scheme val="none"/></font></dxf>'
    )

    def _log_sync_apply_standard_sort(ws) -> None:
        from openpyxl.styles import Color, Font
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.filters import SortCondition, SortState
        last_row = ws.max_row
        last_col = get_column_letter(ws.max_column)
        header = {str(c.value).strip(): c.column
                  for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value}
        wb = ws.parent
        conds = []
        for name, kind in _LOG_SYNC_SORT_SPEC:
            col = header.get(name)
            if not col:
                continue
            cl = get_column_letter(col)
            ref = f"{cl}2:{cl}{last_row}"
            if isinstance(kind, tuple) and kind[0] == "list":
                conds.append(SortCondition(ref=ref, customList=kind[1]))
            elif kind == "fontColor":
                # Sort by font color, Automatic on top (dxf carries the auto color).
                dxf_id = wb._differential_styles.add(
                    DifferentialStyle(font=Font(color=Color(auto=True))))
                conds.append(SortCondition(ref=ref, sortBy="fontColor", dxfId=dxf_id))
            else:
                conds.append(SortCondition(ref=ref))
        if not conds:
            return
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.auto_filter.sortState = SortState(ref=f"A2:{last_col}{last_row}")
        ws.auto_filter.sortState.sortCondition = conds

    def _log_sync_apply_conditional_formatting(ws) -> None:
        """Re-assert the log's two conditional-formatting rules so they can never
        disappear, collapsing Excel's range fragmentation to one clean rule each:
          1. current-broadcast-week highlight on Time In/out (light-red fill +
             dark-red font);
          2. affidavit 'NO' → solid red on the affidavit-status column (AD).
        Captured from the reference log 2026-07-22. The log's ONLY CF rules are
        these two families, so replacing the whole list is safe and de-fragments.
        """
        from openpyxl.formatting.formatting import ConditionalFormattingList
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        last_row = ws.max_row
        header = {str(c.value).strip(): c.column
                  for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value}
        ws.conditional_formatting = ConditionalFormattingList()
        tin, tout = header.get("Time In"), header.get("Time out")
        if tin and tout:
            a, b = get_column_letter(tin), get_column_letter(tout)
            ws.conditional_formatting.add(
                f"{a}2:{b}{last_row}",
                FormulaRule(
                    formula=[f"AND(TODAY()-ROUNDDOWN({a}2,0)>=(WEEKDAY(TODAY())),"
                             f"TODAY()-ROUNDDOWN({a}2,0)<(WEEKDAY(TODAY())+7))"],
                    fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE",
                                     fill_type="solid"),
                    font=Font(color="FF9C0006")))
        # Affidavit-status column is unlabeled in the log; the reference uses AD.
        ws.conditional_formatting.add(
            f"AD2:AD{last_row}",
            CellIsRule(operator="equal", formula=['"NO"'],
                       fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000",
                                        fill_type="solid")))

    def _log_sync_finalize_xlsx(xlsx_bytes: bytes, fontcolor_col: str = "F") -> bytes:
        """Post-process the openpyxl-saved log to re-add two things openpyxl drops,
        each injected independently and SELF-VALIDATED (well-formed XML) so a
        problem in one skips only that piece and can never corrupt the file:

        1. Custom-color picker swatches — styles.xml <mruColors> (the 10 market
           colors; Excel caps recent colors at 10, so never more). openpyxl drops
           these on every save.
        2. The "Time out → Font Color (Automatic on top)" sort level. Excel keeps
           this in a worksheet-level <sortState> (sibling of <autoFilter>) that
           openpyxl doesn't model, so it's lost on save. We append Excel's exact
           full-font auto-color dxf to <dxfs> and inject a worksheet <sortState>
           mirroring the autoFilter one plus the font-color condition after </autoFilter>.

        Passed (via a lambda binding fontcolor_col) as the _wb_save_fast transform.
        """
        import io as _io
        import re as _re
        import xml.etree.ElementTree as _ET
        import zipfile as _zip
        _NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        try:
            with _zip.ZipFile(_io.BytesIO(xlsx_bytes)) as z:
                names = z.namelist()
                data = {n: z.read(n) for n in names}
        except Exception:
            return xlsx_bytes
        changed = False
        did = None  # index of the appended font-color dxf

        # ── (1) styles.xml: mruColors + append the font-color dxf ──────────────
        styles = data.get("xl/styles.xml", b"").decode("utf-8", "replace")
        if styles:
            palette = list(dict.fromkeys(_LOG_SYNC_MARKET_COLORS.values()))[:10]
            s2 = _re.sub(r"<colors\b[^>]*/>|<colors\b[^>]*>.*?</colors>", "",
                         styles, flags=_re.S)
            block = ("<colors><mruColors>"
                     + "".join(f'<color rgb="{c}"/>' for c in palette)
                     + "</mruColors></colors>")
            # CT_Stylesheet order: ... dxfs, tableStyles, COLORS, extLst.
            mt = _re.search(r"<tableStyles\b[^>]*/>|<tableStyles\b.*?</tableStyles>",
                            s2, _re.S)
            if mt:
                s2 = s2[:mt.end()] + block + s2[mt.end():]
            elif "<extLst" in s2:
                i = s2.find("<extLst")
                s2 = s2[:i] + block + s2[i:]
            elif "</styleSheet>" in s2:
                i = s2.find("</styleSheet>")
                s2 = s2[:i] + block + s2[i:]
            else:
                s2 = None
            if s2 is not None:
                # Append Excel's font-color dxf and remember its index.
                dm = _re.search(r'(<dxfs\b[^>]*\bcount=")(\d+)("[^>]*>)', s2)
                _try_did = None
                if dm:
                    cnt = int(dm.group(2))
                    _try_did = cnt
                    s2 = s2[:dm.start()] + dm.group(1) + str(cnt + 1) + dm.group(3) + s2[dm.end():]
                    ci = s2.find("</dxfs>")
                    if ci != -1:
                        s2 = s2[:ci] + _LOG_SYNC_FONTCOLOR_DXF + s2[ci:]
                    else:
                        _try_did = None
                # Validate: well-formed, exactly one <colors> in the right slot.
                try:
                    kids = [c.tag.replace(_NS, "") for c in _ET.fromstring(s2)]
                    ok = (kids.count("colors") == 1
                          and not ("tableStyles" in kids and kids.index("tableStyles") > kids.index("colors"))
                          and not ("extLst" in kids and kids.index("extLst") < kids.index("colors")))
                    if ok:
                        data["xl/styles.xml"] = s2.encode("utf-8")
                        did = _try_did
                        changed = True
                except Exception:
                    pass

        # ── (2) sheet: worksheet-level <sortState> with the font-color level ──
        if did is not None:
            for n in names:
                if not (n.startswith("xl/worksheets/sheet") and n.endswith(".xml")):
                    continue
                sx = data[n].decode("utf-8", "replace")
                af = _re.search(r"<autoFilter\b.*?</autoFilter>", sx, _re.S)
                if not af or "<sortState" not in af.group(0):
                    continue
                ss = _re.search(r"<sortState\b.*?</sortState>", af.group(0), _re.S).group(0)
                rm = _re.search(r'ref="A2:([A-Z]+)(\d+)"', ss)
                cs = _re.findall(r"<sortCondition\b[^>]*/>", ss)
                if not rm or not cs:
                    break
                lastc, lrr = rm.group(1), rm.group(2)
                fc = (f'<sortCondition sortBy="fontColor" '
                      f'ref="{fontcolor_col}2:{fontcolor_col}{lrr}" dxfId="{did}"/>')
                new, ins = [], False
                for c in cs:
                    new.append(c)
                    if "customList=" in c and not ins:
                        new.append(fc)
                        ins = True
                if not ins:
                    new = cs[:4] + [fc] + cs[4:]
                wss = ('<sortState ref="A2:%s%s" xmlns:xlrd2="http://schemas.microsoft.com/'
                       'office/spreadsheetml/2017/richdata2">%s</sortState>'
                       % (lastc, lrr, "".join(new)))
                aidx = sx.find("</autoFilter>") + len("</autoFilter>")
                cand = sx[:aidx] + wss + sx[aidx:]
                try:
                    _ET.fromstring(cand)  # well-formed?
                    data[n] = cand.encode("utf-8")
                    changed = True
                except Exception:
                    pass
                break

        if not changed:
            return xlsx_bytes
        try:
            out = _io.BytesIO()
            with _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED) as z:
                for n in names:
                    z.writestr(n, data[n])
            return out.getvalue()
        except Exception:
            return xlsx_bytes

    def _log_sync_path(raw: str) -> Path:
        import re as _re
        p = (raw or "").strip() or _LOG_SYNC_DEFAULT_PATH
        m = _re.match(r"^([A-Za-z]):[\\/](.*)$", p)
        if m and os.name != "nt":   # WSL/dev: translate C:\ → /mnt/c/
            p = f"/mnt/{m.group(1).lower()}/" + m.group(2).replace("\\", "/")
        return Path(p)

    def _log_sync_format_copy(title: str, code: str) -> str:
        # Canonical spot-name format (mirrors eterebridge_runner._format_copy)
        if not code:
            return "NEED COPY"
        if not title or title == code:
            return code
        prefix = title.split(":")[0].strip() if ":" in title else ""
        if prefix == code:
            return title
        return f"{title} ({code})"

    def _log_sync_compute(cur, contract_id: int, ws,
                          date_from=None, date_to=None) -> dict:
        from collections import defaultdict

        cur.execute(
            "SELECT ID_CONTRATTIRIGHE AS line,"
            " COALESCE(CONTROLLACAPOFILA, 0) AS capo,"
            " COALESCE(CONTROLLAFINEFILA, 0) AS fine"
            " FROM CONTRATTIRIGHE WHERE ID_CONTRATTITESTATA = %s",
            (contract_id,),
        )
        _line_rows = cur.fetchall()
        line_ids = {r["line"] for r in _line_rows}
        # bookend line (top/bottom) → log convention paints G+H pink
        bookend_lines = {r["line"] for r in _line_rows if r["capo"] and r["fine"]}
        if not line_ids:
            raise ValueError(f"Contract {contract_id} has no lines")

        date_clause, date_params = "", []
        if date_from:
            date_clause += " AND t.DATA >= %s"
            date_params.append(str(date_from))
        if date_to:
            date_clause += " AND t.DATA <= %s"
            date_params.append(str(date_to))

        ids = ",".join(str(li) for li in sorted(line_ids))
        cur.execute(f"""
            SELECT tp.ID_ContrattiRighe AS line, t.DATA, t.ORA,
                   RTRIM(ISNULL(f.COD_PROGRA, ''))                    AS code,
                   RTRIM(ISNULL(NULLIF(f.DESCRIZIO, ''), f.COD_PROGRA)) AS title
            FROM trafficPalinse tp
            JOIN TPALINSE t   ON t.ID_TPALINSE = tp.id_tpalinse
            LEFT JOIN FILMATI f ON f.ID_FILMATI = t.ID_FILMATI
            WHERE tp.ID_ContrattiRighe IN ({ids}) AND t.LIVELLO = 0{date_clause}
            ORDER BY tp.ID_ContrattiRighe, t.DATA, t.ORA
        """, tuple(date_params))
        etere = defaultdict(list)
        etere_spots = 0
        for r in cur.fetchall():
            etere[(r["line"], r["DATA"].date())].append(
                _log_sync_format_copy((r["title"] or "").strip(), (r["code"] or "").strip()))
            etere_spots += 1

        by_group = defaultdict(list)
        log_rows = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                line = int(row[12])
            except (TypeError, ValueError):
                continue
            if line not in line_ids:
                continue
            d = row[1].date() if hasattr(row[1], "date") else row[1]
            if d is not None and ((date_from and d < date_from) or (date_to and d > date_to)):
                continue
            by_group[(line, d)].append({"xlrow": i, "old": (row[7] or "").strip()})
            log_rows += 1

        changes, mismatches = [], []
        for key in sorted(by_group.keys() | etere.keys(),
                          key=lambda k: (k[1] or _date_cls.min, k[0])):
            rows = by_group.get(key, [])
            new_names = etere.get(key, [])
            if len(rows) != len(new_names):
                mismatches.append({
                    "line": key[0],
                    "date": key[1].isoformat() if key[1] else "?",
                    "log_rows": len(rows), "etere_spots": len(new_names),
                })
                continue
            for r, new in zip(rows, new_names):
                if r["old"] != new:
                    # Commercial-log color rules (Lee 2026-07-17): a line that
                    # goes to NEED COPY turns red in A-H; a NEED COPY line that
                    # receives traffic gets A-H restored to its network color
                    # (still present in column K — only A-H ever turn red, and
                    # J can be pink for added value).
                    # Same-to-same traffic swaps keep their existing fill.
                    if new == "NEED COPY" and r["old"] != "NEED COPY":
                        color = "red"
                    elif r["old"] == "NEED COPY" and new != "NEED COPY":
                        color = "native"
                    else:
                        color = None
                    changes.append({
                        "xlrow": r["xlrow"], "line": key[0],
                        "date": key[1].isoformat() if key[1] else "?",
                        "old": r["old"], "new": new, "color": color,
                        "bookend": key[0] in bookend_lines,
                    })
        return {
            "log_rows": log_rows,
            "etere_spots": etere_spots,
            "groups": len(by_group),
            "changes": changes,
            "mismatches": mismatches,
        }

    @router.get("/traffic/log-sync", response_class=HTMLResponse)
    async def traffic_log_sync_page(request: Request):
        return templates.TemplateResponse(request, "traffic/log_sync.html")

    def _log_sync_parse_date(raw: str):
        raw = (raw or "").strip()
        return _date_cls.fromisoformat(raw) if raw else None

    # Writable-workbook warm cache. Parsing the commercial log takes ~40 s of
    # pure CPU (13.8k x 50 cells), which dominated apply time. The preview
    # endpoint warms the writable copy in the background while the user reads
    # the preview, so Apply usually only pays for the edit + save (~10 s).
    # Entries are keyed by resolved path and validated by mtime; popped on use.
    _log_sync_wb_cache: dict = {}
    _log_sync_wb_lock = _threading.Lock()
    _log_sync_warming: dict = {}   # key -> threading.Event (set when done)

    def _log_sync_warm(log_path: Path) -> None:
        key = str(log_path)
        try:
            mtime = log_path.stat().st_mtime_ns
            with _log_sync_wb_lock:
                cached = _log_sync_wb_cache.get(key)
                if (cached and cached[0] == mtime) or key in _log_sync_warming:
                    return
                done = _threading.Event()
                _log_sync_warming[key] = done
            try:
                wb = _wb_load_fast(log_path)
                with _log_sync_wb_lock:
                    _log_sync_wb_cache.clear()   # keep at most one warmed workbook
                    _log_sync_wb_cache[key] = (mtime, wb)
            finally:
                with _log_sync_wb_lock:
                    _log_sync_warming.pop(key, None)
                done.set()
        except Exception:
            pass  # warming is best-effort; apply falls back to a cold load

    def _log_sync_take_warm(log_path: Path):
        """Return the warmed workbook if it still matches the file on disk.
        If a warm-up for this file is in flight (user clicked Apply quickly),
        wait for it instead of duplicating the ~40 s parse."""
        key = str(log_path)
        with _log_sync_wb_lock:
            in_flight = _log_sync_warming.get(key)
        if in_flight is not None:
            in_flight.wait(timeout=180)
        try:
            mtime = log_path.stat().st_mtime_ns
        except OSError:
            return None
        with _log_sync_wb_lock:
            cached = _log_sync_wb_cache.pop(key, None)
        return cached[1] if cached and cached[0] == mtime else None

    @router.get("/api/traffic/log-sync/preview")
    async def traffic_log_sync_preview(contract_id: int, path: str = "",
                                       date_from: str = "", date_to: str = ""):
        def _run():

            from browser_automation.etere_direct_client import connect as _connect
            log_path = _log_sync_path(path)
            if not log_path.exists():
                raise FileNotFoundError(f"Commercial log not found: {log_path}")
            d_from, d_to = _log_sync_parse_date(date_from), _log_sync_parse_date(date_to)
            wb = _wb_load_fast(log_path, read_only=True, data_only=True)
            try:
                if _LOG_SYNC_SHEET not in wb.sheetnames:
                    raise ValueError(f"No '{_LOG_SYNC_SHEET}' sheet in {log_path.name}")
                ws = wb[_LOG_SYNC_SHEET]
                with _connect() as conn:
                    result = _log_sync_compute(conn.cursor(as_dict=True), contract_id, ws,
                                               d_from, d_to)
            finally:
                wb.close()
            result["path"] = str(log_path)
            return result

        try:
            loop = asyncio.get_running_loop()
            result = await loop.run_in_executor(None, _run)
            # warm the writable workbook while the user reads the preview
            loop.run_in_executor(None, _log_sync_warm, _log_sync_path(path))
            return JSONResponse(result)
        except (FileNotFoundError, ValueError) as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.get("/api/traffic/log-sync/browse")
    async def traffic_log_sync_browse(dir: str = ""):
        """Server-side folder listing for the Browse modal (a web page cannot
        read local file paths, so navigation happens on the server)."""
        def _run():
            base = _log_sync_path(dir) if dir.strip() else _log_sync_path("").parent
            if base.is_file():
                base = base.parent
            if not base.is_dir():
                raise FileNotFoundError(f"Folder not found: {base}")
            folders, files = [], []
            for entry in sorted(base.iterdir(), key=lambda p: p.name.lower()):
                try:
                    if entry.name.startswith(("~$", ".")):
                        continue
                    if entry.is_dir():
                        folders.append({"name": entry.name, "path": str(entry)})
                    elif entry.suffix.lower() in (".xlsx", ".xlsm"):
                        files.append({"name": entry.name, "path": str(entry)})
                except OSError:
                    continue
            parent = str(base.parent) if base.parent != base else None
            return {"dir": str(base), "parent": parent, "folders": folders, "files": files}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except FileNotFoundError as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except PermissionError as e:
            return JSONResponse({"error": f"No access: {e}"}, status_code=403)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.post("/api/traffic/log-sync/apply")
    async def traffic_log_sync_apply(body: dict = Body(...)):
        contract_id = int(body.get("contract_id") or 0)
        if not contract_id:
            raise HTTPException(status_code=400, detail="contract_id required")

        def _run():

            from browser_automation.etere_direct_client import connect as _connect
            log_path = _log_sync_path(body.get("path") or "")
            if not log_path.exists():
                raise FileNotFoundError(f"Commercial log not found: {log_path}")
            d_from = _log_sync_parse_date(body.get("date_from") or "")
            d_to   = _log_sync_parse_date(body.get("date_to") or "")
            # full (non-read-only) load so formulas elsewhere survive the save;
            # usually pre-warmed by the preview endpoint (see _log_sync_warm)
            wb = _log_sync_take_warm(log_path) or _wb_load_fast(log_path)
            try:
                ws = wb[_LOG_SYNC_SHEET]
                with _connect() as conn:
                    result = _log_sync_compute(conn.cursor(as_dict=True), contract_id, ws,
                                               d_from, d_to)
                from copy import copy as _style_copy

                from openpyxl.styles import PatternFill
                _red_fill = PatternFill(fill_type="solid",
                                        start_color="FFFF0000", end_color="FFFF0000")
                _pink_fill = PatternFill(fill_type="solid",
                                         start_color="FFFF66FF", end_color="FFFF66FF")
                for ch in result["changes"]:
                    ws.cell(row=ch["xlrow"], column=8).value = ch["new"]
                    if ch["color"] == "red":
                        for col in range(1, 9):   # A-H
                            ws.cell(row=ch["xlrow"], column=col).fill = _red_fill
                    elif ch["color"] == "native":
                        # Restore the row's base network color to A-H. Prefer the
                        # AUTHORITATIVE market→color table (keyed on col AC) so the
                        # restore works even if col K was stripped/blank. Fall back
                        # to col K's live fill only when the market is unknown.
                        _mkt = str(ws.cell(row=ch["xlrow"], column=29).value or "").strip().upper()
                        _argb = _LOG_SYNC_MARKET_COLORS.get(_mkt)
                        native_fill = None
                        if _argb:
                            native_fill = PatternFill(fill_type="solid",
                                                      start_color=_argb, end_color=_argb)
                        else:
                            # col K carries the row's network color (col J can be
                            # pink for added value, col I is no-fill); skip if K is
                            # red/empty rather than guess.
                            k = ws.cell(row=ch["xlrow"], column=11).fill
                            _rgb = getattr(getattr(k, "start_color", None), "rgb", None)
                            if k is not None and k.patternType and _rgb != "FFFF0000":
                                native_fill = _style_copy(k)
                        if native_fill is not None:
                            for col in range(1, 9):
                                ws.cell(row=ch["xlrow"], column=col).fill = _style_copy(native_fill)
                            if ch.get("bookend"):
                                # bookend pair convention: G + H in pink
                                ws.cell(row=ch["xlrow"], column=7).fill = _pink_fill
                                ws.cell(row=ch["xlrow"], column=8).fill = _style_copy(_pink_fill)
                if result["changes"]:
                    # openpyxl handles the 6 value/custom-list sort levels + the CF
                    # rules; the transform re-injects what openpyxl drops — the 10
                    # picker colors and the font-color sort level (as a worksheet
                    # sortState) — completing the team's 7-level custom sort.
                    from openpyxl.utils import get_column_letter as _gcl
                    _log_sync_apply_standard_sort(ws)
                    _log_sync_apply_conditional_formatting(ws)
                    _hdr = {str(c.value).strip(): c.column
                            for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value}
                    _tout = _gcl(_hdr["Time out"]) if "Time out" in _hdr else "F"
                    _wb_save_fast(wb, log_path,
                                  transform=lambda b: _log_sync_finalize_xlsx(b, _tout))
            finally:
                wb.close()
            return {"written": len(result["changes"]),
                    "mismatches": result["mismatches"], "path": str(log_path)}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except FileNotFoundError as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except PermissionError:
            return JSONResponse(
                {"error": "Could not save — the commercial log is open in Excel. Close it and try again."},
                status_code=409)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.get("/api/traffic/diagnose-missing")
    async def diagnose_missing(
        date_from: str = Query(...),
        date_to:   str = Query(...),
    ):
        from datetime import datetime as _dt

        def _parse_date(s: str):
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(s.strip(), fmt)
                except ValueError:
                    pass
            raise ValueError(f"Unrecognized date: {s!r}")

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            out = {}
            dt_from = _parse_date(date_from)
            dt_to   = _parse_date(date_to)

            # --- connection 1: metadata only ---
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute("SELECT DISTINCT cod_user, nome FROM users ORDER BY cod_user")
                stations = [(row[0], row[1]) for row in cur.fetchall()]
                out["stations"] = [{"cod_user": c, "nome": n} for c, n in stations]

                try:
                    cur2 = conn.cursor()
                    cur2.execute("SELECT SUSER_SNAME(), USER_NAME(), IS_MEMBER('db_owner'), IS_MEMBER('db_datareader')")
                    r = cur2.fetchone()
                    out["login"] = {"suser": str(r[0]), "user": str(r[1]), "db_owner": r[2], "db_datareader": r[3]}
                except Exception as e:
                    out["login_error"] = str(e)

                # Get FILMATI and FS_FILMATI column names so we know the real schema
                try:
                    cols_cur = conn.cursor()
                    cols_cur.execute("""
                        SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE
                        FROM INFORMATION_SCHEMA.COLUMNS
                        WHERE TABLE_NAME IN ('FILMATI','FS_FILMATI','CONTRATTIFILMATI')
                        ORDER BY TABLE_NAME, ORDINAL_POSITION
                    """)
                    schema = {}
                    for tbl, col, dtype in cols_cur.fetchall():
                        schema.setdefault(tbl, []).append(f"{col} ({dtype})")
                    out["material_schema"] = schema
                except Exception as e:
                    out["material_schema_error"] = str(e)

                # Direct spot count for CVC in range
                try:
                    cnt_cur = conn.cursor()
                    cnt_cur.execute("""
                        SELECT COUNT(*) as total_spots,
                               SUM(CASE WHEN ID_FILMATI IS NULL OR ID_FILMATI = 0 THEN 1 ELSE 0 END) as no_material,
                               SUM(CASE WHEN ID_FILMATI IS NOT NULL AND ID_FILMATI <> 0 THEN 1 ELSE 0 END) as has_material
                        FROM TPalinseSpotsInCluster
                        WHERE DATA BETWEEN ? AND ? AND COD_USER = 7
                    """, dt_from, dt_to)
                    r = cnt_cur.fetchone()
                    out["cvc_spot_counts"] = {"total": r[0], "no_material": r[1], "has_material": r[2]}
                except Exception as e:
                    out["cvc_spot_counts_error"] = str(e)

                # Check STATUS_MM distribution on linked FILMATI records for CVC
                try:
                    smm_cur = conn.cursor()
                    smm_cur.execute("""
                        SELECT f.STATUS_MM, COUNT(*) as cnt
                        FROM TPalinseSpotsInCluster s
                        JOIN FILMATI f ON f.ID_FILMATI = s.ID_FILMATI
                        WHERE s.DATA BETWEEN ? AND ? AND s.COD_USER = 7
                        GROUP BY f.STATUS_MM
                        ORDER BY cnt DESC
                    """, dt_from, dt_to)
                    out["cvc_status_mm"] = [{"status_mm": str(r[0]), "count": r[1]} for r in smm_cur.fetchall()]
                except Exception as e:
                    out["cvc_status_mm_error"] = str(e)

                # Check FS_FILMATI.FLAG_OFFLINE for CVC spots
                try:
                    off_cur = conn.cursor()
                    off_cur.execute("""
                        SELECT fs.FLAG_OFFLINE, COUNT(*) as cnt
                        FROM TPalinseSpotsInCluster s
                        JOIN FS_FILMATI fs ON fs.ID_FILMATI = s.ID_FILMATI
                        WHERE s.DATA BETWEEN ? AND ? AND s.COD_USER = 7
                        GROUP BY fs.FLAG_OFFLINE
                    """, dt_from, dt_to)
                    out["cvc_flag_offline"] = [{"flag_offline": str(r[0]), "count": r[1]} for r in off_cur.fetchall()]
                except Exception as e:
                    out["cvc_flag_offline_error"] = str(e)

                # Try connecting to the SSRS server (EC2AMAZ-6J2KLLI) directly
                # to test if it has different data than our Tailscale endpoint
                try:
                    import os as _os

                    import pyodbc as _pyodbc
                    user = _os.getenv("ETERE_DB_USER")
                    pwd  = _os.getenv("ETERE_DB_PASSWORD")
                    cs = (
                        f"DRIVER={{SQL Server}};SERVER=EC2AMAZ-6J2KLLI;"
                        f"DATABASE=Etere_crossing;UID={user};PWD={pwd};"
                        if user and pwd else
                        "DRIVER={SQL Server};SERVER=EC2AMAZ-6J2KLLI;"
                        "DATABASE=Etere_crossing;Trusted_Connection=yes;"
                    )
                    ec2_conn = _pyodbc.connect(cs, timeout=5)
                    ec2_cur = ec2_conn.cursor()
                    ec2_cur.execute(
                        "EXEC dbo.rpt_trf_missing_material_list "
                        "@cod_user=7, @startDate=?, @endDate=?, "
                        "@viewNM='1', @viewNA='1', @viewNR='1', @orderBy='3', @codeStart=NULL",
                        dt_from, dt_to,
                    )
                    while ec2_cur.description is None:
                        if not ec2_cur.nextset():
                            break
                    rows = ec2_cur.fetchall() if ec2_cur.description else []
                    out["ec2_sp_row_count"] = len(rows)
                    out["ec2_sp_first_row"] = [str(v) for v in rows[0]] if rows else None
                    ec2_conn.close()
                except Exception as e:
                    out["ec2_sp_error"] = str(e)

            # --- connection 2: SP calls with ARITHABORT ON (clean slate) ---
            out["markets"] = []
            with _db_connect() as sp_conn:
                _set = sp_conn.cursor()
                _set.execute("SET ARITHABORT ON")
                _set.execute("SET ANSI_NULLS ON")
                _set.execute("SET ANSI_WARNINGS ON")
                _set.execute("SET QUOTED_IDENTIFIER ON")
                for cod_user, nome in stations:
                    try:
                        mc = sp_conn.cursor()
                        mc.execute(
                            "EXEC dbo.rpt_trf_missing_material_list ?, ?, ?, ?, ?, ?, ?, ?",
                            cod_user, dt_from, dt_to, "1", "1", "1", "3", None
                        )
                        while mc.description is None:
                            if not mc.nextset():
                                break
                        if mc.description:
                            cols = [d[0] for d in mc.description]
                            rows = mc.fetchall()
                            out["markets"].append({
                                "cod_user": cod_user, "nome": nome,
                                "columns": cols,
                                "row_count": len(rows),
                                "first_row": [str(v) for v in rows[0]] if rows else None,
                            })
                        else:
                            out["markets"].append({"cod_user": cod_user, "nome": nome, "error": "no result set"})
                    except Exception as e:
                        out["markets"].append({"cod_user": cod_user, "nome": nome, "error": str(e)})
            return out

        result = await asyncio.to_thread(_run)
        return JSONResponse(content=result)

    @router.get("/traffic/missing-materials", response_class=HTMLResponse)
    async def traffic_missing_materials_page(request: Request):
        return templates.TemplateResponse(request, "traffic/missing_materials.html")

    @router.get("/api/traffic/missing-materials")
    async def get_missing_materials(
        date_from: str = Query(...),
        date_to:   str = Query(...),
    ):
        from datetime import datetime as _dt

        def _parse_date(s: str):
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(s.strip(), fmt)
                except ValueError:
                    pass
            raise ValueError(f"Unrecognized date: {s!r}")

        def _query():
            from browser_automation.etere_direct_client import connect as _db_connect
            _FPS = 29.97
            _MKT_ORDER = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}

            def _to_ampm(frames):
                if frames is None:
                    return ""
                s = round(frames / _FPS)
                h, m = s // 3600, (s % 3600) // 60
                suffix = "AM" if h < 12 else "PM"
                h12 = h % 12 or 12
                return f"{h12}:{m:02d} {suffix}"

            dt_from = _parse_date(date_from)
            dt_to   = _parse_date(date_to)
            with _db_connect() as conn:
                cur = conn.cursor()
                # tpalinse is authoritative. ID_FILMATI <= 0 means no material
                # (COMS uses -1, others use 0). NOOP/PGM excluded — only ad types.
                cur.execute("""
                    SELECT
                        t.COD_USER,
                        t.DATA,
                        t.ORA,
                        ct.COD_CONTRATTO,
                        ct.DESCRIZIONE,
                        t.ID_TPALINSE,
                        t.DURATION
                    FROM tpalinse t
                    LEFT JOIN trafficPalinse tp ON tp.id_tpalinse = t.ID_TPALINSE
                    LEFT JOIN CONTRATTIRIGHE cr ON tp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                    LEFT JOIN CONTRATTITESTATA ct ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                    WHERE t.DATA BETWEEN %s AND %s
                      AND t.LIVELLO = 0
                      AND (t.ID_FILMATI IS NULL OR t.ID_FILMATI <= 0)
                      AND t.NEWTYPE IN ('COM','COMS','BNS','BART','BB','AV','TRD')
                    ORDER BY t.DATA, t.COD_USER, t.ORA
                """, (dt_from, dt_to))
                rows = []
                for cod_user, data, ora, code, name, id_tpalinse, duration in cur.fetchall():
                    mkt_id  = cod_user or 0
                    market  = _MKT_ORDER.get(mkt_id, str(mkt_id))
                    date_s  = f"{data.month}/{data.day:02d}/{str(data.year)[2:]}" if data else ""
                    rows.append({
                        "market":          market,
                        "market_order":    mkt_id,
                        "date":            date_s,
                        "time":            _to_ampm(ora),
                        "contract_code":   code or "— orphaned —",
                        "contract_name":   name or "No contract link — spot will air black",
                        "id_tpalinse":     id_tpalinse,
                        "duration_frames": int(duration or 0),
                    })
            return rows

        rows = await asyncio.to_thread(_query)
        return JSONResponse(content={"rows": rows, "total": len(rows)})

    @router.get("/api/traffic/missing-materials/status")
    async def get_missing_materials_status():
        """Fast COUNT-only check for today + tomorrow used by the portal badge."""
        from datetime import date as _date
        from datetime import timedelta as _td

        def _count():
            from browser_automation.etere_direct_client import connect as _db_connect
            today    = _date.today().isoformat()
            tomorrow = (_date.today() + _td(days=1)).isoformat()
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT COUNT(*)
                    FROM tpalinse
                    WHERE DATA BETWEEN %s AND %s
                      AND LIVELLO = 0
                      AND (ID_FILMATI IS NULL OR ID_FILMATI <= 0)
                      AND NEWTYPE IN ('COM','COMS','BNS','BART','BB','AV','TRD')
                """, (today, tomorrow))
                return cur.fetchone()[0] or 0

        count = await asyncio.to_thread(_count)
        return JSONResponse(content={"count": count})

    @router.post("/api/traffic/blacklist-spot")
    async def blacklist_spot(request: Request):
        body = await request.json()
        id_tpalinse = int(body["id_tpalinse"])
        replace = bool(body.get("replace", True))

        def _run():
            from datetime import datetime as _dt

            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                # Fetch the spot
                cur.execute("""
                    SELECT DATA, COD_USER, ORA, ORA_P, DATA_P, XORDER, ORDINALE, DURATION, LIVELLO
                    FROM TPALINSE WHERE ID_TPALINSE = %s
                """, (id_tpalinse,))
                spot = cur.fetchone()
                if not spot:
                    raise ValueError(f"Spot {id_tpalinse} not found")
                if spot["LIVELLO"] != 0:
                    raise ValueError(f"Spot {id_tpalinse} is not active (LIVELLO={spot['LIVELLO']})")

                # Fetch trafficPalinse for break context + TSL reference ID
                cur.execute("""
                    SELECT id_trafficPalinse, id_palinsesto, id_fascia, clusterIndex,
                           offset, Date, Cod_User, ID_ContrattiRighe
                    FROM trafficPalinse WHERE id_tpalinse = %s
                """, (id_tpalinse,))
                tpa = cur.fetchone()
                if not tpa:
                    # Fully orphaned spot — no contract link, no TSL accounting needed
                    cur.execute("DELETE FROM TPALINSE WHERE ID_TPALINSE=%s", (id_tpalinse,))
                    conn.commit()
                    return {"orphaned_deleted": True, "filler_inserted": False, "no_filler": True}

                line_id      = tpa["ID_ContrattiRighe"]
                tpa_id       = tpa["id_trafficPalinse"]

                # Fetch contract line date range for TSL
                cur.execute(
                    "SELECT DATA_INIZIO, DATA_FINE FROM CONTRATTIRIGHE WHERE ID_CONTRATTIRIGHE=%s",
                    (line_id,)
                )
                cr = cur.fetchone()
                date_start = cr["DATA_INIZIO"] if cr else None
                date_end   = cr["DATA_FINE"]   if cr else None

                # Delete the TPALINSE and trafficPalinse rows — this is what Etere does natively
                cur.execute("DELETE FROM trafficPalinse WHERE id_tpalinse=%s", (id_tpalinse,))
                cur.execute("DELETE FROM TPALINSE WHERE ID_TPALINSE=%s", (id_tpalinse,))

                # Blacklist entry — INSERT first occurrence; INCREMENT PassageMiss for subsequent spots
                cur.execute(
                    "SELECT ID_TrafficScheduleList FROM Traffic_ScheduleList WHERE ID_ContrattiRighe=%s AND BlackList>0",
                    (line_id,)
                )
                existing = cur.fetchone()
                if existing is None:
                    cur.execute("""
                        INSERT INTO Traffic_ScheduleList (
                            ID_ContrattiRighe, BlackList, PassageMiss,
                            ID_TRAFFICPALINSE, Date, ToDate,
                            Notes, Operator,
                            ID_FILMATI, ID_FILMATI_TAIL, ID_FILMATI_MIDDLE,
                            ID_FATTURAEMITTENTE, Split
                        ) VALUES (%s, 1, 1, %s, %s, %s, %s, %s, -1, -1, -1, 0, 0)
                    """, (
                        line_id, tpa_id, date_start, date_end,
                        "Blacklisted - no materials", "ControlRoom",
                    ))
                else:
                    cur.execute(
                        "UPDATE Traffic_ScheduleList SET PassageMiss = PassageMiss + 1 WHERE ID_ContrattiRighe=%s AND BlackList>0",
                        (line_id,)
                    )

                # Find filler: PI first, PSA fallback (only if replace requested)
                dur = spot["DURATION"]
                filler = None
                filler_type = None
                for pattern, label in ([("PI-%%", "PI"), ("PSA-%%", "PSA")] if replace else []):
                    cur.execute("""
                        SELECT TOP 1 ID_FILMATI, COD_PROGRA, DESCRIZIO, DURATA, NEWTYPE
                        FROM FILMATI
                        WHERE DESCRIZIO LIKE %s
                          AND DESCRIZIO NOT LIKE 'DO NOT%%'
                          AND (DATA_SCAD IS NULL OR DATA_SCAD > GETDATE())
                          AND ABS(DURATA - %s) <= 5
                        ORDER BY NEWID()
                    """, (pattern, dur))
                    filler = cur.fetchone()
                    if filler:
                        filler_type = label
                        break

                new_filler_id = None
                if filler:
                    supporto = _pi_filler_supporto(cur, filler["ID_FILMATI"], filler["DESCRIZIO"])
                    newtype  = filler["NEWTYPE"] or "PER"
                    cur.execute("""
                        INSERT INTO TPALINSE (
                            DATA, COD_USER, LIVELLO, SPLIT, XORDER, ORA,
                            ID_FILMATI, COD_PROGRA, NEWTYPE, TITLE, PART, EVENT_TYPE,
                            DURATION, TIMECODE_I, TIMECODE_O, TRANS, TRANS_DUR,
                            GPI, GPI2, GPID1, GPID2, TITLER, TITLER_IN, TITLER_DUR,
                            CRAWL_DESC, CRAWL_VOL, CRAWL_IN, CRAWL_SPEE, RECORDABLE,
                            DSK1, DSK1_IN, DSK1_DUR, DSK2, DSK2_IN, DSK2_DUR,
                            DSK3, DSK3_IN, DSK3_DUR,
                            OUTPUT_M, ASPECT, SUBTITLE, NOTE, TIPO_TC,
                            RIMAND, RIMAND_IN, RIMAND_VOL,
                            DATA_PREV, ORA_PREV, STATUS, STATUS_RC, STATUS_MM, STATUS_CA,
                            SUPPORTO, SUPPORTORC, SUPPORTOB, COSTO, DSKA, DSKB, DSKC,
                            TVGUIDE, PROVENIENZA, ORDINALE,
                            ID_FATTURAEMITTENTE, ID_FATTURAAGENZIA, EXTERNALID, EXTERSTRID,
                            CRYPTATO, ID_SUPPORT, VISIONATO, ID_DOCUMENT, TRAFFICID,
                            TYPE, ERRORCODE, CODE_PRIMECAST, COLOR, AUDIO, AUDIO_TY,
                            VOICEOVR_A, VOICEOVR_B, VOICEOVR_C, VOICEOVR_D,
                            MEDIA_TY, MEDIA_ID, ORA_P, DATA_P, EVENT, EVENT_P,
                            DURATION_P, FADEIN, FADEOUT, INTRO, OUTRO, ICON
                        )
                        VALUES (
                            %s, %s, 0, 0, %s, %s,
                            %s, %s, %s, %s, 0, 'T',
                            %s, 0, %s, 'CT', 0,
                            '000000000', '000000000', 0, 0, 0, 0, 0,
                            '', 0, 0, 0, '',
                            'NN', 0, 0, 'NN', 0, 0,
                            'NN', 0, 0,
                            '', 'H', '', '', 'C',
                            0, 0, 0,
                            %s, %s, 'I', '', '', '',
                            %s, '', '', 0.00,
                            'NA  0000000000000000', 'NA  0000000000000000', 'NA  0000000000000000',
                            'NN000', 'TRAFFIC_NEW', %s,
                            0, 0, 0, '', '', 0, 'X', 0, 0,
                            'T', 0, '1', 'FF000005', '1234', 'M',
                            0, 0, 0, 0,
                            '', '', %s, %s, 0, 0,
                            %s, 0, 0, 0, 0, ''
                        )
                    """, (
                        spot["DATA"], spot["COD_USER"], spot["XORDER"], spot["ORA"],
                        filler["ID_FILMATI"], filler["COD_PROGRA"], newtype, filler["DESCRIZIO"],
                        filler["DURATA"], filler["DURATA"] - 1,
                        spot["DATA"], spot["ORA"],
                        supporto,
                        spot["ORDINALE"],
                        spot["ORA_P"], spot["DATA_P"],
                        filler["DURATA"],
                    ))
                    cur.execute("SELECT SCOPE_IDENTITY() AS new_id")
                    row = cur.fetchone()
                    new_filler_id = int(row["new_id"]) if row and row["new_id"] else None

                    if new_filler_id:
                        cur.execute("""
                            UPDATE TPALINSE SET
                                EVENT    = 100000000000 + %s,
                                EVENT_P  = 100000000000 + %s,
                                TRAFFICID = %s
                            WHERE ID_TPALINSE = %s
                        """, (new_filler_id, new_filler_id, new_filler_id, new_filler_id))

                        cur.execute("""
                            INSERT INTO trafficPalinse (
                                id_tpalinse, id_palinsesto, id_fascia, clusterIndex, offset,
                                tag, scadenza, data_ins, ID_Operation, ID_ContrattiRighe,
                                Date, Cod_User, ID_FATTURAEMITTENTE, ID_FATTURAAGENZIA,
                                TrafficFlag, TrafficNotes, ID_FATTURAEMITTENTELOGO,
                                ID_TRAFFICTRASH, SCHEDULEDMODE, EVENTTYPE,
                                ID_CPEMITTENTE, ID_CPAGENZIA, ID_CPEMITTENTELOGO
                            ) VALUES (
                                %s, %s, %s, %s, %s,
                                0, '1900-01-01', %s, 0, -1,
                                %s, %s, 0, 0,
                                0, '', 0, 0, 0, 1,
                                0, 0, 0
                            )
                        """, (
                            new_filler_id, tpa["id_palinsesto"], tpa["id_fascia"],
                            tpa["clusterIndex"], tpa["offset"],
                            _dt.now(), tpa["Date"], tpa["Cod_User"],
                        ))

                conn.commit()

            return {
                "blacklisted":    True,
                "filler_type":    filler_type,
                "filler_name":    filler["DESCRIZIO"] if filler else None,
                "filler_inserted": new_filler_id is not None,
                "no_filler":      filler is None,
            }

        try:
            result = await asyncio.to_thread(_run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/traffic/no-material")
    async def get_no_material(
        date_from: str = Query(...),
        date_to:   str = Query(...),
    ):
        from datetime import datetime as _dt

        def _parse_date(s: str):
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    return _dt.strptime(s.strip(), fmt)
                except ValueError:
                    pass
            raise ValueError(f"Unrecognized date: {s!r}")

        def _query():
            from browser_automation.etere_direct_client import connect as _db_connect
            dt_from = _parse_date(date_from)
            dt_to   = _parse_date(date_to)
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute("""
                    SELECT
                        CAST(t.DATA AS DATE) AS day,
                        t.COD_USER,
                        COUNT(*) AS spot_count
                    FROM tpalinse t
                    WHERE t.DATA BETWEEN %s AND %s
                      AND t.LIVELLO = 0
                      AND (t.ID_FILMATI IS NULL OR t.ID_FILMATI <= 0)
                      AND t.NEWTYPE IN ('COM','COMS','BNS','BART','BB','AV','TRD')
                    GROUP BY CAST(t.DATA AS DATE), t.COD_USER
                    ORDER BY CAST(t.DATA AS DATE), t.COD_USER
                """, (dt_from, dt_to))
                rows = []
                for row in cur.fetchall():
                    day, cod_user, cnt = row
                    market = _MARKET_NAMES.get(cod_user, str(cod_user))
                    date_str = f"{day.month}/{day.day:02d}/{str(day.year)[2:]}" if day else ""
                    rows.append({
                        "date":       date_str,
                        "market":     market,
                        "spot_count": cnt,
                    })
            return rows

        rows = await asyncio.to_thread(_query)
        return JSONResponse(content={"rows": rows, "total": len(rows)})

    # ------------------------------------------------------------------
    # Customer Database
    # ------------------------------------------------------------------

    _repo = Path(__file__).resolve().parents[3]
    _BILLING_DIR = _repo.parent / "billing"
    _BILLING_PYTHON = _BILLING_DIR / ".venv" / "Scripts" / "python.exe"
    if not _BILLING_PYTHON.exists():
        _BILLING_PYTHON = _BILLING_DIR / ".venv" / "bin" / "python"
    _BILLING_MANAGE  = _BILLING_DIR / "manage_db.py"
    _BILLING_BACKFILL = _BILLING_DIR / "backfill.py"

    async def _run_manage_json(args: list) -> object:
        cmd = [str(_BILLING_PYTHON), str(_BILLING_MANAGE), "--json"] + [str(a) for a in args]
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=str(_BILLING_DIR),
        )
        stdout, stderr = await proc.communicate()
        if proc.returncode != 0:
            raise HTTPException(status_code=500, detail=stderr.decode(errors="replace").strip())
        return json.loads(stdout.decode(errors="replace"))

    async def _run_manage_write(args: list) -> dict:
        cmd = [str(_BILLING_PYTHON), str(_BILLING_MANAGE)] + [str(a) for a in args]
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=str(_BILLING_DIR),
        )
        stdout, stderr = await proc.communicate()
        if proc.returncode != 0:
            return {"ok": False, "error": (stderr.decode(errors="replace") or stdout.decode(errors="replace")).strip()}
        return {"ok": True}

    @router.get("/customers", response_class=HTMLResponse)
    async def customers_page(request: Request):
        return templates.TemplateResponse(request, "customers.html")

    @router.get("/customers/import-db", response_class=HTMLResponse)
    async def customers_import_db_page(request: Request):
        return templates.TemplateResponse(request, "customers/import_db.html")

    @router.post("/api/customers/import-db/preview")
    async def preview_import_db(file: UploadFile):
        """
        Compare an uploaded customers.db against the live one.
        Returns new records and conflicts (same PK, differing field values).
        """
        import os
        import sqlite3
        import tempfile

        _SKIP = {"created_at"}
        target_path = config.customer_db_path
        if not target_path.exists():
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=400, content={"error": f"Target database not found: {target_path.resolve()}"})

        # Save upload to temp file
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".db")
        try:
            with os.fdopen(tmp_fd, "wb") as f:
                f.write(await file.read())

            # Read source DB
            src_conn = sqlite3.connect(tmp_path)
            src_conn.row_factory = sqlite3.Row
            src_rows = src_conn.execute("SELECT * FROM customers").fetchall()
            src_cols = [d[0] for d in src_conn.execute("PRAGMA table_info(customers)").fetchall()]
            src_conn.close()

            # Read target DB
            tgt_conn = sqlite3.connect(str(target_path))
            tgt_conn.row_factory = sqlite3.Row
            tgt_cols = [d[0] for d in tgt_conn.execute("PRAGMA table_info(customers)").fetchall()]
            tgt_index = {
                (r["customer_name"], r["order_type"]): dict(r)
                for r in tgt_conn.execute("SELECT * FROM customers").fetchall()
            }
            tgt_conn.close()

            # Compare columns that exist in BOTH (skip metadata)
            common_cols = [c for c in src_cols if c in tgt_cols and c not in _SKIP]
            if "customer_name" not in common_cols or "order_type" not in common_cols:
                raise ValueError(f"Uploaded DB is missing required columns. Found: {src_cols}")

            new_records = []
            conflicts = []

            for src_row in src_rows:
                row_dict = {c: src_row[c] for c in src_cols if c in common_cols}
                key = (row_dict["customer_name"], row_dict["order_type"])

                if key not in tgt_index:
                    new_records.append(row_dict)
                else:
                    tgt_row = {c: tgt_index[key].get(c) for c in common_cols}
                    diff_fields = [
                        c for c in common_cols
                        if c not in ("customer_name", "order_type")
                        and str(row_dict.get(c) or "") != str(tgt_row.get(c) or "")
                    ]
                    if diff_fields:
                        conflicts.append({
                            "customer_name": key[0],
                            "order_type":    key[1],
                            "current":  tgt_row,
                            "incoming": row_dict,
                            "diff":     diff_fields,
                        })

            return {
                "new":       new_records,
                "conflicts": conflicts,
                "columns":   common_cols,
            }

        except Exception as exc:
            from fastapi.responses import JSONResponse
            return JSONResponse(status_code=500, content={"error": str(exc)})
        finally:
            os.unlink(tmp_path)

    @router.post("/api/customers/import-db/apply")
    async def apply_import_db(body: Any = Body(default={})):
        """
        Apply the user-resolved merge.
        Body: { insert: [row_dicts], upsert: [row_dicts] }
        """
        import sqlite3

        insert_rows = body.get("insert", [])
        upsert_rows = body.get("upsert", [])
        target_path = config.customer_db_path

        inserted = updated = 0
        with sqlite3.connect(str(target_path)) as conn:
            for row in insert_rows:
                cols = list(row.keys())
                placeholders = ",".join(["?"] * len(cols))
                conn.execute(
                    f"INSERT OR IGNORE INTO customers ({','.join(cols)}) VALUES ({placeholders})",
                    [row[c] for c in cols],
                )
                inserted += conn.total_changes

            for row in upsert_rows:
                cols = [c for c in row if c not in ("customer_name", "order_type")]
                if not cols:
                    continue
                set_clause = ", ".join(f"{c}=?" for c in cols)
                conn.execute(
                    f"UPDATE customers SET {set_clause} WHERE customer_name=? AND order_type=?",
                    [row[c] for c in cols] + [row["customer_name"], row["order_type"]],
                )
                updated += 1

        return {"inserted": inserted, "updated": updated}

    @router.get("/customers/backfill", response_class=HTMLResponse)
    async def customers_backfill_page(request: Request):
        return templates.TemplateResponse(request, "customers/backfill.html")

    @router.get("/api/customers/backfill")
    async def run_backfill(
        since: str = Query(...),
        dry_run: str = Query("0"),
    ):
        args = [str(_BILLING_PYTHON), "-u", str(_BILLING_BACKFILL), "--since", since]
        if dry_run == "1":
            args.append("--dry-run")

        async def event_stream():
            process = await asyncio.create_subprocess_exec(
                *args,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.STDOUT,
                cwd=str(_BILLING_DIR),
            )
            async for line in process.stdout:
                text = line.decode(errors="replace").rstrip()
                yield f"data: {text}\n\n"
            await process.wait()
            yield f"data: [EXIT:{process.returncode}]\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.get("/customers/manage", response_class=HTMLResponse)
    async def customers_manage_page(request: Request):
        return templates.TemplateResponse(request, "customers/manage.html")

    @router.get("/api/customers/manage/search")
    async def manage_search(q: str = Query(...)):
        return JSONResponse(content=await _run_manage_json(["search", q]))

    @router.get("/api/customers/manage/agencies")
    async def manage_agencies():
        return JSONResponse(content=await _run_manage_json(["list-agencies"]))

    @router.get("/api/customers/manage/advertisers")
    async def manage_advertisers():
        return JSONResponse(content=await _run_manage_json(["list-advertisers"]))

    @router.get("/api/customers/manage/order/{contract}")
    async def manage_order(contract: int):
        return JSONResponse(content=await _run_manage_json(["show", contract]))

    @router.get("/api/customers/manage/monthly/{contract}")
    async def manage_monthly(contract: int):
        return JSONResponse(content=await _run_manage_json(["monthly", contract]))

    @router.get("/api/customers/manage/affidavits")
    async def manage_affidavits(month: str = Query(...)):
        return JSONResponse(content=await _run_manage_json(["affidavit", "list", month]))

    @router.post("/api/customers/manage/set-agency")
    async def manage_set_agency(body: dict = Body(...)):
        args = ["set-agency", body.get("agency", "")]
        edi = body.get("edi")
        if edi is True:
            args.append("--edi")
        elif edi is False:
            args.append("--no-edi")
        notes = body.get("edi_notes")
        if notes is not None:
            args += ["--edi-notes", notes]
        return JSONResponse(content=await _run_manage_write(args))

    @router.post("/api/customers/manage/set-advertiser")
    async def manage_set_advertiser(body: dict = Body(...)):
        args = ["set-advertiser", body.get("advertiser", "")]
        args.append("--notarized" if body.get("notarized") else "--no-notarized")
        return JSONResponse(content=await _run_manage_write(args))

    @router.post("/api/customers/manage/affidavit-status")
    async def manage_affidavit_status(body: dict = Body(...)):
        args = ["affidavit", "status", body.get("number", ""), body.get("status", "")]
        return JSONResponse(content=await _run_manage_write(args))

    @router.post("/api/customers/manage/remove-order")
    async def manage_remove_order(body: dict = Body(...)):
        args = ["remove-order", "--confirm", str(body.get("contract_number", ""))]
        return JSONResponse(content=await _run_manage_write(args))

    # ── Order-entry customer database (data/customers.db) ─────────────────────

    def _cdb():
        import sqlite3 as _sq
        conn = _sq.connect(str(config.customer_db_path))
        existing = {r[1] for r in conn.execute("PRAGMA table_info(customers)")}
        for col, defn in [
            ("auto_aircheck", "INTEGER DEFAULT 0"),
            ("abbreviation",  "TEXT DEFAULT ''"),
        ]:
            if col not in existing:
                conn.execute(f"ALTER TABLE customers ADD COLUMN {col} {defn}")
        conn.commit()
        return conn

    @router.get("/orders/customers", response_class=HTMLResponse)
    async def order_customers_page(request: Request):
        return templates.TemplateResponse(request, "order_customers.html", {})

    @router.get("/api/orders/customers")
    async def list_order_customers(q: str = ""):
        conn = _cdb()
        try:
            if q:
                rows = conn.execute(
                    "SELECT * FROM customers WHERE LOWER(customer_name) LIKE ? OR LOWER(order_type) LIKE ? ORDER BY customer_name",
                    (f"%{q.lower()}%", f"%{q.lower()}%"),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM customers ORDER BY customer_name"
                ).fetchall()
            cols = [d[0] for d in conn.execute("SELECT * FROM customers LIMIT 0").description]
            return JSONResponse([dict(zip(cols, r)) for r in rows])
        finally:
            conn.close()

    @router.post("/api/orders/customers")
    async def create_order_customer(body: dict = Body(...)):
        import sqlite3 as _sq
        conn = _cdb()
        try:
            conn.execute(
                """INSERT INTO customers
                   (customer_id, customer_name, order_type, code_name, description_name,
                    billing_type, default_market, separation_customer, separation_event,
                    separation_order, include_market_in_code, auto_aircheck, abbreviation)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    body.get("customer_id", ""),
                    body.get("customer_name", ""),
                    body.get("order_type", ""),
                    body.get("code_name", ""),
                    body.get("description_name", ""),
                    body.get("billing_type", "agency"),
                    body.get("default_market") or None,
                    int(body.get("separation_customer", 15)),
                    int(body.get("separation_event", 0)),
                    int(body.get("separation_order", 0)),
                    int(bool(body.get("include_market_in_code", False))),
                    int(bool(body.get("auto_aircheck", False))),
                    body.get("abbreviation", ""),
                ),
            )
            conn.commit()
        except _sq.IntegrityError as exc:
            raise HTTPException(status_code=409, detail=str(exc))
        finally:
            conn.close()
        return JSONResponse({"ok": True})

    @router.put("/api/orders/customers")
    async def update_order_customer(customer_name: str, order_type: str, body: dict = Body(...)):
        conn = _cdb()
        try:
            conn.execute(
                """UPDATE customers SET
                   customer_name=?, customer_id=?, code_name=?, description_name=?,
                   billing_type=?, default_market=?, separation_customer=?,
                   separation_event=?, separation_order=?, include_market_in_code=?,
                   auto_aircheck=?, abbreviation=?
                   WHERE customer_name=? AND order_type=?""",
                (
                    body.get("customer_name", customer_name),
                    body.get("customer_id", ""),
                    body.get("code_name", ""),
                    body.get("description_name", ""),
                    body.get("billing_type", "agency"),
                    body.get("default_market") or None,
                    int(body.get("separation_customer", 15)),
                    int(body.get("separation_event", 0)),
                    int(body.get("separation_order", 0)),
                    int(bool(body.get("include_market_in_code", False))),
                    int(bool(body.get("auto_aircheck", False))),
                    body.get("abbreviation", ""),
                    customer_name,
                    order_type,
                ),
            )
            conn.commit()
        finally:
            conn.close()
        return JSONResponse({"ok": True})

    @router.post("/api/orders/customers/delete")
    async def delete_order_customer(body: dict = Body(...)):
        customer_name = body.get("customer_name", "")
        order_type    = body.get("order_type", "")
        conn = _cdb()
        try:
            conn.execute(
                "DELETE FROM customers WHERE LOWER(customer_name)=LOWER(?) AND LOWER(order_type)=LOWER(?)",
                (customer_name, order_type),
            )
            conn.commit()
        finally:
            conn.close()
        return JSONResponse({"ok": True})

    # ── Master Control ───────────────────────────────────────────────────────────

    @router.get("/master-control", response_class=HTMLResponse)
    async def master_control_page(request: Request):
        return templates.TemplateResponse(request, "master_control.html")

    @router.get("/master-control/logs", response_class=HTMLResponse)
    async def master_control_logs_page(request: Request):
        return templates.TemplateResponse(request, "master_control/log_viewer.html")

    @router.get("/master-control/compile-logs", response_class=HTMLResponse)
    async def compile_logs_page(request: Request):
        return templates.TemplateResponse(request, "master_control/compile_logs.html")

    @router.get("/api/master-control/compile-logs/ls")
    async def compile_logs_ls(date: str = Query(...)):
        def _run():
            target = _date_cls.fromisoformat(date)
            monday = target - timedelta(days=target.weekday())
            month_folder = monday.strftime("%m %B %Y")
            folder = _TRAFFIC_LOG_ROOT / str(monday.year) / month_folder
            if not folder.exists():
                return {"error": f"Folder not found: {folder}", "monday": monday.isoformat()}
            files = sorted(f.name for f in folder.iterdir() if f.suffix in (".xlsm", ".xlsx", ".xls"))
            return {"folder": str(folder), "monday": monday.isoformat(), "files": files}
        result = await asyncio.get_running_loop().run_in_executor(None, _run)
        return JSONResponse(result)

    @router.get("/api/master-control/compile-logs/run")
    async def compile_logs_run(date: str = Query(...)):
        import json as _json

        MARKETS = ["NYC", "CMP", "HOU", "SFO", "SEA", "LAX", "CVC", "WDC", "MMT", "DAL"]

        def _monday(d: _date_cls) -> _date_cls:
            return d - timedelta(days=d.weekday())

        def _to_win(p: Path) -> str:
            return str(p).replace("/", "\\")

        target = _date_cls.fromisoformat(date)
        monday = _monday(target)

        async def event_stream():
            yield f"data: {_json.dumps({'type': 'week', 'monday': monday.isoformat()})}\n\n"

            # Resolve all paths first
            found: list[tuple[str, Path]] = []
            for mkt in MARKETS:
                log_path = await asyncio.get_running_loop().run_in_executor(
                    None, _find_traffic_log, mkt, monday
                )
                if log_path is None:
                    fname = f"{mkt} Log - {monday.strftime('%y%m%d')}.xlsm"
                    folder = f"K:\\Traffic\\logs\\{monday.year}\\{monday.strftime('%m %B %Y')}"
                    yield f"data: {_json.dumps({'type': 'market', 'market': mkt, 'status': 'missing', 'file': fname, 'msg': folder})}\n\n"
                else:
                    found.append((mkt, log_path))

            if not found:
                yield f"data: {_json.dumps({'type': 'done'})}\n\n"
                return

            # Build one PowerShell script: single Excel instance, loop all files
            file_entries = ";".join(
                f"@{{mkt='{mkt}';path='{_to_win(p)}'}}"
                for mkt, p in found
            )
            ps = (
                "$ErrorActionPreference='Continue';"
                "$xl=New-Object -ComObject Excel.Application;"
                "$xl.Visible=$false;$xl.DisplayAlerts=$false;"
                f"$files=@({file_entries});"
                "try{"
                "foreach($f in $files){"
                "Write-Output \"RUNNING:$($f.mkt)\";"
                "[Console]::Out.Flush();"
                "try{"
                "$wb=$xl.Workbooks.Open($f.path);"
                "$xl.Run('BillingMacro');"
                "$wb.Save();$wb.Close($false);"
                "Write-Output \"OK:$($f.mkt)\""
                "}catch{Write-Output \"ERR:$($f.mkt):$_\"}"
                "[Console]::Out.Flush()"
                "}"
                "}finally{"
                "$xl.Quit();"
                "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($xl)|Out-Null"
                "}"
            )

            # Seed all found markets as "waiting" so the UI shows them
            for mkt, p in found:
                yield f"data: {_json.dumps({'type': 'market', 'market': mkt, 'status': 'waiting', 'file': p.name})}\n\n"

            try:
                proc = await asyncio.create_subprocess_exec(
                    "powershell.exe", "-NoProfile", "-NonInteractive", "-Command", ps,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.STDOUT,
                )
                async for raw in proc.stdout:
                    line = raw.decode(errors="replace").strip()
                    if not line:
                        continue
                    if line.startswith("RUNNING:"):
                        mkt = line[8:]
                        fname = next((p.name for m, p in found if m == mkt), "")
                        yield f"data: {_json.dumps({'type': 'market', 'market': mkt, 'status': 'running', 'file': fname})}\n\n"
                    elif line.startswith("OK:"):
                        mkt = line[3:]
                        fname = next((p.name for m, p in found if m == mkt), "")
                        yield f"data: {_json.dumps({'type': 'market', 'market': mkt, 'status': 'ok', 'file': fname})}\n\n"
                    elif line.startswith("ERR:"):
                        parts = line[4:].split(":", 1)
                        mkt = parts[0]
                        msg = parts[1] if len(parts) > 1 else ""
                        fname = next((p.name for m, p in found if m == mkt), "")
                        yield f"data: {_json.dumps({'type': 'market', 'market': mkt, 'status': 'error', 'file': fname, 'msg': msg})}\n\n"
                await proc.wait()
            except Exception as exc:
                yield f"data: {_json.dumps({'type': 'error', 'msg': str(exc)})}\n\n"

            yield f"data: {_json.dumps({'type': 'done'})}\n\n"

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @router.get("/api/master-control/logs/load")
    async def load_traffic_log(date: str = Query(...), market: str = Query(...)):

        def _fmt_t(val) -> str:
            if val is None:
                return ""
            if isinstance(val, timedelta):
                total = int(val.total_seconds())
                h, rem = divmod(total, 3600)
                m, s = divmod(rem, 60)
                return f"{h}:{m:02d}:{s:02d}"
            if hasattr(val, "strftime"):
                return val.strftime("%-H:%M:%S")
            return str(val)

        def _run():
            target = _date_cls.fromisoformat(date)
            log_path = _find_traffic_log(market, target)
            if log_path is None:
                return None
            day_name = target.strftime("%A")
            wb = _wb_load_fast(log_path, keep_vba=True, data_only=True)
            if day_name not in wb.sheetnames:
                return None
            ws = wb[day_name]
            programs: list = []
            current_prg: Optional[dict] = None
            orphans: list = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not any(c is not None for c in row):
                    continue
                rtype = str(row[13] or "").strip()
                if not rtype:
                    continue
                r = {
                    "excel_row":   i + 1,
                    "bill_code":   str(row[0] or ""),
                    "time_in":     _fmt_t(row[4]),
                    "time_out":    _fmt_t(row[5]),
                    "length":      _fmt_t(row[6]),
                    "show_name":   str(row[7] or ""),
                    "actual_time": _fmt_t(row[8]),
                    "language":    str(row[9] or ""),
                    "type":        rtype,
                    "gross":       float(row[15] or 0),
                    "net":         float(row[21] or 0),
                    "affidavit":   str(row[26] or ""),
                    "revenue_type": str(row[23] or ""),
                }
                if rtype == "PRG":
                    current_prg = {**r, "spots": []}
                    programs.append(current_prg)
                else:
                    if current_prg is not None:
                        current_prg["spots"].append(r)
                    else:
                        orphans.append(r)
            return {
                "date": date,
                "market": market,
                "day": day_name,
                "file": log_path.name,
                "programs": programs,
                "orphans": orphans,
            }

        result = await asyncio.get_running_loop().run_in_executor(None, _run)
        if result is None:
            return JSONResponse({"error": "Log file not found"}, status_code=404)
        return JSONResponse(result)

    @router.post("/api/master-control/logs/fill-program")
    async def fill_program_times(body: dict = Body(...)):

        date_str = body["date"]
        market   = body["market"]
        spots             = body["spots"]  # [{excel_row, show_name, actual_time}]
        time_in           = body.get("time_in", "")
        time_out          = body.get("time_out", "")
        program_language  = body.get("language", "")
        market_id = _MC_MARKET_IDS.get(market)
        if market_id is None:
            return JSONResponse({"error": f"Unknown market: {market}"}, status_code=400)

        target = _date_cls.fromisoformat(date_str)

        def _run():
            log_path = _find_traffic_log(market, target)
            if log_path is None:
                raise FileNotFoundError(f"Log not found for {market} {date_str}")
            day_name = target.strftime("%A")
            wb = _wb_load_fast(log_path, keep_vba=True)
            try:
                ws = wb[day_name]
                project_root = Path(__file__).parent.parent.parent.parent
                sys.path.insert(0, str(project_root))
                from browser_automation.etere_direct_client import connect

                with connect() as conn:
                    cur = conn.cursor(as_dict=True)
                    results = _mc_fill_program_spots(
                        ws, target, market_id, spots, time_in, time_out, cur,
                        program_language=program_language,
                    )
                _wb_save_fast(wb, log_path)
                return results
            finally:
                wb.close()

        try:
            results = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse({"results": results})
        except FileNotFoundError as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.post("/api/master-control/logs/fill-all")
    async def fill_all_program_times(body: dict = Body(...)):

        date_str = body["date"]
        market = body["market"]
        programs = body.get("programs") or []
        if not programs:
            return JSONResponse({"error": "No programs provided"}, status_code=400)

        market_id = _MC_MARKET_IDS.get(market)
        if market_id is None:
            return JSONResponse({"error": f"Unknown market: {market}"}, status_code=400)

        target = _date_cls.fromisoformat(date_str)

        def _run():
            log_path = _find_traffic_log(market, target)
            if log_path is None:
                raise FileNotFoundError(f"Log not found for {market} {date_str}")
            day_name = target.strftime("%A")
            wb = _wb_load_fast(log_path, keep_vba=True)
            try:
                ws = wb[day_name]
                project_root = Path(__file__).parent.parent.parent.parent
                sys.path.insert(0, str(project_root))
                from browser_automation.etere_direct_client import connect

                all_results = []
                with connect() as conn:
                    cur = conn.cursor(as_dict=True)
                    for prg in programs:
                        spots = prg.get("spots") or []
                        if not spots:
                            continue
                        all_results.extend(
                            _mc_fill_program_spots(
                                ws,
                                target,
                                market_id,
                                spots,
                                prg.get("time_in", ""),
                                prg.get("time_out", ""),
                                cur,
                                program_language=prg.get("language", ""),
                            )
                        )
                _wb_save_fast(wb, log_path)
                return all_results
            finally:
                wb.close()

        try:
            results = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse({"results": results})
        except FileNotFoundError as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.post("/api/master-control/logs/save-airtime")
    async def save_airtime(body: dict = Body(...)):

        def _run():
            target = _date_cls.fromisoformat(body["date"])
            log_path = _find_traffic_log(body["market"], target)
            if log_path is None:
                raise FileNotFoundError(f"Log not found for {body['market']} {body['date']}")
            day_name = target.strftime("%A")
            import datetime as _dt2
            wb = _wb_load_fast(log_path, keep_vba=True)
            ws = wb[day_name]
            raw = str(body["actual_time"]).strip()
            try:
                parts = raw.split(":")
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0
                cell_val = _dt2.timedelta(hours=h, minutes=m, seconds=s)
            except Exception:
                cell_val = raw
            ws.cell(row=int(body["excel_row"]), column=9).value = cell_val
            _wb_save_fast(wb, log_path)

        try:
            await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse({"ok": True})
        except FileNotFoundError as e:
            return JSONResponse({"error": str(e)}, status_code=404)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    # ── Break Optimization ─────────────────────────────────────────────────────

    @router.get("/master-control/break-optimization", response_class=HTMLResponse)
    async def break_optimization_page(request: Request):
        return templates.TemplateResponse(request, "master_control/break_optimization.html")

    # Log-style variant (user email 2026-07-15): Edit Logs-like program list,
    # expand a show to see its break optimizations; everything auto-loads
    # (pill click / date change / typed time window), no Load button.
    # Kept alongside the classic page so the team can compare and pick one.
    @router.get("/master-control/break-optimization-log", response_class=HTMLResponse)
    async def break_optimization_log_page(request: Request):
        return templates.TemplateResponse(request, "master_control/break_optimization_log.html")

    # ── Set up Daily Programming ────────────────────────────────────────────────

    # Network → set of station COD_USERs the programming fans out across.
    _DP_NETWORK_CODUSERS = {
        "CTV": [1, 2, 3, 4, 5, 6, 7, 8, 9],  # Crossings TV: all markets except DAL
        "TAC": [10],                          # The Asian Channel: DAL only
    }

    @router.get("/master-control/daily-programming", response_class=HTMLResponse)
    async def daily_programming_page(request: Request):
        return templates.TemplateResponse(request, "master_control/daily_programming.html")

    @router.get("/api/master-control/daily-programming/grid")
    async def daily_programming_grid(network: str, date: str):
        """Return the program lineup for a network/day from the K: weekly grid."""
        import datetime as _dt

        from src.business_logic.services.programming_grid import get_day_programs
        net = (network or "").upper()
        if net not in _DP_NETWORK_CODUSERS:
            return {"found": False, "programs": [], "error": f"Unknown network '{network}'"}
        try:
            d = _dt.datetime.strptime(date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"found": False, "programs": [], "error": "Invalid date (expected YYYY-MM-DD)"}
        try:
            return get_day_programs(net, d)
        except Exception as exc:  # noqa: BLE001 - surface grid-read errors to the UI
            return {"found": False, "programs": [], "error": f"Grid read failed: {exc}"}

    @router.get("/api/master-control/daily-programming/search-files")
    async def daily_programming_search_files(q: str, limit: int = 25):
        """Super-search the FILMATI media library by code/description."""
        from browser_automation.etere_direct_client import connect as _db_connect
        term = (q or "").strip()
        if len(term) < 2:
            return {"files": []}
        n = max(1, min(int(limit), 100))
        like = f"%{term}%"
        try:
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                # Programs only: FILMATI.TIPO is always 'T'; NEWTYPE='PGM' = program
                # (vs 'COM' commercial, 'PER'/'PSA'/etc.). PGMX is a rarer program variant.
                cur.execute(
                    f"""SELECT TOP {n} ID_FILMATI, COD_PROGRA, DESCRIZIO, DURATA, NEWTYPE
                        FROM FILMATI
                        WHERE NEWTYPE = 'PGM'
                          AND (DATA_SCAD IS NULL OR DATA_SCAD >= CAST(GETDATE() AS DATE))
                          AND (COD_PROGRA LIKE %s OR DESCRIZIO LIKE %s)
                        ORDER BY ID_FILMATI DESC""",
                    (like, like),
                )
                rows = cur.fetchall()
            return {"files": [
                {"id": r["ID_FILMATI"], "code": r["COD_PROGRA"], "desc": r["DESCRIZIO"],
                 "durata": r["DURATA"], "tipo": r["NEWTYPE"]}
                for r in rows]}
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"files": [], "error": f"Search failed: {exc}"}

    @router.get("/api/master-control/daily-programming/edl-status")
    async def daily_programming_edl_status(filmati: int):
        """Auto-detect EDL on a file via the same dbo.ExplodeEdl the engine uses.
        >1 segment ⇒ the file has EDL splits and should be exploded."""
        from browser_automation.etere_direct_client import connect as _db_connect
        try:
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT COUNT(*) FROM dbo.ExplodeEdl(%s, 0, N'eeAutomatic', 1, dbo.sch_GetInfDigit(%s, 1))",
                    (filmati, filmati),
                )
                n = int(cur.fetchone()[0])
            return {"filmati": filmati, "has_edl": n > 1, "segments": n}
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"filmati": filmati, "has_edl": False, "segments": 0, "error": f"EDL check failed: {exc}"}

    # Count of PRGS (program) segments in a market/date time-window.
    _DP_PRGS_COUNT_SQL = """SELECT COUNT(*) AS n FROM (
        SELECT DISTINCT seg.ID_TrafficSegment
        FROM traffic_calendar ca WITH(NOLOCK)
        JOIN traffic_scheduleblock sb WITH(NOLOCK) ON sb.ID_TrafficSchedule=ca.ID_TrafficSchedule
        JOIN traffic_block bl WITH(NOLOCK) ON bl.ID_TrafficBlock=sb.ID_TrafficBlock
        OUTER APPLY (
          SELECT si.ID_TrafficSegment,si.Offset,si.Type FROM trf_instancesegment si WITH(NOLOCK)
            WHERE si.ID_TrafficBlock=bl.ID_TrafficBlock AND si.COD_USER=ca.Cod_User AND si.INSTANCEDATE=ca.Date AND si.visible=1
          UNION
          SELECT se.ID_TrafficSegment,se.Offset,se.Type FROM traffic_segment se WITH(NOLOCK)
            WHERE se.ID_TrafficBlock=bl.ID_TrafficBlock AND se.visible=1
            AND (SELECT COUNT(*) FROM trf_instancesegment WITH(NOLOCK) WHERE ID_TrafficBlock=bl.ID_TrafficBlock AND COD_USER=ca.Cod_User AND INSTANCEDATE=ca.Date)=0
        ) seg
        WHERE ca.Cod_User=%s AND ca.Date=%s AND bl.expired=0 AND seg.Type='PRGS'
          AND (sb.offset+seg.Offset)>=%s AND (sb.offset+seg.Offset)<%s
    ) x"""

    @router.get("/api/master-control/daily-programming/program-pieces")
    async def daily_programming_program_pieces(code: str, date: str, coduser: int,
                                               start: str = "", end: str = ""):
        """No-EDL path: find the show's a/b/c/d pieces and compare the count to the
        PRGS program-break count in that market/time-window (fewer pieces ⇒ fillers)."""
        import datetime as _dt

        from browser_automation.etere_direct_client import connect as _db_connect

        # Broadcast-day-aware frame window (06:00→30:00; post-midnight = 24:00–29:59).
        from src.business_logic.services.daily_programming_run import _window as _bcast_window

        try:
            d = _dt.datetime.strptime(date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"error": "Invalid date", "pieces": []}

        base = code or ""
        if base[-1:].isalpha() and base[-1:].isupper():
            base = base[:-1]  # strip the piece letter (A/B/C/…) to get the show base
        lo, hi = _bcast_window(start, end)
        try:
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    "SELECT ID_FILMATI, COD_PROGRA, DURATA FROM FILMATI "
                    "WHERE NEWTYPE='PGM' AND COD_PROGRA LIKE %s ORDER BY COD_PROGRA",
                    (base + "%",),
                )
                rows = cur.fetchall()
                pieces = [r for r in rows
                          if len(r["COD_PROGRA"]) == len(base) + 1 and r["COD_PROGRA"][-1:].isalpha()]
                has_pieces = bool(pieces)   # real a/b/c… siblings exist (vs single whole file)
                if not pieces:  # no a/b/c siblings → single whole-file program
                    cur.execute("SELECT ID_FILMATI, COD_PROGRA, DURATA FROM FILMATI WHERE COD_PROGRA=%s", (code,))
                    pieces = cur.fetchall()
                slots = None
                if lo is not None and hi is not None:
                    cur.execute(_DP_PRGS_COUNT_SQL, (coduser, d, lo, hi))
                    slots = int(cur.fetchone()["n"])
            n = len(pieces)
            out = {"base": base, "count": n, "has_pieces": has_pieces, "prgs_slots": slots,
                   "pieces": [{"id": r["ID_FILMATI"], "code": r["COD_PROGRA"], "durata": r["DURATA"]} for r in pieces]}
            if slots is not None:
                out["fillers_needed"] = max(0, slots - n)
                out["status"] = "match" if n == slots else ("fillers_needed" if n < slots else "too_many")
            return out
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"error": f"Piece check failed: {exc}", "pieces": []}

    @router.post("/api/master-control/daily-programming/import-edl")
    async def daily_programming_import_edl(body: dict = Body(...)):
        """Write an EDL to a single-file program from a dropped EDIUS marker CSV,
        then self-validate via dbo.ExplodeEdl. Commits only if it explodes into
        exactly the expected parts; otherwise rolls back untouched."""
        from browser_automation.etere_direct_client import connect as _db_connect
        from src.business_logic.services.edl_import import apply_edl_from_csv, parse_edius_csv
        try:
            filmati = int(body.get("filmati"))
        except (TypeError, ValueError):
            return {"ok": False, "error": "Missing/invalid filmati"}
        coduser = int(body.get("coduser") or 1)
        try:
            splits, eom = parse_edius_csv(body.get("csv") or "")
        except ValueError as exc:
            return {"ok": False, "error": str(exc)}
        try:
            with _db_connect() as conn:
                res = apply_edl_from_csv(conn, filmati, splits, eom, coduser)
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"ok": False, "error": f"EDL import failed: {exc}"}
        res["splits"] = splits
        res["eom"] = eom
        res["count"] = len(splits) + 1
        return res

    @router.get("/api/master-control/daily-programming/placement")
    async def daily_programming_placement(network: str, date: str):
        """Pre-flight: which markets already have live program content placed, for the
        day. Returns the network's markets + all live (LIVELLO=0, non-bump) PGM
        placements as (cu, ora) so the UI can bucket per-program."""
        import datetime as _dt

        from browser_automation.etere_direct_client import connect as _db_connect
        net = (network or "").upper()
        cus = _DP_NETWORK_CODUSERS.get(net)
        if not cus:
            return {"markets": [], "placed": [], "error": f"Unknown network '{network}'"}
        try:
            d = _dt.datetime.strptime(date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"markets": cus, "placed": [], "error": "Invalid date"}
        inlist = ",".join(str(int(c)) for c in cus)
        try:
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    f"""SELECT COD_USER, ORA, COD_PROGRA FROM TPALINSE
                        WHERE COD_USER IN ({inlist}) AND DATA=%s AND NEWTYPE='PGM'
                          AND LIVELLO=0 AND ID_FILMATI>0 AND COD_PROGRA NOT LIKE 'BUMP%%'""",
                    (d,),
                )
                placed = [{"cu": int(r[0]), "ora": int(r[1]), "code": (r[2] or "").strip()}
                          for r in cur.fetchall()]
            return {"markets": cus, "placed": placed}
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"markets": cus, "placed": [], "error": f"Placement check failed: {exc}"}

    @router.get("/api/master-control/daily-programming/placed-pieces")
    async def daily_programming_placed_pieces(network: str, date: str):
        """All live pieces-mode program rows for the network's markets on one
        date, grouped client-side for the Replace-a-piece panel. Explode-mode
        parts (PART>=1) are excluded by design — replacing those means
        re-importing the EDL, not swapping a file."""
        import datetime as _dt

        from browser_automation.etere_direct_client import connect as _db_connect
        from src.business_logic.services.daily_programming_run import list_placed_pieces
        cus = _DP_NETWORK_CODUSERS.get((network or "").upper())
        if not cus:
            return {"pieces": [], "error": f"Unknown network '{network}'"}
        try:
            d = _dt.datetime.strptime(date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"pieces": [], "error": "Invalid date"}

        def _run():
            out = []
            with _db_connect() as conn:
                cur = conn.cursor()
                for cu in cus:
                    for p in list_placed_pieces(cur, int(cu), d):
                        p["cu"] = int(cu)
                        out.append(p)
            return out

        loop = asyncio.get_running_loop()
        try:
            return {"pieces": await loop.run_in_executor(None, _run)}
        except Exception as exc:  # noqa: BLE001 - surface DB errors to the UI
            return {"pieces": [], "error": f"Lookup failed: {exc}"}

    @router.post("/api/master-control/daily-programming/replace-piece")
    async def daily_programming_replace_piece(body: dict = Body(...)):
        """Replace one placed piece (e.g. revised Piece B) with a new asset in
        the selected markets, leaving the rest of the show untouched. Markets
        run sequentially — a piece swap is quick and a lone transaction never
        deadlocks."""
        import datetime as _dt

        from browser_automation.etere_direct_client import connect as _db_connect
        from src.business_logic.services.daily_programming_run import (
            _drain_pending_filmati_syncs,
            _window,
            replace_piece,
        )
        try:
            d = _dt.datetime.strptime(body.get("date") or "", "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"results": [], "error": "Invalid date"}
        try:
            old_fid = int(body.get("oldFileId"))
            new_fid = int(body.get("newFileId"))
            markets = [int(c) for c in (body.get("markets") or [])]
        except (TypeError, ValueError):
            return {"results": [], "error": "oldFileId, newFileId and markets are required"}
        lo, hi = _window(body.get("start") or "", body.get("end") or "")
        if lo is None or hi is None:
            return {"results": [], "error": "Invalid show time window"}
        if not markets:
            return {"results": [], "error": "No markets selected"}
        if old_fid == new_fid:
            return {"results": [], "error": "Replacement is the same asset as the placed piece"}

        def _run():
            results = []
            pending = []
            with _db_connect() as conn:
                for cu in markets:
                    r = replace_piece(conn, cu, d, lo, hi, old_fid, new_fid, pending)
                    r.pop("_deadlock", None)
                    results.append(r)
                if pending:
                    _drain_pending_filmati_syncs(conn.cursor(), conn, pending)
            return results

        loop = asyncio.get_running_loop()
        try:
            return {"results": await loop.run_in_executor(None, _run)}
        except Exception as exc:  # noqa: BLE001 - surface connection-level errors
            return {"results": [], "error": f"Replace failed: {exc}"}

    @router.post("/api/master-control/daily-programming/run")
    async def daily_programming_run(body: dict = Body(...)):
        """Set up across markets: place each assignment into each target market
        (transaction-safe per market, skips already-placed). Returns per
        (program, market) results.

        Each market gets its own DB connection and runs on its own thread.
        Every TPALINSE query in run_market() is scoped to a single COD_USER,
        so the placement itself never crosses markets. Running them in
        parallel (instead of one long serial loop over every program x market
        pair) keeps a full-network run well under the reverse proxy's read
        timeout.

        FILMATI (the asset library) IS shared across markets — a show file
        exploded into every market, a shared bumper, a shared FCC-ID element
        is one row multiple markets' threads touch. run_market()'s checksum
        sync guards that with a per-asset lock and defers to `pending` on
        contention instead of blocking (2026-07-09, fixed the 1205 deadlock
        this parallelism introduced); each market thread drains its own
        pending list once it's placed everything it was given.

        Markets that still exhaust run_market()'s deadlock retries during the
        parallel fan-out get one solo rerun at the end, after every sibling
        thread has finished (2026-07-13) — see the second-pass loop below."""
        import datetime as _dt
        from concurrent.futures import ThreadPoolExecutor

        from browser_automation.etere_direct_client import connect as _db_connect
        from src.business_logic.services.daily_programming_run import (
            _drain_pending_filmati_syncs,
            run_market,
            sweep_daily_ids,
        )
        date = body.get("date")
        codusers = body.get("codUsers") or []
        assignments = body.get("assignments") or []
        try:
            d = _dt.datetime.strptime(date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"results": [], "error": "Invalid date"}
        if not codusers or not assignments:
            return {"results": [], "error": "Nothing to set up (no markets or assignments)"}

        def _run_market_all(cu):
            cu = int(cu)
            out = []
            pending = []
            try:
                with _db_connect() as conn:
                    for a in assignments:
                        if a.get("mode") == "shoplc" and cu == 10:
                            continue  # Shop LC never airs on DAL
                        label = a.get("title") or ("program " + str(a.get("programIndex")))
                        r = run_market(conn, cu, d, a, pending)
                        r["program"] = label
                        if r.get("_deadlock"):
                            r["_assignment"] = a  # for the sequential second pass
                        out.append(r)
                    if pending:
                        _drain_pending_filmati_syncs(conn.cursor(), conn, pending)
            except Exception as exc:  # noqa: BLE001 - surface connection-level errors for this market
                out.append({"cu": cu, "ok": False, "skipped": False, "message": f"Run failed: {exc}"})
            return out

        results = []
        with ThreadPoolExecutor(max_workers=len(codusers)) as ex:
            for out in ex.map(_run_market_all, codusers):
                results.extend(out)

        # Sequential second pass: a market that lost the 1205 deadlock coin flip
        # on every retry during the parallel fan-out reruns here one at a time,
        # after all sibling threads have finished — a lone market has nothing of
        # ours to deadlock with (verified 2026-07-13: HOU+SFO failed the parallel
        # run 5/5 attempts, then placed cleanly on a solo rerun).
        for i, r in enumerate(results):
            if not r.pop("_deadlock", False):
                continue
            a = r.pop("_assignment")
            cu = int(r["cu"])
            print(f"[daily-programming] cu={cu} rerunning solo after parallel-pass deadlock")
            pending = []
            try:
                with _db_connect() as conn:
                    retry = run_market(conn, cu, d, a, pending)
                    if pending:
                        _drain_pending_filmati_syncs(conn.cursor(), conn, pending)
            except Exception as exc:  # noqa: BLE001 - surface connection-level errors for this market
                retry = {"cu": cu, "ok": False, "skipped": False, "message": f"Run failed: {exc}"}
            retry.pop("_deadlock", None)
            retry["program"] = r["program"]
            results[i] = retry

        # Standing daily elements (the OTA end-of-day FCC IDs): any run that
        # includes a market with daily elements also sweeps today → today+2
        # (plus the run's own date), filling whichever days are published and
        # missing the element. Idempotent — unpublished days are reported and
        # picked up by a later run's sweep. Runs after all placement threads
        # so it never contends with them.
        for cu in {int(c) for c in codusers}:
            sweep_dates = [_dt.date.today() + _dt.timedelta(days=i) for i in range(3)]
            if d not in sweep_dates:
                sweep_dates.append(d)
            pending = []
            try:
                with _db_connect() as conn:
                    sweep = sweep_daily_ids(conn, cu, sweep_dates, pending)
                    if pending:
                        _drain_pending_filmati_syncs(conn.cursor(), conn, pending)
            except Exception as exc:  # noqa: BLE001 - surface sweep-level errors as a result row
                sweep = [{"date": "", "ok": False, "skipped": False,
                          "message": f"sweep failed: {exc}"}]
            if not sweep:
                continue  # market has no daily elements configured — no row

            def _md(iso):  # 'YYYY-MM-DD' → 'M/D' (calendar-safe, no %-d)
                try:
                    _y, m, day = iso.split("-")
                    return f"{int(m)}/{int(day)}"
                except ValueError:
                    return iso
            by_msg = {}
            for r in sweep:
                by_msg.setdefault(r["message"], []).append(_md(r["date"]))
            results.append({
                "cu": cu,
                "ok": all(r["ok"] for r in sweep),
                "skipped": all(r["skipped"] for r in sweep),
                "program": "End-of-day FCC ID",
                "message": " · ".join(f"{m}: {', '.join(ds)}" for m, ds in by_msg.items()),
            })
        return {"results": results}

    _BO_MARKET_IDS = {"NYC": 1, "CMP": 2, "HOU": 3, "SFO": 4, "SEA": 5, "LAX": 6, "CVC": 7, "WDC": 8, "MMT": 9, "DAL": 10}
    _BO_FPS = 29.97

    def _bo_frames_to_time(frames: int) -> str:
        secs = round(frames / _BO_FPS)
        h, rem = divmod(secs, 3600)
        mn, s = divmod(rem, 60)
        return f"{h}:{mn:02d}:{s:02d}"

    def _bo_frames_to_hhmm(frames: int) -> str:
        secs = round(frames / _BO_FPS)
        h, mn = divmod(secs, 3600)
        return f"{h:02d}:{mn // 60:02d}"

    def _bo_time_to_frames(t: str) -> int:
        return _bcast_time_to_frames(t, _BO_FPS)  # broadcast-day aware (post-midnight = 24:00–29:59)

    def _bo_classify(newtype: str, capo, fine, is_wl: bool, prev_label: str, prev_contract: str = "", contract: str = ""):
        if capo and fine:
            return 1, "BOOKEND"
        if capo and not fine:
            return 2, "BILLBOARD"
        if prev_label == "BILLBOARD" and newtype in ("COM", "BNS") and contract and contract == prev_contract:
            return 3, "COMPANION"
        if newtype in ("COM", "BNS") and not is_wl:
            return 4, "PAYING"
        if newtype in ("COM", "BNS") and is_wl:
            return 5, "WORLDLINK"
        if newtype == "PER":
            return 6, "PI"
        if newtype == "PSA":
            return 7, "PSA"
        if newtype == "ID":
            return 8, "STATION ID"
        return 0, newtype or "OTHER"

    def _pi_product_key(title: str) -> str:
        """'PI-504-030: ...' → 'PI-504'; unrecognised titles return the full title."""
        import re as _re
        m = _re.match(r'^(PI-\d+)-\d+', (title or "").strip(), _re.IGNORECASE)
        return m.group(1).upper() if m else (title or "").strip().upper()

    def _bo_optimize(spots: list) -> list:
        skip = set()
        pairs = []
        bookend_count = 0
        for j, s in enumerate(spots):
            if j in skip:
                continue
            if s["label"] == "BILLBOARD":
                pair = [s]
                if j + 1 < len(spots) and spots[j + 1]["label"] == "COMPANION":
                    pair.append(spots[j + 1])
                    skip.add(j + 1)
                pairs.append((2, pair))
            elif s["label"] == "BOOKEND":
                bookend_count += 1
                # First bookend → top of break; second bookend → bottom (after everything else)
                prio = 1 if bookend_count == 1 else 999
                pairs.append((prio, [s]))
            else:
                pairs.append((s["priority"], [s]))
        pairs.sort(key=lambda x: x[0])
        result = [s for _, grp in pairs for s in grp]

        # Reorder PI spots to avoid same-product adjacency (e.g. PI-504-030 and PI-504-060)
        pi_indices = [i for i, s in enumerate(result) if s["label"] == "PI"]
        if len(pi_indices) > 1:
            from collections import Counter
            pi_spots = [result[i] for i in pi_indices]
            counts = Counter(_pi_product_key(s["title"]) for s in pi_spots)
            reordered, last_key, remaining = [], None, list(pi_spots)
            while remaining:
                best = max(
                    (s for s in remaining if _pi_product_key(s["title"]) != last_key),
                    key=lambda s: counts[_pi_product_key(s["title"])],
                    default=remaining[0],
                )
                reordered.append(best)
                last_key = _pi_product_key(best["title"])
                counts[last_key] -= 1
                remaining.remove(best)
            for idx, pi_idx in enumerate(pi_indices):
                result[pi_idx] = reordered[idx]

        return result

    def _bo_pi_library_pick(cur, duration: int, exclude_keys: set):
        """Pick a valid replacement filler of the same length (±5 frames) from the
        FILMATI library — PI first, PSA fallback — whose product key is not already
        used in the break. Returns a filmati dict or None. Mirrors the proven
        blacklist filler-selection query."""
        for pattern in ("PI-%%", "PSA-%%"):
            cur.execute(
                "SELECT TOP 20 ID_FILMATI, COD_PROGRA, DESCRIZIO, DURATA, NEWTYPE"
                " FROM FILMATI"
                " WHERE DESCRIZIO LIKE %s"
                "   AND DESCRIZIO NOT LIKE 'DO NOT%%'"
                "   AND (DATA_SCAD IS NULL OR DATA_SCAD > GETDATE())"
                "   AND ABS(DURATA - %s) <= 5"
                " ORDER BY NEWID()",
                (pattern, duration),
            )
            for row in cur.fetchall():
                if _pi_product_key(row["DESCRIZIO"]) not in exclude_keys:
                    return dict(row)
        return None

    def _bo_resolve_pi_duplicates(cur, breaks: list) -> None:
        """When a break contains two PI spots of the same product (e.g. two PI-504),
        replace one of them IN PLACE with a different valid same-length PI from the
        FILMATI library (PSA fallback). Nothing moves between breaks or time slots —
        the switch stays inside the pod/program. Modifies breaks in-place: sets the
        replacement payload on the affected optimized spot and updates the
        'changed' / 'violation' / 'pi_unresolvable' flags."""

        def _pi_keys(brk):
            return [_pi_product_key(s["title"]) for s in brk["optimized"] if s["label"] == "PI"]

        for brk in breaks:
            keys = _pi_keys(brk)
            if len(keys) == len(set(keys)):
                brk["pi_unresolvable"] = False
                continue

            # Extra occurrences (2nd, 3rd, … of any product key) are the duplicates
            # to replace. Positions never shift (replacement is in-place), so a single
            # pass over the pre-computed indices is safe.
            seen: set = set()
            dup_indices: list = []
            for j, s in enumerate(brk["optimized"]):
                if s["label"] != "PI":
                    continue
                k = _pi_product_key(s["title"])
                if k in seen:
                    dup_indices.append(j)
                else:
                    seen.add(k)

            replacements: list = []
            for dup_idx in dup_indices:
                dup_spot = brk["optimized"][dup_idx]
                # Replacement must differ from every OTHER PI product left in the break
                exclude = {_pi_product_key(s["title"])
                           for m, s in enumerate(brk["optimized"])
                           if s["label"] == "PI" and m != dup_idx}
                pick = _bo_pi_library_pick(cur, dup_spot["duration"], exclude)
                if not pick:
                    continue                        # nothing fits at this length
                old_title = dup_spot["title"]
                new_title = (pick["DESCRIZIO"] or "").strip()
                brk["optimized"][dup_idx] = {
                    **dup_spot,
                    "title":              new_title,
                    "replace_filmati_id": int(pick["ID_FILMATI"]),
                    "replace_title":      new_title,
                    "replace_cod_progra": (pick["COD_PROGRA"] or "").strip(),
                    "replace_newtype":    (pick["NEWTYPE"] or "PER").strip(),
                    "pi_replacement":     f"{old_title} → {new_title}",
                }
                replacements.append(f"{old_title} → {new_title}")

            new_keys = _pi_keys(brk)
            brk["pi_replacements"] = replacements
            # Any product key still duplicated had no same-length filler → manual fix
            brk["pi_unresolvable"] = len(new_keys) != len(set(new_keys))
            # A creative swap keeps the id order, so flag 'changed' explicitly.
            cur_ids = [s["id"] for s in brk["current"]]
            opt_ids = [s["id"] for s in brk["optimized"]]
            brk["changed"] = brk["changed"] or bool(replacements) or cur_ids != opt_ids
            brk["ordering_violation"] = brk["violation"] = brk["changed"] or brk["pi_unresolvable"]

    def _bo_apply_pi_replacement(cur, u: dict, id_tpalinse: int) -> None:
        """If an optimized-spot payload carries a library PI/PSA replacement, swap the
        creative on its TPALINSE row in place (ID_FILMATI + identity fields). ORA /
        XORDER / DURATION are left untouched — the library match is within ±5 frames,
        so downstream pod timing is unaffected. Mirrors the auto-assign creative
        update (ID_FILMATI) and the blacklist filler SUPPORTO convention."""
        fid = u.get("replace_filmati_id")
        if not fid:
            return
        title    = (u.get("replace_title") or "").strip()
        supporto = _pi_filler_supporto(cur, fid, title)
        cur.execute(
            "UPDATE TPALINSE SET ID_FILMATI = %d, COD_PROGRA = %s, NEWTYPE = %s,"
            " TITLE = %s, SUPPORTO = %s WHERE ID_TPALINSE = %d",
            (int(fid), (u.get("replace_cod_progra") or ""), (u.get("replace_newtype") or "PER"),
             title, supporto, id_tpalinse),
        )
        # Yellow-triangle removal: the swap leaves the row's stored
        # SCHEDULE_CHECKSUM reflecting the OLD creative, so EE flags the spot.
        # PI/PSA fillers rotate near-daily, so the file already exists locally
        # on the playout servers — freezing the checksum fully clears the
        # triangle here (unlike fresh-from-S3 shows, where the per-CIB download
        # re-stales it). Same recipe as daily programming's _sync_checksums:
        # normalize the checksum-input FILMATI fields to canonical settled
        # values — never on live-feed assets (LIVE_ID guard) — then re-store.
        cur.execute(
            "UPDATE FILMATI SET INF_DIGIT = 0, AUDIO = NULL, AUDIO_LANGUAGE = NULL"
            " WHERE ID_FILMATI = %d AND LIVE_ID IS NULL",
            (int(fid),),
        )
        cur.execute(
            "UPDATE TPALINSE SET SCHEDULE_CHECKSUM = dbo.sch_getFilmatiCheckSum(%d)"
            " WHERE ID_TPALINSE = %d",
            (id_tpalinse, id_tpalinse),
        )

    def _bo_fetch_sep_context(cur, market_id: int, date: str, from_frames: int, to_frames: int) -> list:
        """COM/BNS spots in a ±1-hr window — used for separation checking."""
        one_hour = round(3600 * _BO_FPS)
        ext_from = max(0, from_frames - one_hour)
        ext_to   = to_frames + one_hour
        cur.execute(
            "SELECT t.ID_TPALINSE, t.ORA, t.TITLE,"
            " ct.COMMITTENTE, ct.ID_CONTRATTITESTATA AS contract_id,"
            " COALESCE(cr.Interv_Committente, 0)    AS cust_sep,"
            " COALESCE(cr.INTERV_CONTRATTO, 0)      AS order_sep,"
            " COALESCE(cr.CONTROLLACAPOFILA, 0)     AS capofila,"
            " COALESCE(cr.CONTROLLAFINEFILA, 0)     AS finefila,"
            " COALESCE(cr.ORA_INIZIOF, cr.ORA_INIZIO) AS line_time_from,"
            " COALESCE(cr.ORA_FINEF,   cr.ORA_FINE)   AS line_time_to"
            " FROM TPALINSE t"
            " LEFT JOIN trafficTPalinse tp ON tp.ID_TPalinse = t.ID_TPALINSE"
            " LEFT JOIN CONTRATTIRIGHE cr ON cr.ID_CONTRATTIRIGHE = tp.ID_ContrattiRighe"
            " LEFT JOIN CONTRATTITESTATA ct ON ct.ID_CONTRATTITESTATA = tp.ID_CONTRATTITESTATA"
            " WHERE t.DATA = %s AND t.COD_USER = %d"
            " AND t.NEWTYPE IN ('COM', 'BNS')"
            " AND t.ORA >= %d AND t.ORA < %d"
            " AND t.LIVELLO = 0",
            (date, market_id, ext_from, ext_to),
        )
        return [dict(r) for r in cur.fetchall()]

    def _bo_check_separation(breaks: list, sep_spots: list) -> None:
        """
        For each COM/BNS spot inside a break, check two separation rules:
          - Customer separation (Interv_Committente): gap between any two spots for
            the same customer (COMMITTENTE), regardless of which contract they're on.
          - Order separation (INTERV_CONTRATTO): gap between spots under the same
            contract header (ID_CONTRATTITESTATA). Catches cases like Admerasia
            where customer sep = 0 but order sep > 0.
        Attaches 'sep_violations' list to every break; sets violation=True when found.
        """
        from collections import defaultdict

        by_cust:     dict = defaultdict(list)
        by_contract: dict = defaultdict(list)
        id_to_meta:  dict = {}

        for s in sep_spots:
            sid       = s["ID_TPALINSE"]
            cid       = s.get("COMMITTENTE")
            ctr_id    = s.get("contract_id")
            cust_sep  = int(s.get("cust_sep")  or 0)
            order_sep = int(s.get("order_sep") or 0)
            is_bookend  = bool(s.get("capofila")) and bool(s.get("finefila"))
            is_billboard = bool(s.get("capofila")) and not bool(s.get("finefila"))
            ctr_id_int  = int(ctr_id) if ctr_id is not None else None
            entry = {
                "id":           sid,
                "ora":          s["ORA"],
                "title":        (s.get("TITLE") or "").strip(),
                "is_bookend":   is_bookend,
                "is_billboard": is_billboard,
                "ctr_id":       ctr_id_int,
            }
            if cid is not None:
                by_cust[int(cid)].append(entry)
            if ctr_id_int is not None:
                by_contract[ctr_id_int].append(entry)
            id_to_meta[sid] = {
                "cust_id":     int(cid) if cid is not None else None,
                "ctr_id":      ctr_id_int,
                "cust_sep":    cust_sep,
                "order_sep":   order_sep,
                "is_bookend":  is_bookend,
                "is_billboard": is_billboard,
                "time_from":   _bo_frames_to_hhmm(int(s["line_time_from"])) if s.get("line_time_from") is not None else None,
                "time_to":     _bo_frames_to_hhmm(int(s["line_time_to"]))   if s.get("line_time_to")   is not None else None,
            }

        def _check_group(spot, group_list, req, seen_pairs, violations,
                         spot_is_bookend=False, spot_is_billboard=False, spot_ctr_id=None):
            sid = spot["id"]
            for other in group_list:
                if other["id"] == sid:
                    continue
                # Bookend pairs intentionally share a break — not a separation violation
                if spot_is_bookend and other.get("is_bookend"):
                    continue
                # Billboard+companion pairs are by design adjacent in the same contract
                other_ctr = other.get("ctr_id")
                same_contract = (spot_ctr_id is not None and spot_ctr_id == other_ctr)
                if same_contract and (spot_is_billboard or other.get("is_billboard")):
                    continue
                pair_key = (min(sid, other["id"]), max(sid, other["id"]))
                if pair_key in seen_pairs:
                    continue
                gap = abs(spot["ora"] - other["ora"])
                if gap < req:
                    seen_pairs.add(pair_key)
                    spot_meta  = id_to_meta.get(sid, {})
                    other_meta = id_to_meta.get(other["id"], {})
                    violations.append({
                        "spot_id":             sid,
                        "spot_title":          spot["title"],
                        "spot_time":           spot["time"],
                        "spot_valid_from":     spot_meta.get("time_from"),
                        "spot_valid_to":       spot_meta.get("time_to"),
                        "conflict_id":         other["id"],
                        "conflict_title":      other["title"],
                        "conflict_time":       _bo_frames_to_time(other["ora"]),
                        "conflict_valid_from": other_meta.get("time_from"),
                        "conflict_valid_to":   other_meta.get("time_to"),
                        "req_mins":            round(req / (_BO_FPS * 60), 1),
                        "actual_mins":         round(gap / (_BO_FPS * 60), 1),
                    })

        for brk in breaks:
            violations = []
            seen_pairs: set = set()
            for spot in brk["current"]:
                sid  = spot["id"]
                meta = id_to_meta.get(sid)
                if not meta:
                    continue
                is_be = meta["is_bookend"]
                is_bb = meta["is_billboard"]
                ctr   = meta["ctr_id"]
                if meta["cust_sep"] > 0 and meta["cust_id"] is not None:
                    _check_group(spot, by_cust[meta["cust_id"]], meta["cust_sep"],
                                 seen_pairs, violations, is_be, is_bb, ctr)
                if meta["order_sep"] > 0 and ctr is not None:
                    _check_group(spot, by_contract[ctr], meta["order_sep"],
                                 seen_pairs, violations, is_be, is_bb, ctr)
            brk["sep_violations"] = violations
            if violations:
                brk["violation"] = True

    def _bo_process_market(cur, market_id: int, date: str, from_frames: int, to_frames: int) -> list:
        """Fetch, annotate, segment, and optimise all breaks for one market. Returns break list."""
        _BO_BUFFER = round(3 * 60 * _BO_FPS)
        cur.execute(
            "SELECT t.ID_TPALINSE, t.ORA, t.XORDER, t.TITLE, t.COD_PROGRA, t.NEWTYPE, t.DURATION,"
            " cr.CONTROLLACAPOFILA, cr.CONTROLLAFINEFILA, ct.COD_CONTRATTO"
            " FROM TPALINSE t"
            " LEFT JOIN trafficTPalinse tp ON tp.ID_TPalinse = t.ID_TPALINSE"
            " LEFT JOIN CONTRATTIRIGHE cr ON cr.ID_CONTRATTIRIGHE = tp.ID_ContrattiRighe"
            " LEFT JOIN CONTRATTITESTATA ct ON ct.ID_CONTRATTITESTATA = tp.ID_CONTRATTITESTATA"
            " WHERE t.DATA = %s AND t.COD_USER = %d"
            " AND t.ORA >= %d AND t.ORA < %d"
            " AND t.LIVELLO = 0"
            " ORDER BY t.XORDER, t.ORA",
            (date, market_id, from_frames, to_frames + _BO_BUFFER),
        )
        rows = cur.fetchall()

        prev_label, prev_contract = None, ""
        annotated = []
        for r in rows:
            nt = (r["NEWTYPE"] or "").strip()
            contract = (r["COD_CONTRATTO"] or "").strip()
            is_wl = contract.startswith("WL")
            pri, label = _bo_classify(nt, r["CONTROLLACAPOFILA"], r["CONTROLLAFINEFILA"],
                                      is_wl, prev_label, prev_contract, contract)
            prev_label, prev_contract = label, contract
            annotated.append({
                "id":        r["ID_TPALINSE"],
                "ora":       r["ORA"],
                "time":      _bo_frames_to_time(r["ORA"]),
                "title":     (r["TITLE"] or "").strip(),
                "cod_progra":(r["COD_PROGRA"] or "").strip(),
                "newtype":   nt,
                "label":     label,
                "priority":  pri,
                "duration":  r["DURATION"] or 0,
                "contract":  contract,
                "is_fixed":  pri == 0,
            })

        breaks, i = [], 0
        while i < len(annotated):
            if annotated[i]["is_fixed"]:
                i += 1
            else:
                block = []
                while i < len(annotated):
                    row = annotated[i]
                    if row["newtype"] == "NOOP":
                        i += 1
                    elif row["is_fixed"]:
                        break
                    else:
                        block.append(row)
                        i += 1
                if not block:
                    continue
                optimized = _bo_optimize(block)
                cur_pos = block[0]["ora"]
                opt_timed = []
                for s in optimized:
                    opt_timed.append({**s, "new_ora": cur_pos, "new_time": _bo_frames_to_time(cur_pos)})
                    cur_pos += s["duration"]
                orig_ids = [s["id"] for s in block]
                pri_viol = orig_ids != [s["id"] for s in optimized]
                pi_keys  = [_pi_product_key(s["title"]) for s in block if s["label"] == "PI"]
                violation = pri_viol or len(pi_keys) != len(set(pi_keys))
                if block[0]["ora"] < to_frames:
                    breaks.append({
                        "current":            block,
                        "optimized":          opt_timed,
                        "violation":          violation,
                        "ordering_violation": violation,
                        "bookend_warning":    sum(1 for s in block if s["label"] == "BOOKEND") % 2 != 0,
                        "changed":            orig_ids != [s["id"] for s in opt_timed],
                        "pi_unresolvable":    False,
                    })

        _bo_resolve_pi_duplicates(cur, breaks)
        _bo_check_separation(breaks, _bo_fetch_sep_context(cur, market_id, date, from_frames, to_frames))
        return breaks

    @router.get("/api/master-control/break-optimization/load")
    async def load_break_optimization(
        market: str = Query(...),
        date: str = Query(...),
        time_from: str = Query(...),
        time_to: str = Query(...),
    ):
        market_id = _BO_MARKET_IDS.get(market.upper())
        if not market_id:
            return JSONResponse({"error": f"Unknown market: {market}"}, status_code=400)
        from_frames = _bo_time_to_frames(time_from)
        to_frames   = _bo_time_to_frames(time_to)

        def _run():
            from browser_automation.etere_direct_client import connect as _connect
            with _connect() as conn:
                return _bo_process_market(conn.cursor(as_dict=True), market_id, date, from_frames, to_frames)

        try:
            breaks = await asyncio.get_running_loop().run_in_executor(None, _run)
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

        return JSONResponse({
            "market": market, "date": date,
            "time_from": time_from, "time_to": time_to,
            "breaks": breaks,
        })

    @router.post("/api/master-control/break-optimization/apply")
    async def apply_break_optimization(body: dict = Body(...)):
        updates = body.get("updates", [])
        if not updates:
            return JSONResponse({"ok": True, "updated": 0})

        def _run():
            from browser_automation.etere_direct_client import connect as _connect
            ids = [int(u["id_tpalinse"]) for u in updates]
            id_placeholders = ",".join(["%d"] * len(ids))
            with _connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    f"SELECT ID_TPALINSE, XORDER FROM TPALINSE WHERE ID_TPALINSE IN ({id_placeholders})",
                    tuple(ids),
                )
                xorder_map = {r["ID_TPALINSE"]: r["XORDER"] for r in cur.fetchall()}
                # Distribute current XORDERs (sorted) to spots in their new optimized order
                sorted_xorders = sorted(v for v in xorder_map.values() if v is not None)
                cur2 = conn.cursor()
                for i, u in enumerate(updates):
                    new_ora = int(u["new_ora"])
                    new_xorder = sorted_xorders[i] if i < len(sorted_xorders) else None
                    if new_xorder is not None:
                        cur2.execute(
                            "UPDATE TPALINSE SET ORA = %d, ORA_P = %d, XORDER = %d WHERE ID_TPALINSE = %d",
                            (new_ora, new_ora, new_xorder, int(u["id_tpalinse"])),
                        )
                    else:
                        cur2.execute(
                            "UPDATE TPALINSE SET ORA = %d, ORA_P = %d WHERE ID_TPALINSE = %d",
                            (new_ora, new_ora, int(u["id_tpalinse"])),
                        )
                    _bo_apply_pi_replacement(cur2, u, int(u["id_tpalinse"]))
                conn.commit()

        try:
            await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse({"ok": True, "updated": len(updates)})
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    @router.post("/api/master-control/break-optimization/bulk-apply")
    async def bulk_apply_break_optimization(body: dict = Body(...)):
        date      = body.get("date", "")
        time_from = body.get("time_from", "")
        time_to   = body.get("time_to", "")
        if not date or not time_from or not time_to:
            return JSONResponse({"error": "date, time_from, time_to required"}, status_code=400)

        from_frames = _bo_time_to_frames(time_from)
        to_frames   = _bo_time_to_frames(time_to)
        # All Crossings TV markets — DAL excluded (Asian Channel, different break structure)
        bulk_markets = {k: v for k, v in _BO_MARKET_IDS.items() if k != "DAL"}

        def _run():
            from browser_automation.etere_direct_client import connect as _connect
            results = []
            with _connect() as conn:
                for market_name, market_id in bulk_markets.items():
                    cur = conn.cursor(as_dict=True)
                    breaks = _bo_process_market(cur, market_id, date, from_frames, to_frames)
                    changed_breaks = [b for b in breaks if b["changed"]]
                    all_updates = []
                    for brk in changed_breaks:
                        all_updates.extend(brk["optimized"])
                    if not all_updates:
                        results.append({
                            "market": market_name, "breaks_total": len(breaks),
                            "breaks_changed": 0, "spots_updated": 0,
                        })
                        continue
                    ids = [int(u["id"]) for u in all_updates]
                    id_ph = ",".join(["%d"] * len(ids))
                    cur2 = conn.cursor(as_dict=True)
                    cur2.execute(
                        f"SELECT ID_TPALINSE, XORDER FROM TPALINSE WHERE ID_TPALINSE IN ({id_ph})",
                        tuple(ids),
                    )
                    xorder_map = {r["ID_TPALINSE"]: r["XORDER"] for r in cur2.fetchall()}
                    sorted_xorders = sorted(v for v in xorder_map.values() if v is not None)
                    cur3 = conn.cursor()
                    for i, u in enumerate(all_updates):
                        new_ora    = int(u["new_ora"])
                        new_xorder = sorted_xorders[i] if i < len(sorted_xorders) else None
                        if new_xorder is not None:
                            cur3.execute(
                                "UPDATE TPALINSE SET ORA = %d, ORA_P = %d, XORDER = %d WHERE ID_TPALINSE = %d",
                                (new_ora, new_ora, new_xorder, int(u["id"])),
                            )
                        else:
                            cur3.execute(
                                "UPDATE TPALINSE SET ORA = %d, ORA_P = %d WHERE ID_TPALINSE = %d",
                                (new_ora, new_ora, int(u["id"])),
                            )
                        _bo_apply_pi_replacement(cur3, u, int(u["id"]))
                    conn.commit()
                    results.append({
                        "market":         market_name,
                        "breaks_total":   len(breaks),
                        "breaks_changed": len(changed_breaks),
                        "spots_updated":  len(all_updates),
                    })
            return results

        try:
            results = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse({"ok": True, "results": results})
        except Exception as e:
            return JSONResponse({"error": str(e)}, status_code=500)

    # ── Assign Traffic ─────────────────────────────────────────────────────────

    @router.get("/traffic/assign-assets", response_class=HTMLResponse)
    async def traffic_assign_assets_page(request: Request):
        return templates.TemplateResponse(request, "traffic/asset_assignment.html",
                                          {"traffic_formats": _TRAFFIC_FORMAT_LABELS})

    @router.get("/api/traffic/contract-search")
    async def traffic_contract_search(q: str = Query("")):
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                term = f"%{q.upper()}%"
                id_clause = "OR ct.ID_CONTRATTITESTATA = %d" % int(q) if q.isdigit() else ""
                cur.execute(f"""
                    SELECT TOP 20
                        ct.ID_CONTRATTITESTATA                      AS id,
                        ct.COD_CONTRATTO                            AS code,
                        ct.DESCRIZIONE                              AS description,
                        CONVERT(VARCHAR(10), ct.DATA_INIZIO,   101) AS date_start,
                        CONVERT(VARCHAR(10), ct.DATA_TERMINE,  101) AS date_end
                    FROM CONTRATTITESTATA ct
                    WHERE UPPER(ct.COD_CONTRATTO) LIKE %s
                       OR UPPER(ct.DESCRIZIONE)   LIKE %s
                       {id_clause}
                    ORDER BY ct.DATA_INIZIO DESC
                """, (term, term))
                return [dict(r) for r in cur.fetchall()]
        try:
            rows = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(rows)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/traffic/contract/{contract_id}/assignment")
    async def traffic_contract_assignment(
        contract_id: int,
        date_from: str = Query(""),
        date_to: str = Query(""),
    ):
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute("""
                    SELECT COD_CONTRATTO AS code, DESCRIZIONE AS description,
                           CONVERT(VARCHAR(10), DATA_INIZIO,  101) AS date_start,
                           CONVERT(VARCHAR(10), DATA_TERMINE, 101) AS date_end
                    FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = %d
                """ % contract_id)
                hdr = cur.fetchone()
                if not hdr:
                    return None

                cur.execute("""
                    SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                           cr.DESCRIZIONE       AS description,
                           cr.DURATA            AS duration_frames,
                           COUNT(tp.ID_TPALINSE) AS spot_count
                    FROM CONTRATTIRIGHE cr
                    JOIN trafficPalinse tpa ON tpa.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                    JOIN TPALINSE tp        ON tp.ID_TPALINSE = tpa.id_tpalinse
                    WHERE cr.ID_CONTRATTITESTATA = %d
                    GROUP BY cr.ID_CONTRATTIRIGHE, cr.DESCRIZIONE, cr.DURATA
                    ORDER BY cr.ID_CONTRATTIRIGHE
                """ % contract_id)
                lines = [dict(r) for r in cur.fetchall()]

                def _parse_date(s):
                    if not s:
                        return None
                    from datetime import datetime
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
                        try:
                            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                        except ValueError:
                            pass
                    return None

                df = _parse_date(date_from)
                dt = _parse_date(date_to)
                date_clause = ""
                if df and dt:
                    date_clause = f" AND tp.DATA BETWEEN '{df}' AND '{dt}'"
                elif df:
                    date_clause = f" AND tp.DATA >= '{df}'"
                elif dt:
                    date_clause = f" AND tp.DATA <= '{dt}'"

                cur.execute("""
                    SELECT f.ID_FILMATI AS filmati_id, f.COD_PROGRA AS code,
                           f.DESCRIZIO  AS title,      f.DURATA     AS durata,
                           COUNT(*)     AS count
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse       = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE  = tpa.id_contrattirighe
                    LEFT JOIN FILMATI f     ON f.ID_FILMATI = tp.ID_FILMATI
                    WHERE cr.ID_CONTRATTITESTATA = %d%s
                    GROUP BY f.ID_FILMATI, f.COD_PROGRA, f.DESCRIZIO, f.DURATA
                    ORDER BY COUNT(*) DESC
                """ % (contract_id, date_clause))
                rotation_rows = cur.fetchall()
                total = sum(r["count"] for r in rotation_rows) if rotation_rows else 0

                cur.execute("""
                    SELECT DISTINCT cr.COD_USER AS cod_user
                    FROM CONTRATTIRIGHE cr
                    JOIN trafficPalinse tpa ON tpa.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                    WHERE cr.ID_CONTRATTITESTATA = %d
                """ % contract_id)
                _code_by_id = {v: k for k, v in _MARKET_CODES.items()}
                markets = [
                    _code_by_id[r["cod_user"]]
                    for r in cur.fetchall()
                    if r["cod_user"] in _code_by_id
                ]

                cur.execute("""
                    SELECT DISTINCT cr.DURATA AS durata
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = %d AND cr.DURATA > 0
                    ORDER BY cr.DURATA
                """ % contract_id)
                durations = [
                    {"frames": r["durata"], "label": f":{round(r['durata'] / _FPS_GLOBAL)}"}
                    for r in cur.fetchall()
                ]

                return {
                    "header": dict(hdr),
                    "lines": lines,
                    "markets": markets,
                    "durations": durations,
                    "rotation": [
                        {**dict(r), "pct": round(r["count"] / total * 100, 1) if total else 0}
                        for r in rotation_rows
                    ],
                    "total_spots": total,
                }

        result = await asyncio.get_running_loop().run_in_executor(None, _run)
        if result is None:
            raise HTTPException(status_code=404, detail="Contract not found")
        return JSONResponse(result)

    @router.get("/api/traffic/spots/search")
    async def traffic_spots_search(
        q: str = Query(""),
        duration: Optional[int] = Query(None),
        prefix: str = Query(""),
    ):
        if not prefix and len(q) < 2:
            return JSONResponse([])

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                if prefix:
                    # Prefix mode: filter by code prefix, contains-search on title or code suffix
                    pfx_term  = f"{prefix.upper()}%"
                    contains  = f"%{q.upper()}%" if q else "%"
                    dur_clause = (
                        f" AND DURATA BETWEEN {duration - 5} AND {duration + 5}"
                        if duration is not None
                        else " AND DURATA <= 1800"
                    )
                    cur.execute(f"""
                        SELECT ID_FILMATI AS id, COD_PROGRA AS code,
                               DESCRIZIO AS title, DURATA AS durata
                        FROM FILMATI
                        WHERE UPPER(COD_PROGRA) LIKE %s
                          AND (UPPER(DESCRIZIO) LIKE %s OR UPPER(COD_PROGRA) LIKE %s)
                          AND TIPO = 'T'
                          {dur_clause}
                        ORDER BY DESCRIZIO
                    """, (pfx_term, contains, contains))
                else:
                    term = f"{q.upper()}%"
                    if duration is not None:
                        cur.execute("""
                            SELECT ID_FILMATI AS id, COD_PROGRA AS code,
                                   DESCRIZIO AS title, DURATA AS durata
                            FROM FILMATI
                            WHERE UPPER(DESCRIZIO) LIKE %s
                              AND DURATA BETWEEN %s AND %s
                              AND TIPO = 'T'
                            ORDER BY DESCRIZIO
                        """, (term, duration - 5, duration + 5))
                    else:
                        cur.execute("""
                            SELECT ID_FILMATI AS id, COD_PROGRA AS code,
                                   DESCRIZIO AS title, DURATA AS durata
                            FROM FILMATI
                            WHERE UPPER(DESCRIZIO) LIKE %s
                              AND TIPO = 'T'
                              AND DURATA <= 1800
                            ORDER BY DESCRIZIO
                        """, (term,))
                rows = cur.fetchall()
            for r in rows:
                r["duration_sec"] = round(r["durata"] / 30) if r["durata"] else 0
            return rows

        try:
            rows = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(rows)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/contract/{contract_id}/assign")
    async def traffic_contract_assign(contract_id: int, body: dict = Body(...)):
        spots = body.get("spots", [])
        filters = body.get("filters", {})
        spots = [s for s in spots if s.get("weight", 0) > 0]
        if not spots:
            raise HTTPException(status_code=400, detail="No spots provided")
        total_weight = sum(s["weight"] for s in spots)
        if total_weight <= 0:
            raise HTTPException(status_code=400, detail="Weights must be > 0")

        def _run():
            from collections import defaultdict

            from browser_automation.etere_direct_client import (
                ETERE_WEB_URL,
            )
            from browser_automation.etere_direct_client import (
                connect as _db_connect,
            )

            filmati_ids = [s["id"] for s in spots]
            filter_sql = _build_spot_filter(filters)

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(f"""
                    SELECT tpa.id_contrattirighe AS line_id,
                           tp.ID_TPALINSE        AS tp_id,
                           cr.ID_BOOKINGCODE     AS booking_code
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse       = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE  = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                    {filter_sql}
                    ORDER BY tp.DATA, tp.ORA
                """)
                all_rows = cur.fetchall()

                placeholders = ",".join(str(i) for i in filmati_ids)
                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI WHERE ID_FILMATI IN ({placeholders})"
                )
                filmati_rows = cur.fetchall()
                filmati_cod_map = {r["ID_FILMATI"]: (r["COD_PROGRA"] or "") for r in filmati_rows}
                filmati_title_map = {r["ID_FILMATI"]: (r["DESCRIZIO"] or "") for r in filmati_rows}

                # Build SUPPORTO string per filmati from FS_FILMATI + FS_METADEVICE.
                # Format: LEGACY_BASESUPP + FILE_ID  (e.g. "0ETX      TOY30M1206")
                cur.execute(f"""
                    SELECT ff.ID_FILMATI, ff.FILE_ID, ff.VIDEOSTANDARD, ff.DUR,
                           ISNULL(d.LEGACY_BASESUPP,
                                  CAST(d.LEGACY_MEDIAID AS VARCHAR) + 'ETX      ') AS supporto_prefix
                    FROM FS_FILMATI ff
                    JOIN FS_METADEVICE d ON d.ID_METADEVICE = ff.ID_METADEVICE
                    WHERE ff.ID_FILMATI IN ({placeholders})
                      AND d.LEGACY_MEDIAID IS NOT NULL
                """)
                # VIDEOSTANDARD "D" (digital HD) and null both map to ASPECT "H"
                _VS_TO_ASPECT = {"D": "H"}
                filmati_supporto_map: dict = {}
                filmati_aspect_map: dict = {}
                filmati_duration_map: dict = {}
                for r in cur.fetchall():
                    fid = r["ID_FILMATI"]
                    if fid not in filmati_supporto_map:
                        filmati_supporto_map[fid] = (r["supporto_prefix"] or "") + (r["FILE_ID"] or "")
                        filmati_aspect_map[fid] = _VS_TO_ASPECT.get(r["VIDEOSTANDARD"], "H")
                        filmati_duration_map[fid] = r["DUR"] or 0

                cur.execute("SELECT id_bookingcode, code FROM trf_bookingcode")
                bookingcode_to_newtype = {r["id_bookingcode"]: r["code"] for r in cur.fetchall()}

            if not all_rows:
                raise ValueError("No scheduled spots found matching the selected filters")

            line_tp_map = defaultdict(list)
            tp_newtype_map: dict = {}
            for r in all_rows:
                line_tp_map[r["line_id"]].append(r["tp_id"])
                tp_newtype_map[r["tp_id"]] = bookingcode_to_newtype.get(r["booking_code"], "COM")

            # Calculate rotation % per filmati (sum = 100, last spot absorbs rounding remainder)
            perc_map: dict = {}
            remaining = 100
            for i, s in enumerate(spots):
                if i == len(spots) - 1:
                    perc_map[s["id"]] = remaining
                else:
                    p = round(s["weight"] / total_weight * 100)
                    perc_map[s["id"]] = p
                    remaining -= p

            total_spots = len(all_rows)

            # Build chronologically-interleaved rotation list (Bresenham-style)
            # e.g. 50/50 → A,B,A,B,...  70/30 → A,A,A,B,A,A,A,B,...
            rotation_list = []
            accum = {s["id"]: 0.0 for s in spots}
            for _ in range(total_spots):
                for s in spots:
                    accum[s["id"]] += s["weight"] / total_weight
                chosen = max(accum, key=accum.__getitem__)
                rotation_list.append(chosen)
                accum[chosen] -= 1.0

            # Map each tp_id to its rotation filmati using global chronological order
            # so the A,B,A,B interleave is correct across the whole contract, not per-line
            tp_filmati_map = {row["tp_id"]: rotation_list[i] for i, row in enumerate(all_rows)}

            # Check which filmati are already in the pool so we don't create duplicates
            with _db_connect() as conn_pool:
                cur_pool = conn_pool.cursor(as_dict=True)
                cur_pool.execute(
                    f"SELECT DISTINCT ID_FILMATI FROM CONTRATTIFILMATI"
                    f" WHERE ID_CONTRATTIRIGHE IN ("
                    f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                    f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                    f" )"
                )
                existing_pool = {r["ID_FILMATI"] for r in cur_pool.fetchall()}

            session = _get_etere_session()
            lines_updated = spots_updated = 0
            tp_assignments = []  # (tp_id, filmati_id) pairs
            try:
                # Only add filmati not already in the pool — MaterialAddToAssetListC is not idempotent
                for fid in filmati_ids:
                    if fid in existing_pool:
                        continue
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAddToAssetListC",
                        json={"idFilmatiList": [fid], "idct": contract_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    add_result = r.json()
                    if not add_result.get("IsOk"):
                        raise ValueError(f"MaterialAddToAssetListC failed for {fid}: {add_result}")

                for line_id, tp_ids in line_tp_map.items():
                    n = len(tp_ids)
                    slice_ = [tp_filmati_map[tp_id] for tp_id in tp_ids]
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAssignAssetRotation",
                        json={
                            "idp": list(tp_ids),
                            "idf": list(slice_),
                            "idcr": line_id,
                        },
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    assign_result = r.json()
                    if not assign_result.get("IsOk"):
                        raise ValueError(f"MaterialAssignAssetRotation line {line_id} failed: {assign_result}")
                    lines_updated += 1
                    spots_updated += n
                    tp_assignments.extend(zip(tp_ids, slice_))

            except Exception:
                # Invalidate cache so next call gets a fresh login
                _invalidate_etere_session()
                raise

            # Sync COD_PROGRA and PERCROTATION so the native app reflects the rotation
            if tp_assignments:
                with _db_connect() as conn:
                    cur = conn.cursor()
                    for tp_id, filmati_id in tp_assignments:
                        cod = filmati_cod_map.get(filmati_id, "")
                        title = filmati_title_map.get(filmati_id, "")
                        newtype = tp_newtype_map.get(tp_id, "COM")
                        supporto = filmati_supporto_map.get(filmati_id, "")
                        aspect = filmati_aspect_map.get(filmati_id, "H")
                        duration_p = filmati_duration_map.get(filmati_id, 0)
                        cur.execute(
                            "UPDATE TPALINSE SET COD_PROGRA = %s, TITLE = %s, ID_FILMATI = %d,"
                            " NEWTYPE = %s, SUPPORTO = %s, ASPECT = %s, DURATION_P = %d"
                            " WHERE ID_TPALINSE = %d",
                            (cod, title, filmati_id, newtype, supporto, aspect, duration_p, tp_id),
                        )
                    # Ensure every assigned line has pool rows in CONTRATTIFILMATI.
                    # PERCROTATION is left 0 — actual rotation is driven by TPALINSE.
                    for line_id in line_tp_map.keys():
                        for filmati_id in perc_map:
                            cur.execute(
                                "DELETE FROM CONTRATTIFILMATI"
                                " WHERE ID_CONTRATTIRIGHE = %d AND ID_FILMATI = %d",
                                (line_id, filmati_id),
                            )
                            cur.execute(
                                "INSERT INTO CONTRATTIFILMATI"
                                " (ID_CONTRATTIRIGHE, ID_FILMATI, PERCROTATION)"
                                " VALUES (%d, %d, 0)",
                                (line_id, filmati_id),
                            )
                    # Remove pool rows for non-assigned lines (MaterialAddToAssetListC
                    # adds to every line; clean up unused ones — but EXCLUDE assigned
                    # lines or we delete the rows we just inserted above).
                    if filmati_ids and line_tp_map:
                        fid_str      = ",".join(str(f) for f in filmati_ids)
                        assigned_str = ",".join(str(lid) for lid in line_tp_map.keys())
                        cur.execute(
                            f"DELETE FROM CONTRATTIFILMATI"
                            f" WHERE ID_FILMATI IN ({fid_str})"
                            f" AND PERCROTATION = 0"
                            f" AND ID_CONTRATTIRIGHE NOT IN ({assigned_str})"
                            f" AND ID_CONTRATTIRIGHE IN ("
                            f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                            f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                            f" )"
                        )
                    conn.commit()

            # Check if this contract's customer requires airchecks
            needs_airchecks = False
            try:
                import sqlite3 as _sqlite3
                with _db_connect() as _conn:
                    _cur = _conn.cursor(as_dict=True)
                    _cur.execute(
                        "SELECT COMMITTENTE FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = %d" % contract_id
                    )
                    _hdr = _cur.fetchone()
                if _hdr and _hdr.get("COMMITTENTE"):
                    _cid = str(int(_hdr["COMMITTENTE"]))
                    _db_path = config.customer_db_path
                    if _db_path.exists():
                        with _sqlite3.connect(str(_db_path)) as _sdb:
                            _row = _sdb.execute(
                                "SELECT auto_aircheck FROM customers WHERE customer_id = ?", (_cid,)
                            ).fetchone()
                            needs_airchecks = bool(_row and _row[0])
            except Exception:
                pass

            return {"ok": True, "spots_updated": spots_updated, "lines_updated": lines_updated, "needs_airchecks": needs_airchecks, "contract_id": contract_id}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/contract/{contract_id}/preview")
    async def traffic_contract_preview(contract_id: int, body: dict = Body({})):
        filters = body.get("filters", {})

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            filter_sql = _build_spot_filter(filters)
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(f"""
                    SELECT COUNT(*) AS n
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                    {filter_sql}
                """)
                return {"count": cur.fetchone()["n"]}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/contract/{contract_id}/clear")
    async def traffic_contract_clear(contract_id: int, body: dict = Body({})):
        filters = body.get("filters", {})

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            filter_sql = _build_spot_filter(filters)
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(f"""
                    UPDATE tp
                    SET tp.COD_PROGRA = '', tp.TITLE = cr.DESCRIZIONE, tp.ID_FILMATI = 0
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                    {filter_sql}
                """)
                cleared = cur.rowcount
                conn.commit()
            return {"ok": True, "cleared": cleared}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Auto-assign by client (Sky River Casino, etc.) ────────────────────────

    @router.get("/api/traffic/contract/{contract_id}/auto-scan")
    async def traffic_auto_scan(contract_id: int, client: str = Query("sky_river")):
        """
        Scan contract lines for [SPOT_CODE] in descriptions and find matching
        FILMATI variants.  Returns per-line candidates with preferred variant
        pre-flagged based on the client's language preference rules.
        """
        LANG_PREFS: dict = {
            "filipino":   ["T", "E"],
            "vietnamese": ["V", "E"],
            "hmong":      ["H", "E"],
        }

        def _detect_lang(desc: str) -> str:
            d = desc.lower()
            for lang in ("filipino", "vietnamese", "hmong"):
                if lang in d:
                    return lang
            return ""

        def _suffix(cod: str) -> str:
            parts = cod.rsplit("-", 1)
            return parts[-1].upper() if len(parts) > 1 else ""

        def _run():
            import re

            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(f"""
                    SELECT ID_CONTRATTIRIGHE, DESCRIZIONE
                    FROM CONTRATTIRIGHE
                    WHERE ID_CONTRATTITESTATA = {contract_id}
                    ORDER BY ID_CONTRATTIRIGHE
                """)
                lines = cur.fetchall()

                result = []
                for line in lines:
                    desc = line["DESCRIZIONE"] or ""
                    codes = re.findall(r'\[([A-Za-z0-9]+)\]', desc)
                    if not codes:
                        result.append({"line_id": line["ID_CONTRATTIRIGHE"],
                                       "description": desc, "code": None, "variants": []})
                        continue

                    lang = _detect_lang(desc)
                    prefs = LANG_PREFS.get(lang, ["E"])
                    all_variants = []
                    seen_fids: set = set()
                    for code in codes:
                        safe_code = code.replace("'", "''")
                        cur.execute(f"""
                            SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO
                            FROM FILMATI
                            WHERE COD_PROGRA LIKE '{safe_code}%'
                              AND ANNULLATO = 0 AND SOSPESO = 0
                            ORDER BY COD_PROGRA
                        """)
                        code_has_preferred = False
                        for f in cur.fetchall():
                            fid = f["ID_FILMATI"]
                            if fid in seen_fids:
                                continue
                            seen_fids.add(fid)
                            is_pref = (_suffix(f["COD_PROGRA"]) == prefs[0]) if prefs else False
                            if is_pref:
                                code_has_preferred = True
                            all_variants.append({
                                "filmati_id": fid,
                                "cod_progra":  f["COD_PROGRA"],
                                "descrizio":   f["DESCRIZIO"] or "",
                                "preferred":   is_pref,
                            })
                        # If no variant matched the language preference, pre-check the first one
                        # for this code so rotation always has something selected per code.
                        if not code_has_preferred and all_variants:
                            # find the last-appended group for this code and mark first as preferred
                            for v in reversed(all_variants):
                                if v["cod_progra"].startswith(code):
                                    v["preferred"] = True
                                    break

                    result.append({"line_id": line["ID_CONTRATTIRIGHE"],
                                   "description": desc, "code": codes[0], "variants": all_variants})

            return result

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/contract/{contract_id}/auto-assign")
    async def traffic_auto_assign(contract_id: int, body: dict = Body(...)):
        """
        Apply confirmed auto-assign selections.
        Body: {assignments: [{line_id, filmati_ids: [...]}, ...]}
        Each line's spots get a round-robin of its filmati_ids.
        Uses the same MaterialAddToAssetListC + MaterialAssignAssetRotation
        HTTP flow as the manual rotation builder, then syncs TPALINSE directly.
        """
        assignments = body.get("assignments", [])
        date_from   = body.get("date_from")   # optional 'YYYY-MM-DD' — restrict to instruction period
        date_to     = body.get("date_to")
        if not assignments:
            raise HTTPException(status_code=400, detail="No assignments provided")

        def _run():
            from collections import defaultdict

            from browser_automation.etere_direct_client import (
                ETERE_WEB_URL,
            )
            from browser_automation.etere_direct_client import (
                connect as _db_connect,
            )

            # Support both legacy filmati_ids and new spots format
            all_filmati_ids = list({
                s["filmati_id"] if isinstance(s, dict) else s
                for a in assignments
                for s in (a.get("spots") or [{"filmati_id": f} for f in a.get("filmati_ids", [])])
            })
            if not all_filmati_ids:
                raise ValueError("No filmati selected")

            placeholders = ",".join(str(f) for f in all_filmati_ids)
            line_ids_str  = ",".join(str(a["line_id"]) for a in assignments)

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE ID_FILMATI IN ({placeholders})"
                )
                frows = cur.fetchall()
                filmati_cod_map   = {r["ID_FILMATI"]: (r["COD_PROGRA"] or "") for r in frows}
                filmati_title_map = {r["ID_FILMATI"]: (r["DESCRIZIO"] or "") for r in frows}

                cur.execute(f"""
                    SELECT ff.ID_FILMATI, ff.FILE_ID, ff.VIDEOSTANDARD, ff.DUR,
                           ISNULL(d.LEGACY_BASESUPP,
                                  CAST(d.LEGACY_MEDIAID AS VARCHAR) + 'ETX      ') AS supporto_prefix
                    FROM FS_FILMATI ff
                    JOIN FS_METADEVICE d ON d.ID_METADEVICE = ff.ID_METADEVICE
                    WHERE ff.ID_FILMATI IN ({placeholders})
                      AND d.LEGACY_MEDIAID IS NOT NULL
                """)
                _VS_TO_ASPECT = {"D": "H"}
                filmati_supporto_map: dict = {}
                filmati_aspect_map:   dict = {}
                filmati_duration_map: dict = {}
                for r in cur.fetchall():
                    fid = r["ID_FILMATI"]
                    if fid not in filmati_supporto_map:
                        filmati_supporto_map[fid] = (r["supporto_prefix"] or "") + (r["FILE_ID"] or "")
                        filmati_aspect_map[fid]   = _VS_TO_ASPECT.get(r["VIDEOSTANDARD"], "H")
                        filmati_duration_map[fid] = r["DUR"] or 0

                cur.execute("SELECT id_bookingcode, code FROM trf_bookingcode")
                bookingcode_to_newtype = {r["id_bookingcode"]: r["code"] for r in cur.fetchall()}

                date_clause = " AND tp.DATA >= CAST(GETDATE() AS DATE)"
                if date_from:
                    date_clause += f" AND tp.DATA >= '{date_from}'"
                if date_to:
                    date_clause += f" AND tp.DATA <= '{date_to}'"
                cur.execute(f"""
                    SELECT tpa.id_contrattirighe AS line_id,
                           tp.ID_TPALINSE        AS tp_id,
                           cr.ID_BOOKINGCODE     AS booking_code
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE tpa.id_contrattirighe IN ({line_ids_str})
                    {date_clause}
                    ORDER BY tp.DATA, tp.ORA
                """)
                line_tp_map: dict = defaultdict(list)
                tp_newtype_map: dict = {}
                all_rows_ordered: list = []  # global chronological order preserved
                for r in cur.fetchall():
                    line_tp_map[r["line_id"]].append(r["tp_id"])
                    tp_newtype_map[r["tp_id"]] = bookingcode_to_newtype.get(r["booking_code"], "COM")
                    all_rows_ordered.append((r["line_id"], r["tp_id"]))

                # Check the entire contract's pool (not just this batch's lines) so
                # that when the Lexus JS loop calls us once per duration group on the
                # same contract, the second call sees filmati already added by the first.
                cur.execute(
                    f"SELECT DISTINCT ID_FILMATI FROM CONTRATTIFILMATI"
                    f" WHERE ID_CONTRATTIRIGHE IN ("
                    f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                    f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                    f" )"
                )
                existing_pool = {r["ID_FILMATI"] for r in cur.fetchall()}

            session = _get_etere_session()
            lines_updated = spots_updated = 0
            tp_assignments: list = []

            try:
                for fid in all_filmati_ids:
                    if fid in existing_pool:
                        continue
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAddToAssetListC",
                        json={"idFilmatiList": [fid], "idct": contract_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(f"MaterialAddToAssetListC failed for filmati {fid}: {r.json()}")

                # Normalize: convert legacy filmati_ids list → weighted spots format
                for a in assignments:
                    if "filmati_ids" in a and "spots" not in a:
                        a["spots"] = [{"filmati_id": fid, "weight": 1} for fid in a["filmati_ids"]]

                def _bresenham(spots_list: list, count: int) -> list:
                    """Return `count` filmati_ids in Bresenham-weighted order."""
                    total_w = sum(s.get("weight", 1) for s in spots_list)
                    accum = {s["filmati_id"]: 0.0 for s in spots_list}
                    result = []
                    for _ in range(count):
                        for s in spots_list:
                            accum[s["filmati_id"]] += s.get("weight", 1) / total_w
                        chosen = max(accum, key=accum.__getitem__)
                        result.append(chosen)
                        accum[chosen] -= 1.0
                    return result

                # Use global chronological interleave when all lines share identical spots+weights
                # (common for Daviselen/Lexus). Per-line Bresenham otherwise.
                asgn_spots_key = [
                    tuple((s["filmati_id"], s.get("weight", 1)) for s in a.get("spots", []))
                    for a in assignments
                ]
                common_spots = assignments[0].get("spots", []) if assignments else []
                use_global = bool(
                    common_spots
                    and len(set(asgn_spots_key)) == 1
                    and len(all_rows_ordered) > 0
                )
                if use_global:
                    global_rotation = _bresenham(common_spots, len(all_rows_ordered))
                    tp_filmati_global = {
                        tp_id: global_rotation[i]
                        for i, (_, tp_id) in enumerate(all_rows_ordered)
                    }

                asgn_map = {a["line_id"]: a for a in assignments}
                for line_id, tp_ids in line_tp_map.items():
                    asgn = asgn_map.get(line_id)
                    if not asgn:
                        continue
                    spots = asgn.get("spots", [])
                    if not spots or not tp_ids:
                        continue
                    if use_global:
                        idf = [tp_filmati_global[tp_id] for tp_id in tp_ids]
                    else:
                        idf = _bresenham(spots, len(tp_ids))
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAssignAssetRotation",
                        json={"idp": list(tp_ids), "idf": idf, "idcr": line_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(f"MaterialAssignAssetRotation failed for line {line_id}: {r.json()}")
                    lines_updated += 1
                    spots_updated += len(tp_ids)
                    tp_assignments.extend(zip(tp_ids, idf))

            except Exception:
                _invalidate_etere_session()
                raise

            if tp_assignments:
                with _db_connect() as conn:
                    cur = conn.cursor()
                    for tp_id, filmati_id in tp_assignments:
                        cur.execute(
                            "UPDATE TPALINSE SET COD_PROGRA = %s, TITLE = %s, ID_FILMATI = %d,"
                            " NEWTYPE = %s, SUPPORTO = %s, ASPECT = %s, DURATION_P = %d"
                            " WHERE ID_TPALINSE = %d",
                            (
                                filmati_cod_map.get(filmati_id, ""),
                                filmati_title_map.get(filmati_id, ""),
                                filmati_id,
                                tp_newtype_map.get(tp_id, "COM"),
                                filmati_supporto_map.get(filmati_id, ""),
                                filmati_aspect_map.get(filmati_id, "H"),
                                filmati_duration_map.get(filmati_id, 0),
                                tp_id,
                            ),
                        )
                    # Ensure every assigned line has pool rows in CONTRATTIFILMATI.
                    # PERCROTATION is left 0 — the pool list just needs to exist;
                    # actual rotation is driven by TPALINSE.ID_FILMATI.
                    for line_id in line_tp_map.keys():
                        asgn = asgn_map.get(line_id)
                        if not asgn:
                            continue
                        filmati_ids_line = asgn.get("filmati_ids", [])
                        if not filmati_ids_line:
                            continue
                        for fid in filmati_ids_line:
                            cur.execute(
                                "DELETE FROM CONTRATTIFILMATI"
                                " WHERE ID_CONTRATTIRIGHE = %d AND ID_FILMATI = %d",
                                (line_id, fid),
                            )
                            cur.execute(
                                "INSERT INTO CONTRATTIFILMATI"
                                " (ID_CONTRATTIRIGHE, ID_FILMATI, PERCROTATION)"
                                " VALUES (%d, %d, 0)",
                                (line_id, fid),
                            )
                    conn.commit()

            needs_airchecks = False
            try:
                import sqlite3 as _sqlite3
                with _db_connect() as _conn:
                    _cur = _conn.cursor(as_dict=True)
                    _cur.execute(
                        "SELECT COMMITTENTE FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = %d" % contract_id
                    )
                    _hdr = _cur.fetchone()
                if _hdr and _hdr.get("COMMITTENTE"):
                    _cid = str(int(_hdr["COMMITTENTE"]))
                    _db_path = config.customer_db_path
                    if _db_path.exists():
                        with _sqlite3.connect(str(_db_path)) as _sdb:
                            _row = _sdb.execute(
                                "SELECT auto_aircheck FROM customers WHERE customer_id = ?", (_cid,)
                            ).fetchone()
                            needs_airchecks = bool(_row and _row[0])
            except Exception:
                pass

            return {"ok": True, "lines_updated": lines_updated, "spots_updated": spots_updated, "needs_airchecks": needs_airchecks, "contract_id": contract_id}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Bookend Pair Assignment ───────────────────────────────────────────────

    @router.get("/traffic/bookend-pairs", response_class=HTMLResponse)
    async def bookend_pairs_page(request: Request):
        return templates.TemplateResponse(request, "traffic/bookend_pairs.html")

    @router.get("/api/traffic/bookend-pairs/{contract_id}")
    async def bookend_pairs_load(
        contract_id: int,
        date_from: str = Query(""),
        date_to: str = Query(""),
    ):
        import re as _re2

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    "SELECT COD_CONTRATTO AS code, DESCRIZIONE AS description,"
                    " CONVERT(VARCHAR(10), DATA_INIZIO, 101) AS date_start,"
                    " CONVERT(VARCHAR(10), DATA_TERMINE, 101) AS date_end"
                    f" FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = {contract_id}"
                )
                hdr = cur.fetchone()
                if not hdr:
                    return None

                dc = ""
                if date_from and _re2.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
                    dc += f" AND tp.DATA >= '{date_from}'"
                if date_to and _re2.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
                    dc += f" AND tp.DATA <= '{date_to}'"

                cur.execute(f"""
                    SELECT COUNT(*) AS cnt
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id} {dc}
                """)
                cnt = cur.fetchone()["cnt"]

                return {
                    "header": dict(hdr),
                    "total_spots": cnt,
                    "break_count": cnt // 2,
                    "odd_warning": cnt % 2 != 0,
                }

        result = await asyncio.get_running_loop().run_in_executor(None, _run)
        if result is None:
            raise HTTPException(status_code=404, detail="Contract not found")
        return JSONResponse(result)

    @router.post("/api/traffic/bookend-pairs/{contract_id}/assign")
    async def bookend_pairs_assign(contract_id: int, body: dict = Body(...)):
        date_from = body.get("date_from", "")
        date_to   = body.get("date_to",   "")
        pairs     = body.get("pairs", [])

        if not pairs:
            raise HTTPException(status_code=400, detail="No pairs provided")
        total_pct = sum(p.get("pct", 0) for p in pairs)
        if total_pct <= 0:
            raise HTTPException(status_code=400, detail="Pair percentages must sum to > 0")

        def _run():
            import re as _re2
            from collections import defaultdict

            from browser_automation.etere_direct_client import (
                ETERE_WEB_URL,
            )
            from browser_automation.etere_direct_client import (
                connect as _db_connect,
            )

            weights = [p["pct"] / total_pct for p in pairs]

            dc = " AND tp.DATA >= CAST(GETDATE() AS DATE)"
            if date_from and _re2.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
                dc += f" AND tp.DATA >= '{date_from}'"
            if date_to and _re2.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
                dc += f" AND tp.DATA <= '{date_to}'"

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(f"""
                    SELECT tpa.id_contrattirighe AS line_id,
                           tp.ID_TPALINSE        AS tp_id,
                           tp.DATA               AS data,
                           tp.ORA                AS ora,
                           cr.ID_BOOKINGCODE     AS booking_code
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id} {dc}
                    ORDER BY tp.DATA, tp.ORA
                """)
                all_rows = cur.fetchall()
                cur.execute("SELECT id_bookingcode, code FROM trf_bookingcode")
                bookingcode_to_newtype = {r["id_bookingcode"]: r["code"] for r in cur.fetchall()}

            if not all_rows:
                raise ValueError("No scheduled spots found in the selected date range")

            # Group into break pairs by ORA proximity.
            # Top-of-break (A) has lower ORA, bottom-of-break (B) has higher ORA.
            # Threshold: 18000 frames ≈ 600 sec at 30fps (safely spans any break size)
            BREAK_THRESHOLD = 18000
            breaks: list[tuple] = []
            current: list = [all_rows[0]]
            for row in all_rows[1:]:
                prev = current[-1]
                same_date   = str(row["data"])[:10] == str(prev["data"])[:10]
                close_enough = (row["ora"] - prev["ora"]) <= BREAK_THRESHOLD
                if same_date and close_enough:
                    current.append(row)
                else:
                    breaks.append(tuple(current))
                    current = [row]
            breaks.append(tuple(current))

            bad = [(i + 1, b[0]["data"], len(b)) for i, b in enumerate(breaks) if len(b) != 2]
            if bad:
                detail = "; ".join(f"break {n} on {d}: {c} spot(s)" for n, d, c in bad[:5])
                raise ValueError(
                    f"Bookend breaks must have exactly 2 spots each. Problem breaks: {detail}"
                )

            n_breaks = len(breaks)

            # Bresenham distribution: assign a pair index to each break
            accum = [0.0] * len(pairs)
            rotation_list = []
            for _ in range(n_breaks):
                for i, w in enumerate(weights):
                    accum[i] += w
                chosen = max(range(len(accum)), key=lambda i: accum[i])
                rotation_list.append(chosen)
                accum[chosen] -= 1.0

            all_filmati_ids = list({p["a_id"] for p in pairs} | {p["b_id"] for p in pairs})
            fid_str = ",".join(str(f) for f in all_filmati_ids)

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI WHERE ID_FILMATI IN ({fid_str})"
                )
                filmati_cod_map   = {r["ID_FILMATI"]: (r["COD_PROGRA"] or "") for r in cur.fetchall()}
                cur.execute(
                    f"SELECT ID_FILMATI, DESCRIZIO FROM FILMATI WHERE ID_FILMATI IN ({fid_str})"
                )
                filmati_title_map = {r["ID_FILMATI"]: (r["DESCRIZIO"] or "") for r in cur.fetchall()}

                cur.execute(f"""
                    SELECT ff.ID_FILMATI, ff.FILE_ID, ff.VIDEOSTANDARD, ff.DUR,
                           ISNULL(d.LEGACY_BASESUPP,
                                  CAST(d.LEGACY_MEDIAID AS VARCHAR) + 'ETX      ') AS supporto_prefix
                    FROM FS_FILMATI ff
                    JOIN FS_METADEVICE d ON d.ID_METADEVICE = ff.ID_METADEVICE
                    WHERE ff.ID_FILMATI IN ({fid_str}) AND d.LEGACY_MEDIAID IS NOT NULL
                """)
                _VS_ASPECT = {"D": "H"}
                filmati_supporto_map: dict = {}
                filmati_aspect_map:   dict = {}
                filmati_duration_map: dict = {}
                for r in cur.fetchall():
                    fid = r["ID_FILMATI"]
                    if fid not in filmati_supporto_map:
                        filmati_supporto_map[fid] = (r["supporto_prefix"] or "") + (r["FILE_ID"] or "")
                        filmati_aspect_map[fid]   = _VS_ASPECT.get(r["VIDEOSTANDARD"], "H")
                        filmati_duration_map[fid] = r["DUR"] or 0

                all_line_ids = list({r["line_id"] for r in all_rows})
                line_ids_str = ",".join(str(lid) for lid in all_line_ids)
                cur.execute(
                    f"SELECT DISTINCT ID_FILMATI FROM CONTRATTIFILMATI"
                    f" WHERE ID_CONTRATTIRIGHE IN ("
                    f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                    f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                    f" )"
                )
                existing_pool = {r["ID_FILMATI"] for r in cur.fetchall()}

            tp_filmati_map: dict = {}
            tp_newtype_map: dict = {}
            line_tp_map: dict    = defaultdict(list)
            for k, brk in enumerate(breaks):
                top_row, bot_row = brk[0], brk[1]
                pair_idx  = rotation_list[k]
                a_filmati = pairs[pair_idx]["a_id"]
                b_filmati = pairs[pair_idx]["b_id"]
                tp_filmati_map[top_row["tp_id"]] = a_filmati
                tp_filmati_map[bot_row["tp_id"]] = b_filmati
                tp_newtype_map[top_row["tp_id"]] = bookingcode_to_newtype.get(top_row["booking_code"], "COM")
                tp_newtype_map[bot_row["tp_id"]] = bookingcode_to_newtype.get(bot_row["booking_code"], "COM")
                line_tp_map[top_row["line_id"]].append(top_row["tp_id"])
                line_tp_map[bot_row["line_id"]].append(bot_row["tp_id"])

            session = _get_etere_session()
            lines_updated = spots_updated = 0
            tp_assignments: list = []
            try:
                for fid in all_filmati_ids:
                    if fid in existing_pool:
                        continue
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAddToAssetListC",
                        json={"idFilmatiList": [fid], "idct": contract_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(f"MaterialAddToAssetListC failed for filmati {fid}")

                for line_id, tp_ids in line_tp_map.items():
                    filmati_for_line = [tp_filmati_map[tp] for tp in tp_ids]
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAssignAssetRotation",
                        json={"idp": tp_ids, "idf": filmati_for_line, "idcr": line_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(f"MaterialAssignAssetRotation failed for line {line_id}")
                    lines_updated += 1
                    spots_updated += len(tp_ids)
                    tp_assignments.extend(zip(tp_ids, filmati_for_line))
            except Exception:
                _invalidate_etere_session()
                raise

            if tp_assignments:
                filmati_count: dict = defaultdict(int)
                for _, fid in tp_assignments:
                    filmati_count[fid] += 1
                total_assigned = sum(filmati_count.values())

                with _db_connect() as conn:
                    cur = conn.cursor()
                    for tp_id, filmati_id in tp_assignments:
                        cur.execute(
                            "UPDATE TPALINSE SET COD_PROGRA = %s, TITLE = %s, ID_FILMATI = %d,"
                            " NEWTYPE = %s, SUPPORTO = %s, ASPECT = %s, DURATION_P = %d"
                            " WHERE ID_TPALINSE = %d",
                            (
                                filmati_cod_map.get(filmati_id, ""),
                                filmati_title_map.get(filmati_id, ""),
                                filmati_id,
                                tp_newtype_map.get(tp_id, "COM"),
                                filmati_supporto_map.get(filmati_id, ""),
                                filmati_aspect_map.get(filmati_id, "H"),
                                filmati_duration_map.get(filmati_id, 0),
                                tp_id,
                            ),
                        )
                    for line_id in all_line_ids:
                        for filmati_id, count in filmati_count.items():
                            perc = round(count / total_assigned * 100)
                            cur.execute(
                                "UPDATE CONTRATTIFILMATI SET PERCROTATION = %d"
                                " WHERE ID_CONTRATTIRIGHE = %d AND ID_FILMATI = %d",
                                (perc, line_id, filmati_id),
                            )
                    if all_filmati_ids:
                        cur.execute(
                            f"DELETE FROM CONTRATTIFILMATI"
                            f" WHERE ID_FILMATI IN ({fid_str})"
                            f" AND PERCROTATION = 0"
                            f" AND ID_CONTRATTIRIGHE IN ({line_ids_str})"
                        )
                    conn.commit()

            needs_airchecks = False
            try:
                import sqlite3 as _sqlite3
                with _db_connect() as _conn:
                    _cur = _conn.cursor(as_dict=True)
                    _cur.execute(
                        "SELECT COMMITTENTE FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = %d" % contract_id
                    )
                    _hdr = _cur.fetchone()
                if _hdr and _hdr.get("COMMITTENTE"):
                    _cid = str(int(_hdr["COMMITTENTE"]))
                    _db_path = config.customer_db_path
                    if _db_path.exists():
                        with _sqlite3.connect(str(_db_path)) as _sdb:
                            _row = _sdb.execute(
                                "SELECT auto_aircheck FROM customers WHERE customer_id = ?", (_cid,)
                            ).fetchone()
                            needs_airchecks = bool(_row and _row[0])
            except Exception:
                pass

            return {"ok": True, "breaks_assigned": n_breaks, "spots_updated": spots_updated, "needs_airchecks": needs_airchecks, "contract_id": contract_id}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Admerasia IO traffic ─────────────────────────────────────────────────

    @router.post("/api/traffic/admerasia/parse-io")
    async def admerasia_parse_io(file: UploadFile):
        """Parse ISCI key from an Admerasia IO PDF and resolve filmati IDs."""
        pdf_bytes = await file.read()

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.admerasia_traffic_parser import (
                parse_admerasia_io_iscis,
                parse_admerasia_io_lines,
            )

            raw = parse_admerasia_io_iscis(pdf_bytes)
            if not raw:
                raise ValueError("No ISCI codes found in PDF")

            io_lines = parse_admerasia_io_lines(pdf_bytes)

            isci_codes   = [r["isci_code"] for r in raw]
            placeholders = ",".join(f"'{c}'" for c in isci_codes)

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE COD_PROGRA IN ({placeholders})"
                )
                rows = cur.fetchall()

            filmati_map = {
                r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"], "db_title": r["DESCRIZIO"] or ""}
                for r in rows
            }

            isci_options: dict = {"15": [], "30": []}
            not_found: list = []
            for item in raw:
                code    = item["isci_code"]
                dur_key = str(item["duration_sec"])
                if dur_key not in isci_options:
                    isci_options[dur_key] = []
                if code in filmati_map:
                    isci_options[dur_key].append({
                        "title":      item["title"] or filmati_map[code]["db_title"],
                        "isci":       code,
                        "filmati_id": filmati_map[code]["filmati_id"],
                    })
                else:
                    not_found.append(code)

            return {"isci_options": isci_options, "not_found": not_found, "io_lines": io_lines}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── TCAA (Toyota) IO traffic ─────────────────────────────────────────────

    @router.post("/api/traffic/tcaa/parse-io")
    async def tcaa_parse_io(file: UploadFile, contract_id: int = Query(...)):
        """Parse a TCAA CABLE Traffic Instructions PDF for a pre-selected contract.

        TCAA's printed estimate does not match the Etere contract, so the operator
        selects the contract first; this endpoint reads the creatives + rotation %
        and resolves the contract's duration-matched lines to assign onto.
        Returns {is_tcaa: False} fast if the PDF is not a TCAA sheet."""
        pdf_bytes = await file.read()
        filename = file.filename or "io.pdf"

        def _run():
            import io as _io

            import pdfplumber

            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.tcaa_traffic_parser import (
                parse_tcaa_traffic_pdf,
            )

            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            except Exception:
                return {"is_tcaa": False, "filename": filename}
            if "TCAA" not in text.upper():
                return {"is_tcaa": False, "filename": filename}

            instr = parse_tcaa_traffic_pdf(pdf_bytes)
            if not instr.spots:
                raise ValueError("No creative rows found in TCAA PDF")

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                isci_codes   = [s.isci for s in instr.spots]
                placeholders = ",".join(f"'{c}'" for c in isci_codes)
                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE COD_PROGRA IN ({placeholders})"
                )
                filmati_map = {
                    r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                      "db_title":   r["DESCRIZIO"] or ""}
                    for r in cur.fetchall()
                }

                spots_out, not_found = [], []
                for s in instr.spots:
                    found = s.isci in filmati_map
                    spots_out.append({
                        "isci":         s.isci,
                        "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                        "duration_sec": s.duration_sec,
                        "rotation_pct": s.rotation_pct,
                        "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                        "found":        found,
                    })
                    if not found:
                        not_found.append(s.isci)

                # Contract's lines matching the sheet's duration(s), with scheduled
                # spot counts inside the campaign window. "RUN IN ALL PROGRAMMING" →
                # every duration-matched line (paid + bonus) receives the rotation.
                durations = sorted({s["duration_sec"] for s in spots_out})
                dur_clause = " OR ".join(
                    f"CAST(ROUND(CAST(cr.DURATA AS FLOAT) / {_FPS_GLOBAL}, 0) AS INT) = {d}"
                    for d in durations
                )
                date_filter = ""
                if instr.date_from_sql:
                    date_filter += f" AND tp.DATA >= '{instr.date_from_sql}'"
                if instr.date_to_sql:
                    date_filter += f" AND tp.DATA <= '{instr.date_to_sql}'"
                cur.execute(f"""
                    SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                           cr.DESCRIZIONE       AS description,
                           ISNULL(sc.spot_count, 0) AS spot_count
                    FROM CONTRATTIRIGHE cr
                    LEFT JOIN (
                        SELECT tpa.id_contrattirighe, COUNT(*) AS spot_count
                        FROM trafficPalinse tpa
                        JOIN TPALINSE tp ON tp.ID_TPALINSE = tpa.id_tpalinse
                        WHERE 1=1 {date_filter}
                        GROUP BY tpa.id_contrattirighe
                    ) sc ON sc.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                    WHERE cr.ID_CONTRATTITESTATA = {int(contract_id)}
                      AND ({dur_clause})
                    ORDER BY cr.ID_CONTRATTIRIGHE
                """)
                lines = [{"line_id": r["line_id"], "description": r["description"],
                          "spot_count": r["spot_count"]} for r in cur.fetchall()]

            return {
                "is_tcaa":       True,
                "filename":      filename,
                "estimate":      instr.estimate,
                "product":       instr.product,
                "station":       instr.station,
                "warning":       instr.warning,
                "duration_sec":  durations[0] if durations else None,
                "date_range":    f"{instr.start_date}–{instr.end_date}",
                "date_from_sql": instr.date_from_sql,
                "date_to_sql":   instr.date_to_sql,
                "spots":         spots_out,
                "not_found":     not_found,
                "lines":         lines,
            }

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/tcaa/auto-match")
    async def tcaa_auto_match(file: UploadFile):
        """Drop-first TCAA auto-assigner.

        Finds the contract itself by 'TCAA' + campaign-date overlap (the printed
        estimate never matches ours, so it isn't usable), then returns the matched
        contract for the operator to confirm plus the creative→line assignment
        preview. Returns {is_tcaa: False} fast if the PDF is not a TCAA sheet."""
        pdf_bytes = await file.read()
        filename = file.filename or "io.pdf"

        def _run():
            import io as _io

            import pdfplumber

            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.tcaa_traffic_parser import (
                parse_tcaa_traffic_pdf,
            )

            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            except Exception:
                return {"is_tcaa": False, "filename": filename}
            if "TCAA" not in text.upper():
                return {"is_tcaa": False, "filename": filename}

            instr = parse_tcaa_traffic_pdf(pdf_bytes)
            if not instr.spots:
                return {"is_tcaa": True, "filename": filename,
                        "error": "No creative rows found on the Crossings TV page"}

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                isci_codes   = [s.isci for s in instr.spots]
                placeholders = ",".join(f"'{c}'" for c in isci_codes)
                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE COD_PROGRA IN ({placeholders})"
                )
                filmati_map = {
                    r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                      "db_title":   r["DESCRIZIO"] or ""}
                    for r in cur.fetchall()
                }

                spots_out, not_found = [], []
                for s in instr.spots:
                    found = s.isci in filmati_map
                    spots_out.append({
                        "isci":         s.isci,
                        "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                        "duration_sec": s.duration_sec,
                        "rotation_pct": s.rotation_pct,
                        "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                        "found":        found,
                    })
                    if not found:
                        not_found.append(s.isci)

                # Find the contract by "TCAA" + campaign-date overlap. Verified to
                # resolve uniquely per monthly campaign; candidates are returned so
                # the operator can confirm (or spot an ambiguous multi-month case).
                date_filter = ""
                if instr.date_from_sql:
                    date_filter += f" AND cr.DATA_FINE   >= '{instr.date_from_sql}'"
                if instr.date_to_sql:
                    date_filter += f" AND cr.DATA_INIZIO <= '{instr.date_to_sql}'"
                cur.execute(f"""
                    SELECT TOP 10
                        ct.ID_CONTRATTITESTATA AS id,
                        ct.COD_CONTRATTO       AS code,
                        ct.DESCRIZIONE         AS description,
                        CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                        CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end
                    FROM CONTRATTITESTATA ct
                    JOIN CONTRATTIRIGHE cr
                      ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                    WHERE (UPPER(ct.COD_CONTRATTO) LIKE '%TCAA%'
                        OR UPPER(ct.DESCRIZIONE)   LIKE '%TCAA%')
                    {date_filter}
                    GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                    ORDER BY ct.ID_CONTRATTITESTATA DESC
                """)
                candidates = [dict(r) for r in cur.fetchall()]
                contract = candidates[0] if candidates else None

                lines = []
                if contract:
                    durations = sorted({s["duration_sec"] for s in spots_out})
                    dur_clause = " OR ".join(
                        f"CAST(ROUND(CAST(cr.DURATA AS FLOAT) / {_FPS_GLOBAL}, 0) AS INT) = {d}"
                        for d in durations
                    )
                    line_date_filter = ""
                    if instr.date_from_sql:
                        line_date_filter += f" AND tp.DATA >= '{instr.date_from_sql}'"
                    if instr.date_to_sql:
                        line_date_filter += f" AND tp.DATA <= '{instr.date_to_sql}'"
                    cur.execute(f"""
                        SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                               cr.DESCRIZIONE       AS description,
                               ISNULL(sc.spot_count, 0) AS spot_count
                        FROM CONTRATTIRIGHE cr
                        LEFT JOIN (
                            SELECT tpa.id_contrattirighe, COUNT(*) AS spot_count
                            FROM trafficPalinse tpa
                            JOIN TPALINSE tp ON tp.ID_TPALINSE = tpa.id_tpalinse
                            WHERE 1=1 {line_date_filter}
                            GROUP BY tpa.id_contrattirighe
                        ) sc ON sc.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                        WHERE cr.ID_CONTRATTITESTATA = {int(contract['id'])}
                          AND ({dur_clause})
                        ORDER BY cr.ID_CONTRATTIRIGHE
                    """)
                    lines = [{"line_id": r["line_id"], "description": r["description"],
                              "spot_count": r["spot_count"]} for r in cur.fetchall()]

            durations = sorted({s["duration_sec"] for s in spots_out})
            warning = instr.warning
            if len(candidates) > 1:
                warning = (warning + " " if warning else "") + (
                    f"{len(candidates)} TCAA contracts overlap these dates — "
                    f"using {contract['code']}; verify it's the right one."
                )
            if not contract:
                warning = (warning + " " if warning else "") + (
                    "No TCAA contract found overlapping the campaign dates — "
                    "select the contract manually via the Auto-Assign dropdown."
                )

            return {
                "is_tcaa":             True,
                "filename":            filename,
                "estimate":            instr.estimate,
                "product":             instr.product,
                "station":             instr.station,
                "warning":             warning,
                "duration_sec":        durations[0] if durations else None,
                "date_range":          f"{instr.start_date}–{instr.end_date}",
                "date_from_sql":       instr.date_from_sql,
                "date_to_sql":         instr.date_to_sql,
                "spots":               spots_out,
                "not_found":           not_found,
                "contract":            contract,
                "contract_candidates": candidates,
                "lines":               lines,
            }

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/admerasia/auto-color")
    async def admerasia_auto_color(file: UploadFile):
        """Drop-first Admerasia colour-match auto-assigner.

        Finds the entered contract by the IO order number (CUSTOMERREF), reads each
        grid cell's colour + the vision ISCI legend, and returns a per-spot creative
        assignment preview. The write goes through /assign-spots. Returns
        {is_admerasia: False} fast if the PDF isn't an Admerasia IO (no vision spent)."""
        pdf_bytes = await file.read()
        filename = file.filename or "io.pdf"

        def _run():
            import io as _io
            import os
            import tempfile
            from collections import Counter

            import pdfplumber

            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.admerasia_traffic import resolve_traffic

            # Cheap gate: every Admerasia IO carries the agency name. Avoids spending
            # a vision read on non-Admerasia drops routed here.
            try:
                with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                    text = pdf.pages[0].extract_text() or ""
            except Exception:
                return {"is_admerasia": False, "filename": filename}
            if "admerasia" not in text.lower():
                return {"is_admerasia": False, "filename": filename}

            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            try:
                tmp.write(pdf_bytes)
                tmp.close()
                with _db_connect() as conn:
                    res = resolve_traffic(tmp.name, conn.cursor(as_dict=True))
            finally:
                for p in (tmp.name, tmp.name + ".adm.json", tmp.name + ".adm-legend.json"):
                    try:
                        os.unlink(p)
                    except OSError:
                        pass

            ok_asg = [a for a in res.assignments if a.ok]
            per = Counter(a.isci for a in ok_asg)
            legend_map = {isci: {"duration": d, "rgb": list(rgb), "name": name}
                          for isci, d, rgb, name in res.legend}
            return {
                "is_admerasia":  True,
                "filename":      filename,
                "order_number":  res.order_number,
                "contract_id":   res.contract_id,
                "contract_code": res.contract_code,
                "ok":            res.ok,
                "warnings":      res.warnings,
                "total":         len(res.assignments),
                "assignable":    len(ok_asg),
                "assignments":   [{"tp_id": a.tp_id, "filmati_id": a.filmati_id, "isci": a.isci,
                                   "duration_ok": a.duration_ok, "ok": a.ok, "reason": a.reason}
                                  for a in res.assignments],
                "summary":       [{"isci": isci, "count": per[isci],
                                   "name": legend_map.get(isci, {}).get("name"),
                                   "rgb":  legend_map.get(isci, {}).get("rgb"),
                                   "duration": legend_map.get(isci, {}).get("duration")}
                                  for isci in sorted(per, key=lambda k: -per[k])],
            }

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/daviselen/parse-instructions")
    async def daviselen_parse_instructions(files: List[UploadFile] = File(...)):
        """Parse one or more Davis Elen traffic instruction PDFs.

        For each PDF: extracts the estimate number and ISCI codes, searches
        CONTRATTITESTATA for a matching contract, resolves FILMATI records, and
        returns all contract line IDs so the caller can immediately assign.
        """
        parsed_files = []
        for f in files:
            parsed_files.append((f.filename, await f.read()))

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.daviselen_traffic_parser import (
                parse_daviselen_traffic_pdf,
            )

            results = []
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                for filename, pdf_bytes in parsed_files:
                    instr = parse_daviselen_traffic_pdf(pdf_bytes)
                    if not instr.estimate:
                        results.append({"filename": filename, "error": "No estimate number found"})
                        continue

                    # Find matching contract by estimate number
                    term = f"%{instr.estimate}%"
                    cur.execute("""
                        SELECT TOP 5
                            ct.ID_CONTRATTITESTATA AS id,
                            ct.COD_CONTRATTO       AS code,
                            ct.DESCRIZIONE         AS description,
                            CONVERT(VARCHAR(10), ct.DATA_INIZIO,  101) AS date_start,
                            CONVERT(VARCHAR(10), ct.DATA_TERMINE, 101) AS date_end
                        FROM CONTRATTITESTATA ct
                        WHERE UPPER(ct.COD_CONTRATTO) LIKE %s
                           OR UPPER(ct.DESCRIZIONE)   LIKE %s
                        ORDER BY ct.DATA_INIZIO DESC
                    """, (term, term))
                    contracts = [dict(r) for r in cur.fetchall()]

                    contract_id = contracts[0]["id"] if contracts else None

                    # Resolve FILMATI for each ISCI code
                    isci_codes   = [s.isci for s in instr.spots]
                    placeholders = ",".join(f"'{c}'" for c in isci_codes)
                    filmati_map  = {}
                    if isci_codes:
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        for r in cur.fetchall():
                            filmati_map[r["COD_PROGRA"]] = {
                                "filmati_id": r["ID_FILMATI"],
                                "db_title":   r["DESCRIZIO"] or "",
                            }

                    spots_out = []
                    not_found = []
                    for s in instr.spots:
                        if s.isci in filmati_map:
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title or filmati_map[s.isci]["db_title"],
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   filmati_map[s.isci]["filmati_id"],
                                "found":        True,
                            })
                        else:
                            not_found.append(s.isci)
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title,
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   None,
                                "found":        False,
                            })

                    # Fetch line IDs for this contract (filtered by duration).
                    # Spot count is restricted to the instruction date range so the
                    # UI shows exactly how many spots will be touched.
                    line_ids = []
                    if contract_id:
                        date_filter = ""
                        if instr.date_from_sql:
                            date_filter += f" AND tp.DATA >= '{instr.date_from_sql}'"
                        if instr.date_to_sql:
                            date_filter += f" AND tp.DATA <= '{instr.date_to_sql}'"
                        cur.execute(f"""
                            SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                                   cr.DESCRIZIONE       AS description,
                                   ISNULL(sc.spot_count, 0) AS spot_count
                            FROM CONTRATTIRIGHE cr
                            LEFT JOIN (
                                SELECT tpa.id_contrattirighe, COUNT(*) AS spot_count
                                FROM trafficPalinse tpa
                                JOIN TPALINSE tp ON tp.ID_TPALINSE = tpa.id_tpalinse
                                WHERE 1=1 {date_filter}
                                GROUP BY tpa.id_contrattirighe
                            ) sc ON sc.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                            WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                              AND CAST(ROUND(CAST(cr.DURATA AS FLOAT) / {_FPS_GLOBAL}, 0) AS INT)
                                  = {instr.duration_sec}
                            ORDER BY cr.ID_CONTRATTIRIGHE
                        """)
                        line_ids = [{"line_id": r["line_id"], "description": r["description"],
                                     "spot_count": r["spot_count"]} for r in cur.fetchall()]

                    results.append({
                        "filename":     filename,
                        "estimate":     instr.estimate,
                        "product":      f"{instr.product_code} {instr.product_name}".strip(),
                        "duration_sec": instr.duration_sec,
                        "date_range":   f"{instr.start_date}–{instr.end_date}",
                        "date_from_sql": instr.date_from_sql,
                        "date_to_sql":   instr.date_to_sql,
                        "contract":    contracts[0] if contracts else None,
                        "contract_candidates": contracts,
                        "spots":       spots_out,
                        "not_found":   not_found,
                        "lines":       line_ids,
                    })

            return results

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/parse-instructions")
    async def parse_traffic_instructions(files: List[UploadFile] = File(...)):
        """Auto-detect agency format and parse one or more traffic instruction PDFs.

        Supports:
          - Davis Elen  → auto-finds contract by estimate number
          - IW Group (Lexus) → returns ISCI periods; contract selected by user
        """
        parsed_files = []
        for f in files:
            parsed_files.append((f.filename, await f.read()))

        def _detect_format(text: str) -> str:
            upper = text.upper()
            if "DAVIS ELEN ADVERTISING" in upper:
                return "daviselen"
            if "IW GROUP" in upper and "TRAFFIC SHEET" in upper:
                return "lexus"
            if "TRAFFIC INSTRUCTIONS" in upper and "tatari" in text.lower():
                return "tatari"
            if "marketing architects" in text.lower():
                return "ma"
            if "icon media direct" in text.lower():
                return "imd"
            if "hl.agency" in text.lower() or ("ESTIMATE NUMBER:" in upper and "ISCI/Ad-ID" in text):
                return "hl"
            if "RPM ADVERTISING" in upper and "TRAFFIC INSTRUCTIONS" in upper:
                return "rpm"
            return "unknown"

        def _run():
            import io as _io

            import pdfplumber

            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.parsers.daviselen_traffic_parser import (
                parse_daviselen_traffic_pdf,
            )
            from browser_automation.parsers.lexus_traffic_parser import (
                parse_lexus_traffic_pdf,
            )
            from browser_automation.parsers.tatari_traffic_parser import (
                parse_tatari_traffic_pdf,
            )

            items = []
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                for filename, pdf_bytes in parsed_files:
                    # ODS files are Direct Donor TV; PDFs need text-based detection
                    if filename.lower().endswith((".ods", ".xlsx", ".xls")):
                        fmt = "directdonor"
                        text = ""
                    else:
                        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
                            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                        fmt = _detect_format(text)

                    if fmt == "daviselen":
                        instr = parse_daviselen_traffic_pdf(pdf_bytes)
                        if not instr.estimate:
                            items.append({"filename": filename, "format": "daviselen",
                                          "error": "No estimate number found"})
                            continue

                        term = f"%{instr.estimate}%"
                        cur.execute("""
                            SELECT TOP 5
                                ct.ID_CONTRATTITESTATA AS id,
                                ct.COD_CONTRATTO       AS code,
                                ct.DESCRIZIONE         AS description,
                                CONVERT(VARCHAR(10), ct.DATA_INIZIO,  101) AS date_start,
                                CONVERT(VARCHAR(10), ct.DATA_TERMINE, 101) AS date_end
                            FROM CONTRATTITESTATA ct
                            WHERE UPPER(ct.COD_CONTRATTO) LIKE %s
                               OR UPPER(ct.DESCRIZIONE)   LIKE %s
                            ORDER BY ct.DATA_INIZIO DESC
                        """, (term, term))
                        contracts = [dict(r) for r in cur.fetchall()]
                        contract_id = contracts[0]["id"] if contracts else None

                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes)
                        filmati_map  = {}
                        if isci_codes:
                            cur.execute(
                                f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                                f" WHERE COD_PROGRA IN ({placeholders})"
                            )
                            for r in cur.fetchall():
                                filmati_map[r["COD_PROGRA"]] = {
                                    "filmati_id": r["ID_FILMATI"],
                                    "db_title":   r["DESCRIZIO"] or "",
                                }

                        spots_out, not_found = [], []
                        for s in instr.spots:
                            if s.isci in filmati_map:
                                spots_out.append({
                                    "isci": s.isci, "title": s.title or filmati_map[s.isci]["db_title"],
                                    "rotation_pct": s.rotation_pct,
                                    "filmati_id": filmati_map[s.isci]["filmati_id"], "found": True,
                                })
                            else:
                                not_found.append(s.isci)
                                spots_out.append({
                                    "isci": s.isci, "title": s.title, "rotation_pct": s.rotation_pct,
                                    "filmati_id": None, "found": False,
                                })

                        line_ids = []
                        if contract_id:
                            date_filter = ""
                            if instr.date_from_sql:
                                date_filter += f" AND tp.DATA >= '{instr.date_from_sql}'"
                            if instr.date_to_sql:
                                date_filter += f" AND tp.DATA <= '{instr.date_to_sql}'"
                            cur.execute(f"""
                                SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                                       cr.DESCRIZIONE       AS description,
                                       ISNULL(sc.spot_count, 0) AS spot_count
                                FROM CONTRATTIRIGHE cr
                                LEFT JOIN (
                                    SELECT tpa.id_contrattirighe, COUNT(*) AS spot_count
                                    FROM trafficPalinse tpa
                                    JOIN TPALINSE tp ON tp.ID_TPALINSE = tpa.id_tpalinse
                                    WHERE 1=1 {date_filter}
                                    GROUP BY tpa.id_contrattirighe
                                ) sc ON sc.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                                WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                                  AND CAST(ROUND(CAST(cr.DURATA AS FLOAT) / {_FPS_GLOBAL}, 0) AS INT)
                                      = {instr.duration_sec}
                                ORDER BY cr.ID_CONTRATTIRIGHE
                            """)
                            line_ids = [{"line_id": r["line_id"], "description": r["description"],
                                         "spot_count": r["spot_count"]} for r in cur.fetchall()]

                        items.append({
                            "filename": filename, "format": "daviselen",
                            "estimate": instr.estimate,
                            "product":  f"{instr.product_code} {instr.product_name}".strip(),
                            "duration_sec": instr.duration_sec,
                            "date_range":   f"{instr.start_date}–{instr.end_date}",
                            "date_from_sql": instr.date_from_sql,
                            "date_to_sql":   instr.date_to_sql,
                            "contract":  contracts[0] if contracts else None,
                            "contract_candidates": contracts,
                            "spots":     spots_out,
                            "not_found": not_found,
                            "lines":     line_ids,
                        })

                    elif fmt == "lexus":
                        instr = parse_lexus_traffic_pdf(pdf_bytes)
                        all_isci = list({s.isci for p in instr.periods for s in p.spots})
                        placeholders = ",".join(f"'{c}'" for c in all_isci) if all_isci else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        periods_out = []
                        for p in instr.periods:
                            spots_out = []
                            for s in p.spots:
                                found = s.isci in filmati_map
                                spots_out.append({
                                    "isci":         s.isci,
                                    "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                    "rotation_pct": s.rotation_pct,
                                    "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                                    "found":        found,
                                    "notes":        s.notes,
                                })
                            periods_out.append({
                                "duration_sec":  p.duration_sec,
                                "date_from_sql": p.date_from_sql,
                                "date_to_sql":   p.date_to_sql,
                                "date_label":    p.date_label,
                                "spots":         spots_out,
                                "all_found":     all(s["found"] for s in spots_out),
                            })

                        items.append({
                            "filename":         filename,
                            "format":           "lexus",
                            "advertiser":       instr.advertiser,
                            "campaign":         instr.campaign,
                            "coverage_area":    instr.coverage_area,
                            "market_code":      instr.market_code,
                            "search_suggestion": instr.search_suggestion,
                            "periods":          periods_out,
                        })

                    elif fmt == "tatari":
                        instr = parse_tatari_traffic_pdf(pdf_bytes)

                        # ISCI → FILMATI lookup (exact COD_PROGRA match)
                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out = []
                        for s in instr.spots:
                            found = s.isci in filmati_map
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "duration_sec": s.duration_sec,
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                                "found":        found,
                            })

                        # Group spots by duration for UI display
                        from collections import defaultdict as _dd
                        by_dur: dict = _dd(list)
                        for s in spots_out:
                            by_dur[s["duration_sec"]].append(s)
                        duration_groups = [
                            {"duration_sec": dur, "spots": grp,
                             "all_found": all(s["found"] for s in grp)}
                            for dur, grp in sorted(by_dur.items())
                        ]

                        # Fuzzy-search contracts by advertiser name + date overlap
                        term = f"%{instr.search_suggestion}%"
                        date_filter = ""
                        if instr.date_from_sql:
                            date_filter += f" AND cr.DATA_FINE >= '{instr.date_from_sql}'"
                        if instr.date_to_sql:
                            date_filter += f" AND cr.DATA_INIZIO <= '{instr.date_to_sql}'"
                        cur.execute(f"""
                            SELECT TOP 10
                                ct.ID_CONTRATTITESTATA AS id,
                                ct.COD_CONTRATTO       AS code,
                                ct.DESCRIZIONE         AS description,
                                CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                                CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end,
                                COUNT(DISTINCT cr.ID_CONTRATTIRIGHE) AS line_count
                            FROM CONTRATTITESTATA ct
                            JOIN CONTRATTIRIGHE cr
                              ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                            WHERE (ct.DESCRIZIONE LIKE %s OR ct.COD_CONTRATTO LIKE %s)
                              {date_filter}
                            GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                            ORDER BY ct.ID_CONTRATTITESTATA DESC
                        """, (term, term))
                        contracts = [dict(r) for r in cur.fetchall()]

                        items.append({
                            "filename":            filename,
                            "format":              "tatari",
                            "advertiser":          instr.advertiser,
                            "search_suggestion":   instr.search_suggestion,
                            "date_from_sql":       instr.date_from_sql,
                            "date_to_sql":         instr.date_to_sql,
                            "date_from_display":   instr.date_from_display,
                            "date_to_display":     instr.date_to_display,
                            "spots":               spots_out,
                            "duration_groups":     duration_groups,
                            "contract_candidates": contracts,
                        })

                    elif fmt == "directdonor":
                        from browser_automation.parsers.directdonor_traffic_parser import (
                            parse_directdonor_traffic_ods,
                        )
                        instr = parse_directdonor_traffic_ods(pdf_bytes, filename)

                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out = []
                        for s in instr.spots:
                            found = s.isci in filmati_map
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "duration_sec": s.duration_sec,
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                                "found":        found,
                            })

                        from collections import defaultdict as _dd
                        by_dur: dict = _dd(list)
                        for s in spots_out:
                            by_dur[s["duration_sec"]].append(s)
                        duration_groups = [
                            {"duration_sec": dur, "spots": grp,
                             "all_found": all(s["found"] for s in grp)}
                            for dur, grp in sorted(by_dur.items())
                        ]

                        term = f"%{instr.search_suggestion}%"
                        date_filter = ""
                        if instr.date_from_sql:
                            date_filter += f" AND cr.DATA_FINE >= '{instr.date_from_sql}'"
                        if instr.date_to_sql:
                            date_filter += f" AND cr.DATA_INIZIO <= '{instr.date_to_sql}'"

                        # Filter to contracts with lines matching the file's duration so
                        # e.g. the CVH :120 contract never appears when a :60 file is parsed.
                        durations = list({s["duration_sec"] for s in spots_out})
                        dur_having = ""
                        if len(durations) == 1:
                            dur_having = (
                                f" HAVING SUM(CASE WHEN CAST(ROUND(CAST(cr.DURATA AS FLOAT)"
                                f" / {_FPS_GLOBAL}, 0) AS INT) = {durations[0]} THEN 1 ELSE 0 END) > 0"
                            )

                        cur.execute(f"""
                            SELECT TOP 10
                                ct.ID_CONTRATTITESTATA AS id,
                                ct.COD_CONTRATTO       AS code,
                                ct.DESCRIZIONE         AS description,
                                CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                                CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end,
                                COUNT(DISTINCT cr.ID_CONTRATTIRIGHE) AS line_count
                            FROM CONTRATTITESTATA ct
                            JOIN CONTRATTIRIGHE cr
                              ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                            WHERE (ct.DESCRIZIONE LIKE %s OR ct.COD_CONTRATTO LIKE %s)
                              {date_filter}
                            GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                            {dur_having}
                            ORDER BY ct.ID_CONTRATTITESTATA DESC
                        """, (term, term))
                        contracts = [dict(r) for r in cur.fetchall()]

                        items.append({
                            "filename":            filename,
                            "format":              "directdonor",
                            "advertiser":          instr.advertiser,
                            "search_suggestion":   instr.search_suggestion,
                            "date_from_sql":       instr.date_from_sql,
                            "date_to_sql":         instr.date_to_sql,
                            "date_from_display":   instr.date_from_display,
                            "date_to_display":     instr.date_to_display,
                            "spots":               spots_out,
                            "duration_groups":     duration_groups,
                            "contract_candidates": contracts,
                        })

                    elif fmt == "rpm":
                        from browser_automation.parsers.rpm_traffic_parser import (
                            parse_rpm_traffic_pdf,
                        )
                        instr = parse_rpm_traffic_pdf(pdf_bytes)
                        if not instr.estimate:
                            items.append({"filename": filename, "format": "rpm",
                                          "error": "No estimate number found"})
                            continue

                        # Extract system_dialect ("Cantonese", "Mandarin", …) from spot title
                        _RPM_LANG_KW = {
                            "cantonese": "Cantonese", "mandarin": "Mandarin",
                            "vietnamese": "Vietnamese", "korean": "Korean",
                            "punjabi": "Punjabi", "hindi": "Hindi",
                            "south asian": "SouthAsian", "filipino": "Filipino",
                            "hmong": "Hmong",
                        }
                        def _rpm_dialect(title: str) -> str:
                            low = title.lower()
                            for kw, d in _RPM_LANG_KW.items():
                                if kw in low:
                                    return d
                            return ""

                        # Market label → Etere short code for contract filtering
                        _RPM_MKT = {
                            "sacramento": "CV", "san francisco": "SF",
                            "seattle": "SEA", "los angeles": "LA",
                            "houston": "HOU", "chicago": "CMP",
                            "washington": "WDC", "new york": "NYC",
                        }
                        mkt_low = instr.market.lower()
                        market_short = next((v for k, v in _RPM_MKT.items() if k in mkt_low), "")

                        # ISCI → FILMATI lookup
                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out, not_found = [], []
                        dialect_to_filmati: dict = {}  # system_dialect → filmati_id
                        for s in instr.spots:
                            found      = s.isci in filmati_map
                            fid        = filmati_map[s.isci]["filmati_id"] if found else None
                            sys_dialect = _rpm_dialect(s.title)
                            if found and sys_dialect:
                                dialect_to_filmati[sys_dialect] = fid
                            spots_out.append({
                                "isci":           s.isci,
                                "title":          s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "dialect":        sys_dialect or s.title,
                                "system_dialect": sys_dialect,
                                "rotation_pct":   s.rotation_pct,
                                "filmati_id":     fid,
                                "found":          found,
                            })
                            if not found:
                                not_found.append(s.isci)

                        # Find contract by estimate + market short code
                        term = f"%{instr.estimate}%"
                        mkt_filter = (
                            f" AND (UPPER(ct.COD_CONTRATTO) LIKE '%{market_short}%'"
                            f"   OR UPPER(ct.DESCRIZIONE)   LIKE '%{market_short}%')"
                            if market_short else ""
                        )
                        cur.execute(f"""
                            SELECT ct.ID_CONTRATTITESTATA AS id,
                                   ct.COD_CONTRATTO       AS code,
                                   ct.DESCRIZIONE         AS description,
                                   CONVERT(VARCHAR(10), ct.DATA_INIZIO,  101) AS date_start,
                                   CONVERT(VARCHAR(10), ct.DATA_TERMINE, 101) AS date_end
                            FROM CONTRATTITESTATA ct
                            WHERE (UPPER(ct.COD_CONTRATTO) LIKE %s
                               OR  UPPER(ct.DESCRIZIONE)   LIKE %s)
                            {mkt_filter}
                            ORDER BY ct.DATA_INIZIO DESC
                        """, (term, term))
                        contracts_raw = [dict(r) for r in cur.fetchall()]

                        # Per contract × dialect: count scheduled spots via language time windows
                        contracts_out = []
                        for ct in contracts_raw:
                            dialect_assignments = []
                            for sys_dialect, fid in dialect_to_filmati.items():
                                filters_dict: dict = {
                                    "languages": [sys_dialect],
                                    "duration":  instr.duration_sec,
                                }
                                if instr.date_to_sql:
                                    filters_dict["date_to"] = instr.date_to_sql
                                filter_sql = _build_spot_filter(filters_dict)
                                cur.execute(f"""
                                    SELECT COUNT(*) AS cnt
                                    FROM TPALINSE tp
                                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                                    WHERE cr.ID_CONTRATTITESTATA = {ct['id']}
                                    {filter_sql}
                                """)
                                spot_count = (cur.fetchone() or {}).get("cnt", 0) or 0
                                raw_dialect = next(
                                    (s["dialect"] for s in spots_out if s["system_dialect"] == sys_dialect),
                                    sys_dialect,
                                )
                                dialect_assignments.append({
                                    "system_dialect": sys_dialect,
                                    "dialect":        raw_dialect,
                                    "filmati_id":     fid,
                                    "isci":           next(s["isci"] for s in spots_out
                                                          if s["system_dialect"] == sys_dialect),
                                    "spot_count":     spot_count,
                                    "filters":        filters_dict,
                                })
                            contracts_out.append({**ct, "dialect_assignments": dialect_assignments})

                        items.append({
                            "filename":        filename,
                            "format":          "rpm",
                            "advertiser":      instr.advertiser,
                            "estimate":        instr.estimate,
                            "market":          instr.market,
                            "duration_sec":    instr.duration_sec,
                            "date_to_sql":     instr.date_to_sql,
                            "date_to_display": instr.date_to_display,
                            "spots":           spots_out,
                            "not_found":       not_found,
                            "contracts":       contracts_out,
                        })

                    elif fmt == "ma":
                        from browser_automation.parsers.ma_traffic_parser import (
                            parse_ma_traffic_pdf,
                        )
                        instr = parse_ma_traffic_pdf(pdf_bytes)

                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out = []
                        for s in instr.spots:
                            found = s.isci in filmati_map
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "duration_sec": s.duration_sec,
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                                "found":        found,
                            })

                        from collections import defaultdict as _dd
                        by_dur: dict = _dd(list)
                        for s in spots_out:
                            by_dur[s["duration_sec"]].append(s)
                        duration_groups = [
                            {"duration_sec": dur, "spots": grp,
                             "all_found": all(s["found"] for s in grp)}
                            for dur, grp in sorted(by_dur.items())
                        ]

                        term = f"%{instr.search_suggestion}%"
                        date_filter = ""
                        if instr.date_from_sql:
                            date_filter += f" AND cr.DATA_FINE >= '{instr.date_from_sql}'"
                        if instr.date_to_sql:
                            date_filter += f" AND cr.DATA_INIZIO <= '{instr.date_to_sql}'"

                        durations = list({s["duration_sec"] for s in spots_out})
                        dur_having = ""
                        if len(durations) == 1:
                            dur_having = (
                                f" HAVING SUM(CASE WHEN CAST(ROUND(CAST(cr.DURATA AS FLOAT)"
                                f" / {_FPS_GLOBAL}, 0) AS INT) = {durations[0]} THEN 1 ELSE 0 END) > 0"
                            )

                        cur.execute(f"""
                            SELECT TOP 10
                                ct.ID_CONTRATTITESTATA AS id,
                                ct.COD_CONTRATTO       AS code,
                                ct.DESCRIZIONE         AS description,
                                CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                                CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end,
                                COUNT(DISTINCT cr.ID_CONTRATTIRIGHE) AS line_count
                            FROM CONTRATTITESTATA ct
                            JOIN CONTRATTIRIGHE cr
                              ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                            WHERE (ct.DESCRIZIONE LIKE %s OR ct.COD_CONTRATTO LIKE %s)
                              {date_filter}
                            GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                            {dur_having}
                            ORDER BY ct.ID_CONTRATTITESTATA DESC
                        """, (term, term))
                        contracts = [dict(r) for r in cur.fetchall()]

                        items.append({
                            "filename":            filename,
                            "format":              "ma",
                            "advertiser":          instr.advertiser,
                            "client_code":         instr.client_code,
                            "product_code":        instr.product_code,
                            "search_suggestion":   instr.search_suggestion,
                            "date_from_sql":       instr.date_from_sql,
                            "date_to_sql":         instr.date_to_sql,
                            "date_from_display":   instr.date_from_display,
                            "date_to_display":     instr.date_to_display,
                            "spots":               spots_out,
                            "duration_groups":     duration_groups,
                            "contract_candidates": contracts,
                        })

                    elif fmt == "imd":
                        from browser_automation.parsers.imd_traffic_parser import (
                            parse_imd_traffic_pdf,
                        )
                        instr = parse_imd_traffic_pdf(pdf_bytes)

                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out = []
                        for s in instr.spots:
                            found = s.isci in filmati_map
                            spots_out.append({
                                "isci":         s.isci,
                                "title":        s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "duration_sec": s.duration_sec,
                                "rotation_pct": s.rotation_pct,
                                "filmati_id":   filmati_map[s.isci]["filmati_id"] if found else None,
                                "found":        found,
                            })

                        from collections import defaultdict as _dd
                        by_dur: dict = _dd(list)
                        for s in spots_out:
                            by_dur[s["duration_sec"]].append(s)
                        duration_groups = [
                            {"duration_sec": dur, "spots": grp,
                             "all_found": all(s["found"] for s in grp)}
                            for dur, grp in sorted(by_dur.items())
                        ]

                        term = f"%{instr.search_suggestion}%"
                        date_filter = ""
                        if instr.date_from_sql:
                            date_filter += f" AND cr.DATA_FINE >= '{instr.date_from_sql}'"
                        if instr.date_to_sql:
                            date_filter += f" AND cr.DATA_INIZIO <= '{instr.date_to_sql}'"

                        durations = list({s["duration_sec"] for s in spots_out})
                        dur_having = ""
                        if len(durations) == 1:
                            dur_having = (
                                f" HAVING SUM(CASE WHEN CAST(ROUND(CAST(cr.DURATA AS FLOAT)"
                                f" / {_FPS_GLOBAL}, 0) AS INT) = {durations[0]} THEN 1 ELSE 0 END) > 0"
                            )

                        cur.execute(f"""
                            SELECT TOP 10
                                ct.ID_CONTRATTITESTATA AS id,
                                ct.COD_CONTRATTO       AS code,
                                ct.DESCRIZIONE         AS description,
                                CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                                CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end,
                                COUNT(DISTINCT cr.ID_CONTRATTIRIGHE) AS line_count
                            FROM CONTRATTITESTATA ct
                            JOIN CONTRATTIRIGHE cr
                              ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                            WHERE (ct.DESCRIZIONE LIKE %s OR ct.COD_CONTRATTO LIKE %s)
                              {date_filter}
                            GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                            {dur_having}
                            ORDER BY ct.ID_CONTRATTITESTATA DESC
                        """, (term, term))
                        contracts = [dict(r) for r in cur.fetchall()]

                        items.append({
                            "filename":            filename,
                            "format":              "imd",
                            "advertiser":          instr.advertiser,
                            "client_code":         instr.client_code,
                            "product_code":        instr.product_code,
                            "search_suggestion":   instr.search_suggestion,
                            "date_from_sql":       instr.date_from_sql,
                            "date_to_sql":         instr.date_to_sql,
                            "date_from_display":   instr.date_from_display,
                            "date_to_display":     instr.date_to_display,
                            "spots":               spots_out,
                            "duration_groups":     duration_groups,
                            "contract_candidates": contracts,
                        })

                    elif fmt == "hl":
                        from browser_automation.parsers.hl_traffic_parser import (
                            parse_hl_traffic_pdf,
                        )
                        instr = parse_hl_traffic_pdf(pdf_bytes)

                        # ISCI → filmati lookup
                        isci_codes   = [s.isci for s in instr.spots]
                        placeholders = ",".join(f"'{c}'" for c in isci_codes) if isci_codes else "''"
                        cur.execute(
                            f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                            f" WHERE COD_PROGRA IN ({placeholders})"
                        )
                        filmati_map = {
                            r["COD_PROGRA"]: {"filmati_id": r["ID_FILMATI"],
                                              "db_title":   r["DESCRIZIO"] or ""}
                            for r in cur.fetchall()
                        }

                        spots_out, not_found = [], []
                        # One PDF can carry several flights (e.g. 6/2–6/8, 6/9–6/30,
                        # 6/30–7/6), each with its own ISCI per dialect. Group the found
                        # spots by (system_dialect, date_from, date_to) so each flight's
                        # creative is assigned only to spots inside its OWN window —
                        # never collapsed to a single dialect→filmati map.
                        flight_groups: dict = {}  # (sys_dialect, from, to) → spot info
                        for s in instr.spots:
                            found = s.isci in filmati_map
                            fid   = filmati_map[s.isci]["filmati_id"] if found else None
                            spots_out.append({
                                "isci":           s.isci,
                                "title":          s.title or (filmati_map[s.isci]["db_title"] if found else ""),
                                "dialect":        s.dialect,
                                "system_dialect": s.system_dialect,
                                "rotation_pct":   s.rotation_pct,
                                "date_from_sql":  s.date_from_sql,
                                "date_to_sql":    s.date_to_sql,
                                "start_date":     s.start_date,
                                "end_date":       s.end_date,
                                "filmati_id":     fid,
                                "found":          found,
                            })
                            if not found:
                                not_found.append(s.isci)
                                continue
                            key = (s.system_dialect, s.date_from_sql, s.date_to_sql)
                            if key not in flight_groups:
                                flight_groups[key] = {
                                    "system_dialect": s.system_dialect,
                                    "dialect":        s.dialect,
                                    "isci":           s.isci,
                                    "filmati_id":     fid,
                                    "duration_sec":   s.duration_sec,
                                    "date_from_sql":  s.date_from_sql or instr.date_from_sql,
                                    "date_to_sql":    s.date_to_sql or instr.date_to_sql,
                                    "start_date":     s.start_date,
                                    "end_date":       s.end_date,
                                }

                        # Find ALL contracts matching the estimate number
                        term = f"%{instr.estimate}%"
                        cur.execute("""
                            SELECT ct.ID_CONTRATTITESTATA AS id,
                                   ct.COD_CONTRATTO       AS code,
                                   ct.DESCRIZIONE         AS description,
                                   CONVERT(VARCHAR(10), ct.DATA_INIZIO,  101) AS date_start,
                                   CONVERT(VARCHAR(10), ct.DATA_TERMINE, 101) AS date_end
                            FROM CONTRATTITESTATA ct
                            WHERE UPPER(ct.COD_CONTRATTO) LIKE %s
                               OR UPPER(ct.DESCRIZIONE)   LIKE %s
                            ORDER BY ct.DATA_INIZIO DESC
                        """, (term, term))
                        contracts_raw = [dict(r) for r in cur.fetchall()]

                        # For each contract × dialect, count scheduled spots using the
                        # language time windows — same engine as the manual assign page.
                        # No line-description guessing; spot position determines language.
                        contracts_out = []
                        for ct in contracts_raw:
                            dialect_assignments = []
                            for grp in flight_groups.values():
                                filters_dict = {
                                    "languages": [grp["system_dialect"]],
                                    "duration":  grp["duration_sec"],
                                }
                                if grp["date_from_sql"]:
                                    filters_dict["date_from"] = grp["date_from_sql"]
                                if grp["date_to_sql"]:
                                    filters_dict["date_to"] = grp["date_to_sql"]
                                filter_sql = _build_spot_filter(filters_dict)
                                cur.execute(f"""
                                    SELECT COUNT(*) AS cnt
                                    FROM TPALINSE tp
                                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                                    WHERE cr.ID_CONTRATTITESTATA = {ct['id']}
                                    {filter_sql}
                                """)
                                spot_count = (cur.fetchone() or {}).get("cnt", 0) or 0
                                dialect_assignments.append({
                                    "system_dialect": grp["system_dialect"],
                                    "dialect":        grp["dialect"],
                                    "filmati_id":     grp["filmati_id"],
                                    "isci":           grp["isci"],
                                    "start_date":     grp["start_date"],
                                    "end_date":       grp["end_date"],
                                    "spot_count":     spot_count,
                                    "filters":        filters_dict,
                                })
                            contracts_out.append({**ct, "dialect_assignments": dialect_assignments})

                        items.append({
                            "filename":      filename,
                            "format":        "hl",
                            "advertiser":    instr.advertiser,
                            "estimate":      instr.estimate,
                            "duration_sec":  instr.duration_sec,
                            "date_range":    f"{instr.start_date}–{instr.end_date}",
                            "date_from_sql": instr.date_from_sql,
                            "date_to_sql":   instr.date_to_sql,
                            "spots":         spots_out,
                            "not_found":     not_found,
                            "contracts":     contracts_out,
                        })

                    else:
                        items.append({"filename": filename, "format": "unknown",
                                      "error": "Unrecognised traffic instruction format"})

            return items

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/traffic/tatari/search")
    async def tatari_contract_search(
        q: str = Query(""),
        date_from: str = Query(""),
        date_to: str = Query(""),
    ):
        """Fuzzy-search contracts by advertiser name/code with optional date-overlap filter."""
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                term = f"%{q}%"
                date_filter = ""
                if date_from:
                    date_filter += f" AND cr.DATA_FINE >= '{date_from}'"
                if date_to:
                    date_filter += f" AND cr.DATA_INIZIO <= '{date_to}'"
                id_filter = ""
                params: tuple = (term, term)
                if q.isdigit():
                    id_filter = "OR ct.ID_CONTRATTITESTATA = %s"
                    params = (term, term, int(q))
                cur.execute(f"""
                    SELECT TOP 20
                        ct.ID_CONTRATTITESTATA AS id,
                        ct.COD_CONTRATTO       AS code,
                        ct.DESCRIZIONE         AS description,
                        CONVERT(VARCHAR(10), MIN(cr.DATA_INIZIO), 101) AS date_start,
                        CONVERT(VARCHAR(10), MAX(cr.DATA_FINE),   101) AS date_end,
                        COUNT(DISTINCT cr.ID_CONTRATTIRIGHE) AS line_count
                    FROM CONTRATTITESTATA ct
                    JOIN CONTRATTIRIGHE cr
                      ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                    WHERE (ct.DESCRIZIONE LIKE %s OR ct.COD_CONTRATTO LIKE %s {id_filter})
                      {date_filter}
                    GROUP BY ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO, ct.DESCRIZIONE
                    ORDER BY ct.ID_CONTRATTITESTATA DESC
                """, params)
                return [dict(r) for r in cur.fetchall()]
        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/traffic/contract/{contract_id}/lines-by-duration")
    async def traffic_lines_by_duration(
        contract_id: int,
        duration_sec: int = Query(30),
        date_from: str = Query(""),
        date_to: str = Query(""),
    ):
        """Return contract lines matching a given duration, with scheduled spot counts in the date window."""
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                date_filter = ""
                if date_from:
                    date_filter += f" AND tp.DATA >= '{date_from}'"
                if date_to:
                    date_filter += f" AND tp.DATA <= '{date_to}'"
                cur.execute(f"""
                    SELECT cr.ID_CONTRATTIRIGHE AS line_id,
                           cr.DESCRIZIONE       AS description,
                           ISNULL(sc.spot_count, 0) AS spot_count
                    FROM CONTRATTIRIGHE cr
                    LEFT JOIN (
                        SELECT tpa.id_contrattirighe, COUNT(*) AS spot_count
                        FROM trafficPalinse tpa
                        JOIN TPALINSE tp ON tp.ID_TPALINSE = tpa.id_tpalinse
                        WHERE 1=1 {date_filter}
                        GROUP BY tpa.id_contrattirighe
                    ) sc ON sc.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                      AND CAST(ROUND(CAST(cr.DURATA AS FLOAT) / {_FPS_GLOBAL}, 0) AS INT)
                          = {duration_sec}
                    ORDER BY cr.ID_CONTRATTIRIGHE
                """)
                return [dict(r) for r in cur.fetchall()]
        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/traffic/contract/{contract_id}/tpalinse-spots")
    async def traffic_tpalinse_spots(contract_id: int):
        """Return all individual TPALINSE entries for a contract, ordered by date/time."""
        def _run():
            from datetime import datetime as _dt

            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(f"""
                    SELECT
                        tp.ID_TPALINSE          AS tp_id,
                        tp.DATA                 AS spot_date,
                        tp.ORA                  AS spot_time_frames,
                        tpa.id_contrattirighe   AS line_id,
                        cr.DURATA               AS duration_frames,
                        cr.DESCRIZIONE          AS line_description,
                        cr.COD_USER             AS market_id,
                        COALESCE(cr.ORA_INIZIOF, cr.ORA_INIZIO) AS line_time_from_frames,
                        COALESCE(cr.ORA_FINEF,   cr.ORA_FINE)   AS line_time_to_frames,
                        tp.ID_FILMATI           AS current_filmati_id,
                        f.COD_PROGRA            AS current_filmati_code,
                        ISNULL(NULLIF(f.DESCRIZIO, ''), f.COD_PROGRA) AS current_filmati_title
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    LEFT JOIN FILMATI f     ON f.ID_FILMATI          = tp.ID_FILMATI
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                    ORDER BY cr.ID_CONTRATTIRIGHE ASC, tp.DATA, tp.ORA
                """)
                rows = cur.fetchall()

            FPS = 29.97

            def _frames_to_hhmm(frames):
                if not frames:
                    return ""
                secs = frames / FPS
                h = int(secs // 3600)
                m = int((secs % 3600) // 60)
                return f"{h:02d}:{m:02d}"

            _SKIP_COMPOSITE = {"Chinese", "SouthAsian"}

            def _get_language_for_time(time_from: str, mkt_id: int) -> str:
                if not time_from:
                    return ""
                lang_table = _DAL_LANG_WINDOWS if mkt_id == 10 else _CTV_LANG_WINDOWS
                for lang, windows in lang_table.items():
                    if lang in _SKIP_COMPOSITE:
                        continue
                    for _days, w_from, w_to in windows:
                        if w_from <= time_from < w_to:
                            return lang
                return ""

            result = []
            for r in rows:
                d            = r["spot_date"]
                date_str     = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
                total_secs   = (r["spot_time_frames"] or 0) / FPS
                hours        = int(total_secs // 3600)
                minutes      = int((total_secs % 3600) // 60)
                time_str     = f"{hours:02d}:{minutes:02d}"
                raw_dur_sec  = round((r["duration_frames"] or 0) / FPS)
                # Snap to nearest 15s — Etere Selenium path stores duration at
                # a different fps than 29.97 (web UI uses ~36fps), so raw_dur_sec
                # lands at :18 instead of :15. Snapping mirrors EtereBridge's
                # round_to_nearest_increment behaviour.
                dur_sec      = round(raw_dur_sec / 15) * 15 if raw_dur_sec > 0 else 0
                try:
                    day_name = _dt.strptime(date_str, "%Y-%m-%d").strftime("%A")
                except Exception:
                    day_name = ""

                result.append({
                    "tp_id":                 r["tp_id"],
                    "spot_date":             date_str,
                    "spot_time":             time_str,
                    "day_name":              day_name,
                    "line_id":               r["line_id"],
                    "line_description":      r["line_description"] or "",
                    "line_time_from":        _frames_to_hhmm(r["line_time_from_frames"]),
                    "line_time_to":          _frames_to_hhmm(r["line_time_to_frames"]),
                    "line_language":         _get_language_for_time(
                                                _frames_to_hhmm(r["line_time_from_frames"]),
                                                r["market_id"] or 0,
                                             ),
                    "duration_sec":          dur_sec,
                    "current_filmati_id":    r["current_filmati_id"],
                    "current_filmati_code":  r["current_filmati_code"],
                    "current_filmati_title": r["current_filmati_title"],
                })
            return result

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/contract/{contract_id}/assign-spots")
    async def traffic_assign_spots(contract_id: int, body: dict = Body(...)):
        """
        Assign specific filmati to individual TPALINSE spots.
        Body: {assignments: [{tp_id: int, filmati_id: int}, ...]}
        Each tp_id gets exactly the filmati_id specified — no round-robin.
        """
        assignments = body.get("assignments", [])
        if not assignments:
            raise HTTPException(status_code=400, detail="No assignments provided")

        def _run():
            from collections import defaultdict

            from browser_automation.etere_direct_client import ETERE_WEB_URL
            from browser_automation.etere_direct_client import connect as _db_connect

            tp_ids_list  = [a["tp_id"]      for a in assignments]
            fid_list_raw = [a["filmati_id"] for a in assignments]
            all_filmati  = list(set(fid_list_raw))
            tp_ph        = ",".join(str(t) for t in tp_ids_list)
            fid_ph       = ",".join(str(f) for f in all_filmati)

            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                cur.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE ID_FILMATI IN ({fid_ph})"
                )
                frows = cur.fetchall()
                filmati_cod_map   = {r["ID_FILMATI"]: (r["COD_PROGRA"] or "") for r in frows}
                filmati_title_map = {r["ID_FILMATI"]: (r["DESCRIZIO"] or "") for r in frows}

                cur.execute(f"""
                    SELECT ff.ID_FILMATI, ff.FILE_ID, ff.VIDEOSTANDARD, ff.DUR,
                           ISNULL(d.LEGACY_BASESUPP,
                                  CAST(d.LEGACY_MEDIAID AS VARCHAR) + 'ETX      ') AS supporto_prefix
                    FROM FS_FILMATI ff
                    JOIN FS_METADEVICE d ON d.ID_METADEVICE = ff.ID_METADEVICE
                    WHERE ff.ID_FILMATI IN ({fid_ph})
                      AND d.LEGACY_MEDIAID IS NOT NULL
                """)
                _VS_TO_ASPECT: dict = {"D": "H"}
                filmati_supporto_map: dict = {}
                filmati_aspect_map:   dict = {}
                filmati_duration_map: dict = {}
                for r in cur.fetchall():
                    fid = r["ID_FILMATI"]
                    if fid not in filmati_supporto_map:
                        filmati_supporto_map[fid] = (r["supporto_prefix"] or "") + (r["FILE_ID"] or "")
                        filmati_aspect_map[fid]   = _VS_TO_ASPECT.get(r["VIDEOSTANDARD"], "H")
                        filmati_duration_map[fid] = r["DUR"] or 0

                cur.execute("SELECT id_bookingcode, code FROM trf_bookingcode")
                bookingcode_to_newtype = {r["id_bookingcode"]: r["code"] for r in cur.fetchall()}

                cur.execute(f"""
                    SELECT tp.ID_TPALINSE        AS tp_id,
                           tpa.id_contrattirighe AS line_id,
                           cr.ID_BOOKINGCODE     AS booking_code
                    FROM TPALINSE tp
                    JOIN trafficPalinse tpa ON tpa.id_tpalinse      = tp.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE = tpa.id_contrattirighe
                    WHERE tp.ID_TPALINSE IN ({tp_ph})
                      AND tp.DATA >= CAST(GETDATE() AS DATE)
                """)
                tp_line_map:    dict = {}
                tp_newtype_map: dict = {}
                for r in cur.fetchall():
                    tp_line_map[r["tp_id"]]    = r["line_id"]
                    tp_newtype_map[r["tp_id"]] = bookingcode_to_newtype.get(r["booking_code"], "COM")

                cur.execute(
                    f"SELECT DISTINCT ID_FILMATI FROM CONTRATTIFILMATI"
                    f" WHERE ID_CONTRATTIRIGHE IN ("
                    f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                    f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                    f" )"
                )
                existing_pool = {r["ID_FILMATI"] for r in cur.fetchall()}

            # Group by line: line_id → [(tp_id, filmati_id)]
            line_tp_filmati: dict = defaultdict(list)
            for a in assignments:
                tp_id      = a["tp_id"]
                filmati_id = a["filmati_id"]
                line_id    = tp_line_map.get(tp_id)
                if line_id:
                    line_tp_filmati[line_id].append((tp_id, filmati_id))

            session      = _get_etere_session()
            lines_updated = spots_updated = 0
            tp_assignments: list = []

            try:
                for fid in all_filmati:
                    if fid in existing_pool:
                        continue
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAddToAssetListC",
                        json={"idFilmatiList": [fid], "idct": contract_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(f"MaterialAddToAssetListC failed for filmati {fid}: {r.json()}")

                for line_id, pairs in line_tp_filmati.items():
                    idp = [p[0] for p in pairs]
                    idf = [p[1] for p in pairs]
                    r = session.post(
                        f"{ETERE_WEB_URL}/Sales/MaterialAssignAssetRotation",
                        json={"idp": idp, "idf": idf, "idcr": line_id},
                        headers={"X-Requested-With": "XMLHttpRequest"},
                        timeout=30,
                    )
                    r.raise_for_status()
                    if not r.json().get("IsOk"):
                        raise ValueError(
                            f"MaterialAssignAssetRotation failed for line {line_id}: {r.json()}"
                        )
                    lines_updated += 1
                    spots_updated += len(idp)
                    tp_assignments.extend(pairs)

            except Exception:
                _invalidate_etere_session()
                raise

            if tp_assignments:
                with _db_connect() as conn:
                    cur = conn.cursor()
                    for tp_id, filmati_id in tp_assignments:
                        cur.execute(
                            "UPDATE TPALINSE SET COD_PROGRA = %s, TITLE = %s, ID_FILMATI = %d,"
                            " NEWTYPE = %s, SUPPORTO = %s, ASPECT = %s, DURATION_P = %d"
                            " WHERE ID_TPALINSE = %d",
                            (
                                filmati_cod_map.get(filmati_id, ""),
                                filmati_title_map.get(filmati_id, ""),
                                filmati_id,
                                tp_newtype_map.get(tp_id, "COM"),
                                filmati_supporto_map.get(filmati_id, ""),
                                filmati_aspect_map.get(filmati_id, "H"),
                                filmati_duration_map.get(filmati_id, 0),
                                tp_id,
                            ),
                        )
                    # Ensure every assigned line has pool rows in CONTRATTIFILMATI.
                    # PERCROTATION is left 0 — actual rotation is driven by TPALINSE.
                    for line_id, pairs in line_tp_filmati.items():
                        fids = list({p[1] for p in pairs})
                        for fid in fids:
                            cur.execute(
                                "DELETE FROM CONTRATTIFILMATI"
                                " WHERE ID_CONTRATTIRIGHE = %d AND ID_FILMATI = %d",
                                (line_id, fid),
                            )
                            cur.execute(
                                "INSERT INTO CONTRATTIFILMATI"
                                " (ID_CONTRATTIRIGHE, ID_FILMATI, PERCROTATION)"
                                " VALUES (%d, %d, 0)",
                                (line_id, fid),
                            )
                    # Remove pool entries that weren't assigned to any spot on any line.
                    # Must exclude assigned lines or we delete the rows just inserted above.
                    if line_tp_filmati:
                        assigned_line_ph = ",".join(str(lid) for lid in line_tp_filmati.keys())
                        cur.execute(
                            f"DELETE FROM CONTRATTIFILMATI"
                            f" WHERE ID_FILMATI IN ({fid_ph})"
                            f" AND PERCROTATION = 0"
                            f" AND ID_CONTRATTIRIGHE NOT IN ({assigned_line_ph})"
                            f" AND ID_CONTRATTIRIGHE IN ("
                            f"   SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE"
                            f"   WHERE ID_CONTRATTITESTATA = {contract_id}"
                            f" )"
                        )
                    conn.commit()

            needs_airchecks = False
            try:
                import sqlite3 as _sqlite3
                with _db_connect() as _conn:
                    _cur = _conn.cursor(as_dict=True)
                    _cur.execute(
                        "SELECT COMMITTENTE FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = %d" % contract_id
                    )
                    _hdr = _cur.fetchone()
                if _hdr and _hdr.get("COMMITTENTE"):
                    _cid = str(int(_hdr["COMMITTENTE"]))
                    _db_path = config.customer_db_path
                    if _db_path.exists():
                        with _sqlite3.connect(str(_db_path)) as _sdb:
                            _row = _sdb.execute(
                                "SELECT auto_aircheck FROM customers WHERE customer_id = ?", (_cid,)
                            ).fetchone()
                            needs_airchecks = bool(_row and _row[0])
            except Exception:
                pass

            return {"ok": True, "assigned": spots_updated, "lines_updated": lines_updated, "needs_airchecks": needs_airchecks, "contract_id": contract_id}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Move Contract Lines ────────────────────────────────────────────────
    @router.get("/traffic/move-lines", response_class=HTMLResponse)
    async def move_lines_page(request: Request):
        return templates.TemplateResponse(request, "traffic/move_lines.html")

    @router.get("/api/traffic/contract/{contract_id}/lines")
    async def traffic_contract_lines(contract_id: int):
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute("""
                    SELECT cr.ID_CONTRATTIRIGHE                         AS id,
                           cr.DESCRIZIONE                               AS description,
                           CONVERT(VARCHAR(10), cr.DATA_INIZIO,  101)  AS dt_start,
                           CONVERT(VARCHAR(10), cr.DATA_FINE,    101)  AS dt_end,
                           cr.DURATA                                    AS duration_frames,
                           cr.PRENOTAZIONE                              AS prenotazione,
                           cr.CONTROLLACAPOFILA                         AS capofila,
                           cr.CONTROLLAFINEFILA                         AS finefila,
                           COUNT(tp.ID_TPALINSE)                        AS spot_count
                    FROM CONTRATTIRIGHE cr
                    LEFT JOIN trafficPalinse tpa
                           ON tpa.id_contrattirighe = cr.ID_CONTRATTIRIGHE
                    LEFT JOIN TPALINSE tp
                           ON tp.ID_TPALINSE = tpa.id_tpalinse
                    WHERE cr.ID_CONTRATTITESTATA = %d
                    GROUP BY cr.ID_CONTRATTIRIGHE, cr.DESCRIZIONE,
                             cr.DATA_INIZIO, cr.DATA_FINE, cr.DURATA,
                             cr.PRENOTAZIONE, cr.CONTROLLACAPOFILA, cr.CONTROLLAFINEFILA
                    ORDER BY cr.ID_CONTRATTIRIGHE
                """ % contract_id)
                rows = cur.fetchall()
                result = []
                for r in rows:
                    frames = r["duration_frames"] or 0
                    secs = round(frames / 30)
                    result.append({
                        "id":           r["id"],
                        "description":  r["description"] or "",
                        "dt_start":     r["dt_start"] or "",
                        "dt_end":       r["dt_end"] or "",
                        "duration_sec": secs,
                        "prenotazione": r["prenotazione"],
                        "capofila":     r["capofila"],
                        "finefila":     r["finefila"],
                        "spot_count":   r["spot_count"],
                    })
                return result

        try:
            rows = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(rows)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/traffic/move-lines")
    async def traffic_move_lines(body: dict = Body(...)):
        line_ids      = body.get("line_ids", [])
        to_contract   = body.get("to_contract_id")

        if not line_ids:
            raise HTTPException(status_code=400, detail="No lines selected")
        if not to_contract:
            raise HTTPException(status_code=400, detail="No destination contract specified")

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor()
                ids_ph = ",".join(str(int(i)) for i in line_ids)
                cur.execute(
                    f"UPDATE CONTRATTIRIGHE"
                    f" SET ID_CONTRATTITESTATA = {int(to_contract)}"
                    f" WHERE ID_CONTRATTIRIGHE IN ({ids_ph})"
                )
                moved = cur.rowcount
                conn.commit()
            return {"ok": True, "moved": moved}

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Max Spots per Day / Week ──────────────────────────────────────────────

    @router.get("/scripts/max-spots", response_class=HTMLResponse)
    async def scripts_max_spots(request: Request):
        return templates.TemplateResponse(request, "scripts/max_spots.html")

    @router.get("/api/scripts/max-spots/lines")
    async def get_max_spots_lines(contract_id: int = Query(..., gt=0)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT ID_CONTRATTIRIGHE, DESCRIZIONE,
                           COALESCE(DATESTART, DATA_INIZIO), COALESCE(DATEEND, DATA_FINE),
                           COALESCE(ORA_INIZIOF, ORA_INIZIO), COALESCE(ORA_FINEF, ORA_FINE),
                           LUNEDI, MARTEDI, MERCOLEDI, GIOVEDI, VENERDI, SABATO, DOMENICA,
                           DURATA, PASSAGGI_SETTIMANALI, PASSAGGI_GIORNALIERI,
                           COD_USER
                    FROM   CONTRATTIRIGHE
                    WHERE  ID_CONTRATTITESTATA = %s
                    ORDER  BY ID_CONTRATTIRIGHE
                """, [contract_id])
                rows = cursor.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail=f"No lines found for contract {contract_id}.")

            lines = []
            for row in rows:
                (line_id, desc, date_from, date_to, ora_in, ora_out,
                 lun, mar, mer, gio, ven, sab, dom,
                 durata, spots_pw, max_daily,
                 cod_user) = row
                lines.append({
                    "line_id":      line_id,
                    "description":  desc or "",
                    "market":       _MARKET_NAMES.get(cod_user, str(cod_user) if cod_user else "—"),
                    "date_from":    f"{date_from.month}/{date_from.day}/{date_from.year}" if date_from else "",
                    "date_to":      f"{date_to.month}/{date_to.day}/{date_to.year}" if date_to else "",
                    "time_from":    _frames_to_ampm(ora_in),
                    "time_to":      _frames_to_ampm(ora_out),
                    "days":         _days_str(lun, mar, mer, gio, ven, sab, dom),
                    "duration_sec": _frames_to_sec(durata),
                    "spots_pw":     spots_pw or 0,
                    "max_daily":    max_daily or 0,
                    "max_weekly":   spots_pw or 0,
                })
            return JSONResponse({"contract_id": contract_id, "lines": lines})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/max-spots/apply")
    async def apply_max_spots(payload: dict = Body(...)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            line_ids   = [int(x) for x in payload.get("line_ids", [])]
            max_daily  = payload.get("max_daily")
            max_weekly = payload.get("max_weekly")

            if not line_ids:
                raise HTTPException(status_code=400, detail="No lines selected.")
            if max_daily is None and max_weekly is None:
                raise HTTPException(status_code=400, detail="Provide at least one value to update.")

            set_parts, params = [], []
            if max_daily is not None:
                set_parts.append("PASSAGGI_GIORNALIERI = %s")
                params.append(int(max_daily))
            if max_weekly is not None:
                set_parts.append("PASSAGGI_SETTIMANALI = %s")
                params.append(int(max_weekly))

            placeholders = ",".join(["%s"] * len(line_ids))
            params.extend(line_ids)

            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    f"UPDATE CONTRATTIRIGHE SET {', '.join(set_parts)}"
                    f" WHERE ID_CONTRATTIRIGHE IN ({placeholders})",
                    params
                )
                conn.commit()
                updated = cursor.rowcount

            return JSONResponse({"updated": updated})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/scripts/spot-validation", response_class=HTMLResponse)
    async def scripts_spot_validation(request: Request):
        return templates.TemplateResponse(request, "scripts/spot_validation.html")

    @router.get("/api/scripts/spot-validation/lines")
    async def get_spot_validation_lines(contract_id: int = Query(..., gt=0)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT cr.ID_CONTRATTIRIGHE, cr.DESCRIZIONE, cr.COD_USER,
                           COALESCE(cr.DATESTART, cr.DATA_INIZIO),
                           COALESCE(cr.DATEEND, cr.DATA_FINE),
                           cr.N_PASSAGGI,
                           cr.CONTROLLACAPOFILA, cr.CONTROLLAFINEFILA,
                           COUNT(tp.ID_ContrattiRighe) AS scheduled
                    FROM   CONTRATTIRIGHE cr
                    LEFT JOIN trafficTPalinse tp
                           ON tp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                    WHERE  cr.ID_CONTRATTITESTATA = %s
                    GROUP  BY cr.ID_CONTRATTIRIGHE, cr.DESCRIZIONE, cr.COD_USER,
                              cr.DATESTART, cr.DATA_INIZIO, cr.DATEEND, cr.DATA_FINE,
                              cr.N_PASSAGGI, cr.CONTROLLACAPOFILA, cr.CONTROLLAFINEFILA
                    ORDER  BY cr.ID_CONTRATTIRIGHE
                """, [contract_id])
                rows = cursor.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail=f"No lines found for contract {contract_id}.")

            lines = []
            total_ordered = total_expected = total_scheduled = mismatches = 0
            for (lid, desc, cod_user, date_from, date_to, n_pass, capofila, finefila, sched) in rows:
                ordered = n_pass or 0
                # Bookend lines (scheduling type 6: top AND bottom of break) air each
                # ordered passage as a top/bottom pair, so 2 scheduled spots per N_PASSAGGI
                # is correct — not an over-schedule. Compare against the doubled expectation.
                is_bookend = bool(capofila) and bool(finefila)
                expected   = ordered * 2 if is_bookend else ordered
                diff       = sched - expected
                total_ordered   += ordered
                total_expected  += expected
                total_scheduled += sched
                if diff != 0:
                    mismatches += 1
                lines.append({
                    "line_id":    lid,
                    "description": desc or "",
                    "market":     _MARKET_NAMES.get(cod_user, str(cod_user) if cod_user else "—"),
                    "date_from":  f"{date_from.month}/{date_from.day}/{date_from.year}" if date_from else "",
                    "date_to":    f"{date_to.month}/{date_to.day}/{date_to.year}" if date_to else "",
                    "ordered":    ordered,
                    "expected":   expected,
                    "is_bookend": is_bookend,
                    "scheduled":  sched,
                    "diff":       diff,
                })

            return JSONResponse({
                "contract_id":     contract_id,
                "lines":           lines,
                "total_ordered":   total_ordered,
                "total_expected":  total_expected,
                "total_scheduled": total_scheduled,
                "mismatches":      mismatches,
            })

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/scripts/rename-assets", response_class=HTMLResponse)
    async def scripts_rename_assets(request: Request):
        return templates.TemplateResponse(request, "scripts/rename_assets.html")

    @router.post("/api/scripts/rename-assets/preview")
    async def rename_assets_preview(payload: dict = Body(...)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            # pairs: [{code, title, new_code}]
            # new_code is None/null when no ISCI parenthetical was found
            pairs = payload.get("pairs", [])
            if not pairs:
                raise HTTPException(status_code=400, detail="No pairs provided.")

            codes = [p["code"] for p in pairs]
            placeholders = ",".join(["%s"] * len(codes))

            # Candidate new codes that actually change the spot code — these must
            # not already belong to a *different* asset (COD_PROGRA must stay unique).
            new_codes = sorted({
                p["new_code"] for p in pairs
                if p.get("new_code") and p["new_code"] != p["code"]
            })

            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    f" WHERE COD_PROGRA IN ({placeholders})",
                    codes
                )
                found = {r[1]: {"id": r[0], "current_desc": r[2] or ""} for r in cursor.fetchall()}

                # Map each candidate new code → assets already using it (id, desc).
                existing_new: dict = {}
                if new_codes:
                    nc_ph = ",".join(["%s"] * len(new_codes))
                    cursor.execute(
                        f"SELECT ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                        f" WHERE COD_PROGRA IN ({nc_ph})",
                        new_codes
                    )
                    for r in cursor.fetchall():
                        existing_new.setdefault(r[1], []).append({"id": r[0], "desc": r[2] or ""})

            results = []
            for p in pairs:
                code     = p["code"]
                title    = p["title"]
                new_code = p.get("new_code")  # None = no ISCI, keep COD_PROGRA as-is
                if code not in found:
                    results.append({
                        "code":     code,
                        "new_code": new_code,
                        "title":    title,
                        "found":    False,
                    })
                    continue

                target_id = found[code]["id"]
                # Conflict: the new code is already held by some OTHER asset.
                conflict = None
                if new_code and new_code != code:
                    others = [a for a in existing_new.get(new_code, []) if a["id"] != target_id]
                    if others:
                        conflict = {"id": others[0]["id"], "desc": others[0]["desc"]}

                results.append({
                    "code":          code,
                    "new_code":      new_code,
                    "title":         title,
                    "current_desc":  found[code]["current_desc"],
                    "asset_id":      target_id,
                    "found":         True,
                    "conflict":      bool(conflict),
                    "conflict_id":   conflict["id"]   if conflict else None,
                    "conflict_desc": conflict["desc"] if conflict else None,
                })

            return JSONResponse({"results": results})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/rename-assets/apply")
    async def rename_assets_apply(payload: dict = Body(...)):
        try:
            project_root = Path(__file__).parent.parent.parent.parent
            if str(project_root) not in sys.path:
                sys.path.insert(0, str(project_root))
            from browser_automation.etere_direct_client import connect as _db_connect

            pairs = payload.get("pairs", [])
            if not pairs:
                raise HTTPException(status_code=400, detail="No pairs to apply.")

            with _db_connect() as conn:
                cursor = conn.cursor()
                updated = 0
                conflicts: list = []   # new_code already held by a different asset
                skipped:   list = []   # source code not found in the library
                for p in pairs:
                    code     = p["code"]
                    title    = p["title"]
                    new_code = p.get("new_code")

                    # Description-only update — code unchanged, no uniqueness risk.
                    if not new_code or new_code == code:
                        cursor.execute(
                            "UPDATE FILMATI SET DESCRIZIO = %s WHERE COD_PROGRA = %s",
                            [title, code]
                        )
                        updated += cursor.rowcount
                        continue

                    # Resolve the asset(s) being renamed (by current code).
                    cursor.execute(
                        "SELECT ID_FILMATI FROM FILMATI WHERE COD_PROGRA = %s", [code]
                    )
                    target_ids = [r[0] for r in cursor.fetchall()]
                    if not target_ids:
                        skipped.append(code)
                        continue

                    # HARD GUARD: refuse if new_code is already used by any OTHER asset.
                    # Same-connection read also sees earlier renames in this batch,
                    # so two pairs targeting the same new_code can't both succeed.
                    tgt_ph = ",".join(["%s"] * len(target_ids))
                    cursor.execute(
                        f"SELECT TOP 1 ID_FILMATI FROM FILMATI"
                        f" WHERE COD_PROGRA = %s AND ID_FILMATI NOT IN ({tgt_ph})",
                        [new_code, *target_ids]
                    )
                    if cursor.fetchone():
                        conflicts.append({"code": code, "new_code": new_code})
                        continue

                    cursor.execute(
                        "UPDATE FILMATI SET DESCRIZIO = %s, COD_PROGRA = %s"
                        " WHERE COD_PROGRA = %s",
                        [title, new_code, code]
                    )
                    updated += cursor.rowcount
                conn.commit()

            return JSONResponse({"updated": updated, "conflicts": conflicts, "skipped": skipped})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/scripts/rename-programming", response_class=HTMLResponse)
    async def scripts_rename_programming(request: Request):
        return templates.TemplateResponse(request, "scripts/rename_programming.html")

    @router.post("/api/scripts/rename-programming/search")
    async def rename_programming_search(payload: dict = Body(...)):
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            prefix = (payload.get("prefix") or "").strip().upper()
            if not prefix:
                raise HTTPException(status_code=400, detail="Prefix required.")
            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT TOP 200 ID_FILMATI, COD_PROGRA, DESCRIZIO FROM FILMATI"
                    " WHERE COD_PROGRA LIKE %s ORDER BY COD_PROGRA",
                    [prefix + "%"],
                )
                assets = [{"id": r[0], "code": r[1] or "", "title": r[2] or ""} for r in cursor.fetchall()]
            return JSONResponse({"assets": assets})
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/rename-programming/apply")
    async def rename_programming_apply(payload: dict = Body(...)):
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            pairs = payload.get("pairs", [])
            if not pairs:
                raise HTTPException(status_code=400, detail="No pairs to apply.")
            with _db_connect() as conn:
                cursor = conn.cursor()
                updated = 0
                schedule_updated = 0
                for p in pairs:
                    cursor.execute(
                        "UPDATE FILMATI SET COD_PROGRA = %s, DESCRIZIO = %s WHERE ID_FILMATI = %d",
                        (p["new_code"], p["new_title"], p["asset_id"]),
                    )
                    updated += cursor.rowcount
                    cursor.execute(
                        "UPDATE TPALINSE SET COD_PROGRA = %s, TITLE = %s WHERE ID_FILMATI = %d",
                        (p["new_code"], p["new_title"], p["asset_id"]),
                    )
                    schedule_updated += cursor.rowcount
                    cursor.execute(
                        "UPDATE trafficTPalinse SET Cod_Progra = %s, Title = %s WHERE ID_Filmati = %d",
                        (p["new_code"], p["new_title"], p["asset_id"]),
                    )
                    schedule_updated += cursor.rowcount
                conn.commit()
            return JSONResponse({"updated": updated, "schedule_updated": schedule_updated})
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Fix Overscheduled Contract ────────────────────────────────────────────

    @router.get("/orders/fix-overscheduled", response_class=HTMLResponse)
    async def fix_overscheduled_page(request: Request):
        return templates.TemplateResponse(request, "fix_overscheduled.html")

    @router.get("/api/orders/fix-overscheduled/search")
    async def fix_overscheduled_search(q: str = ""):
        if not q or len(q) < 2:
            return JSONResponse([])

        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    "SELECT TOP 20"
                    "  ID_CONTRATTITESTATA AS id,"
                    "  COD_CONTRATTO AS code,"
                    "  DESCRIZIONE AS description,"
                    "  CONVERT(VARCHAR(10), DATA_INIZIO, 101) AS date_start,"
                    "  CONVERT(VARCHAR(10), DATA_TERMINE, 101) AS date_end"
                    " FROM CONTRATTITESTATA"
                    " WHERE COD_CONTRATTO LIKE %s OR DESCRIZIONE LIKE %s"
                    " ORDER BY DATA_INIZIO DESC",
                    (f"%{q}%", f"%{q}%"),
                )
                return [dict(r) for r in cur.fetchall()]

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/orders/fix-overscheduled/{contract_id}/preview")
    async def fix_overscheduled_preview(contract_id: int):
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)
                cur.execute(
                    "SELECT COD_CONTRATTO AS code, DESCRIZIONE AS description,"
                    " CONVERT(VARCHAR(10), DATA_INIZIO, 101) AS date_start,"
                    " CONVERT(VARCHAR(10), DATA_TERMINE, 101) AS date_end"
                    f" FROM CONTRATTITESTATA WHERE ID_CONTRATTITESTATA = {contract_id}"
                )
                header = cur.fetchone()
                if not header:
                    raise ValueError(f"Contract {contract_id} not found")

                # Lines that have Traffic_ScheduleList blacklist entries
                cur.execute(f"""
                    SELECT
                        cr.ID_CONTRATTIRIGHE AS line_id,
                        cr.DESCRIZIONE       AS description,
                        cr.COD_USER          AS market_id,
                        cr.N_PASSAGGI        AS ordered,
                        CONVERT(VARCHAR(10), cr.DATA_INIZIO, 101) AS date_start,
                        CONVERT(VARCHAR(10), cr.DATA_FINE,  101) AS date_end,
                        (SELECT COUNT(*) FROM trafficPalinse tpa
                         WHERE tpa.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE) AS placed,
                        SUM(tsl.PassageMiss) AS missed
                    FROM CONTRATTIRIGHE cr
                    JOIN Traffic_ScheduleList tsl ON tsl.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                      AND tsl.BlackList > 0
                    GROUP BY cr.ID_CONTRATTIRIGHE, cr.DESCRIZIONE, cr.COD_USER,
                             cr.N_PASSAGGI, cr.DATA_INIZIO, cr.DATA_FINE
                    ORDER BY cr.COD_USER, cr.ID_CONTRATTIRIGHE
                """)
                lines = [dict(r) for r in cur.fetchall()]
                return {"header": dict(header), "lines": lines}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/orders/fix-overscheduled/{contract_id}/apply")
    async def fix_overscheduled_apply(contract_id: int):
        def _run():
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cur = conn.cursor()

                # Lines with blacklist entries and their placed counts
                cur.execute(f"""
                    SELECT cr.ID_CONTRATTIRIGHE,
                           (SELECT COUNT(*) FROM trafficPalinse tpa
                            WHERE tpa.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE) AS placed
                    FROM CONTRATTIRIGHE cr
                    WHERE cr.ID_CONTRATTITESTATA = {contract_id}
                      AND EXISTS (
                          SELECT 1 FROM Traffic_ScheduleList tsl
                          WHERE tsl.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                            AND tsl.BlackList > 0
                      )
                """)
                placed_by_line = {r[0]: r[1] for r in cur.fetchall()}

                if not placed_by_line:
                    return {"ok": True, "blacklist_rows_deleted": 0, "lines_updated": 0}

                line_ph = ",".join(str(lid) for lid in placed_by_line)

                # Delete Traffic_ScheduleList blacklist rows for these lines
                cur.execute(
                    f"DELETE FROM Traffic_ScheduleList"
                    f" WHERE ID_ContrattiRighe IN ({line_ph}) AND BlackList > 0"
                )
                bl_deleted = cur.rowcount

                # Update N_PASSAGGI on each affected line to match placed count
                lines_updated = 0
                for line_id, placed in placed_by_line.items():
                    cur.execute(
                        "UPDATE CONTRATTIRIGHE SET N_PASSAGGI = %d WHERE ID_CONTRATTIRIGHE = %d",
                        (placed, line_id),
                    )
                    lines_updated += 1

                conn.commit()
                return {"ok": True, "blacklist_rows_deleted": bl_deleted, "lines_updated": lines_updated}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Booked Business ──────────────────────────────────────────────────────

    import calendar as _cal_mod

    def _bb_broadcast_month_start(yr: int, mo: int) -> _date_cls:
        first = _date_cls(yr, mo, 1)
        return first - timedelta(days=first.weekday())

    def _bb_month_end(yr: int, mo: int) -> _date_cls:
        return _date_cls(yr, mo, _cal_mod.monthrange(yr, mo)[1])

    @router.get("/master-control/booked-business")
    async def booked_business_page(request: Request):
        return templates.TemplateResponse(request, "master_control/booked_business.html")

    @router.get("/api/master-control/booked-business/load")
    async def booked_business_load(year: int, month: int, show_trade: bool = False):
        def _run():
            import calendar as _cal
            from collections import defaultdict

            from browser_automation.etere_direct_client import connect as _connect

            bcast_start = _bb_broadcast_month_start(year, month)
            cal_start   = _date_cls(year, month, 1)
            month_end   = _bb_month_end(year, month)

            _ny, _nm  = (year, month + 1) if month < 12 else (year + 1, 1)
            bcast_end = _bb_broadcast_month_start(_ny, _nm) - timedelta(days=1)

            trade_guard = (
                "-- show all"
                if show_trade else
                "AND cr.NEWTYPE NOT LIKE '%%TRD%%' AND (ct.CAMBIOMERCE = 0 OR ct.CAMBIOMERCE IS NULL) AND ct.ID_PAGAMENTI != 4"
            )

            with _connect() as conn:
                cur = conn.cursor(as_dict=True)
                # Revenue is counted from TPALINSE (actual scheduled spots) at the
                # contract line rate (cr.IMPORTO).  This matches the commercial log
                # exactly — one spot = one rate unit, no proration across flight days.
                # The CASE expression applies the correct billing window per contract:
                # Broadcast (316) → bcast_start; Calendar (317) or Unset → cal_start.
                cur.execute(f"""
                    SELECT
                        ct.ID_CONTRATTITESTATA             AS id,
                        ct.CENTROMEDIA, ct.P_AGENZIA, ct.COD_CONTRATTO,
                        ct.CAMBIOMERCE, ct.ID_PAGAMENTI,
                        t.COD_USER                         AS cod_user,
                        LTRIM(RTRIM(
                            CASE WHEN ae.Nome IS NOT NULL AND ae.Nome != ''
                                 THEN ae.Nome + ' ' + ae.RAG_SOCIAL
                                 ELSE ae.RAG_SOCIAL END
                        ))                                 AS ae_name,
                        ag.RAG_SOCIAL                      AS buying_agency,
                        cl.RAG_SOCIAL                      AS client_name,
                        ISNULL(SUM(
                            CASE WHEN cr.CONTROLLACAPOFILA = 1 AND cr.CONTROLLAFINEFILA = 1
                                 THEN cr.IMPORTO * 0.5
                                 ELSE cr.IMPORTO
                            END
                        ), 0)                              AS gross
                    FROM TPALINSE t
                    JOIN trafficTPalinse tp
                      ON tp.ID_TPalinse = t.ID_TPALINSE
                    JOIN CONTRATTIRIGHE cr
                      ON cr.ID_CONTRATTIRIGHE = tp.ID_ContrattiRighe
                    JOIN CONTRATTITESTATA ct
                      ON ct.ID_CONTRATTITESTATA = tp.ID_CONTRATTITESTATA
                    LEFT JOIN ANAGRAF ae ON ae.ID_ANAGRAF = ct.AGENTE1
                    LEFT JOIN ANAGRAF ag ON ag.ID_ANAGRAF = ct.AGENZIA
                    LEFT JOIN ANAGRAF cl ON cl.ID_ANAGRAF = ct.COMMITTENTE
                    WHERE cr.NEWTYPE LIKE '%%COM%%'
                      AND cr.IMPORTO > 0
                      AND t.LIVELLO = 0
                      {trade_guard}
                      AND t.DATA >= CASE WHEN ct.CENTROMEDIA = 316 THEN %s ELSE %s END
                      AND t.DATA <= CASE WHEN ct.CENTROMEDIA = 316 THEN %s ELSE %s END
                    GROUP BY
                        ct.ID_CONTRATTITESTATA, ct.CENTROMEDIA, ct.P_AGENZIA, ct.COD_CONTRATTO,
                        ct.CAMBIOMERCE, ct.ID_PAGAMENTI, t.COD_USER,
                        ae.Nome, ae.RAG_SOCIAL, ag.RAG_SOCIAL, cl.RAG_SOCIAL
                """, (str(bcast_start), str(cal_start), str(bcast_end), str(month_end)))
                rows = cur.fetchall()

                # Production / non-airtime charges (CONTRATTISPESE) in the same
                # billing window. Shown as separate "<client> PROD" rows (Lee's
                # existing tagging convention) so charge money is visibly
                # distinct from airtime; may merge once entry is automated.
                # Charge date falls back to the carrier line's flight start
                # when the Orders app left DATA null. Net uses the header
                # agency % — by convention a net-only production deal carries
                # P_AGENZIA=0 on its (standalone) contract.
                cur.execute("""
                    SELECT
                        ct.ID_CONTRATTITESTATA             AS id,
                        ct.CENTROMEDIA, ct.P_AGENZIA, ct.COD_CONTRATTO,
                        ct.CAMBIOMERCE, ct.ID_PAGAMENTI,
                        cr.COD_USER                        AS cod_user,
                        LTRIM(RTRIM(
                            CASE WHEN ae.Nome IS NOT NULL AND ae.Nome != ''
                                 THEN ae.Nome + ' ' + ae.RAG_SOCIAL
                                 ELSE ae.RAG_SOCIAL END
                        ))                                 AS ae_name,
                        ag.RAG_SOCIAL                      AS buying_agency,
                        cl.RAG_SOCIAL                      AS client_name,
                        ISNULL(SUM(s.IMPORTO), 0)          AS gross
                    FROM CONTRATTISPESE s
                    JOIN CONTRATTIRIGHE cr
                      ON cr.ID_CONTRATTIRIGHE = s.ID_CONTRATTIRIGHE
                    JOIN CONTRATTITESTATA ct
                      ON ct.ID_CONTRATTITESTATA = cr.ID_CONTRATTITESTATA
                    LEFT JOIN ANAGRAF ae ON ae.ID_ANAGRAF = ct.AGENTE1
                    LEFT JOIN ANAGRAF ag ON ag.ID_ANAGRAF = ct.AGENZIA
                    LEFT JOIN ANAGRAF cl ON cl.ID_ANAGRAF = ct.COMMITTENTE
                    WHERE COALESCE(s.DATA, cr.DATA_INIZIO)
                              >= CASE WHEN ct.CENTROMEDIA = 316 THEN %s ELSE %s END
                      AND COALESCE(s.DATA, cr.DATA_INIZIO)
                              <= CASE WHEN ct.CENTROMEDIA = 316 THEN %s ELSE %s END
                    GROUP BY
                        ct.ID_CONTRATTITESTATA, ct.CENTROMEDIA, ct.P_AGENZIA, ct.COD_CONTRATTO,
                        ct.CAMBIOMERCE, ct.ID_PAGAMENTI, cr.COD_USER,
                        ae.Nome, ae.RAG_SOCIAL, ag.RAG_SOCIAL, cl.RAG_SOCIAL
                """, (str(bcast_start), str(cal_start), str(bcast_end), str(month_end)))
                prod_rows = cur.fetchall()

            def _is_trade(r):
                return r["CAMBIOMERCE"] or r["ID_PAGAMENTI"] == 4

            _MKT_CODE = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}
            _MKT_ORDER = ["NYC","CMP","HOU","SFO","SEA","LAX","CVC","WDC","MMT","DAL"]

            clients: dict = defaultdict(
                lambda: {"gross": 0.0, "net": 0.0, "centromedia": None, "unset": False,
                         "markets": set(), "by_market": defaultdict(lambda: {"gross": 0.0, "net": 0.0})}
            )
            trade_clients: dict = defaultdict(
                lambda: {"gross": 0.0, "net": 0.0, "markets": set(),
                         "by_market": defaultdict(lambda: {"gross": 0.0, "net": 0.0})}
            )

            wl_fee_by_ae: dict = defaultdict(float)
            wl_fee_by_ae_mkt: dict = defaultdict(float)

            for r in rows:
                cm    = r["CENTROMEDIA"] or 0
                gross = float(r["gross"])
                net   = gross * (1 - float(r["P_AGENZIA"] or 0) / 100)

                ae     = r["ae_name"]      or "Unknown AE"
                agency = (r["buying_agency"] or "").strip()
                client = (r["client_name"]   or "").strip()
                if agency and client and agency != client:
                    cli = f"{agency}:{client}"
                else:
                    cli = client or agency or r["COD_CONTRATTO"] or "Unknown"

                mkt = _MKT_CODE.get(r.get("cod_user") or 0, "")

                if show_trade and _is_trade(r):
                    key = (ae, cli)
                    trade_clients[key]["gross"] += gross
                    trade_clients[key]["net"]   += net
                    if mkt:
                        trade_clients[key]["markets"].add(mkt)
                        trade_clients[key]["by_market"][mkt]["gross"] += gross
                        trade_clients[key]["by_market"][mkt]["net"]   += net
                else:
                    key = (ae, cli)
                    clients[key]["gross"] += gross
                    clients[key]["net"]   += net
                    if clients[key]["centromedia"] is None:
                        clients[key]["centromedia"] = cm
                    if cm == 0:
                        clients[key]["unset"] = True
                    if mkt:
                        clients[key]["markets"].add(mkt)
                        clients[key]["by_market"][mkt]["gross"] += gross
                        clients[key]["by_market"][mkt]["net"]   += net
                    if agency == "Worldlink":
                        # Round fee per contract (matches spreadsheet per-contract rounding).
                        # Rows arrive per (contract, market), so the per-market fee
                        # split uses the same rounding as the total.
                        wl_fee_by_ae[ae] += round(-0.10 * net, 2)
                        if mkt:
                            wl_fee_by_ae_mkt[(ae, mkt)] += round(-0.10 * net, 2)

            # Production / non-airtime charges → "<client> PROD" rows.
            # Same AE/client keying as airtime so the PROD row sorts right
            # under its airtime sibling; trade-contract charges follow the
            # same show_trade rules as airtime revenue.
            for r in prod_rows:
                gross = float(r["gross"])
                if not gross:
                    continue
                if _is_trade(r) and not show_trade:
                    continue
                cm  = r["CENTROMEDIA"] or 0
                net = gross * (1 - float(r["P_AGENZIA"] or 0) / 100)
                ae     = r["ae_name"]      or "Unknown AE"
                agency = (r["buying_agency"] or "").strip()
                client = (r["client_name"]   or "").strip()
                if agency and client and agency != client:
                    cli = f"{agency}:{client}"
                else:
                    cli = client or agency or r["COD_CONTRATTO"] or "Unknown"
                cli += " PROD"
                mkt = _MKT_CODE.get(r.get("cod_user") or 0, "")
                key = (ae, cli)
                bucket = trade_clients if (show_trade and _is_trade(r)) else clients
                bucket[key]["gross"] += gross
                bucket[key]["net"]   += net
                if mkt:
                    bucket[key]["markets"].add(mkt)
                    bucket[key]["by_market"][mkt]["gross"] += gross
                    bucket[key]["by_market"][mkt]["net"]   += net
                if bucket is clients:
                    if clients[key]["centromedia"] is None:
                        clients[key]["centromedia"] = cm
                    if cm == 0:
                        clients[key]["unset"] = True

            # Inject WorldLink broker fee line (DO NOT INVOICE)
            for ae, fee in wl_fee_by_ae.items():
                key = (ae, "WorldLink Broker Fees (DO NOT INVOICE)")
                clients[key]["gross"] += fee
                clients[key]["net"]   += fee
                if clients[key]["centromedia"] is None:
                    clients[key]["centromedia"] = 316
            for (ae, mkt), fee in wl_fee_by_ae_mkt.items():
                key = (ae, "WorldLink Broker Fees (DO NOT INVOICE)")
                clients[key]["markets"].add(mkt)
                clients[key]["by_market"][mkt]["gross"] += fee
                clients[key]["by_market"][mkt]["net"]   += fee

            def _build_ae_groups(client_map, include_billing=True):
                ae_map: dict = defaultdict(list)
                for (ae, cli), data in client_map.items():
                    row = {
                        "client":  cli,
                        "gross":   round(data["gross"], 2),
                        "net":     round(data["net"],   2),
                        "markets": sorted(data.get("markets", set()),
                                          key=lambda m: _MKT_ORDER.index(m) if m in _MKT_ORDER else 99),
                        # Per-market split so the UI's market pills can show
                        # per-market REVENUE, not just filter row visibility.
                        "by_market": {
                            m: {"gross": round(v["gross"], 2), "net": round(v["net"], 2)}
                            for m, v in data.get("by_market", {}).items()
                        },
                    }
                    if include_billing:
                        cm = data.get("centromedia") or 0
                        row["billing"] = "Broadcast" if cm == 316 else ("Calendar" if cm == 317 else "—")
                        row["unset"]   = data.get("unset", False)
                    else:
                        row["billing"] = "Trade"
                        row["unset"]   = False
                    ae_map[ae].append(row)
                groups = []
                for ae in sorted(ae_map, key=str.casefold):
                    rows_out = sorted(ae_map[ae], key=lambda x: x["client"].casefold())
                    groups.append({
                        "ae":      ae,
                        "clients": rows_out,
                        "gross":   round(sum(c["gross"] for c in rows_out), 2),
                        "net":     round(sum(c["net"]   for c in rows_out), 2),
                    })
                return groups

            ae_groups    = _build_ae_groups(clients,       include_billing=True)
            trade_groups = _build_ae_groups(trade_clients, include_billing=False) if show_trade else []

            grand_gross = round(sum(g["gross"] for g in ae_groups), 2)
            grand_net   = round(sum(g["net"]   for g in ae_groups), 2)

            all_markets = sorted(
                {m for g in ae_groups for c in g["clients"] for m in c["markets"]},
                key=lambda m: _MKT_ORDER.index(m) if m in _MKT_ORDER else 99,
            )

            def _md(d):
                return f"{d.strftime('%b')} {d.day}"

            month_label = f"{_cal.month_name[month]} {year}"
            bcast_label = f"{_md(bcast_start)} – {_md(bcast_end)}, {bcast_end.year}"
            cal_label   = f"{_md(cal_start)} – {_md(month_end)}, {month_end.year}"

            return {
                "month_label":   month_label,
                "bcast_bounds":  bcast_label,
                "cal_bounds":    cal_label,
                "ae_groups":     ae_groups,
                "grand_gross":   grand_gross,
                "grand_net":     grand_net,
                "trade_groups":  trade_groups,
                "trade_gross":   round(sum(g["gross"] for g in trade_groups), 2),
                "trade_net":     round(sum(g["net"]   for g in trade_groups), 2),
                "all_markets":   all_markets,
            }

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Billing Type Cleanup ──────────────────────────────────────────────────

    @router.get("/master-control/billing-type-fix")
    async def billing_type_fix_page(request: Request):
        return templates.TemplateResponse(request, "master_control/billing_type_fix.html")

    @router.get("/api/master-control/billing-type-fix/clients")
    async def billing_type_fix_clients():
        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            conn = _connect()
            cur = conn.cursor(as_dict=True)
            cur.execute("""
                SELECT
                    a.ID_ANAGRAF,
                    a.RAG_SOCIAL                       AS client_name,
                    ISNULL(a.CENTROMEDIA, 0)           AS default_billing,
                    COUNT(ct.ID_CONTRATTITESTATA)      AS unset_count
                FROM ANAGRAF a
                JOIN CONTRATTITESTATA ct
                     ON ct.COMMITTENTE = a.ID_ANAGRAF
                    AND ct.CENTROMEDIA = 0
                    AND (ct.CAMBIOMERCE = 0 OR ct.CAMBIOMERCE IS NULL)
                    AND ct.ID_PAGAMENTI != 4
                GROUP BY a.ID_ANAGRAF, a.RAG_SOCIAL, a.CENTROMEDIA
                ORDER BY a.RAG_SOCIAL
            """)
            rows = cur.fetchall()
            conn.close()

            clients = []
            for r in rows:
                db = int(r["default_billing"] or 0)
                clients.append({
                    "id":              r["ID_ANAGRAF"],
                    "name":            (r["client_name"] or "").strip(),
                    "default_billing": db,
                    "unset_count":     r["unset_count"],
                })
            return {"clients": clients}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/api/master-control/billing-type-fix/suggest")
    async def billing_type_fix_suggest():
        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            conn = _connect()
            cur = conn.cursor(as_dict=True)

            # All clients with unset contracts + ANAGRAF default + most-recent historical billing
            cur.execute("""
                SELECT
                    a.ID_ANAGRAF,
                    ISNULL(a.CENTROMEDIA, 0) AS anagraf_billing,
                    hist.CENTROMEDIA         AS hist_billing
                FROM ANAGRAF a
                JOIN CONTRATTITESTATA ct
                     ON ct.COMMITTENTE = a.ID_ANAGRAF
                    AND ct.CENTROMEDIA = 0
                    AND (ct.CAMBIOMERCE = 0 OR ct.CAMBIOMERCE IS NULL)
                    AND ct.ID_PAGAMENTI != 4
                OUTER APPLY (
                    SELECT TOP 1 CENTROMEDIA
                    FROM CONTRATTITESTATA h
                    WHERE h.COMMITTENTE = a.ID_ANAGRAF
                      AND h.CENTROMEDIA IN (316, 317)
                    ORDER BY h.ID_CONTRATTITESTATA DESC
                ) hist
                GROUP BY a.ID_ANAGRAF, a.CENTROMEDIA, hist.CENTROMEDIA
            """)
            suggestions: dict = {}
            for r in cur.fetchall():
                cid = r["ID_ANAGRAF"]
                ab  = int(r["anagraf_billing"] or 0)
                hb  = r["hist_billing"]
                if ab in (316, 317):
                    suggestions[cid] = {"billing": ab,  "source": "anagraf"}
                elif hb in (316, 317):
                    suggestions[cid] = {"billing": hb,  "source": "history"}
                else:
                    suggestions[cid] = {"billing": None, "source": "none"}

            conn.close()
            return {"suggestions": {str(k): v for k, v in suggestions.items()}}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/master-control/billing-type-fix/apply")
    async def billing_type_fix_apply(request: Request):
        body = await request.json()
        updates = body.get("updates", [])
        if not updates:
            return JSONResponse({"updated": 0})

        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            conn = _connect()
            cur = conn.cursor()
            total_contracts = 0
            for u in updates:
                cid     = int(u["client_id"])
                billing = int(u["billing"])
                if billing not in (316, 317):
                    continue
                cur.execute(
                    "UPDATE ANAGRAF SET CENTROMEDIA = %s WHERE ID_ANAGRAF = %s",
                    (billing, cid),
                )
                cur.execute(
                    """UPDATE CONTRATTITESTATA
                          SET CENTROMEDIA = %s
                        WHERE COMMITTENTE = %s
                          AND CENTROMEDIA = 0
                          AND (CAMBIOMERCE = 0 OR CAMBIOMERCE IS NULL)
                          AND ID_PAGAMENTI != 4""",
                    (billing, cid),
                )
                total_contracts += cur.rowcount
            conn.commit()
            conn.close()
            return {"updated_contracts": total_contracts, "updated_clients": len(updates)}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Per-contract unset billing fix ───────────────────────────────────────

    @router.get("/api/master-control/booked-business/unset-contracts")
    async def booked_business_unset_contracts(year: int, month: int):
        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            bcast_start = _bb_broadcast_month_start(year, month)
            month_end   = _bb_month_end(year, month)

            conn = _connect()
            cur = conn.cursor(as_dict=True)
            cur.execute("""
                SELECT DISTINCT
                    ct.ID_CONTRATTITESTATA,
                    ct.COD_CONTRATTO,
                    ct.DESCRIZIONE,
                    LTRIM(RTRIM(
                        CASE WHEN ae.Nome IS NOT NULL AND ae.Nome != ''
                             THEN ae.Nome + ' ' + ae.RAG_SOCIAL
                             ELSE ae.RAG_SOCIAL END
                    ))             AS ae_name,
                    ag.RAG_SOCIAL  AS buying_agency,
                    cl.RAG_SOCIAL  AS client_name
                FROM CONTRATTITESTATA ct
                JOIN CONTRATTIRIGHE cr
                     ON cr.ID_CONTRATTITESTATA = ct.ID_CONTRATTITESTATA
                    AND cr.DATA_INIZIO <= %s
                    AND cr.DATA_FINE   >= %s
                LEFT JOIN ANAGRAF ae ON ae.ID_ANAGRAF = ct.AGENTE1
                LEFT JOIN ANAGRAF ag ON ag.ID_ANAGRAF = ct.AGENZIA
                LEFT JOIN ANAGRAF cl ON cl.ID_ANAGRAF = ct.COMMITTENTE
                WHERE ct.CENTROMEDIA = 0
                  AND (ct.CAMBIOMERCE = 0 OR ct.CAMBIOMERCE IS NULL)
                  AND ct.ID_PAGAMENTI != 4
                ORDER BY ae_name, buying_agency, client_name, ct.COD_CONTRATTO
            """, (str(month_end), str(bcast_start)))
            rows = cur.fetchall()
            conn.close()

            contracts = []
            for r in rows:
                agency = (r["buying_agency"] or "").strip()
                client = (r["client_name"]   or "").strip()
                if agency and client and agency != client:
                    display_client = f"{agency}:{client}"
                else:
                    display_client = client or agency or r["COD_CONTRATTO"] or "Unknown"
                contracts.append({
                    "id":          r["ID_CONTRATTITESTATA"],
                    "code":        r["COD_CONTRATTO"] or "",
                    "description": (r["DESCRIZIONE"] or "").strip(),
                    "ae":          (r["ae_name"] or "Unknown AE").strip(),
                    "client":      display_client,
                })
            return {"contracts": contracts}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/master-control/booked-business/set-contract-billing")
    async def booked_business_set_contract_billing(request: Request):
        body = await request.json()
        updates = body.get("updates", [])  # [{contract_id, billing}, ...]
        if not updates:
            return JSONResponse({"updated": 0})

        def _run():
            from browser_automation.etere_direct_client import connect as _connect

            conn = _connect()
            cur = conn.cursor()
            count = 0
            for u in updates:
                cid     = int(u["contract_id"])
                billing = int(u["billing"])
                if billing not in (316, 317):
                    continue
                cur.execute(
                    "UPDATE CONTRATTITESTATA SET CENTROMEDIA = %s WHERE ID_CONTRATTITESTATA = %s AND CENTROMEDIA = 0",
                    (billing, cid),
                )
                count += cur.rowcount
            conn.commit()
            conn.close()
            return {"updated": count}

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Trade Entry ────────────────────────────────────────────────────────────

    @router.get("/orders/trade-entry", response_class=HTMLResponse)
    async def trade_entry_page(request: Request):
        return templates.TemplateResponse(request, "trade/trade_entry.html")

    @router.get("/api/trade/search-customer")
    async def trade_search_customer(q: str = Query(..., min_length=2)):
        def _run():
            from browser_automation.etere_direct_client import connect as _connect
            conn = _connect()
            cur = conn.cursor(as_dict=True)
            like = f"%{q}%"
            cur.execute(
                "SELECT TOP 20 ID_ANAGRAF, RAG_SOCIAL FROM ANAGRAF "
                "WHERE RAG_SOCIAL LIKE %s ORDER BY RAG_SOCIAL",
                (like,),
            )
            rows = cur.fetchall()
            conn.close()
            return [{"id": r["ID_ANAGRAF"], "name": r["RAG_SOCIAL"]} for r in rows]

        try:
            return JSONResponse(await asyncio.get_running_loop().run_in_executor(None, _run))
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/trade/create")
    async def trade_create_contract(request: Request):
        body = await request.json()

        customer_id  = int(body["customer_id"])
        code         = str(body["code"]).strip()
        description  = str(body["description"]).strip()
        date_from_s  = str(body["date_from"])
        date_to_s    = str(body["date_to"])
        note         = str(body.get("note", ""))
        separation   = int(body.get("separation", 15))
        lines_data   = body["lines"]  # [{market, description, daypart, days, duration_sec, total_spots, rate}]

        def _run():
            from datetime import date as _d

            from browser_automation.etere_direct_client import (
                EtereDirectClient,
            )
            from browser_automation.etere_direct_client import (
                connect as _connect,
            )

            date_from = _d.fromisoformat(date_from_s)
            date_to   = _d.fromisoformat(date_to_s)

            conn = _connect()
            try:
                client = EtereDirectClient(conn, owner="HOUSE", autocommit=False)
                client.set_master_market("NYC")

                contract_id = client.create_contract_header(
                    code=code,
                    description=description,
                    customer_id=customer_id,
                    agency_id=0,
                    media_center_id=0,
                    contract_date=_d.today(),
                    contract_end_date=date_to,
                    payment_id=4,
                    contract_type=1,
                    note=note,
                )

                # Mark as Exchange for Goods (not exposed through the SP)
                cur = conn.cursor()
                cur.execute(
                    "UPDATE CONTRATTITESTATA SET CAMBIOMERCE = 1 "
                    "WHERE ID_CONTRATTITESTATA = %s",
                    (contract_id,),
                )

                markets_created = []
                for ld in lines_data:
                    dur_sec   = int(ld["duration_sec"])
                    dur_str   = f"00:00:{dur_sec:02d}:00"
                    spots     = int(ld["total_spots"])
                    rate      = float(ld["rate"])
                    market    = str(ld["market"])
                    daypart   = str(ld.get("daypart", "06:00-23:59"))
                    days      = str(ld.get("days", "M-Su"))
                    line_desc = str(ld.get("description", description))
                    line_from = _d.fromisoformat(ld["date_from"]) if ld.get("date_from") else date_from
                    line_to   = _d.fromisoformat(ld["date_to"])   if ld.get("date_to")   else date_to

                    client.add_contract_line(
                        market=market,
                        days=days,
                        time_range=daypart,
                        description=line_desc,
                        rate=rate,
                        total_spots=spots,
                        spots_per_week=0,   # monthly rotation
                        date_from=line_from,
                        date_to=line_to,
                        duration=dur_str,
                        is_trade=True,
                        separation_intervals=(separation, 0, 0),
                        contract_id=contract_id,
                        priority=600,
                        whitelist_priority=60,
                        booking_code=3,     # TRD
                    )
                    markets_created.append(market)

                conn.commit()
                conn.close()
                return {
                    "contract_id":   contract_id,
                    "code":          code,
                    "lines_created": len(lines_data),
                    "markets":       markets_created,
                }
            except Exception:
                conn.rollback()
                conn.close()
                raise

        try:
            result = await asyncio.get_running_loop().run_in_executor(None, _run)
            return JSONResponse(result)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Release Blacklist ─────────────────────────────────────────────────────

    @router.get("/scripts/spot-relocator", response_class=HTMLResponse)
    async def scripts_spot_relocator(request: Request):
        return templates.TemplateResponse(request, "scripts/spot_relocator.html")

    @router.get("/api/scripts/spot-relocator/analyze")
    async def spot_relocator_analyze(contract_id: int = Query(..., gt=0)):
        # READ-ONLY: compute a feasible relocation plan per stuck line. No DB writes.
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            from browser_automation.spot_relocator import analyze_contract
            with _db_connect() as conn:
                return JSONResponse(analyze_contract(conn.cursor(), int(contract_id)))
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.get("/scripts/release-blacklist", response_class=HTMLResponse)
    async def scripts_release_blacklist(request: Request):
        return templates.TemplateResponse(request, "scripts/release_blacklist.html")

    @router.get("/api/scripts/release-blacklist/preview")
    async def release_blacklist_preview(contract_id: int = Query(..., gt=0)):
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT
                        cr.ID_CONTRATTIRIGHE,
                        cr.COD_USER,
                        cr.DESCRIZIONE,
                        cr.N_PASSAGGI,
                        COUNT(tsl.ID_ContrattiRighe) AS blacklist_rows,
                        SUM(tsl.PassageMiss) AS missed_spots,
                        (SELECT COUNT(*) FROM trafficPalinse tp
                         WHERE tp.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                         AND (tp.ID_TRAFFICTRASH = 0 OR tp.ID_TRAFFICTRASH IS NULL)) AS placed_spots
                    FROM CONTRATTIRIGHE cr
                    JOIN Traffic_ScheduleList tsl ON tsl.ID_ContrattiRighe = cr.ID_CONTRATTIRIGHE
                    WHERE cr.ID_CONTRATTITESTATA = %s AND tsl.BlackList > 0
                    GROUP BY cr.ID_CONTRATTIRIGHE, cr.COD_USER, cr.DESCRIZIONE, cr.N_PASSAGGI
                    ORDER BY cr.COD_USER, cr.ID_CONTRATTIRIGHE
                """, [contract_id])
                rows = cursor.fetchall()

            if not rows:
                return JSONResponse({"lines": [], "total_missed": 0})

            _mn = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}
            lines = []
            total_missed = 0
            for (lid, cod_user, desc, n_pass, _bl_rows, missed, placed) in rows:
                m = missed or 0
                total_missed += m
                lines.append({
                    "line_id":     lid,
                    "market":      _mn.get(cod_user, str(cod_user) if cod_user else "—"),
                    "description": desc or "",
                    "ordered":     n_pass or 0,
                    "placed":      placed or 0,
                    "missed":      m,
                })

            return JSONResponse({"contract_id": contract_id, "lines": lines, "total_missed": total_missed})
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    @router.post("/api/scripts/release-blacklist/apply")
    async def release_blacklist_apply(body: dict = Body(...)):
        contract_id = body.get("contract_id")
        if not contract_id:
            raise HTTPException(status_code=400, detail="contract_id required")
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    DELETE FROM Traffic_ScheduleList
                    WHERE ID_ContrattiRighe IN (
                        SELECT ID_CONTRATTIRIGHE FROM CONTRATTIRIGHE
                        WHERE ID_CONTRATTITESTATA = %s
                    )
                    AND BlackList > 0
                """, [contract_id])
                deleted = cursor.rowcount
                # Reset line status to Ready (0) so the scheduler can pick them up again
                cursor.execute("""
                    UPDATE CONTRATTIRIGHE
                    SET ROWSTATUS = 0
                    WHERE ID_CONTRATTITESTATA = %s
                      AND ROWSTATUS != 0
                """, [contract_id])
                updated_lines = cursor.rowcount
                conn.commit()
            return JSONResponse({"deleted": deleted, "lines_reset": updated_lines})
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    # ── Make Goods ──────────────────────────────────────────────────────────

    @router.get("/orders/make-goods", response_class=HTMLResponse)
    async def make_goods_page(request: Request):
        return templates.TemplateResponse(request, "make_goods.html")

    @router.get("/api/orders/make-goods")
    async def get_make_goods(date_from: str = Query(...), date_to: str = Query(...)):
        try:
            from browser_automation.etere_direct_client import connect as _db_connect
            with _db_connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT
                        ct.ID_CONTRATTITESTATA,
                        ct.COD_CONTRATTO,
                        ct.DESCRIZIONE              AS contract_desc,
                        ISNULL(a.RAG_SOCIAL, '')    AS client_name,
                        ISNULL(ag.RAG_SOCIAL, '')   AS agency_name,
                        ct.P_AGENZIA,
                        cr.ID_CONTRATTIRIGHE,
                        cr.COD_USER                 AS market_id,
                        cr.DESCRIZIONE              AS line_desc,
                        CONVERT(varchar, cr.DATA_INIZIO, 101) AS date_start,
                        CONVERT(varchar, cr.DATA_FINE,   101) AS date_end,
                        cr.N_PASSAGGI               AS ordered,
                        cr.IMPORTO,
                        cr.ID_BOOKINGCODE,
                        cr.ORA_INIZIO,
                        cr.ORA_FINE,
                        SUM(tsl.PassageMiss)        AS missed
                    FROM Traffic_ScheduleList tsl
                    JOIN CONTRATTIRIGHE cr
                        ON cr.ID_CONTRATTIRIGHE = tsl.ID_ContrattiRighe
                    JOIN CONTRATTITESTATA ct
                        ON ct.ID_CONTRATTITESTATA = cr.ID_CONTRATTITESTATA
                    LEFT JOIN ANAGRAF a
                        ON a.ID_ANAGRAF = ct.COMMITTENTE
                    LEFT JOIN ANAGRAF ag
                        ON ag.ID_ANAGRAF = ct.AGENZIA
                    WHERE tsl.BlackList > 0
                      AND ISNULL(tsl.Date,   cr.DATA_INIZIO) <= %s
                      AND ISNULL(tsl.ToDate, cr.DATA_FINE)   >= %s
                    GROUP BY
                        ct.ID_CONTRATTITESTATA, ct.COD_CONTRATTO,
                        ct.DESCRIZIONE, ct.P_AGENZIA, a.RAG_SOCIAL, ag.RAG_SOCIAL,
                        cr.ID_CONTRATTIRIGHE, cr.COD_USER, cr.DESCRIZIONE,
                        cr.DATA_INIZIO, cr.DATA_FINE, cr.N_PASSAGGI,
                        cr.IMPORTO, cr.ID_BOOKINGCODE, cr.ORA_INIZIO, cr.ORA_FINE
                    ORDER BY a.RAG_SOCIAL, ct.COD_CONTRATTO, cr.COD_USER, cr.DATA_INIZIO
                """, [date_to, date_from])
                rows = cursor.fetchall()

        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

        _mn    = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}
        _fps   = 29.97

        def _fr2hm(frames):
            if not frames:
                return "00:00"
            total_sec = round(frames / _fps)
            return f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"

        contracts: dict = {}
        for (ct_id, code, ct_desc, client_name, agency_name, p_agenzia, line_id, market_id, line_desc,
             d_start, d_end, ordered, importo, id_bookingcode,
             ora_inizio, ora_fine, missed) in rows:
            if ct_id not in contracts:
                contracts[ct_id] = {
                    "contract_id":  ct_id,
                    "code":         code or "",
                    "description":  ct_desc or "",
                    "client":       client_name or "",
                    "agency":       agency_name or "",
                    "total_missed": 0,
                    "lines":        [],
                }
            m = missed or 0
            agency_pct  = float(p_agenzia or 0)
            gross_rate  = float(importo or 0)
            net_rate    = round(gross_rate * (1 - agency_pct / 100), 2)
            spot_type   = "BNS" if id_bookingcode == 10 else "Paid"
            contracts[ct_id]["total_missed"] += m
            contracts[ct_id]["lines"].append({
                "line_id":     line_id,
                "market":      _mn.get(market_id, str(market_id) if market_id else "—"),
                "description": line_desc or "",
                "date_start":  d_start or "",
                "date_end":    d_end or "",
                "ordered":     ordered or 0,
                "missed":      m,
                "spot_type":   spot_type,
                "gross_rate":  gross_rate,
                "net_rate":    net_rate,
                "time_from":   _fr2hm(ora_inizio),
                "time_to":     _fr2hm(ora_fine),
            })

        contract_list = list(contracts.values())
        total_missed  = sum(c["total_missed"] for c in contract_list)

        return JSONResponse({
            "contracts":       contract_list,
            "total_contracts": len(contract_list),
            "total_missed":    total_missed,
        })

    @router.post("/api/orders/make-goods/apply")
    async def apply_make_good(body: dict = Body(...)):
        line_id       = int(body["line_id"])
        spots         = int(body["spots"])
        date_from_str = body["date_from"]   # MM/DD/YYYY
        date_to_str   = body["date_to"]
        time_from_str = body["time_from"]   # HH:MM
        time_to_str   = body["time_to"]     # HH:MM

        from datetime import datetime as _dt

        from browser_automation.etere_direct_client import (
            EtereDirectClient,
        )
        from browser_automation.etere_direct_client import (
            connect as _db_connect,
        )

        _mn_rev = {1:"NYC",2:"CMP",3:"HOU",4:"SFO",5:"SEA",6:"LAX",7:"CVC",8:"WDC",9:"MMT",10:"DAL"}
        _fps    = 29.97

        try:
            with _db_connect() as conn:
                cur = conn.cursor(as_dict=True)

                # Load original line
                cur.execute("""
                    SELECT cr.*, ct.ID_CONTRATTITESTATA AS contract_id
                    FROM CONTRATTIRIGHE cr
                    JOIN CONTRATTITESTATA ct ON ct.ID_CONTRATTITESTATA = cr.ID_CONTRATTITESTATA
                    WHERE cr.ID_CONTRATTIRIGHE = %s
                """, [line_id])
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail=f"Line {line_id} not found")

                # Reconstruct day string from Italian weekday flags
                day_parts = []
                for flag, code in [("LUNEDI","M"),("MARTEDI","Tu"),("MERCOLEDI","W"),
                                    ("GIOVEDI","Th"),("VENERDI","F"),("SABATO","Sa"),("DOMENICA","Su")]:
                    if row.get(flag):
                        day_parts.append(code)
                days = ",".join(day_parts) or "M-Su"

                # Duration frames → "HH:MM:SS:FF"
                dur_sec = round(row["DURATA"] / _fps)
                duration_str = f"00:00:{dur_sec:02d}:00"

                # Separation frames → minutes
                sep_cust = round(row["Interv_Committente"] / (60 * _fps))
                sep_ord  = round(row["INTERVALLO"]         / (60 * _fps))
                sep_evt  = round(row["INTERV_CONTRATTO"]   / (60 * _fps))

                # Parse user-supplied dates
                df = _dt.strptime(date_from_str, "%m/%d/%Y").date()
                dt = _dt.strptime(date_to_str,   "%m/%d/%Y").date()

                market      = _mn_rev.get(row["COD_USER"], "NYC")
                contract_id = row["contract_id"]

                # Insert make-good line (row_status=2 → Change Data, requires approval before scheduling)
                client = EtereDirectClient(conn, autocommit=False)
                client._contract_id = contract_id
                new_line_id = client.add_contract_line(
                    market             = market,
                    days               = days,
                    time_range         = f"{time_from_str}-{time_to_str}",
                    description        = f"-MG- {(row['DESCRIZIONE'] or '').strip()}",
                    rate               = float(row["IMPORTO"] or 0),
                    total_spots        = spots,
                    spots_per_week     = 0,
                    max_daily_run      = int(row["PASSAGGI_GIORNALIERI"] or 1),
                    date_from          = df,
                    date_to            = dt,
                    duration           = duration_str,
                    is_bonus           = (row["ID_BOOKINGCODE"] == 10),
                    separation_intervals = (sep_cust, sep_ord, sep_evt),
                    contract_id        = contract_id,
                    priority           = int(row["PRIORITA"] or 500),
                    whitelist_priority = int(row["PrioritaWhiteList"] or 50),
                    booking_code       = int(row["ID_BOOKINGCODE"] or 2),
                    scheduling_type    = int(row["PRENOTAZIONE"] or 1),
                    row_status         = 2,
                )

                # Refresh blocks for the new line
                client.assign_blocks_for_existing_line(new_line_id)

                # Decrement N_PASSAGGI on original line
                cur.execute(
                    "UPDATE CONTRATTIRIGHE SET N_PASSAGGI = N_PASSAGGI - %s WHERE ID_CONTRATTIRIGHE = %s",
                    [spots, line_id]
                )

                # Update TSL PassageMiss (delete row if it reaches 0)
                cur.execute(
                    "SELECT PassageMiss FROM Traffic_ScheduleList WHERE ID_ContrattiRighe=%s AND BlackList>0",
                    [line_id]
                )
                tsl = cur.fetchone()
                if tsl:
                    new_miss = (tsl["PassageMiss"] or 0) - spots
                    if new_miss <= 0:
                        cur.execute(
                            "DELETE FROM Traffic_ScheduleList WHERE ID_ContrattiRighe=%s AND BlackList>0",
                            [line_id]
                        )
                    else:
                        cur.execute(
                            "UPDATE Traffic_ScheduleList SET PassageMiss=%s WHERE ID_ContrattiRighe=%s AND BlackList>0",
                            [new_miss, line_id]
                        )

                conn.commit()
                return JSONResponse({"ok": True, "new_line_id": new_line_id, "spots": spots})

        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    return router
