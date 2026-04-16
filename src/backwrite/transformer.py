"""
Backwrite transformer: Etere placement confirmation CSV → three-tab Excel

Tabs produced:
  Sales Confirmation  — one row per contract line (grouped)
  Run Sheet           — one row per spot that aired
  Sheet1              — monthly pivot (Gross Rate / Station Net)

CSV format expected (Etere export):
  Row 1:  column metadata labels (ignored)
  Row 2:  header values (agency, contract_code, date, description, address, client, city)
  Row 3:  blank
  Row 4:  data column headers (COD_CONTRATTO1, committente, ...)
  Row 5+: one row per spot
"""

import calendar
import copy
import csv
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

REVENUE_TYPES = [
    "Branded Content",
    "Direct Response Sales",
    "Internal Ad Sales",
    "Paid Programming",
    "Trade",
]

_LANG_KEYWORDS: List[Tuple[str, str]] = [
    ("cantonese",   "C"),
    ("mandarin",    "M"),
    ("chinese",     "M"),
    ("south asian", "SA"),
    ("hindi",       "SA"),
    ("punjabi",     "SA"),
    ("filipino",    "T"),
    ("tagalog",     "T"),
    ("vietnamese",  "V"),
    ("hmong",       "Hm"),
    ("korean",      "K"),
    ("japanese",    "J"),
]


# ─────────────────────────────────────────────────────────────────────────────
# DATA CLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CsvHeader:
    agency:        str
    client:        str
    contract_code: str
    description:   str
    order_date:    str
    address:       str
    city:          str


@dataclass
class SpotRow:
    contract_code:   str
    client:          str
    line_id:         int
    priority:        int
    duration_s:      int
    flight_start:    date
    time_from:       str   # "HH:MM"
    time_to:         str   # "HH:MM"
    gross_rate:      float
    days_pattern:    str
    market:          str
    air_date:        date
    air_time:        str   # "HH:MM:SS"
    copy_code:       str
    row_description: str


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def detect_language(text: str) -> str:
    t = text.lower()
    for keyword, code in _LANG_KEYWORDS:
        if keyword in t:
            return code
    return "E"


def _strip_line_prefix(desc: str) -> str:
    """Remove '(Line N) ' prefix from rowdescription."""
    return re.sub(r'^\(Line \d+\)\s*', '', desc)


def _parse_date(s: str) -> date:
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Cannot parse date: {s!r}")


def _hhmm_to_timedelta(s: str) -> timedelta:
    """Convert 'HH:MM' or 'HH:MM:SS' string to timedelta for Excel time serial."""
    try:
        parts = s.strip().split(":")
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        sec = int(parts[2]) if len(parts) > 2 else 0
        return timedelta(hours=h, minutes=m, seconds=sec)
    except (ValueError, IndexError):
        return timedelta(0)


def compute_broadcast_month(d: date) -> date:
    """Return the first of the broadcast month containing date d.

    Broadcast month = month of the Sunday that ends the week containing d.
    Broadcast weeks run Monday–Sunday.
    """
    days_until_sunday = (6 - d.weekday()) % 7
    next_sunday = d + timedelta(days=days_until_sunday)
    year = d.year
    if d.month == 12 and next_sunday.month == 1:
        year += 1
    return date(year, next_sunday.month, 1)


# ─────────────────────────────────────────────────────────────────────────────
# CSV PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_csv(data: bytes) -> Tuple[CsvHeader, List[SpotRow]]:
    """Parse an Etere placement confirmation CSV.

    Returns (CsvHeader, list of SpotRow).
    Raises ValueError if the data header row cannot be found.
    """
    text = data.decode("utf-8-sig")
    lines = text.splitlines()

    # ── Header values from row 2 ──────────────────────────────────────────
    h_parts: List[str] = []
    if len(lines) >= 2:
        reader = csv.reader([lines[1]])
        h_parts = next(reader, [])

    csv_header = CsvHeader(
        agency        = h_parts[0].strip() if len(h_parts) > 0 else "",
        contract_code = h_parts[1].strip() if len(h_parts) > 1 else "",
        order_date    = h_parts[2].strip() if len(h_parts) > 2 else "",
        description   = h_parts[3].strip() if len(h_parts) > 3 else "",
        address       = h_parts[4].strip() if len(h_parts) > 4 else "",
        client        = h_parts[5].strip() if len(h_parts) > 5 else "",
        city          = h_parts[6].strip() if len(h_parts) > 6 else "",
    )

    # ── Find data column header row ───────────────────────────────────────
    data_start: Optional[int] = None
    for i, line in enumerate(lines):
        if "COD_CONTRATTO1" in line or "dateschedule" in line.lower():
            data_start = i
            break

    if data_start is None:
        return csv_header, []

    # ── Parse data rows ───────────────────────────────────────────────────
    data_text = "\n".join(lines[data_start:])
    reader = csv.DictReader(data_text.splitlines())

    spots: List[SpotRow] = []
    for row in reader:
        air_date_str = (row.get("dateschedule") or "").strip()
        if not air_date_str:
            continue

        try:
            air_date = _parse_date(air_date_str)
        except ValueError:
            continue

        datestart = row.get("DATESTART2", "").strip()
        try:
            flight_start = _parse_date(datestart) if datestart else air_date
        except ValueError:
            flight_start = air_date

        # Time range "13:00-14:00"
        timerange = row.get("timerange2", "").strip()
        time_parts = timerange.split("-", 1)
        time_from = time_parts[0].strip() if time_parts else ""
        time_to   = time_parts[1].strip() if len(time_parts) > 1 else ""

        try:
            gross_rate = float(row.get("IMPORTO2", 0) or 0)
        except (ValueError, TypeError):
            gross_rate = 0.0

        try:
            duration_s = int(row.get("duration3", 30) or 30)
        except (ValueError, TypeError):
            duration_s = 30

        try:
            line_id = int(row.get("id_contrattirighe", 0) or 0)
        except (ValueError, TypeError):
            line_id = 0

        try:
            priority = int(row.get("Textbox14", 4) or 4)
        except (ValueError, TypeError):
            priority = 4

        spots.append(SpotRow(
            contract_code   = row.get("COD_CONTRATTO1", "").strip(),
            client          = row.get("committente",    "").strip(),
            line_id         = line_id,
            priority        = priority,
            duration_s      = duration_s,
            flight_start    = flight_start,
            time_from       = time_from,
            time_to         = time_to,
            gross_rate      = gross_rate,
            days_pattern    = row.get("Textbox25", "").strip(),
            market          = row.get("nome2",       "").strip(),
            air_date        = air_date,
            air_time        = row.get("airtimep",    "").strip(),
            copy_code       = row.get("bookingcode2","").strip(),
            row_description = row.get("rowdescription","").strip(),
        ))

    return csv_header, spots


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────────────────────────────────────

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# PLACEHOLDER HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _phone_to_int(val: str) -> object:
    """Strip non-digits and return int so Excel phone-number formats apply."""
    if not val:
        return ""
    digits = re.sub(r'\D', '', str(val))
    if len(digits) >= 7:
        try:
            return int(digits)
        except ValueError:
            pass
    return val


def _replace_placeholder(cell, ctx: dict) -> bool:
    """Replace <field> in a single cell using ctx. Returns True if modified.

    If cell value is exactly '<field>' → replace with the typed Python value.
    If '<field>' is embedded in a larger string → replace with str(value).
    """
    if not isinstance(cell.value, str):
        return False
    text = cell.value
    keys = re.findall(r'<([^>]+)>', text)
    if not keys:
        return False
    stripped = text.strip()
    if len(keys) == 1 and stripped == f'<{keys[0]}>' and keys[0] in ctx:
        cell.value = ctx[keys[0]]
        return True
    changed = False
    for k in keys:
        if k in ctx:
            v = ctx[k]
            if isinstance(v, (date, datetime)):
                v = v.strftime('%m/%d/%Y')
            elif v is None:
                v = ''
            text = text.replace(f'<{k}>', str(v))
            changed = True
    if changed:
        cell.value = text.strip()
    return changed


def _copy_row_format(ws, src_row: int, dst_row: int) -> None:
    """Copy cell styles from src_row to dst_row."""
    src = list(ws.iter_rows(min_row=src_row, max_row=src_row))[0]
    dst = list(ws.iter_rows(min_row=dst_row, max_row=dst_row))[0]
    for s, d in zip(src, dst):
        if s.has_style:
            d.font         = copy.copy(s.font)
            d.border       = copy.copy(s.border)
            d.fill         = copy.copy(s.fill)
            d.number_format = s.number_format
            d.alignment    = copy.copy(s.alignment)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET FILLERS
# ─────────────────────────────────────────────────────────────────────────────

def _snapshot_row(ws, row_num: int) -> List[dict]:
    """Capture all cells in a row into a list of dicts before any insertions."""
    max_col = ws.max_column
    snaps = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        snaps.append({
            'col':           col_idx,
            'value':         cell.value,
            'has_style':     cell.has_style,
            'font':          copy.copy(cell.font)      if cell.has_style else None,
            'border':        copy.copy(cell.border)    if cell.has_style else None,
            'fill':          copy.copy(cell.fill)      if cell.has_style else None,
            'number_format': cell.number_format        if cell.has_style else None,
            'alignment':     copy.copy(cell.alignment) if cell.has_style else None,
        })

    # Capture intra-row merged ranges (e.g. B16:G16) so inserted rows stay consistent
    row_merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == mr.max_row == row_num:
            row_merges.append((mr.min_col, mr.max_col))

    return snaps, row_merges


def _apply_snapshot(
    ws, snapshot_pair, new_row: int, ref_row: int
) -> None:
    """Write a row snapshot to new_row, updating formula row references."""
    snapshot, row_merges = snapshot_pair
    for snap in snapshot:
        col_idx  = snap['col']
        val      = snap['value']
        dst_cell = ws.cell(row=new_row, column=col_idx)
        if isinstance(val, str) and val.startswith('='):
            val = re.sub(
                r'([A-Z]+)' + str(ref_row),
                lambda m, nr=new_row: f'{m.group(1)}{nr}',
                val,
            )
        dst_cell.value = val
        if snap['has_style']:
            dst_cell.font          = copy.copy(snap['font'])
            dst_cell.border        = copy.copy(snap['border'])
            dst_cell.fill          = copy.copy(snap['fill'])
            dst_cell.number_format = snap['number_format']
            dst_cell.alignment     = copy.copy(snap['alignment'])

    # Re-apply same intra-row merges to the new row
    for col_start, col_end in row_merges:
        ws.merge_cells(
            start_row=new_row, start_column=col_start,
            end_row=new_row,   end_column=col_end,
        )


def _fill_monthly_breakdown(
    ws, monthly_gross: dict, monthly_net: dict, last_line_row: int
) -> None:
    """Replace the hardcoded monthly breakdown table with actual monthly totals."""
    # Find "Month" header row below the data lines
    month_hdr_row: Optional[int] = None
    for row in ws.iter_rows(min_row=last_line_row + 1):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == 'month':
                month_hdr_row = row[0].row
                break
        if month_hdr_row:
            break
    if month_hdr_row is None:
        return

    # Determine column positions from header row
    month_col = gross_col = net_col = None
    for cell in ws[month_hdr_row]:
        v = str(cell.value or '').strip().lower()
        if v == 'month':
            month_col = cell.column
        elif v == 'gross' and month_col is not None and gross_col is None:
            gross_col = cell.column
        elif v == 'net' and gross_col is not None and net_col is None:
            net_col = cell.column

    if month_col is None:
        return

    # Find existing data rows and the Total row
    data_rows: List[int] = []
    total_row: Optional[int] = None
    for row in ws.iter_rows(min_row=month_hdr_row + 1):
        cell = ws.cell(row=row[0].row, column=month_col)
        if isinstance(cell.value, str) and 'total' in cell.value.lower():
            total_row = row[0].row
            break
        data_rows.append(row[0].row)

    if not data_rows:
        return

    # Sort months in calendar order
    month_order = {m: i for i, m in enumerate(calendar.month_name) if m}
    sorted_months = sorted(monthly_gross.keys(), key=lambda m: month_order.get(m, 99))
    n_needed   = len(sorted_months)
    n_existing = len(data_rows)
    ref_snap   = _snapshot_row(ws, data_rows[0])

    # Expand rows if needed
    if n_needed > n_existing:
        for _ in range(n_needed - n_existing):
            new_r = data_rows[-1] + 1
            ws.insert_rows(new_r)
            _apply_snapshot(ws, ref_snap, new_r, data_rows[0])
            data_rows.append(new_r)
            if total_row:
                total_row += 1

    # Fill month rows
    for i, month_name_str in enumerate(sorted_months):
        rn = data_rows[i]
        if month_col:
            ws.cell(row=rn, column=month_col).value = month_name_str
        if gross_col:
            ws.cell(row=rn, column=gross_col).value = round(monthly_gross[month_name_str], 2)
        if net_col:
            ws.cell(row=rn, column=net_col).value = round(monthly_net.get(month_name_str, 0), 2)

    # Delete extra template rows (shrink to exactly n_needed data rows)
    extra_rows = data_rows[n_needed:]
    for rn in reversed(extra_rows):
        ws.delete_rows(rn)
        if total_row and total_row > rn:
            total_row -= 1

    # Update Total row
    if total_row and n_needed > 0:
        if gross_col:
            ws.cell(row=total_row, column=gross_col).value = round(sum(monthly_gross.values()), 2)
        if net_col:
            ws.cell(row=total_row, column=net_col).value = round(sum(monthly_net.values()), 2)


def _fill_sales_confirmation(
    ws, ctx: dict, sc_lines: List[dict],
    monthly_gross: dict, monthly_net: dict,
) -> None:
    """Fill the Sales Confirmation sheet using placeholder replacement."""
    # Detect line template rows: any row that contains '<date_range_start>'
    line_rows: List[int] = []
    for row in ws.iter_rows():
        if any(isinstance(c.value, str) and '<date_range_start>' in c.value for c in row):
            line_rows.append(row[0].row)

    n_tmpl  = len(line_rows)
    n_lines = len(sc_lines)
    max_col = ws.max_column

    # Detect weeks column from header row above the first line row
    weeks_col: Optional[int] = None
    if line_rows:
        hdr_row_num = line_rows[0] - 1
        for cell in ws[hdr_row_num]:
            if isinstance(cell.value, str) and 'days' in cell.value.lower():
                weeks_col = cell.column
                break

    # Snapshot first template row BEFORE any insertions (avoids read-after-insert issues)
    ref_row  = line_rows[0] if line_rows else None
    snapshot = _snapshot_row(ws, ref_row) if ref_row else ([], [])

    # Expand: insert rows when we need more lines than template provides
    if n_lines > n_tmpl and n_tmpl > 0:
        insert_start = line_rows[-1] + 1
        n_inserts    = n_lines - n_tmpl

        # Explicitly manage merges that sit at or after the insertion zone.
        # openpyxl's auto-shift is unreliable near merged areas, so we
        # unmerge them first, insert rows, then re-merge at the correct rows.
        saved_merges = [
            (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
            for mr in list(ws.merged_cells.ranges)
            if mr.min_row >= insert_start
        ]
        for min_r, max_r, min_c, max_c in saved_merges:
            ws.unmerge_cells(start_row=min_r, start_column=min_c,
                             end_row=max_r,   end_column=max_c)

        for _ in range(n_inserts):
            new_row = line_rows[-1] + 1
            ws.insert_rows(new_row)
            _apply_snapshot(ws, snapshot, new_row, ref_row)
            line_rows.append(new_row)

        # Re-merge at shifted positions
        for min_r, max_r, min_c, max_c in saved_merges:
            ws.merge_cells(start_row=min_r + n_inserts, start_column=min_c,
                           end_row=max_r + n_inserts,   end_column=max_c)

    # Update SUM formulas that span the line range
    if line_rows:
        first_line   = line_rows[0]
        last_line    = line_rows[-1]
        line_row_set = set(line_rows)
        for row in ws.iter_rows():
            if row[0].row in line_row_set:
                continue
            for cell in row:
                if isinstance(cell.value, str) and '=SUM(' in cell.value:
                    cell.value = re.sub(
                        r'([A-Z]+)\d+:([A-Z]+)\d+',
                        lambda m: f'{m.group(1)}{first_line}:{m.group(2)}{last_line}',
                        cell.value,
                    )

    line_row_set = set(line_rows)

    # Fill line rows — use explicit ws.cell() to avoid row-iteration truncation
    for i, row_num in enumerate(line_rows):
        if i < n_lines:
            line_ctx = {**ctx, **sc_lines[i]}
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                _replace_placeholder(cell, line_ctx)
                if weeks_col is not None and col_idx == weeks_col:
                    cell.value = sc_lines[i].get('weeks', 1)
            # Sequential line number in column B (handles hardcoded "1" in template)
            ws.cell(row=row_num, column=2).value = i + 1
        else:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell.value, str) and '<' in cell.value:
                    cell.value = None

    # Fill all non-line rows with single-value context
    for row in ws.iter_rows():
        if row[0].row in line_row_set:
            continue
        for cell in row:
            _replace_placeholder(cell, ctx)

    # Fill the monthly breakdown section
    last_line_row = line_rows[-1] if line_rows else 20
    _fill_monthly_breakdown(ws, monthly_gross, monthly_net, last_line_row)


def _fill_run_sheet(ws, run_rows: List[dict]) -> None:
    """Fill the Run Sheet with one row per spot using the template."""
    # Find template rows (rows 2+ that contain any <placeholder>)
    tmpl_rows: List[int] = []
    for row in ws.iter_rows(min_row=2):
        if any(isinstance(c.value, str) and '<' in c.value for c in row):
            tmpl_rows.append(row[0].row)

    if not tmpl_rows:
        return

    ref_row  = tmpl_rows[0]
    max_col  = ws.max_column
    snapshot = _snapshot_row(ws, ref_row)
    snaps, _ = snapshot

    # Build col → field map from the first template row
    col_map: dict = {}
    for snap in snaps:
        val = snap['value']
        if isinstance(val, str):
            m = re.search(r'<([^>]+)>', val)
            if m:
                col_map[snap['col']] = m.group(1)
            elif re.match(r'=([A-Z]+)\d+$', val.strip()):
                col_map[snap['col']] = ('formula', val)

    n_tmpl  = len(tmpl_rows)
    n_spots = len(run_rows)

    # Expand using snapshot
    if n_spots > n_tmpl:
        for _ in range(n_spots - n_tmpl):
            new_row = tmpl_rows[-1] + 1
            ws.insert_rows(new_row)
            _apply_snapshot(ws, snapshot, new_row, ref_row)
            tmpl_rows.append(new_row)

    # Fill data rows
    for spot_idx, rr in enumerate(run_rows):
        if spot_idx >= len(tmpl_rows):
            break
        row_num = tmpl_rows[spot_idx]
        for col_idx, spec in col_map.items():
            cell = ws.cell(row=row_num, column=col_idx)
            if isinstance(spec, tuple) and spec[0] == 'formula':
                cell.value = re.sub(
                    r'([A-Z]+)\d+',
                    lambda m, nr=row_num: f'{m.group(1)}{nr}',
                    spec[1],
                )
            elif spec in rr:
                cell.value = rr[spec]

    # Clear unused template rows
    for row_num in tmpl_rows[n_spots:]:
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_num, column=col_idx).value = None


def _fill_pivot(ws, run_rows: List[dict]) -> None:
    """Fill the Sheet1 pivot with monthly gross/net totals."""
    from collections import defaultdict
    monthly_gross: dict = defaultdict(float)
    monthly_net:   dict = defaultdict(float)
    for rr in run_rows:
        m = rr.get('month')
        if isinstance(m, datetime):
            key = m.strftime('%B %Y')
            monthly_gross[key] += rr.get('gross_rate', 0) or 0
            monthly_net[key]   += rr.get('station_net', 0) or 0

    months = sorted(monthly_gross)
    if not months:
        return

    # Find the <month> placeholder row
    month_row: Optional[int] = None
    for row in ws.iter_rows():
        if any(isinstance(c.value, str) and '<month>' in c.value for c in row):
            month_row = row[0].row
            break
    if month_row is None:
        return

    # Determine column positions from the placeholder row
    ph_cells  = list(ws.iter_rows(min_row=month_row, max_row=month_row))[0]
    month_col = gross_col = net_col = None
    for cell in ph_cells:
        if cell.value is None:
            continue
        if isinstance(cell.value, str) and '<month>' in cell.value:
            month_col = cell.column - 1
        elif month_col is not None and gross_col is None:
            gross_col = cell.column - 1
        elif gross_col is not None and net_col is None:
            net_col = cell.column - 1

    # Find the Grand Total row (first row after month_row with 'Grand' in month_col)
    grand_row: Optional[int] = None
    for row in ws.iter_rows(min_row=month_row + 1):
        cell = row[month_col] if month_col is not None else row[0]
        if isinstance(cell.value, str) and 'Grand' in cell.value:
            grand_row = cell.row
            break

    # Insert extra rows for additional months (copy format from month_row)
    extra = len(months) - 1
    for i in range(extra):
        ws.insert_rows(month_row + i + 1)
        _copy_row_format(ws, month_row, month_row + i + 1)

    # Fill month rows
    for i, month_name in enumerate(months):
        rn    = month_row + i
        cells = list(ws.iter_rows(min_row=rn, max_row=rn))[0]
        for cell in cells:
            col = cell.column - 1
            if col == month_col:
                cell.value = month_name
            elif col == gross_col:
                cell.value = round(monthly_gross[month_name], 2)
            elif col == net_col:
                cell.value = round(monthly_net[month_name], 2)

    # Update Grand Total row
    if grand_row is not None:
        grand_row += extra
        for cell in list(ws.iter_rows(min_row=grand_row, max_row=grand_row))[0]:
            col = cell.column - 1
            if col == gross_col:
                cell.value = round(sum(monthly_gross.values()), 2)
            elif col == net_col:
                cell.value = round(sum(monthly_net.values()), 2)


# ─────────────────────────────────────────────────────────────────────────────
# EXISTING ORDER READER
# ─────────────────────────────────────────────────────────────────────────────

def read_existing_order_fields(data: bytes) -> dict:
    """Extract header field values from an existing Sales Confirmation Excel.

    Returns a dict of field names → string values.  Unknown fields are "".
    Raises ValueError if no 'Sales Confirmation' tab is found.
    """
    wb = openpyxl.load_workbook(BytesIO(data))
    if "Sales Confirmation" not in wb.sheetnames:
        raise ValueError("No 'Sales Confirmation' tab found in uploaded file.")
    ws = wb["Sales Confirmation"]

    def cv(row: int, col: int) -> str:
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None else ""

    # Header section rows 3–13 at fixed column positions (these rows never expand)
    # D=4, F=6, G=7, K=11, L=12
    return {
        "agency":          cv(3,  4),
        "client":          cv(3, 12),
        "contact_person":  cv(4,  4),
        "estimate":        cv(4, 12),
        "address":         cv(5,  4),
        "billing_type":    cv(5, 12),
        "city":            cv(6,  4),
        "state":           cv(6,  6),
        "zip":             cv(6,  7),
        "market":          cv(6, 12),
        "phone":           cv(8,  4),
        "order_date":      cv(8, 12),
        "fax":             cv(9,  4),
        "contract":        cv(9, 12),
        "email_1":         cv(10, 4),
        "email_2":         cv(11, 4),
        "sales_person":    cv(11, 11),
        "email_3":         cv(12, 4),
        "email_4":         cv(13, 4),
    }


# ─────────────────────────────────────────────────────────────────────────────
# ETEREBRIDGE INTEGRATION
# ─────────────────────────────────────────────────────────────────────────────

def _eb_df_to_run_rows(df, agency_fee: float, is_agency: bool) -> List[dict]:
    """Convert an EtereBridge output DataFrame into the run_rows format
    expected by _fill_run_sheet().  Computes broker_fees/station_net/spot_value
    directly since EtereBridge leaves those as None (formula-driven in its own Excel)."""
    import math

    import pandas as pd

    def _clean(val):
        """Return None for pandas NA / float NaN, otherwise the value as-is."""
        if val is None:
            return None
        try:
            if isinstance(val, float) and math.isnan(val):
                return None
        except TypeError:
            pass
        if isinstance(val, float) and pd.isna(val):
            return None
        return val

    rows: List[dict] = []
    for _, row in df.iterrows():
        gross  = float(_clean(row.get("Gross Rate")) or 0)
        broker = round(gross * agency_fee, 2) if is_agency else 0.0
        net    = round(gross - broker, 2)

        air_dt = row.get("Air Date")
        if isinstance(air_dt, pd.Timestamp):
            air_date = air_dt.date()
        elif isinstance(air_dt, datetime):
            air_date = air_dt.date()
        elif isinstance(air_dt, date):
            air_date = air_dt
        else:
            air_date = None

        month_val = row.get("Month")
        if isinstance(month_val, pd.Timestamp):
            month_dt = datetime(month_val.year, month_val.month, month_val.day)
        elif isinstance(month_val, datetime):
            month_dt = month_val
        elif isinstance(month_val, date):
            month_dt = datetime(month_val.year, month_val.month, month_val.day)
        else:
            month_dt = None

        rows.append({
            "bill_code":    _clean(row.get("Bill Code")),
            "air_date":     air_date,
            "day":          air_date.strftime("%A") if air_date else "",
            "time_in":      _hhmm_to_timedelta(str(_clean(row.get("Time In")) or "")),
            "time_out":     _hhmm_to_timedelta(str(_clean(row.get("Time Out")) or "")),
            "length":       timedelta(seconds=int(_clean(row.get("Length")) or 0)),
            "media":        _clean(row.get("Media")),
            "program":      _clean(row.get("Program")),
            "lang":         _clean(row.get("Lang.")),
            "line":         _clean(row.get("Line")),
            "type":         _clean(row.get("Type")),
            "estimate":     _clean(row.get("Estimate")),
            "gross_rate":   gross,
            "spot_value":   gross,
            "month":        month_dt,
            "broker_fees":  broker,
            "priority":     4,
            "station_net":  net,
            "sales_person": _clean(row.get("Sales Person")),
            "revenue_type": _clean(row.get("Revenue Type")),
            "billing_type": _clean(row.get("Billing Type")),
            "agency_flag":  _clean(row.get("Agency?")),
            "affidavit":    _clean(row.get("Affidavit?")),
            "contract":     _clean(row.get("Contract")),
            "market":       _clean(row.get("Market")),
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(header: CsvHeader, spots: List[SpotRow], user_inputs: dict, raw_csv: bytes = b"") -> bytes:
    """Generate backwrite Excel from template and return raw bytes."""
    billing_type = user_inputs.get("billing_type", "Broadcast")
    agency_flag  = user_inputs.get("agency_flag",  "Agency")
    agency_fee   = float(user_inputs.get("agency_fee", 0.15) or 0)
    sales_person = user_inputs.get("sales_person", "")
    revenue_type = user_inputs.get("revenue_type", "Internal Ad Sales")
    affidavit    = user_inputs.get("affidavit",    "Y")
    estimate     = user_inputs.get("estimate",     "")
    contract     = user_inputs.get("contract",     "")

    is_agency = agency_flag == "Agency"
    bill_code = f"{header.agency}:{header.client}" if header.agency else header.client

    # ── Per-spot run rows: try EtereBridge pipeline first ────────────────────
    run_rows: List[dict] = []
    if raw_csv:
        try:
            from .eterebridge_runner import run_eterebridge_pipeline
            eb_df = run_eterebridge_pipeline(raw_csv, user_inputs)
            if eb_df is not None and not eb_df.empty:
                run_rows = _eb_df_to_run_rows(eb_df, agency_fee, is_agency)
        except Exception as _eb_exc:
            print(f"[EtereBridge] Falling back to built-in pipeline: {_eb_exc}")

    if not run_rows:
        for s in spots:
            if billing_type == "Broadcast":
                month_date = compute_broadcast_month(s.air_date)
            else:
                month_date = date(s.air_date.year, s.air_date.month, 1)
            month_dt    = datetime(month_date.year, month_date.month, month_date.day)
            broker_fees = round(s.gross_rate * agency_fee, 2) if is_agency else 0.0
            station_net = round(s.gross_rate - broker_fees, 2)
            run_rows.append({
                "bill_code":    bill_code,
                "air_date":     s.air_date,
                "day":          s.air_date.strftime("%A"),
                "time_in":      _hhmm_to_timedelta(s.time_from),
                "time_out":     _hhmm_to_timedelta(s.time_to),
                "length":       timedelta(seconds=s.duration_s),
                "media":        s.copy_code,
                "program":      s.air_time,
                "lang":         detect_language(s.row_description),
                "line":         s.line_id,
                "type":         "BNS" if s.gross_rate == 0 else "COM",
                "estimate":     estimate,
                "gross_rate":   s.gross_rate,
                "spot_value":   s.gross_rate,
                "month":        month_dt,
                "broker_fees":  broker_fees,
                "priority":     4,
                "station_net":  station_net,
                "sales_person": sales_person,
                "revenue_type": revenue_type,
                "billing_type": billing_type,
                "agency_flag":  agency_flag,
                "affidavit":    affidavit,
                "contract":     contract,
                "market":       s.market,
            })

    # ── Sales Confirmation lines (grouped by line_id) ─────────────────────────
    groups: Dict[int, List[SpotRow]] = OrderedDict()
    for s in spots:
        groups.setdefault(s.line_id, []).append(s)

    sc_lines: List[dict] = []
    for idx, (line_id, group) in enumerate(groups.items()):
        d_start     = min(s.air_date for s in group)
        d_end       = max(s.air_date for s in group)
        total_spots = len(group)
        weeks       = max(1, round((d_end - d_start).days / 7) + 1)
        spw         = max(1, round(total_spots / weeks))
        first       = group[0]
        sc_lines.append({
            "line_number":      idx + 1,
            "date_range_start": d_start.strftime("%m/%d/%Y"),
            "date_range_end":   d_end.strftime("%m/%d/%Y"),
            "spot_count":       spw,         # spots per week → F column
            "per":              "Wk",
            "weeks":            weeks,        # # of weeks   → I column (computed, not a placeholder)
            "line_description": _strip_line_prefix(first.row_description),
            "type":             "BNS" if first.gross_rate == 0 else "COM",
            "length":           f":{first.duration_s}",
            "gross_rate":       first.gross_rate,
        })

    # ── Single-value context ──────────────────────────────────────────────────
    total_gross = sum(s.gross_rate for s in spots) if spots else 0.0
    total_net   = round(total_gross * (1 - agency_fee), 2) if is_agency else total_gross
    markets     = sorted(set(s.market for s in spots))
    d_start_all = min(s.air_date for s in spots) if spots else date.today()
    d_end_all   = max(s.air_date for s in spots) if spots else date.today()

    ctx: dict = {
        "agency":           header.agency,
        "client":           header.client,
        "contract_code":    header.contract_code,
        "description":      header.description,
        "estimate":         estimate,
        "address":          user_inputs.get("address") or header.address,
        "billing_type":     billing_type,
        "city":             user_inputs.get("city") or header.city,
        "order_date":       user_inputs.get("order_date") or f"{date.today().month}/{date.today().day}/{date.today().year}",
        "contract":         contract,
        "sales_person":     sales_person,
        "revenue_type":     revenue_type,
        "agency_flag":      agency_flag,
        "affidavit":        affidavit,
        "date_range_start": d_start_all.strftime("%m/%d/%Y"),
        "date_range_end":   d_end_all.strftime("%m/%d/%Y"),
        "spot_count":       len(spots),
        "total_gross":      total_gross,
        "total_net":        total_net,
        "market":           ", ".join(markets),
        "bill_code":        bill_code,
        # Contact / address fields supplied by user (or pre-filled from existing order)
        "contact_person":   user_inputs.get("contact_person", ""),
        "phone":            _phone_to_int(user_inputs.get("phone", "")),
        "fax":              _phone_to_int(user_inputs.get("fax", "")),
        "email_1":          user_inputs.get("email_1", ""),
        "email_2":          user_inputs.get("email_2", ""),
        "email_3":          user_inputs.get("email_3", ""),
        "email_4":          user_inputs.get("email_4", ""),
        "state":            user_inputs.get("state", ""),
        "zip":              user_inputs.get("zip", ""),
        "notes":            user_inputs.get("notes", ""),
        "per":              "Wk",
    }

    # ── Load template and fill ────────────────────────────────────────────────
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    # Clear hyperlinks (e.g. mailto: links in email cells) so placeholders are visible
    for ws_name in wb.sheetnames:
        for row in wb[ws_name].iter_rows():
            for cell in row:
                if cell.hyperlink:
                    cell.hyperlink = None
                if isinstance(cell.value, str) and cell.value.lower().startswith('mailto:'):
                    cell.value = ""
    # Monthly totals for Sales Confirmation breakdown
    _mg: dict = defaultdict(float)
    for rr in run_rows:
        m = rr.get('month')
        if isinstance(m, datetime):
            _mg[m.strftime('%B')] += rr.get('gross_rate', 0) or 0
    monthly_gross = {k: round(v, 2) for k, v in _mg.items()}
    monthly_net   = {k: round(v * (1 - agency_fee), 2) for k, v in monthly_gross.items()}

    _fill_sales_confirmation(wb["Sales Confirmation"], ctx, sc_lines, monthly_gross, monthly_net)
    _fill_run_sheet(wb["Run Sheet"], run_rows)
    _fill_pivot(wb["Sheet1"], run_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
