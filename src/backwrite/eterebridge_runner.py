"""
EtereBridge integration module.

Runs EtereBridge's CSV processing pipeline (language detection, bill-code
generation, market standardisation, user-input stamping) and returns a
pandas DataFrame suitable for the Run Sheet tab.

Requires EtereBridge as a sibling of this repo (e.g. dev/ctv-orderentry and
dev/EtereBridge).  If EtereBridge is unavailable, run_eterebridge_pipeline()
returns None and the caller falls back to the built-in transformer logic.
"""

import csv
import io
import logging
import os
import sys
import tempfile
from copy import copy
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# EtereBridge lives next to this repo 
_repo_root = Path(__file__).resolve().parents[2]
_eb_path = None
for _name in ("EtereBridge", "eterebridge"):
    _candidate = _repo_root.parent / _name
    if _candidate.exists():
        _eb_path = _candidate
        break
_EB_DIR = str(_eb_path.resolve()) if _eb_path else ""

# Add EtereBridge source dir to path so its modules can be imported.
# config_manager.py resolves config.ini relative to its own __file__, so it
# works correctly regardless of our working directory.
if _EB_DIR and _EB_DIR not in sys.path:
    sys.path.insert(0, _EB_DIR)

try:
    from config_manager import config_manager as _eb_config  # type: ignore[import]
    from file_processor import FileProcessor, transform_month_column  # type: ignore[import]
    from monetary_utils import (  # type: ignore[import]
        format_excel_monetary_columns,
        standardize_monetary_columns,
    )
    from time_utils import transform_times  # type: ignore[import]
    from utils import safe_convert_date  # type: ignore[import]

    _eb_app_config   = _eb_config.get_config()
    _file_processor  = FileProcessor(_eb_app_config)
    _AVAILABLE       = True
except Exception as _exc:
    _AVAILABLE = False
    logging.warning("[EtereBridge] Not available — will fall back to built-in pipeline: %s", _exc)


def is_available() -> bool:
    return _AVAILABLE


def _snap_duration(seconds) -> int:
    """Snap durations that are 1 second below a 5-second commercial boundary.

    Etere stores :25/:15/:30 spots with frame-count rounding that can land
    1 second short (e.g. 749 frames / 29.97fps = 24.99 → 25, but EtereBridge
    rounds 25 → 30 via nearest-15 logic).  We capture raw seconds before
    EtereBridge touches them and apply this correction instead.

    :14 → :15, :24 → :25, :29 → :30, :59 → :60, etc.
    """
    try:
        val = int(float(seconds) if seconds is not None else 0)
        return val + 1 if val % 5 == 4 else val
    except (ValueError, TypeError):
        return int(seconds) if seconds else 0


def _parse_raw_durations_from_csv(csv_bytes: bytes) -> list:
    """Return the raw duration3 values (in seconds) from every data row.

    Called before load_and_clean_data so we get the original values before
    EtereBridge's nearest-15 rounding overwrites them.
    """
    import io as _io_mod
    text = csv_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(_io_mod.StringIO(text))
    rows = list(reader)
    dur_col = None
    data_start = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if str(cell).strip().lower() == "duration3":
                dur_col = j
                data_start = i + 1
                break
        if dur_col is not None:
            break
    if dur_col is None or data_start is None:
        return []
    durations = []
    for row in rows[data_start:]:
        # Blank line separates data rows from the footer summary — stop here.
        if not any(str(cell).strip() for cell in row):
            break
        if len(row) > dur_col:
            try:
                durations.append(float(row[dur_col]))
            except (ValueError, TypeError):
                pass
    return durations


def get_language_options() -> list:
    """Return the list of valid language codes from EtereBridge config."""
    if not _AVAILABLE:
        return ["E", "C", "M", "V", "T", "K", "J", "SA", "Hm"]
    return list(_eb_app_config.language_options)


def get_language_details(csv_bytes: bytes) -> list:
    """
    Run language detection and return per-unique-description results.

    Returns [{"description": str, "lang": str, "count": int}, ...] sorted by
    (lang, description).  Returns [] if EtereBridge is unavailable.
    """
    if not _AVAILABLE:
        return []

    with tempfile.NamedTemporaryFile(
        suffix=".csv", prefix="eterebridge_lang_", delete=False
    ) as tmp:
        tmp.write(csv_bytes)
        tmp_path = tmp.name

    try:
        df = _file_processor.load_and_clean_data(tmp_path)
        df.columns = df.columns.str.strip()
        _detected_counts, row_languages = _file_processor.detect_languages(df)

        # rowdescription stays unrenamed by load_and_clean_data
        if "rowdescription" not in df.columns:
            logging.warning(
                "[EtereBridge] rowdescription column missing from DataFrame — "
                "available: %s", list(df.columns)
            )
            return []

        unique: dict = {}
        for idx, desc in df["rowdescription"].items():
            if not isinstance(desc, str):
                desc = str(desc) if desc is not None else ""
            lang = row_languages.get(idx, "E")
            if desc not in unique:
                unique[desc] = {"lang": lang, "count": 0}
            unique[desc]["count"] += 1

        return sorted(
            [
                {"description": d, "lang": info["lang"], "count": info["count"]}
                for d, info in unique.items()
            ],
            key=lambda x: (x["lang"], x["description"]),
        )
    except Exception as exc:
        logging.warning("[EtereBridge] Language details failed: %s", exc, exc_info=True)
        return []
    finally:
        os.unlink(tmp_path)


def run_eterebridge_pipeline(
    csv_bytes: bytes,
    user_inputs: dict,
) -> Optional[pd.DataFrame]:
    """
    Run EtereBridge's data processing pipeline on an Etere placement CSV.

    Returns a DataFrame with EtereBridge's 29-column output structure, or
    None if EtereBridge is unavailable or an error occurs.

    Expected user_inputs keys (same as the backwrite web form):
        sales_person, billing_type, revenue_type, agency_flag, agency_fee,
        estimate, contract, affidavit,
        gross_up_rates        (optional dict {str(rounded_gross): str(net_rate)})
        language_corrections  (optional dict {description: lang_code})
    """
    if not _AVAILABLE:
        return None

    # Write CSV to a temp file; EtereBridge's load_and_clean_data requires a path.
    with tempfile.NamedTemporaryFile(
        suffix=".csv", prefix="eterebridge_", delete=False
    ) as tmp:
        tmp.write(csv_bytes)
        tmp_path = tmp.name

    try:
        # 1. Extract bill-code parts from CSV row 2 (agency col 1, venue col 6)
        tb180, tb171 = _extract_header_values(tmp_path)

        # Capture raw duration3 values before EtereBridge rounds them to nearest 15.
        # EtereBridge maps :25 → :30 (round(25/15)*15), which is wrong for billboard spots.
        _raw_durations = _parse_raw_durations_from_csv(csv_bytes)

        # 2. Load and clean: skip 3 header rows, rename Etere columns
        df = _file_processor.load_and_clean_data(tmp_path)

        # Strip whitespace from any column names the rename left untouched
        # (Etere CSVs sometimes have leading/trailing spaces in headers)
        df.columns = df.columns.str.strip()
        if "Media" not in df.columns and "bookingcode2" in df.columns:
            df = df.rename(columns={"bookingcode2": "Media"})

        # 3. Language detection — auto, no stdin prompt
        _detected_counts, row_languages = _file_processor.detect_languages(df)

        # 3a. Apply user corrections over auto-detected languages
        language_corrections = user_inputs.get("language_corrections") or {}
        if language_corrections:
            applied = 0
            for idx, desc in df["rowdescription"].items():
                if isinstance(desc, str) and desc in language_corrections:
                    row_languages.at[idx] = language_corrections[desc]
                    applied += 1
            logging.info("[EtereBridge] Applied %d language correction(s): %s", applied, language_corrections)

        # 4. Transformations: bill code, market replacements, gross rate, length
        df = _file_processor.apply_transformations(df, tb180, tb171)
        df = standardize_monetary_columns(df)
        df = transform_times(df)

        # 5. Optional gross-up: replace rounded Etere rates with full-precision values
        agency_fee     = float(user_inputs.get("agency_fee") or 0.15)
        gross_up_rates = user_inputs.get("gross_up_rates") or {}
        if gross_up_rates and user_inputs.get("agency_flag") == "Agency" and (1 - agency_fee) > 0:
            rate_map = {
                float(k): float(v) / (1 - agency_fee)
                for k, v in gross_up_rates.items()
            }
            if rate_map:
                df["Gross Rate"] = df["Gross Rate"].apply(
                    lambda r: rate_map.get(float(r), r)
                )

        # 6. Stamp user inputs onto every row (skips the interactive verify_languages step)
        language_dict = row_languages.to_dict() if not row_languages.empty else {}
        df = _apply_user_inputs(
            df,
            billing_type = user_inputs.get("billing_type", "Broadcast"),
            revenue_type = user_inputs.get("revenue_type", "Internal Ad Sales"),
            agency_flag  = user_inputs.get("agency_flag",  "Agency"),
            sales_person = user_inputs.get("sales_person", ""),
            affidavit    = user_inputs.get("affidavit",    "Y"),
            estimate     = user_inputs.get("estimate",     ""),
            contract     = user_inputs.get("contract",     ""),
            language     = language_dict,
            is_worldlink = user_inputs.get("is_worldlink", False),
        )

        # 7. Compute Month column (Calendar vs. Broadcast logic)
        df = transform_month_column(df)

        # 8. Sort: Line (M) → Air Date (B) → Program/airtime (I)
        sort_cols = [c for c in ["Line", "Air Date", "Program"] if c in df.columns]
        if sort_cols:
            df = df.sort_values(sort_cols).reset_index(drop=True)

        # 9. Restore correctly-snapped durations, overriding EtereBridge's nearest-15
        #    rounding (:25 → :30, etc.).  Only applied when row counts match exactly.
        if _raw_durations and "Length" in df.columns and len(_raw_durations) == len(df):
            df["Length"] = [_snap_duration(d) for d in _raw_durations]

        return df

    except Exception as exc:
        logging.error("[EtereBridge] Pipeline error: %s", exc, exc_info=True)
        return None
    finally:
        os.unlink(tmp_path)


def get_language_counts(csv_bytes: bytes) -> dict:
    """
    Run language detection only and return a {lang_code: count} dict.
    Used by the preview endpoint to show detected languages before generation.
    Returns {} if EtereBridge is unavailable.
    """
    if not _AVAILABLE:
        return {}

    with tempfile.NamedTemporaryFile(
        suffix=".csv", prefix="eterebridge_lang_", delete=False
    ) as tmp:
        tmp.write(csv_bytes)
        tmp_path = tmp.name

    try:
        df = _file_processor.load_and_clean_data(tmp_path)
        detected_counts, _ = _file_processor.detect_languages(df)
        return detected_counts
    except Exception as exc:
        logging.warning("[EtereBridge] Language detection failed: %s", exc)
        return {}
    finally:
        os.unlink(tmp_path)


# ─────────────────────────────────────────────────────────────────────────────
# Private helpers
# ─────────────────────────────────────────────────────────────────────────────

def build_placement_csv_from_db(
    contract_id: int,
    date_from: str = None,
    date_to: str = None,
    isci_only: bool = False,
) -> bytes:
    """
    Generate an EtereBridge-compatible placement confirmation CSV directly
    from the database, bypassing the slow Etere web report fetch.

    date_from / date_to: ISO date strings (YYYY-MM-DD) to filter airings.
    isci_only: write just the ISCI code in the Media/bookingcode2 column
               instead of the formatted "title (code)" string.
    """
    import csv as _csv
    import io as _io

    # Etere COD_USER integer → human-readable market name understood by
    # transformer._normalise_market() (keyword match, case-insensitive).
    # Keys must match EtereBridge's market_replacements dict (case-sensitive).
    _COD_USER_TO_MARKET = {
        1:  "NEW YORK",
        2:  "CHI MSP",
        3:  "HOUSTON",
        4:  "SAN FRANCISCO",
        5:  "SEATTLE",
        6:  "LOS ANGELES",
        7:  "Central Valley",
        8:  "WDC",
        9:  "MMT",
        10: "DALLAS",
    }
    FPS = 29.97

    def _frames_to_hhmm(frames: int) -> str:
        total_s = int(round(frames / FPS))
        return f"{total_s // 3600:02d}:{(total_s % 3600) // 60:02d}"

    def _frames_to_hhmmss(frames: int) -> str:
        total_s = int(round(frames / FPS))
        h = total_s // 3600
        m = (total_s % 3600) // 60
        s = total_s % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _format_copy(title: str, code: str) -> str:
        if not code:
            return "NEED COPY"
        if not title or title == code:
            return code
        prefix = title.split(":")[0].strip() if ":" in title else ""
        if prefix == code:
            return title
        return f"{title} ({code})"

    import sys as _sys
    from pathlib import Path as _Path
    _proj = _Path(__file__).parent.parent.parent
    for _p in [str(_proj), str(_proj / "browser_automation")]:
        if _p not in _sys.path:
            _sys.path.insert(0, _p)

    from browser_automation.etere_direct_client import connect as _db_connect

    with _db_connect() as conn:
        cur = conn.cursor(as_dict=True)

        cur.execute("""
            SELECT ct.COD_CONTRATTO         AS contract_code,
                   ct.DESCRIZIONE           AS description,
                   RTRIM(ag.RAG_SOCIAL)     AS agency_name,
                   RTRIM(comm.RAG_SOCIAL)   AS client_name,
                   ISNULL(ag.VIA,   '')     AS agency_address,
                   ISNULL(ag.CITTA, '')     AS agency_city
            FROM CONTRATTITESTATA ct
            LEFT JOIN ANAGRAF ag   ON ag.ID_ANAGRAF   = ct.AGENZIA
            LEFT JOIN ANAGRAF comm ON comm.ID_ANAGRAF = ct.COMMITTENTE
            WHERE ct.ID_CONTRATTITESTATA = %d
        """ % contract_id)
        hdr = cur.fetchone()
        if not hdr:
            raise ValueError(f"Contract {contract_id} not found")

        cur.execute("""
            SELECT tpa.id_contrattirighe        AS line_id,
                   cr.DESCRIZIONE               AS line_desc,
                   ISNULL(f.DURATA, cr.DURATA)  AS dur_frames,
                   cr.IMPORTO                   AS gross_rate,
                   cr.COD_USER                  AS market_id,
                   cr.ORA_INIZIOF               AS daypart_start,
                   cr.ORA_FINEF                 AS daypart_end,
                   CAST(tp.DATA AS DATE)         AS air_date,
                   tp.ORA                        AS airtime_frames,
                   ISNULL(f.COD_PROGRA, 'NEED COPY') AS copy_code,
                   ISNULL(NULLIF(f.DESCRIZIO, ''), f.COD_PROGRA) AS copy_title
            FROM TPALINSE tp
            JOIN trafficPalinse tpa ON tpa.id_tpalinse         = tp.ID_TPALINSE
            JOIN CONTRATTIRIGHE cr  ON cr.ID_CONTRATTIRIGHE    = tpa.id_contrattirighe
            JOIN CONTRATTITESTATA ct ON ct.ID_CONTRATTITESTATA = cr.ID_CONTRATTITESTATA
            LEFT JOIN FILMATI f ON f.ID_FILMATI = tp.ID_FILMATI
            WHERE ct.ID_CONTRATTITESTATA = %d
            %s
            ORDER BY CAST(tp.DATA AS DATE), tp.ORA
        """ % (contract_id, (
            f"AND CAST(tp.DATA AS DATE) >= '{date_from}' AND CAST(tp.DATA AS DATE) <= '{date_to}'"
            if date_from and date_to else ""
        )))
        spots = cur.fetchall()

    if not spots:
        raise ValueError(f"No placed spots found for contract {contract_id}")

    contract_code   = hdr["contract_code"]   or ""
    description     = hdr["description"]     or ""
    agency_name     = hdr["agency_name"]     or description
    client_name     = hdr["client_name"]     or ""
    agency_address  = hdr["agency_address"]  or ""
    agency_city     = hdr["agency_city"]     or ""

    buf = _io.StringIO()
    w   = _csv.writer(buf)

    # Row 0: dummy (skipped by parser)
    w.writerow([""] * 10)
    # Row 1: bill-code / header values row.
    # parse_csv reads: agency=col0, contract_code=col1, description=col3,
    #                  address=col4, client=col5, city=col6
    # _extract_header_values reads: tb180=col0, tb171=col5 (for EtereBridge bill code)
    w.writerow([agency_name, contract_code, "", description,
                agency_address, client_name, agency_city, "", "", ""])
    # Row 2: dummy (skipped by parser)
    w.writerow([""] * 10)
    # Row 3: column headers (triggers data-section detection via "dateschedule")
    w.writerow(["id_contrattirighe", "Textbox14", "duration3", "IMPORTO2",
                "nome2", "dateschedule", "airtimep", "bookingcode2",
                "timerange2", "rowdescription"])

    for s in spots:
        dur_frames    = int(s["dur_frames"] or 0)
        dur_sec       = int(round(dur_frames / FPS))
        market_name   = _COD_USER_TO_MARKET.get(s["market_id"], "")
        air_date      = s["air_date"]
        date_str      = air_date.isoformat() if hasattr(air_date, "isoformat") else str(air_date)
        airtime_frames = int(s["airtime_frames"] or 0)
        daypart_start  = int(s["daypart_start"] or 0)
        daypart_end    = int(s["daypart_end"]   or 0)
        daypart_range  = f"{_frames_to_hhmm(daypart_start)}-{_frames_to_hhmm(daypart_end)}"

        w.writerow([
            s["line_id"],
            1 if isci_only else 4,               # Textbox14 → # column (1 for WorldLink)
            dur_sec,
            s["gross_rate"] or 0,
            market_name,
            date_str,
            _frames_to_hhmmss(airtime_frames),   # airtimep = actual airtime HH:MM:SS
            (s["copy_code"] or "NEED COPY") if isci_only else _format_copy(s.get("copy_title", ""), s["copy_code"]),  # bookingcode2
            daypart_range,                        # timerange2 = contract line daypart
            s["line_desc"] or "",
        ])

    return buf.getvalue().encode("utf-8")


def save_to_excel_with_template(df: pd.DataFrame, agency_fee: float = 0.15) -> bytes:
    """
    Write df to an xlsx workbook using EtereBridge's template file, applying
    its row-2 formulas and formatting to every data row.  Returns raw bytes
    suitable for streaming as a download.  Mirrors EtereBridge's save_to_excel().
    """
    if not _AVAILABLE:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name="Run Sheet")
        buf.seek(0)
        return buf.read()

    template_path = _eb_app_config.paths.template_path
    workbook = load_workbook(template_path, data_only=False)
    sheet    = workbook.active
    columns  = _eb_app_config.final_columns

    # Extract formulas + formatting from template row 2
    template_formulas   = {}
    template_formatting = {}
    for col in range(1, len(columns) + 1):
        cell = sheet.cell(row=2, column=col)
        if cell.value and str(cell.value).startswith("="):
            template_formulas[col] = cell.value
        template_formatting[col] = {
            "style":         cell.style,
            "number_format": cell.number_format,
            "border":        copy(cell.border),
            "fill":          copy(cell.fill),
            "font":          copy(cell.font),
            "alignment":     copy(cell.alignment),
        }

    # Write headers in row 1
    for col_num, col_title in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_num, value=col_title)

    monetary_columns = ["Gross Rate", "Spot Value", "Station Net", "Broker Fees"]
    gross_col_letter  = get_column_letter(columns.index("Gross Rate") + 1)  if "Gross Rate"  in columns else None
    air_date_letter   = get_column_letter(columns.index("Air Date") + 1)     if "Air Date"    in columns else None
    agency_col_data   = df[columns[columns.index("Agency?")]]                if "Agency?"     in columns else None

    def _parse_time(time_str):
        if not time_str:
            return None
        try:
            dt = pd.to_datetime(time_str, format="%H:%M:%S", errors="raise")
        except ValueError:
            try:
                dt = pd.to_datetime(time_str, format="%I:%M:%S %p", errors="raise")
            except ValueError:
                return None
        return (dt.hour * 3600 + dt.minute * 60 + dt.second) / 86400.0

    # Write data rows starting at row 2
    for row_num, row_data in enumerate(df.values, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell     = sheet.cell(row=row_num, column=col_num)
            col_name = columns[col_num - 1]

            if col_name in ("Time In", "Time Out"):
                ts = _parse_time(cell_value)
                if ts is not None:
                    cell.value = ts
                    cell.number_format = "[h]:mm:ss"
                else:
                    cell.value = cell_value
            elif col_name == "End Date":
                if air_date_letter:
                    cell.value = f"={air_date_letter}{row_num}"
                    cell.number_format = "m/d/yy"
                else:
                    cell.value = cell_value
            elif col_name == "Length":
                try:
                    cell.value = float(cell_value) / 86400 if pd.notna(cell_value) else 0
                    cell.number_format = "[h]:mm:ss"
                    cell.alignment = Alignment(horizontal="center")
                except Exception:
                    cell.value = cell_value
            elif col_name == "Broker Fees" and gross_col_letter and agency_col_data is not None:
                if agency_col_data.iloc[row_num - 2] == "Agency":
                    cell.value = f"={gross_col_letter}{row_num}*{agency_fee}"
                else:
                    cell.value = None
            elif col_num in template_formulas and col_name not in ("Time In", "Time Out", "Length", "End Date", "Broker Fees"):
                cell.value = template_formulas[col_num].replace("2", str(row_num))
            elif col_name in monetary_columns:
                cell.value = cell_value if pd.notna(cell_value) else 0
            else:
                cell.value = cell_value

            # Apply template formatting
            if col_num in template_formatting:
                fmt = template_formatting[col_num]
                cell.fill      = fmt["fill"]
                cell.border    = fmt["border"]
                cell.font      = fmt["font"]
                cell.alignment = fmt["alignment"]
                if col_name not in ("Time In", "Time Out", "Length", "End Date") + tuple(monetary_columns):
                    cell.style         = fmt["style"]
                    cell.number_format = fmt["number_format"]

    format_excel_monetary_columns(sheet, df, monetary_columns)

    # Air Date formatting
    if "Air Date" in columns:
        air_date_col = columns.index("Air Date") + 1
        for row_num in range(2, len(df) + 2):
            cell = sheet.cell(row=row_num, column=air_date_col)
            if cell.value:
                dt = safe_convert_date(cell.value)
                if dt is not None:
                    cell.value = dt
                    cell.number_format = "m/d/yy"

    # Month formatting
    if "Month" in columns:
        month_col = columns.index("Month") + 1
        for row_num in range(2, len(df) + 2):
            cell = sheet.cell(row=row_num, column=month_col)
            mv   = df["Month"].iloc[row_num - 2]
            cell.value         = mv if pd.notna(mv) else None
            cell.number_format = "mmm-yy"

    # Priority = 4
    if "Priority" in columns:
        pri_col = columns.index("Priority") + 1
        for row_num in range(2, len(df) + 2):
            sheet.cell(row=row_num, column=pri_col, value=4)

    # Remove leftover template rows
    if sheet.max_row > len(df) + 1:
        sheet.delete_rows(len(df) + 2, sheet.max_row - (len(df) + 1))

    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return buf.read()


def _extract_header_values(file_path: str) -> tuple[str, str]:
    """
    Read CSV row 2 (zero-indexed row 1) to extract bill-code parts.
    Column 1 = agency/client name, column 6 = venue/site name.
    Matches EtereBridge's extract_header_values() logic exactly.
    """
    try:
        with open(file_path, "r") as f:
            f.readline()           # skip row 1 (column labels)
            second_line = f.readline().strip()
        if not second_line:
            return "", ""
        reader = csv.reader([second_line])
        parts  = next(reader)
        first  = parts[0].strip() if len(parts) > 0 else ""
        second = parts[5].strip() if len(parts) > 5 else ""
        return first, second
    except Exception as exc:
        logging.warning("[EtereBridge] Header extraction failed: %s", exc)
        return "", ""


def _apply_user_inputs(
    df: pd.DataFrame,
    billing_type: str,
    revenue_type: str,
    agency_flag:  str,
    sales_person: str,
    affidavit:    str,
    estimate:     str,
    contract:     str,
    language:     dict,
    is_worldlink: bool = False,
) -> pd.DataFrame:
    """
    Stamp user-provided metadata onto every row and reorder columns to
    EtereBridge's final_columns order.  Mirrors EtereBridge's apply_user_inputs()
    without calling verify_languages().
    """
    df["Billing Type"] = billing_type
    df["Revenue Type"] = revenue_type
    df["Agency?"]      = agency_flag
    df["Sales Person"] = sales_person
    df["Affidavit?"]   = affidavit
    df["Estimate"]     = estimate
    df["Contract"]     = contract
    df["Lang."]        = df.index.map(language).fillna("E")

    df["Type"] = df["Gross Rate"].apply(
        lambda r: "BNS" if (pd.isna(r) or float(r) == 0) else "COM"
    )

    # WorldLink: copy Market → Make Good (per EtereBridge WorldLink branch)
    if is_worldlink and "Market" in df.columns:
        df["Make Good"] = df["Market"]

    # Ensure every final column exists before reordering
    for col in _eb_app_config.final_columns:
        if col not in df.columns:
            df[col] = None

    return df[_eb_app_config.final_columns]
