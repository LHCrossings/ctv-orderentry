"""
WorldLink backwrite transformer - template-based.
Loads worldlink_template.xlsx and fills in order data from the parsed IO PDF.
Generates 2 tabs: Sales Confirmation + Monthly Lines and Broker Fees.
"""

import io
import re
import sys
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

_src = Path(__file__).parent.parent
if str(_src) not in sys.path:
    sys.path.insert(0, str(_src))

from backwrite.transformer import _apply_snapshot, _snapshot_row, compute_broadcast_month

YELLOW_FILL  = PatternFill("solid", fgColor="FFFFFF00")
GREEN_FILL   = PatternFill("solid", fgColor="FF92D050")
AGENCY_FEE   = 0.15
_DATA_FONT   = Font(name="Calibri", size=11)
_BOLD_FONT   = Font(name="Calibri", size=11, bold=True)
_TEMPLATE    = Path(__file__).parent / "templates" / "worldlink_template.xlsx"
_CURRENCY_NF = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_MONTH_NF    = '[$-409]mmm\\-yy;@'   # "Apr-26" — MLBF column S
_TIME_NF     = '[h]:mm:ss;@'          # MLBF time columns E/F/G
_INT_NF      = "0"                    # integer (spots counts)
_PCT_NF      = "0%"                   # agency fee (0.15 → 15%)

_MONTH_NAMES = {
    1: "January",  2: "February",  3: "March",    4: "April",
    5: "May",      6: "June",      7: "July",      8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


# ──────────────────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────────────────

def generate_worldlink_excel(io_data: dict, user_inputs: dict) -> bytes:
    """
    io_data    : output of worldlink_parser.parse_worldlink_pdf
    user_inputs: {"contract_number": str, "revision": int|str}
    Returns xlsx bytes.
    """
    wb = load_workbook(_TEMPLATE)

    for name in ("Run Sheet", "Monthly Totals"):
        if name in wb.sheetnames:
            del wb[name]

    _fill_sc_tab(wb["Sales Confirmation"], io_data, user_inputs)
    _fill_mlbf_tab(wb["MONTHLY LINES AND BROKER FEES"], io_data, user_inputs)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Revision helpers — read back a prior Excel and merge with revision PDF lines
# ──────────────────────────────────────────────────────────────────────────────

def read_sc_lines_from_excel(xlsx_bytes: bytes):
    """Read order lines and revision number from a prior Sales Confirmation tab.

    Returns (lines: list[dict], prev_revision: int).
    Lines have action=None (untouched — no yellow highlight).
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Sales Confirmation"]

    rev_val = ws.cell(10, 12).value
    try:
        prev_revision = int(rev_val or 0)
    except (TypeError, ValueError):
        prev_revision = 0

    def _d(v):
        if v is None:
            return ""
        if hasattr(v, "month"):
            return f"{v.month}/{v.day}/{v.year}"
        return str(v)

    lines = []
    r = 14
    while r <= ws.max_row:
        line_no_val = ws.cell(r, 2).value
        try:
            line_no = int(line_no_val)
        except (TypeError, ValueError):
            break
        rate = 0.0
        try:
            rate = float(ws.cell(r, 15).value or 0)
        except (TypeError, ValueError):
            pass
        dur_raw = ws.cell(r, 14).value or ":30"
        lines.append({
            "line_number":   line_no,
            "action":        None,
            "start_date":    _d(ws.cell(r, 4).value),
            "end_date":      _d(ws.cell(r, 5).value),
            "spots":         int(ws.cell(r, 6).value or 0),
            "rate":          rate,
            "program_label": str(ws.cell(r, 8).value or ""),
            "duration":      str(dur_raw).lstrip(":"),
        })
        r += 1

    return lines, prev_revision


def merge_revision_lines(prev_lines: list, rev_lines: list) -> list:
    """Merge revision PDF lines into the full prev line set.

    CANCEL  → remove line; CHANGE/ADD → full replacement (parser gives complete data).
    Untouched prev lines keep action=None (no yellow).
    """
    prev_by_no = {ln["line_number"]: dict(ln) for ln in prev_lines}
    for line in rev_lines:
        no     = line.get("line_number")
        action = line.get("action", "ADD")
        if action == "CANCEL":
            prev_by_no.pop(no, None)
        else:
            prev_by_no[no] = dict(line)
    return sorted(prev_by_no.values(), key=lambda ln: ln.get("line_number", 0))


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _fmt_time_short(t24: str) -> str:
    """HH:MM → short 12-hour label: '6a', '9a', '5p', '10p', '12a', '12p'."""
    if not t24:
        return ""
    if t24 == "23:59":
        return "12a"
    try:
        h, m = map(int, t24.split(":"))
    except ValueError:
        return t24
    if h == 0 and m == 0:
        return "12a"
    if h == 12 and m == 0:
        return "12p"
    period = "a" if h < 12 else "p"
    h12 = h if h < 12 else (h - 12 if h > 12 else 12)
    return f"{h12}:{m:02d}{period}" if m else f"{h12}{period}"


def _fmt_program(line: dict) -> str:
    """Build 'M-Su 6a-9a' style string from days + time fields."""
    days   = line.get("days_of_week", "M-Su")
    from_t = line.get("from_time", "06:00")
    to_t   = line.get("to_time",   "23:59")
    return f"{days} {_fmt_time_short(from_t)}-{_fmt_time_short(to_t)}"


def _count_weeks(line: dict) -> int:
    """Return week count — prefer 'weeks' field from parser, then derive."""
    if line.get("weeks"):
        return int(line["weeks"])
    spw = line.get("spots", 0) or 0
    tot = line.get("total_spots", 0) or 0
    if spw:
        return max(1, round(tot / spw))
    try:
        start = _parse_date_str(line.get("start_date", ""))
        end   = _parse_date_str(line.get("end_date",   ""))
        if start and end:
            return max(1, round((end - start).days / 7))
    except Exception:
        pass
    return 1


def _clean_org_name(name: str) -> str:
    """Strip trailing corporate suffixes: ', Inc.', ', LLC', etc."""
    return re.sub(
        r",?\s*(Inc\.?|LLC\.?|Ltd\.?|Corp\.?|Co\.?)$", "", name, flags=re.I
    ).strip()


def _make_bill_code(agency: str, advertiser: str) -> str:
    """'WorldLink:CleanAgency AdvertiserWord'"""
    clean = _clean_org_name(agency)
    adv   = (_clean_org_name(advertiser).split()[0] if advertiser else "").rstrip(",")
    return f"WorldLink:{clean} {adv}".strip()


def _parse_date_str(s: str):
    """Parse M/D/YYYY or M/D/YY to date, or return None."""
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_int_if_numeric(val) -> object:
    """Return int when val is purely numeric, else return as-is."""
    try:
        s = str(val).strip()
        if s.isdigit():
            return int(s)
    except Exception:
        pass
    return val


def _compute_monthly_revenue(lines: list) -> Dict[date, float]:
    """
    Iterate each IO line week-by-week; bucket (spots_per_week × rate) into
    broadcast months.  Zero-rate lines (bonus/free) are skipped.
    Returns {broadcast_month_first_of_month: gross_amount}.
    """
    monthly: Dict[date, float] = defaultdict(float)
    for line in lines:
        rate = float(line.get("rate", 0) or 0)
        if rate == 0:
            continue
        spw   = line.get("spots", 0) or 0
        start = _parse_date_str(line.get("start_date", ""))
        end   = _parse_date_str(line.get("end_date",   ""))
        if not start or not end or spw == 0:
            continue
        week_start = start
        while week_start <= end:
            bm = compute_broadcast_month(week_start)
            monthly[bm] += spw * rate
            week_start += timedelta(days=7)
    return {k: round(v, 2) for k, v in sorted(monthly.items())}


def _broadcast_month_formula(r: int) -> str:
    """Broadcast-month DATE formula for MLBF column S (row r)."""
    b = f"B{r}"
    return (
        f'=IF(Y{r}="Calendar",{b},'
        f"DATE(IF(AND(MONTH({b})=12,MONTH({b}+(7-WEEKDAY({b},2)))=1),"
        f"YEAR({b})+1,YEAR({b})),"
        f"MONTH({b}+(7-WEEKDAY({b},2))),1))"
    )


def _wc(ws, row: int, col: int, val, nf: str = "General", font=None) -> None:
    """Write a value to a data cell; always sets number_format to prevent
    inheriting column-level styles (e.g. date format on col L)."""
    cell = ws.cell(row=row, column=col)
    cell.value         = val
    cell.font          = copy(font if font is not None else _DATA_FONT)
    cell.number_format = nf


# ──────────────────────────────────────────────────────────────────────────────
# Sales Confirmation tab
# ──────────────────────────────────────────────────────────────────────────────

def _fill_sc_tab(ws, io_data: dict, user_inputs: dict) -> None:
    agency        = io_data.get("agency", "")
    advertiser    = io_data.get("advertiser", "")
    tracking      = str(io_data.get("tracking_number", "") or "")
    buyer         = io_data.get("buyer", "")
    order_comment = io_data.get("order_comment", "") or ""
    lines         = io_data.get("lines", [])

    agency_street = io_data.get("agency_street", "")
    agency_city   = io_data.get("agency_city", "")
    agency_state  = io_data.get("agency_state", "")
    agency_zip    = io_data.get("agency_zip", "")

    contract_no = str(user_inputs.get("contract_number", "") or "")
    revision    = user_inputs.get("revision", 0)
    is_revision = int(revision) > 0

    # ── Overwrite header value-cells (labels stay from template) ─────────────
    def sv(r, c, v, nf=None):
        cell = ws.cell(row=r, column=c)
        cell.value = v
        if nf:
            cell.number_format = nf

    sv(3,  4,  _clean_org_name(agency))
    sv(3,  12, advertiser)
    sv(4,  4,  buyer)
    sv(4,  12, _to_int_if_numeric(tracking))
    sv(5,  4,  agency_street)
    sv(6,  4,  agency_city)
    sv(6,  6,  agency_state)
    sv(6,  7,  _to_int_if_numeric(agency_zip))
    sv(8,  12, datetime.today(), "mm-dd-yy")
    sv(9,  12, _to_int_if_numeric(contract_no))
    sv(10, 12, str(revision))
    # K11 "House (Worldlink)" stays from template

    # ── Snapshot row 14 before any insertions ────────────────────────────────
    row_snapshot = _snapshot_row(ws, 14)

    # ── Insert extra rows for additional lines, pushing footer/summary down ───
    DATA_START = 14
    n_lines    = max(len(lines), 1)
    n_inserts  = n_lines - 1
    line_rows  = [DATA_START]

    if n_inserts > 0:
        # Save and unmerge all merged ranges that will be pushed down
        insert_start = DATA_START + 1
        saved_merges = [
            (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
            for mr in list(ws.merged_cells.ranges)
            if mr.min_row >= insert_start
        ]
        for min_r, max_r, min_c, max_c in saved_merges:
            ws.unmerge_cells(start_row=min_r, start_column=min_c,
                             end_row=max_r,   end_column=max_c)

        for _ in range(n_inserts):
            new_row = line_rows[-1] + 1
            ws.insert_rows(new_row)
            _apply_snapshot(ws, row_snapshot, new_row, DATA_START)
            line_rows.append(new_row)

        # Re-merge at shifted positions
        for min_r, max_r, min_c, max_c in saved_merges:
            ws.merge_cells(start_row=min_r + n_inserts, start_column=min_c,
                           end_row=max_r   + n_inserts, end_column=max_c)

    last_data = line_rows[-1]

    # ── Write order lines into rows 14..last_data ─────────────────────────────
    for i, line in enumerate(lines):
        r         = line_rows[i]
        action    = line.get("action", "ADD")
        is_cancel = action == "CANCEL"
        spots     = 0 if is_cancel else (line.get("spots", 0) or 0)
        rate      = 0.0 if is_cancel else float(line.get("rate", 0) or 0)
        weeks     = _count_weeks(line)
        start_d   = _parse_date_str(line.get("start_date", ""))
        end_d     = _parse_date_str(line.get("end_date",   ""))
        start_dt  = datetime(start_d.year, start_d.month, start_d.day) if start_d else None
        end_dt    = datetime(end_d.year,   end_d.month,   end_d.day)   if end_d   else None

        _wc(ws, r, 2,  line.get("line_number", i + 1))
        _wc(ws, r, 4,  start_dt, "mm-dd-yy")
        _wc(ws, r, 5,  end_dt,   "mm-dd-yy")
        _wc(ws, r, 6,  spots)
        _wc(ws, r, 7,  "week")
        _wc(ws, r, 8,  line.get("program_label") or _fmt_program(line))
        _wc(ws, r, 9,  weeks)
        _wc(ws, r, 10, "COM")
        _wc(ws, r, 12, f"=F{r}*I{r}")
        _wc(ws, r, 14, f":{line.get('duration', '30')}")
        _wc(ws, r, 15, rate,           _CURRENCY_NF)
        _wc(ws, r, 16, f"=L{r}*O{r}", _CURRENCY_NF)

        ws.row_dimensions[r].height = 16.5

        if is_revision and action in ("ADD", "CHANGE"):
            for col in range(1, 23):
                ws.cell(row=r, column=col).fill = YELLOW_FILL

    # ── Fix summary row formulas (rows shifted by inserts but refs not updated) ─
    sum_row  = last_data + 1
    disc_row = sum_row + 1
    net_row  = disc_row + 1

    ws.cell(sum_row,  12).value = f"=SUM(L{DATA_START}:L{last_data})"
    ws.cell(sum_row,  16).value = f"=SUM(P{DATA_START}:P{last_data})"
    ws.cell(disc_row, 16).value = f"=-1*(L{disc_row}*P{sum_row})"
    ws.cell(net_row,   2).value = order_comment
    ws.cell(net_row,  16).value = f"=SUM(P{sum_row}:P{disc_row})"

    ws.column_dimensions["P"].width = 14

    # ── Rewrite monthly breakdown with actual computed data ───────────────────
    monthly_rev   = _compute_monthly_revenue(lines)
    sorted_months = sorted(monthly_rev.keys())

    # Broadcast months touched by ADD/CHANGE lines — drives yellow fill on revisions
    affected_months: set = set()
    if is_revision:
        for line in lines:
            if line.get("action", "ADD") not in ("ADD", "CHANGE"):
                continue
            rate  = float(line.get("rate", 0) or 0)
            spw   = line.get("spots", 0) or 0
            start = _parse_date_str(line.get("start_date", ""))
            end   = _parse_date_str(line.get("end_date",   ""))
            if not start or not end or spw == 0 or rate == 0:
                continue
            wk = start
            while wk <= end:
                affected_months.add(compute_broadcast_month(wk))
                wk += timedelta(days=7)

    # Scan for "MONTHLY BREAKDOWN" title (position shifts with each inserted row)
    mbr_title = None
    for r in range(net_row + 1, ws.max_row + 1):
        if ws.cell(r, 2).value == "MONTHLY BREAKDOWN":
            mbr_title = r
            break

    if mbr_title is not None and sorted_months:
        mbr_hdr        = mbr_title + 2
        mbr_data_start = mbr_hdr + 1

        # Snapshot first existing month row for formatting
        month_snap = _snapshot_row(ws, mbr_data_start)

        # Remove sample month rows (stop at "Total" or blank col B)
        existing = 0
        while True:
            val = ws.cell(mbr_data_start + existing, 2).value
            if val is None or str(val).strip().lower() == "total":
                break
            existing += 1

        # Snapshot the Total row BEFORE deleting sample rows (preserves italics + double border)
        total_snap = _snapshot_row(ws, mbr_data_start + existing)

        if existing > 0:
            ws.delete_rows(mbr_data_start, existing)

        # Insert and fill actual month rows
        for j, bm in enumerate(sorted_months):
            r = mbr_data_start + j
            if j > 0:
                ws.insert_rows(r)
            _apply_snapshot(ws, month_snap, r, mbr_data_start)
            gross = round(monthly_rev[bm], 2)
            ws.cell(r, 2).value = _MONTH_NAMES[bm.month]
            ws.cell(r, 4).value = gross
            ws.cell(r, 5).value = f"=D{r}*0.85"
            ws.cell(r, 6).value = f"=E{r}*0.1*-1"
            for col in [4, 5, 6]:
                if is_revision and bm in affected_months:
                    ws.cell(r, col).fill = YELLOW_FILL
                ws.cell(r, col).number_format = _CURRENCY_NF

        # Restore Total row formatting from template snapshot, then fix values/formulas
        mbr_last  = mbr_data_start + len(sorted_months) - 1
        total_row = mbr_last + 1
        _apply_snapshot(ws, total_snap, total_row, mbr_data_start + existing)
        ws.cell(total_row, 2).value = "Total"
        ws.cell(total_row, 4).value = f"=SUM(D{mbr_data_start}:D{mbr_last})"
        ws.cell(total_row, 5).value = f"=SUM(E{mbr_data_start}:E{mbr_last})"
        for col in [4, 5]:
            if is_revision and affected_months:
                ws.cell(total_row, col).fill = YELLOW_FILL
            ws.cell(total_row, col).number_format = _CURRENCY_NF


# ──────────────────────────────────────────────────────────────────────────────
# Monthly Lines and Broker Fees tab
# ──────────────────────────────────────────────────────────────────────────────

def _fill_mlbf_tab(ws, io_data: dict, user_inputs: dict) -> None:
    agency      = io_data.get("agency", "")
    advertiser  = io_data.get("advertiser", "")
    tracking    = str(io_data.get("tracking_number", "") or "")
    lines       = io_data.get("lines", [])
    contract_no = user_inputs.get("contract_number", "")

    contract_val = _to_int_if_numeric(contract_no)
    tracking_val = _to_int_if_numeric(tracking)
    bill_code    = _make_bill_code(agency, advertiser)
    monthly_rev  = _compute_monthly_revenue(lines)
    sorted_months = sorted(monthly_rev.keys())

    # Rows 1-4 preserved from template (headers, blank, billing instruction, blank).
    # Delete rows 5 onwards and rebuild.
    ws.delete_rows(5, ws.max_row)

    # ── Billing group (rows 5 … 5+n-1) ───────────────────────────────────────
    BILL_START = 5
    for i, bm in enumerate(sorted_months):
        r            = BILL_START + i
        gross        = round(monthly_rev[bm], 2)
        billing_date = datetime(bm.year, bm.month, 20)

        _wc(ws, r,  1, bill_code)
        _wc(ws, r,  2, billing_date,                  "m/d/yy")
        _wc(ws, r,  3, f"=B{r}",                   "m/d/yy")
        _wc(ws, r,  4, f'=TEXT(B{r},"dddd")')
        _wc(ws, r,  5, timedelta(0),               _TIME_NF)
        _wc(ws, r,  6, timedelta(0),               _TIME_NF)
        _wc(ws, r,  7, timedelta(0),               _TIME_NF)
        _wc(ws, r,  8, f"{tracking} Monthly Charges")
        _wc(ws, r,  9, "BILLING LINE")
        _wc(ws, r, 11, "NX")
        _wc(ws, r, 12, 1,                          _INT_NF)
        _wc(ws, r, 14, "COM")
        _wc(ws, r, 15, tracking_val)
        _wc(ws, r, 16, gross,                      _CURRENCY_NF)
        _wc(ws, r, 18, f"=P{r}",                   _CURRENCY_NF)
        _wc(ws, r, 19, _broadcast_month_formula(r), _MONTH_NF)
        _wc(ws, r, 20, f"=P{r}*0.15",              _CURRENCY_NF)
        _wc(ws, r, 21, 4)
        _wc(ws, r, 22, f"=P{r}-T{r}",             _CURRENCY_NF)
        _wc(ws, r, 23, "House")
        _wc(ws, r, 24, "Direct Response Sales")
        _wc(ws, r, 25, "Broadcast")
        _wc(ws, r, 26, "Agency")
        _wc(ws, r, 27, "Y")
        _wc(ws, r, 28, contract_val)
        _wc(ws, r, 29, "Admin")

    # ── Broker fee instruction row ────────────────────────────────────────────
    fee_instr = BILL_START + len(sorted_months) + 1
    cell = ws.cell(fee_instr, 1)
    cell.value = "These are the lines you will paste in to show the monthly broker fees"
    cell.fill  = GREEN_FILL

    # ── Broker fee group (two blank rows after billing group) ─────────────────
    FEE_START = fee_instr + 2
    for i, bm in enumerate(sorted_months):
        r            = FEE_START + i
        gross        = round(monthly_rev[bm], 2)
        broker_fee   = round(-gross * 0.85 * 0.10, 2)
        billing_date = datetime(bm.year, bm.month, 20)

        _wc(ws, r,  1, "WorldLink Broker Fees (DO NOT INVOICE)")
        _wc(ws, r,  2, billing_date,                  "m/d/yy")
        _wc(ws, r,  3, f"=B{r}",                   "m/d/yy")
        _wc(ws, r,  4, f'=TEXT(B{r},"dddd")')
        _wc(ws, r,  5, timedelta(0),               _TIME_NF)
        _wc(ws, r,  6, timedelta(0),               _TIME_NF)
        _wc(ws, r,  7, timedelta(0),               _TIME_NF)
        _wc(ws, r,  8, f"{tracking} Broker Fees")
        _wc(ws, r,  9, "BILLING LINE")
        _wc(ws, r, 11, "NX")
        _wc(ws, r, 12, 1,                          _INT_NF)
        _wc(ws, r, 14, "COM")
        _wc(ws, r, 15, tracking_val)
        _wc(ws, r, 16, broker_fee,                 _CURRENCY_NF)
        _wc(ws, r, 18, f"=P{r}",                   _CURRENCY_NF)
        _wc(ws, r, 19, _broadcast_month_formula(r), _MONTH_NF)
        _wc(ws, r, 20, 0,                          _CURRENCY_NF)
        _wc(ws, r, 21, 4)
        _wc(ws, r, 22, f"=P{r}-T{r}",             _CURRENCY_NF)
        _wc(ws, r, 23, "House")
        _wc(ws, r, 24, "Direct Response Sales")
        _wc(ws, r, 25, "Broadcast")
        _wc(ws, r, 26, "Agency")
        _wc(ws, r, 27, "Y")
        _wc(ws, r, 28, contract_val)
        _wc(ws, r, 29, "Admin")
