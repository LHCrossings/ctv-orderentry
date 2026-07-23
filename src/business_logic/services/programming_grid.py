"""Read the weekly programming grids (K: drive Excel) for the Daily Programming tool.

The grids are the manual source of truth: one .xlsx per M-Su week, named by the
Monday's date, under K:\\Programming\\! <Network>\\<year>\\<MM yyyy>\\.
This reads a single day's program lineup out of the "Local Channels" sheet.

See tasks/daily_programming_discovery.md for the grid layout details.
"""
from __future__ import annotations

import datetime
import os
import re
import sys
from pathlib import Path

import openpyxl

# K: on Windows, the SMB mount elsewhere; override via PROGRAMMING_GRID_ROOT.
GRID_ROOT = os.environ.get(
    "PROGRAMMING_GRID_ROOT",
    r"K:\Programming" if sys.platform == "win32" else "/mnt/k/Programming",
)

NETWORK_DIR = {"CTV": "! Crossings TV", "TAC": "! The Asian Channel"}
NETWORK_FILE = {"CTV": "Crossings TV", "TAC": "The Asian Channel"}

# Footer / non-program markers that signal the end of the day's lineup.
_FOOTER_RE = re.compile(
    r"local channels|xfinity|spectrum|\bch\.|thick borders denote|"
    r"^\s*legend\s*$|"  # 'LEGEND' key header — anchored so drama titles ("The Legend of…") aren't caught
    r"^\s*\d{1,2}/\d{1,2}/\d{2,4}\s*$", re.I)


def _week_monday(d: datetime.date) -> datetime.date:
    return d - datetime.timedelta(days=d.weekday())


def find_grid_file(network: str, d: datetime.date) -> Path | None:
    """Locate the weekly grid .xlsx for the week containing ``d``.

    Checks only the specific month folders (fast on the network drive) — the
    week's Monday and Sunday months, each with and without the in-progress
    ``!!!`` prefix (e.g. "06 2026" or "!!!08 2026"). Globs a single small
    folder for the Monday-stamped file rather than walking the whole tree.
    """
    netdir = NETWORK_DIR.get(network)
    if not netdir:
        return None
    mon = _week_monday(d)
    stamp = mon.strftime("%Y%m%d")
    seen: set = set()
    for dd in (mon, mon + datetime.timedelta(days=6)):  # broadcast month may follow the week's end
        for prefix in ("", "!!!"):
            month_dir = Path(GRID_ROOT) / netdir / dd.strftime("%Y") / f"{prefix}{dd.month:02d} {dd.year}"
            if month_dir in seen:
                continue
            seen.add(month_dir)
            if month_dir.exists():
                matches = list(month_dir.glob(f"*{stamp}*.xlsx"))
                if matches:
                    return matches[0]
    return None


# The day-grid sheet differs by network: CTV = "Local Channels", TAC = "DALLAS".
_PREFERRED_SHEETS = ("Local Channels", "DALLAS")


def _pick_grid_sheet(wb):
    """Select the sheet holding the day grid (prefers known names, else detects
    the sheet whose row 3 carries the per-day dates)."""
    for name in _PREFERRED_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        sh = wb[name]
        if any(isinstance(sh.cell(3, c).value, datetime.datetime) for c in range(2, 9)):
            return sh
    return None


def _time_label(ws, row: int) -> str | None:
    v = ws.cell(row, 1).value
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None


def _parse_title(raw: str) -> dict:
    """Split a grid cell into title + language + kind from the trailing (Lang Kind)."""
    text = " ".join(str(raw).split())
    language = kind = None
    m = re.search(r"\(([^()]*)\)\s*$", text)
    if m:
        inside = m.group(1).strip()
        text_wo = text[: m.start()].strip()
        parts = inside.split()
        if parts:
            language = parts[0]
            kind = " ".join(parts[1:]) or None
        title = text_wo or text
    else:
        title = text
    return {"title": title, "language": language, "kind": kind, "raw": text}


def _hhmm_to_min(hhmm: str):
    m = re.match(r"\s*(\d{1,2}):(\d{2})", hhmm or "")
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def _min_to_hhmm(mins: int) -> str:
    mins %= 24 * 60
    return f"{mins // 60:02d}:{mins % 60:02d}"


def _norm_time(tok: str):
    """Normalise a single time token to 24-hour 'HH:MM', or None.

    Handles the 24-hour labels from `datetime.time` cells ('22:30') AND the am/pm
    text labels the grids use for the overnight block ('12a', '6a', '1230a', '12n',
    '12p'). Midnight '12a' → '00:00' (the broadcast-day tail; the +24h shift for
    ORA math happens later in the frame converters, not here)."""
    tok = (tok or "").strip().lower().replace(" ", "")
    m = re.match(r"^(\d{1,2})(?::?(\d{2}))?([apn])?m?$", tok)
    if not m:
        return None
    h = int(m.group(1))
    mn = int(m.group(2) or 0)
    suf = m.group(3)
    if suf == "n":          # 12n = noon
        h = 12
    elif suf == "a":        # am; 12a = midnight
        if h == 12:
            h = 0
    elif suf == "p":        # pm; 12p = noon
        if h != 12:
            h += 12
    if h > 23 or mn > 59:
        return None
    return f"{h:02d}:{mn:02d}"


def _split_time_label(label: str):
    """(start, end) for a grid time cell. A range cell like '12a-6a' yields both ends
    ('00:00', '06:00'); a plain cell yields (start, None). Falls back to the raw label
    if it can't be parsed, preserving prior behaviour for unexpected text."""
    if not label:
        return None, None
    if "-" in label:
        a, _, b = label.partition("-")
        na = _norm_time(a)
        if na:
            return na, _norm_time(b)
        return label, None
    return (_norm_time(label) or label), None


def _block_times(ws, this_row: int, next_row: int, span: int):
    """Start/end (HH:MM) for a grid block. Normalises am/pm + range labels, takes the
    end from the block's own range cell if present else the next row's start, and —
    for the day's final block, whose following row has no time label — derives the end
    from the block's 30-min row span (each grid row = 30 min, so 04:30 over 3 rows → 06:00)."""
    own_start, own_end = _split_time_label(_time_label(ws, this_row))
    next_start, _ = _split_time_label(_time_label(ws, next_row))
    start = own_start
    end = own_end or next_start
    if _hhmm_to_min(end) is None and _hhmm_to_min(start) is not None:
        end = _min_to_hhmm(_hhmm_to_min(start) + span * 30)
    return start, end


def _split_shared_block(prog: dict) -> list[dict]:
    """A single grid cell can carry two (or more) shorter shows whose titles are
    joined with '/', e.g. 'Headline News / Culture & Travel (Vietnamese News)' — a
    30-min block holding two 15-min shows. Etere carries them as separate blocks, so
    emit one program per title and divide the block's duration evenly among them.

    Only a '/' in the TITLE separates shows. A '/' inside the trailing
    '(language kind)' — e.g. '(Vietnamese Culture/Education)' — is NOT a separator;
    _parse_title has already split that parenthetical off `title`, so we never see
    it here. Requires a valid start+end, which also filters out section-header rows
    like 'Tagalog / Filipino' / 'MultiAsian/ English' (no time label)."""
    parts = [t.strip() for t in (prog.get("title") or "").split("/")]
    parts = [t for t in parts if t]
    s, e = _hhmm_to_min(prog.get("start")), _hhmm_to_min(prog.get("end"))
    if len(parts) < 2 or s is None or e is None:
        return [prog]
    if e <= s:  # block crosses midnight (e.g. 23:45–00:15)
        e += 24 * 60
    n = len(parts)
    step = (e - s) / n
    out = []
    for i, t in enumerate(parts):
        seg = dict(prog)
        seg["title"] = t
        seg["start"] = _min_to_hhmm(int(round(s + i * step)))
        seg["end"] = _min_to_hhmm(int(round(s + (i + 1) * step)))
        out.append(seg)
    return out


def _coalesce_hmong_weekend(programs: list[dict], d: datetime.date) -> list[dict]:
    """Override: Hmong on Sat/Sun 18:00–20:00 is always TWO one-hour fillable
    blocks (18:00–19:00, 19:00–20:00).

    The K: grids we receive split that window into :30/:30 for public posting,
    but the actual media is delivered as ONE file per hour — so the placement
    tool must show one fillable block per hour, not the :30 breakdown. Only
    Hmong entries fully inside the window are merged; anything else (and any
    weekday) is left untouched. Idempotent: a grid already carrying two 1-hour
    Hmong blocks comes back unchanged. Hmong is CTV-only, so this never fires on
    a DAL grid."""
    if d.weekday() not in (5, 6):  # Saturday=5, Sunday=6
        return programs
    win_s, win_e = 18 * 60, 20 * 60

    def _is_hmong(p: dict) -> bool:
        return (p.get("language") or "").strip().lower().startswith("hm") \
            or "hmong" in (p.get("raw") or "").lower()

    def _inside(p: dict, s: int, e: int) -> bool:
        ps, pe = _hhmm_to_min(p.get("start")), _hhmm_to_min(p.get("end"))
        return ps is not None and pe is not None and ps < pe and s <= ps and pe <= e

    hmong = [p for p in programs if _is_hmong(p) and _inside(p, win_s, win_e)]
    if not hmong:
        return programs

    hmong_ids = {id(p) for p in hmong}
    kept = [p for p in programs if id(p) not in hmong_ids]

    new_blocks: list[dict] = []
    for hs in (win_s, win_s + 60):          # 18:00–19:00, then 19:00–20:00
        he = hs + 60
        covering = [p for p in hmong if _hhmm_to_min(p["start"]) < he and _hhmm_to_min(p["end"]) > hs]
        if not covering:
            continue
        covering.sort(key=lambda p: _hhmm_to_min(p["start"]))
        block = dict(covering[0])            # keep title/language/kind of the hour's first show
        block["start"], block["end"] = _min_to_hhmm(hs), _min_to_hhmm(he)
        new_blocks.append(block)

    out = kept + new_blocks
    out.sort(key=lambda p: (_hhmm_to_min(p.get("start")) if _hhmm_to_min(p.get("start")) is not None else 9999))
    return out


def _split_marketplace(programs: list[dict], network: str) -> list[dict]:
    """Override: a TAC "Marketplace" listing is a long-form paid-programming hour
    built as PRG 28:30 + COM 1:30 + PRG 28:30 + COM 1:30 — i.e. TWO independent
    long-form spots, one per half hour, each in its own PRGS segment in Etere.

    The Dallas grid lists the whole hour as one merged "Marketplace" cell, so the
    parser emits a single 60-min block and the placement tool offers one fillable
    slot expecting one program across both segments. Split each Marketplace block
    into consecutive 30-min slots so each half-hour PRGS segment can be filled
    with its own long-form piece. TAC-only (Crossings TV has no Marketplace hour).
    Idempotent: a Marketplace block already ≤30 min passes through unchanged."""
    if network != "TAC":
        return programs
    out: list[dict] = []
    for p in programs:
        s, e = _hhmm_to_min(p.get("start")), _hhmm_to_min(p.get("end"))
        if (p.get("title") or "").strip().lower() != "marketplace" or s is None or e is None:
            out.append(p)
            continue
        if e <= s:  # crosses midnight (e.g. 23:00–00:00)
            e += 24 * 60
        n = max(1, round((e - s) / 30))
        if n == 1:
            out.append(p)
            continue
        step = (e - s) / n
        for i in range(n):
            seg = dict(p)
            seg["start"] = _min_to_hhmm(int(round(s + i * step)))
            seg["end"] = _min_to_hhmm(int(round(s + (i + 1) * step)))
            out.append(seg)
    return out


def get_day_programs(network: str, d: datetime.date) -> dict:
    """Return the program lineup for one network/day from the K: grid.

    Result: {"found": bool, "file": str|None, "date": iso, "programs": [
        {start, end, title, language, kind, raw}, ...]}.
    """
    path = find_grid_file(network, d)
    if not path:
        return {"found": False, "file": None, "date": d.isoformat(), "programs": [],
                "error": f"No grid file found for {network} week of {_week_monday(d):%Y-%m-%d}"}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_grid_sheet(wb)
    if ws is None:
        return {"found": False, "file": str(path), "date": d.isoformat(), "programs": [],
                "error": f"No day-grid sheet found in {path.name} (sheets: {wb.sheetnames})"}

    # Day column: row 3 holds the per-column dates (Mon..Sun in cols 2..8).
    daycol = None
    for c in range(2, 9):
        v = ws.cell(3, c).value
        if isinstance(v, datetime.datetime) and v.date() == d:
            daycol = c
            break
    if daycol is None:
        return {"found": False, "file": str(path), "date": d.isoformat(), "programs": [],
                "error": f"{d:%Y-%m-%d} not found as a day column in {path.name}"}

    # Pre-index the merged ranges covering the day column (one pass) so we don't
    # rescan every range per row. Some sheets report a hugely inflated max_row
    # from stray formatting, so cap the scan — grids are only ~60 rows.
    range_at: dict[int, object] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= daycol <= mr.max_col and (mr.max_row - mr.min_row) <= 200:
            for rr in range(mr.min_row, mr.max_row + 1):
                range_at[rr] = mr

    last_row = min(ws.max_row, 500)
    programs: list[dict] = []
    seen: set = set()
    empties = 0
    r = 4
    while r <= last_row:
        mr = range_at.get(r)
        if mr is not None:
            key = mr.min_row
            val = ws.cell(mr.min_row, mr.min_col).value
            start, end = _block_times(ws, mr.min_row, mr.max_row + 1, mr.max_row - mr.min_row + 1)
            r = mr.max_row + 1
            if key in seen:
                continue
            seen.add(key)
        else:
            val = ws.cell(r, daycol).value
            start, end = _block_times(ws, r, r + 1, 1)
            r += 1

        if not val:
            empties += 1
            if empties >= 12:  # several blank rows running = end of the lineup
                break
            continue
        empties = 0
        raw = " ".join(str(val).split())
        if _FOOTER_RE.search(raw):
            break  # reached the channel-listing footer
        prog = _parse_title(raw)
        prog["start"], prog["end"] = start, end
        programs.extend(_split_shared_block(prog))

    programs = _split_marketplace(programs, network)
    programs = _coalesce_hmong_weekend(programs, d)
    return {"found": True, "file": str(path), "date": d.isoformat(), "programs": programs}
